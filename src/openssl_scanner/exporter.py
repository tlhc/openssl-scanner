"""
Export module for scan reports.

Supports exporting to:
- Excel (.xlsx) with multiple sheets for full data analysis
- Self-contained HTML with embedded JS/CSS
"""

import json
import os
import logging
from typing import Dict, List, Any, Optional
from importlib import resources

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Export report to Excel with comprehensive data sheets.

    Sheets:
    1. Overview     - Metadata and summary statistics
    2. Files        - All scanned files with full attributes
    3. File-Symbol  - Flat table for pivot analysis (file, symbol, category)
    4. Import Chains - Symbol import paths
    5. By Category  - Category-wise symbol statistics
    6. By Depth     - Depth-wise symbol statistics
    7. Dep Tree     - Flattened dependency tree
    8. Errors       - Scan errors and warnings
    """

    def export(self, report_path: str, output_path: str) -> None:
        """
        Export JSON report to Excel file.

        Args:
            report_path: Path to JSON report file
            output_path: Output Excel file path
        """
        from . import _vendor  # noqa: F401 - adds vendored packages to sys.path
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with open(report_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        wb = Workbook()

        self._header_font = Font(bold=True)
        self._header_fill = PatternFill(start_color="E8F4FC", end_color="E8F4FC", fill_type="solid")

        self._create_overview_sheet(wb, data)
        self._create_files_sheet(wb, data)
        self._create_file_symbol_sheet(wb, data)
        self._create_import_chains_sheet(wb, data)
        self._create_category_sheet(wb, data)
        self._create_depth_sheet(wb, data)
        self._create_dep_tree_sheet(wb, data)
        self._create_errors_sheet(wb, data)

        del wb['Sheet']

        wb.save(output_path)
        logger.info(f"Excel report saved to: {output_path}")

    def _style_header(self, ws, row: int, num_cols: int) -> None:
        """Apply header styling to a row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self._header_font
            cell.fill = self._header_fill

    def _create_overview_sheet(self, wb, data: Dict) -> None:
        """Create overview sheet with metadata and summary."""
        ws = wb.create_sheet("Overview", 0)

        meta = data.get('meta', {})
        summary = data.get('summary', {})
        report_type = meta.get('report_type', 'single')

        is_aggregated = report_type == 'aggregated'

        rows = [
            ("REPORT METADATA", ""),
            ("Report Type", report_type),
            ("Tool Version", meta.get('tool_version', '')),
            ("Scan Time", meta.get('aggregation_time', meta.get('scan_time', ''))),
            ("Scan Root", meta.get('scan_root', '')),
            ("Target Architecture", meta.get('target_arch', '')),
            ("", ""),
        ]

        if is_aggregated:
            rows.extend([
                ("AGGREGATION INFO", ""),
                ("Source Reports", meta.get('source_reports_count', 0)),
                ("Mapping File", meta.get('mapping_file', 'None')),
                ("", ""),
            ])

        rows.extend([
            ("SUMMARY STATISTICS", ""),
            ("Total Files Scanned", summary.get('total_files_scanned', summary.get('total_executables', 0))),
            ("Total ELF Files", summary.get('total_elf_files', summary.get('total_executables', 0))),
            ("Files with OpenSSL Deps", summary.get('files_with_openssl_deps', summary.get('total_components', 0))),
            ("Total OpenSSL Symbols (refs)", summary.get('total_openssl_symbols', 0)),
            ("Unique OpenSSL Symbols", summary.get('unique_openssl_symbols', summary.get('global_unique_symbols', 0))),
            ("", ""),
        ])

        openssl_libs = summary.get('openssl_libs_found', [])
        if openssl_libs:
            rows.append(("DETECTED OPENSSL LIBRARIES", ""))
            for lib in openssl_libs:
                rows.append(("", lib))

        for row_idx, (label, value) in enumerate(rows, 1):
            cell_a = ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=str(value) if value else '')
            if label and label.isupper():
                cell_a.font = self._header_font

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60

    def _create_files_sheet(self, wb, data: Dict) -> None:
        """Create files sheet with complete file information."""
        ws = wb.create_sheet("Files")

        headers = [
            "File Path", "File Name", "Type", "Arch",
            "OpenSSL Direct", "OpenSSL Transitive", "OpenSSL Libs",
            "Symbol Count", "Direct Dependencies"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header(ws, 1, len(headers))

        files_detail = data.get('files_detail', [])

        if not files_detail:
            by_file = data.get('openssl_symbols', {}).get('by_file', {})
            if by_file:
                for path, info in by_file.items():
                    symbols = info.get('symbols', []) if isinstance(info, dict) else []
                    files_detail.append({
                        'path': path,
                        'type': 'shared_library' if '.so' in path else 'executable',
                        'arch': data.get('meta', {}).get('target_arch', ''),
                        'direct_deps': [],
                        'openssl_deps': {'direct': False, 'transitive': True, 'libs': []},
                        'openssl_symbols_used': symbols,
                        'error': None
                    })

            components = data.get('components', {})
            if components and not files_detail:
                for comp_name, comp_data in components.items():
                    executables = comp_data.get('executables', [comp_name])
                    symbols = comp_data.get('unique_symbols', comp_data.get('symbols', []))
                    for exe in executables:
                        files_detail.append({
                            'path': exe,
                            'type': 'component',
                            'arch': '',
                            'direct_deps': [],
                            'openssl_deps': {'direct': False, 'transitive': True, 'libs': []},
                            'openssl_symbols_used': symbols,
                            'error': None
                        })

        row_idx = 2
        for f in files_detail:
            path = f.get('path', '')
            openssl_deps = f.get('openssl_deps', {})
            symbols = f.get('openssl_symbols_used', [])
            direct_deps = f.get('direct_deps', [])

            ws.cell(row=row_idx, column=1, value=path)
            ws.cell(row=row_idx, column=2, value=os.path.basename(path))
            ws.cell(row=row_idx, column=3, value=f.get('type', ''))
            ws.cell(row=row_idx, column=4, value=f.get('arch', ''))
            ws.cell(row=row_idx, column=5, value='Yes' if openssl_deps.get('direct') else 'No')
            ws.cell(row=row_idx, column=6, value='Yes' if openssl_deps.get('transitive') else 'No')
            ws.cell(row=row_idx, column=7, value=', '.join(openssl_deps.get('libs', [])))
            ws.cell(row=row_idx, column=8, value=len(symbols))
            ws.cell(row=row_idx, column=9, value=', '.join(direct_deps) if direct_deps else '')
            row_idx += 1

        if row_idx == 2:
            ws.cell(row=2, column=1, value="No file data available")

        col_widths = [60, 25, 15, 10, 15, 18, 40, 12, 60]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_file_symbol_sheet(self, wb, data: Dict) -> None:
        """
        Create flat file-symbol table for pivot analysis.

        Each row is one (component/file, binary, symbol) tuple - ideal for pivot tables.
        For single scans: Component = File Path, Binary = File Name
        For aggregated scans: Component = component name, Binary = executable name
        """
        ws = wb.create_sheet("File-Symbol")

        report_type = data.get('meta', {}).get('report_type', 'single')
        is_aggregated = report_type == 'aggregated'

        headers = ["Component", "Binary", "Symbol", "Category"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header(ws, 1, len(headers))

        by_category = data.get('openssl_symbols', {}).get('by_category', {})
        symbol_to_category = {}
        for cat, cat_data in by_category.items():
            if isinstance(cat_data, dict):
                for sym in cat_data.get('symbols', []):
                    symbol_to_category[sym] = cat

        components = data.get('components', {})
        if components:
            for comp_name, comp_data in components.items():
                for cat, cat_data in comp_data.get('by_category', {}).items():
                    if isinstance(cat_data, dict):
                        for sym in cat_data.get('symbols', []):
                            symbol_to_category[sym] = cat

        row_idx = 2

        by_file = data.get('openssl_symbols', {}).get('by_file', {})
        if by_file:
            for path, info in by_file.items():
                symbols = info.get('symbols', []) if isinstance(info, dict) else []
                filename = os.path.basename(path)
                for sym in sorted(symbols):
                    category = symbol_to_category.get(sym, 'other')
                    ws.cell(row=row_idx, column=1, value=path)
                    ws.cell(row=row_idx, column=2, value=filename)
                    ws.cell(row=row_idx, column=3, value=sym)
                    ws.cell(row=row_idx, column=4, value=category)
                    row_idx += 1

        files_detail = data.get('files_detail', [])
        if files_detail and row_idx == 2:
            for f in files_detail:
                path = f.get('path', '')
                filename = os.path.basename(path)
                symbols = f.get('openssl_symbols_used', [])
                for sym in sorted(symbols):
                    category = symbol_to_category.get(sym, 'other')
                    ws.cell(row=row_idx, column=1, value=path)
                    ws.cell(row=row_idx, column=2, value=filename)
                    ws.cell(row=row_idx, column=3, value=sym)
                    ws.cell(row=row_idx, column=4, value=category)
                    row_idx += 1

        if components and row_idx == 2:
            for comp_name, comp_data in components.items():
                exec_detail = comp_data.get('executables_detail', {})
                if exec_detail:
                    for bin_name, bin_data in exec_detail.items():
                        for cat, cat_data in bin_data.get('by_category', {}).items():
                            symbols = cat_data.get('symbols', []) if isinstance(cat_data, dict) else []
                            for sym in sorted(symbols):
                                ws.cell(row=row_idx, column=1, value=comp_name)
                                ws.cell(row=row_idx, column=2, value=bin_name)
                                ws.cell(row=row_idx, column=3, value=sym)
                                ws.cell(row=row_idx, column=4, value=cat)
                                row_idx += 1
                else:
                    for cat, cat_data in comp_data.get('by_category', {}).items():
                        symbols = cat_data.get('symbols', []) if isinstance(cat_data, dict) else []
                        for sym in sorted(symbols):
                            ws.cell(row=row_idx, column=1, value=comp_name)
                            ws.cell(row=row_idx, column=2, value=comp_name)
                            ws.cell(row=row_idx, column=3, value=sym)
                            ws.cell(row=row_idx, column=4, value=cat)
                            row_idx += 1

        if row_idx == 2:
            ws.cell(row=2, column=1, value="No file-symbol data available")

        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 18

    def _create_import_chains_sheet(self, wb, data: Dict) -> None:
        """
        Create import chains sheet showing dependency paths.

        Supports both formats:
        - Detailed: {symbol: [{source_file, chain, depth}, ...]}
        - Legacy: {symbol: [chain_string, ...]}
        """
        ws = wb.create_sheet("Import Chains")

        headers = ["Source File", "File Name", "Symbol", "Category", "Import Chain", "Depth"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header(ws, 1, len(headers))

        import_chains = data.get('openssl_symbols', {}).get('import_chains', {})
        by_category = data.get('openssl_symbols', {}).get('by_category', {})

        symbol_to_category = {}
        for cat, cat_data in by_category.items():
            if isinstance(cat_data, dict):
                for sym in cat_data.get('symbols', []):
                    symbol_to_category[sym] = cat

        row_idx = 2
        for symbol, chains in sorted(import_chains.items()):
            category = symbol_to_category.get(symbol, 'other')

            for chain_item in chains:
                if isinstance(chain_item, dict):
                    source_file = chain_item.get('source_file', '')
                    if not source_file:
                        component = chain_item.get('component', '')
                        binary = chain_item.get('binary', '')
                        source_file = f"{component}/{binary}" if component else binary
                    chain_str = chain_item.get('chain', '')
                    depth = chain_item.get('depth', 0)
                else:
                    source_file = ''
                    chain_str = chain_item
                    depth = chain_str.count(' -> ')

                file_name = os.path.basename(source_file) if source_file else ''

                ws.cell(row=row_idx, column=1, value=source_file)
                ws.cell(row=row_idx, column=2, value=file_name)
                ws.cell(row=row_idx, column=3, value=symbol)
                ws.cell(row=row_idx, column=4, value=category)
                ws.cell(row=row_idx, column=5, value=chain_str)
                ws.cell(row=row_idx, column=6, value=depth)
                row_idx += 1

        if row_idx == 2:
            ws.cell(row=2, column=1, value="No import chain data available")

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 80
        ws.column_dimensions['F'].width = 10

    def _create_category_sheet(self, wb, data: Dict) -> None:
        """Create category statistics sheet."""
        ws = wb.create_sheet("By Category")

        headers = ["Category", "Symbol Count", "Percentage", "Symbols"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header(ws, 1, len(headers))

        by_category = data.get('openssl_symbols', {}).get('by_category', {})

        if not by_category:
            components = data.get('components', {})
            category_symbols: Dict[str, set] = {}
            for comp_data in components.values():
                for cat, cat_data in comp_data.get('by_category', {}).items():
                    if cat not in category_symbols:
                        category_symbols[cat] = set()
                    symbols = cat_data.get('symbols', []) if isinstance(cat_data, dict) else []
                    category_symbols[cat].update(symbols)
            by_category = {
                cat: {'count': len(syms), 'symbols': list(syms)}
                for cat, syms in category_symbols.items()
            }

        total = sum(
            cat_data.get('count', len(cat_data.get('symbols', [])))
            if isinstance(cat_data, dict) else 0
            for cat_data in by_category.values()
        ) or 1

        sorted_cats = sorted(
            by_category.items(),
            key=lambda x: x[1].get('count', 0) if isinstance(x[1], dict) else 0,
            reverse=True
        )

        row_idx = 2
        for cat, cat_data in sorted_cats:
            if not isinstance(cat_data, dict):
                continue
            count = cat_data.get('count', len(cat_data.get('symbols', [])))
            symbols = cat_data.get('symbols', [])
            pct = (count / total) * 100

            ws.cell(row=row_idx, column=1, value=cat)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%")
            ws.cell(row=row_idx, column=4, value=', '.join(sorted(symbols)))
            row_idx += 1

        ws.cell(row=row_idx, column=1, value="TOTAL")
        ws.cell(row=row_idx, column=2, value=total)
        ws.cell(row=row_idx, column=3, value="100%")
        ws.cell(row=row_idx, column=1).font = self._header_font

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 100

    def _create_depth_sheet(self, wb, data: Dict) -> None:
        """Create depth statistics sheet."""
        ws = wb.create_sheet("By Depth")

        headers = ["Depth", "Description", "Symbol Count", "File Count", "Symbols"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header(ws, 1, len(headers))

        by_depth = data.get('openssl_symbols', {}).get('by_depth', {})

        if not by_depth:
            ws.cell(row=2, column=1, value="No depth data available")
            ws.column_dimensions['A'].width = 50
            return

        depth_descriptions = {
            1: "Direct OpenSSL dependency (depth 1)",
            2: "One hop to OpenSSL (depth 2)",
            3: "Two hops to OpenSSL (depth 3)",
        }

        row_idx = 2
        for depth_key in sorted(by_depth.keys()):
            depth_data = by_depth[depth_key]
            if not isinstance(depth_data, dict):
                continue

            depth_num = depth_key.replace('depth_', '') if isinstance(depth_key, str) else depth_key
            try:
                depth_int = int(depth_num)
            except (ValueError, TypeError):
                depth_int = -1

            count = depth_data.get('count', len(depth_data.get('symbols', [])))
            symbols = depth_data.get('symbols', [])
            files = depth_data.get('files', [])

            desc = depth_descriptions.get(depth_int, f"{depth_int - 1} hops to OpenSSL (depth {depth_int})")

            ws.cell(row=row_idx, column=1, value=depth_int)
            ws.cell(row=row_idx, column=2, value=desc)
            ws.cell(row=row_idx, column=3, value=count)
            ws.cell(row=row_idx, column=4, value=len(files))
            ws.cell(row=row_idx, column=5, value=', '.join(sorted(symbols)))
            row_idx += 1

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 100

    def _create_dep_tree_sheet(self, wb, data: Dict) -> None:
        """Create flattened dependency tree sheet."""
        ws = wb.create_sheet("Dep Tree")

        headers = ["Parent", "Child", "Depth", "Is OpenSSL Lib", "Symbol Count", "Full Path"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header(ws, 1, len(headers))

        dep_tree = data.get('dependency_tree', {})

        if not dep_tree:
            ws.cell(row=2, column=1, value="No dependency tree data available")
            ws.cell(row=3, column=1, value="(Dependency tree requires single-binary scan mode)")
            ws.column_dimensions['A'].width = 50
            return

        rows = []
        self._flatten_tree(dep_tree, None, 0, rows)

        row_idx = 2
        for item in rows:
            ws.cell(row=row_idx, column=1, value=item['parent'] or '(root)')
            ws.cell(row=row_idx, column=2, value=item['name'])
            ws.cell(row=row_idx, column=3, value=item['depth'])
            ws.cell(row=row_idx, column=4, value='Yes' if item['is_openssl'] else 'No')
            ws.cell(row=row_idx, column=5, value=item['symbol_count'])
            ws.cell(row=row_idx, column=6, value=item['path'])
            row_idx += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 60

    def _flatten_tree(self, node: Dict, parent: Optional[str], depth: int, rows: List[Dict]) -> None:
        """Recursively flatten dependency tree into list of rows."""
        rows.append({
            'parent': parent,
            'name': node.get('name', ''),
            'path': node.get('path', ''),
            'depth': depth,
            'is_openssl': node.get('is_openssl_lib', False),
            'symbol_count': node.get('openssl_symbols_count', 0),
        })

        for child in node.get('children', []):
            self._flatten_tree(child, node.get('name'), depth + 1, rows)

    def _create_errors_sheet(self, wb, data: Dict) -> None:
        """Create errors sheet."""
        ws = wb.create_sheet("Errors")

        headers = ["Severity", "File", "Error Message"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header(ws, 1, len(headers))

        errors = data.get('errors', [])

        row_idx = 2
        for err in errors:
            if isinstance(err, dict):
                ws.cell(row=row_idx, column=1, value=err.get('severity', 'warning').upper())
                ws.cell(row=row_idx, column=2, value=err.get('file', ''))
                ws.cell(row=row_idx, column=3, value=err.get('error', ''))
            else:
                ws.cell(row=row_idx, column=1, value='WARNING')
                ws.cell(row=row_idx, column=2, value='')
                ws.cell(row=row_idx, column=3, value=str(err))
            row_idx += 1

        if row_idx == 2:
            ws.cell(row=2, column=1, value="No errors")

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 60


def get_column_letter(col_idx):
    """Convert column index to letter (1=A, 2=B, etc.)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


class HTMLExporter:
    """Export report to self-contained HTML."""

    def export(self, report_path: str, output_path: str) -> None:
        """
        Export JSON report to self-contained HTML file.

        Args:
            report_path: Path to JSON report file
            output_path: Output HTML file path
        """
        with open(report_path, 'r', encoding='utf-8') as f:
            report_data = json.load(f)

        html_content = self._generate_html(report_data)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        logger.info(f"HTML report saved to: {output_path}")

    def _generate_html(self, data: Dict) -> str:
        """Generate self-contained HTML with embedded data and scripts."""
        css = self._get_css()
        js = self._get_js()
        chartjs = self._get_chartjs_minimal()
        sheetjs = self._get_sheetjs_minimal()

        json_data = json.dumps(data, ensure_ascii=False)

        return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>OpenSSL Dependency Analysis Report</title>
    <style>
{css}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>OpenSSL Dependency Analysis Report</h1>
            <div class="actions">
                <button onclick="exportExcel()">Export Excel</button>
                <button onclick="loadFile()">Load JSON</button>
                <input type="file" id="fileInput" accept=".json" style="display:none" onchange="handleFileSelect(event)">
            </div>
        </header>

        <div class="summary-cards" id="summaryCards"></div>

        <div class="tabs">
            <button class="tab-btn active" onclick="showTab('ranking')">Ranking</button>
            <button class="tab-btn" onclick="showTab('dependencies')">Dependencies</button>
            <button class="tab-btn" onclick="showTab('categories')">Categories</button>
            <button class="tab-btn" onclick="showTab('symbols')">Symbols</button>
        </div>

        <div id="ranking" class="tab-content active">
            <div id="processInfoBar" class="process-info-bar" style="display:none"></div>
            <div class="chart-container">
                <canvas id="rankingChart"></canvas>
            </div>
            <table id="rankingTable"></table>
        </div>

        <div id="dependencies" class="tab-content">
            <div id="depSummaryBar" class="dep-summary-bar"></div>
            <div id="depChains" class="dep-chains-section"></div>
            <div class="dep-graph-section">
                <div class="dep-graph-toolbar">
                    <span class="dep-graph-title">Dependency Graph</span>
                    <button class="dep-filter-btn active" onclick="setDepFilter('ossl',this)">OSSL Path</button>
                    <button class="dep-filter-btn" onclick="setDepFilter('shared',this)">Shared</button>
                    <button class="dep-filter-btn" onclick="setDepFilter('all',this)">All</button>
                </div>
                <div class="dep-graph-viewport" id="depGraphViewport">
                    <canvas id="depGraphCanvas" width="1200" height="500"></canvas>
                </div>
            </div>
            <table id="depStatsTable"></table>
        </div>

        <div id="categories" class="tab-content">
            <div class="chart-container">
                <canvas id="categoryChart"></canvas>
            </div>
            <table id="categoryTable"></table>
        </div>

        <div id="symbols" class="tab-content">
            <div class="search-box">
                <input type="text" id="symbolSearch" placeholder="Search symbols..." oninput="filterSymbols()">
                <select id="categoryFilter" onchange="filterSymbols()">
                    <option value="">All Categories</option>
                </select>
                <select id="componentFilter" onchange="filterSymbols()">
                    <option value="">All Components</option>
                </select>
            </div>
            <table id="symbolsTable"></table>
        </div>
    </div>

    <!-- Component Detail Modal -->
    <div id="componentModal" class="modal">
        <div class="modal-content">
            <div class="modal-header">
                <h2 id="modalTitle">Component Details</h2>
                <button class="modal-close" onclick="closeModal()">&times;</button>
            </div>
            <div class="modal-summary" id="modalSummary"></div>
            <div class="modal-body" id="modalBody"></div>
        </div>
    </div>

    <script>
{chartjs}
    </script>
    <script>
{sheetjs}
    </script>
    <script>
const REPORT_DATA = {json_data};
{js}
document.addEventListener('DOMContentLoaded', function() {{
    try {{
        renderReport(REPORT_DATA);
    }} catch(e) {{
        console.error('Render error:', e);
        document.body.innerHTML += '<div style="color:red;padding:20px;">Error: ' + e.message + '</div>';
    }}
}});
    </script>
</body>
</html>'''

    def _get_css(self) -> str:
        """Get embedded CSS."""
        return '''
* {
    margin: 0;
    padding: 0;
    box-sizing: border-box;
}

body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    background: #f5f5f5;
    color: #333;
    line-height: 1.6;
}

.container {
    max-width: 1400px;
    margin: 0 auto;
    padding: 20px;
}

header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 20px;
    padding: 20px;
    background: #fff;
    border-radius: 8px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}

header h1 {
    font-size: 1.5rem;
    color: #2c3e50;
}

.actions button {
    padding: 8px 16px;
    margin-left: 10px;
    border: none;
    border-radius: 4px;
    background: #3498db;
    color: #fff;
    cursor: pointer;
    font-size: 14px;
}

.actions button:hover {
    background: #2980b9;
}

.summary-cards {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 15px;
    margin-bottom: 20px;
}

.card {
    background: #fff;
    padding: 20px;
    border-radius: 8px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    text-align: center;
}

.card .value {
    font-size: 2rem;
    font-weight: bold;
    color: #3498db;
}

.card .label {
    color: #7f8c8d;
    font-size: 0.9rem;
}

.tabs {
    display: flex;
    gap: 5px;
    margin-bottom: 20px;
}

.tab-btn {
    padding: 10px 20px;
    border: none;
    background: #fff;
    cursor: pointer;
    border-radius: 4px 4px 0 0;
    font-size: 14px;
}

.tab-btn.active {
    background: #3498db;
    color: #fff;
}

.tab-content {
    display: none;
    background: #fff;
    padding: 20px;
    border-radius: 0 8px 8px 8px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}

.tab-content.active {
    display: block;
}

.chart-container {
    height: 300px;
    margin-bottom: 20px;
}

table {
    width: 100%;
    border-collapse: collapse;
}

th, td {
    padding: 12px;
    text-align: left;
    border-bottom: 1px solid #eee;
}

th {
    background: #f8f9fa;
    font-weight: 600;
    color: #2c3e50;
}

tr:hover {
    background: #f8f9fa;
}

.search-box {
    display: flex;
    gap: 10px;
    margin-bottom: 15px;
}

.search-box input,
.search-box select {
    padding: 8px 12px;
    border: 1px solid #ddd;
    border-radius: 4px;
    font-size: 14px;
}

.search-box input {
    flex: 1;
}

.bar {
    height: 20px;
    background: #3498db;
    border-radius: 3px;
    transition: width 0.3s;
}

.bar-container {
    background: #ecf0f1;
    border-radius: 3px;
    overflow: hidden;
}

/* Clickable component links */
.component-link {
    color: #3498db;
    cursor: pointer;
    text-decoration: none;
}
.component-link:hover {
    text-decoration: underline;
}

/* Hierarchical ranking table */
.ranking-exe-row td {
    background: #eaf2f8;
    font-weight: 600;
    border-top: 2px solid #3498db;
}
.ranking-exe-row td:first-child {
    color: #2c3e50;
}
.ranking-lib-row td {
    padding-left: 16px;
    color: #555;
    font-size: 0.93em;
}
.ranking-lib-row td:first-child {
    border-left: 3px solid #dce6f0;
}
.sym-badge {
    display: inline-block;
    background: #3498db;
    color: #fff;
    padding: 2px 8px;
    border-radius: 10px;
    font-size: 0.82em;
    margin-left: 6px;
    font-weight: normal;
}
.sym-badge.zero {
    background: #bdc3c7;
}
.openssl-marker {
    display: inline-block;
    background: #e74c3c;
    color: #fff;
    padding: 1px 6px;
    border-radius: 3px;
    font-size: 0.78em;
    margin-left: 6px;
    font-weight: 600;
}
.circular-tag {
    color: #95a5a6;
    font-style: italic;
    font-size: 0.85em;
}
.cat-dist {
    display: flex;
    flex-wrap: wrap;
    gap: 4px;
    line-height: 1.4;
}
.cat-tag {
    display: inline-block;
    padding: 1px 6px;
    border-radius: 3px;
    font-size: 0.8em;
    white-space: nowrap;
    color: #fff;
}
.cat-tag .cat-count {
    font-weight: 600;
}
.dep-summary-bar {
    display: flex;
    gap: 24px;
    padding: 12px 16px;
    background: #f0f4f8;
    border: 1px solid #d6e0eb;
    border-radius: 6px;
    margin-bottom: 16px;
    font-size: 0.9em;
}
.dep-summary-bar .dep-stat {
    display: flex;
    align-items: baseline;
    gap: 6px;
}
.dep-summary-bar .dep-stat-val {
    font-weight: 700;
    color: #2c3e50;
    font-size: 1.2em;
}
.dep-summary-bar .dep-stat-label {
    color: #7f8c8d;
}
.dep-chains-section {
    margin-bottom: 20px;
}
.dep-chains-title {
    font-weight: 600;
    font-size: 1.05em;
    margin-bottom: 10px;
    color: #2c3e50;
}
.dep-chain-row {
    display: flex;
    align-items: center;
    flex-wrap: wrap;
    gap: 4px;
    padding: 8px 12px;
    margin-bottom: 6px;
    background: #fafbfc;
    border: 1px solid #e8ecf0;
    border-radius: 6px;
}
.dep-chain-node {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 4px;
    font-size: 0.88em;
    font-weight: 500;
    cursor: pointer;
    transition: box-shadow 0.15s;
}
.dep-chain-node:hover {
    box-shadow: 0 0 0 2px rgba(52, 152, 219, 0.4);
}
.dep-chain-node.root {
    background: #eaf2f8;
    color: #2c3e50;
    border: 1px solid #b8d4e8;
}
.dep-chain-node.middle {
    background: #f5f5f5;
    color: #555;
    border: 1px solid #ddd;
}
.dep-chain-node.target {
    background: #fdebd0;
    color: #8a5a00;
    border: 1px solid #f0c27a;
    font-weight: 600;
}
.dep-chain-arrow {
    color: #95a5a6;
    font-size: 1.1em;
    margin: 0 2px;
}
.dep-chain-sym {
    display: flex;
    flex-wrap: wrap;
    gap: 3px;
    align-items: center;
    flex-basis: 100%;
    padding-left: 8px;
    margin-top: 2px;
}
.dep-graph-section {
    margin-bottom: 20px;
    border: 1px solid #e1e4e8;
    border-radius: 6px;
    overflow: hidden;
}
.dep-graph-viewport {
    overflow: auto;
    max-height: 600px;
    cursor: grab;
    position: relative;
}
.dep-graph-viewport.dragging {
    cursor: grabbing;
    user-select: none;
}
.dep-graph-toolbar {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 8px 12px;
    background: #f6f8fa;
    border-bottom: 1px solid #e1e4e8;
}
.dep-graph-title {
    font-weight: 600;
    font-size: 0.95em;
    color: #2c3e50;
    margin-right: auto;
}
.dep-filter-btn {
    padding: 4px 12px;
    border: 1px solid #d1d5da;
    border-radius: 4px;
    background: #fff;
    font-size: 0.82em;
    cursor: pointer;
    color: #555;
}
.dep-filter-btn.active {
    background: #3498db;
    color: #fff;
    border-color: #3498db;
}
#depGraphCanvas {
    width: 100%;
    display: block;
    background: #fff;
}
.process-info-bar {
    display: flex;
    gap: 20px;
    padding: 10px 16px;
    background: #fafbfc;
    border: 1px solid #e1e4e8;
    border-radius: 6px;
    margin-bottom: 16px;
    font-size: 0.9em;
    color: #555;
    flex-wrap: wrap;
}
.process-info-bar .pi-item {
    white-space: nowrap;
}
.process-info-bar .pi-label {
    font-weight: 600;
    color: #2c3e50;
    margin-right: 4px;
}

/* Modal styles */
.modal {
    display: none;
    position: fixed;
    top: 0;
    left: 0;
    width: 100%;
    height: 100%;
    background: rgba(0,0,0,0.5);
    z-index: 10000;
    overflow: auto;
}
.modal.active {
    display: flex;
    justify-content: center;
    align-items: flex-start;
    padding: 40px 20px;
}
.modal-content {
    background: #fff;
    border-radius: 8px;
    width: 100%;
    max-width: 900px;
    max-height: 85vh;
    overflow: hidden;
    display: flex;
    flex-direction: column;
    box-shadow: 0 4px 20px rgba(0,0,0,0.2);
}
.modal-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 15px 20px;
    border-bottom: 1px solid #eee;
    background: #f8f9fa;
}
.modal-header h2 {
    font-size: 1.3rem;
    color: #2c3e50;
}
.modal-close {
    background: none;
    border: none;
    font-size: 28px;
    cursor: pointer;
    color: #7f8c8d;
    line-height: 1;
}
.modal-close:hover {
    color: #e74c3c;
}
.modal-summary {
    padding: 15px 20px;
    background: #fff;
    border-bottom: 1px solid #eee;
    display: flex;
    gap: 20px;
    flex-wrap: wrap;
}
.modal-summary .stat {
    text-align: center;
    padding: 10px 15px;
    background: #f8f9fa;
    border-radius: 6px;
}
.modal-summary .stat .value {
    font-size: 1.5rem;
    font-weight: bold;
    color: #3498db;
}
.modal-summary .stat .label {
    font-size: 0.8rem;
    color: #7f8c8d;
}
.modal-body {
    padding: 20px;
    overflow-y: auto;
    flex: 1;
}
.category-section {
    margin-bottom: 20px;
}
.category-section h3 {
    font-size: 1rem;
    color: #2c3e50;
    padding: 8px 12px;
    background: #e8f4fc;
    border-radius: 4px;
    margin-bottom: 10px;
    display: flex;
    justify-content: space-between;
}
.category-section h3 .count {
    color: #3498db;
    font-weight: normal;
}
.symbol-list {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    padding-left: 12px;
}
.symbol-tag {
    display: inline-block;
    padding: 4px 10px;
    background: #f5f5f5;
    border: 1px solid #ddd;
    border-radius: 4px;
    font-family: monospace;
    font-size: 12px;
    color: #333;
}
.symbol-tag:hover {
    background: #e8f4fc;
    border-color: #3498db;
}

/* Binary-level hierarchy styles */
.binary-section {
    margin-bottom: 15px;
    border: 1px solid #e0e0e0;
    border-radius: 6px;
    overflow: hidden;
}
.binary-header {
    display: flex;
    align-items: center;
    padding: 12px 15px;
    background: #f8f9fa;
    cursor: pointer;
    user-select: none;
    border-bottom: 1px solid #e0e0e0;
}
.binary-header:hover {
    background: #e8f4fc;
}
.binary-header.expanded {
    background: #e8f4fc;
    border-bottom-color: #3498db;
}
.binary-toggle {
    margin-right: 10px;
    color: #7f8c8d;
    font-size: 10px;
    transition: transform 0.2s;
}
.binary-name {
    font-weight: 600;
    color: #2c3e50;
    flex: 1;
}
.binary-stats {
    color: #7f8c8d;
    font-size: 0.85rem;
}
.binary-content {
    padding: 15px;
    background: #fff;
}
.binary-content .category-section {
    margin-bottom: 15px;
}
.binary-content .category-section:last-child {
    margin-bottom: 0;
}

/* Sortable table headers */
th.sortable {
    cursor: pointer;
    user-select: none;
    position: relative;
}
th.sortable:hover {
    background: #e8f4fc;
}
.sort-icon {
    margin-left: 5px;
    color: #bdc3c7;
    font-size: 12px;
}
.sort-icon.active {
    color: #3498db;
}
'''

    def _get_js(self) -> str:
        """Get embedded JavaScript."""
        return '''
let currentData = null;
let rankingChart = null;
let categoryChart = null;

const CATEGORY_COLORS = {
    'ssl_core': '#e74c3c', 'ssl_tls': '#c0392b',
    'crypto_evp': '#3498db', 'crypto_rsa': '#2ecc71',
    'crypto_ec': '#9b59b6', 'crypto_bn': '#f39c12',
    'crypto_dsa': '#1abc9c', 'crypto_dh': '#e67e22',
    'crypto_bio': '#34495e', 'crypto_rand': '#16a085',
    'crypto_aes': '#d35400', 'crypto_sm': '#8e44ad',
    'crypto_x509': '#27ae60', 'crypto_pem': '#2980b9',
    'crypto_err': '#7f8c8d', 'crypto_engine': '#5dade2',
    'crypto_hash': '#2c3e50', 'crypto_hmac': '#1a5276',
    'crypto_des': '#a93226', 'crypto_chacha': '#6c3483',
    'crypto_pkcs': '#117a65', 'crypto_cms': '#0e6655',
    'crypto_ocsp': '#784212', 'crypto_ts': '#7d6608',
    'crypto_obj': '#5b2c6f', 'crypto_kdf': '#154360',
    'crypto_provider': '#0b5345', 'crypto_asn1': '#7b7d7d',
    'openssl_util': '#95a5a6', 'other': '#bdc3c7'
};

function buildCategoryBreakdown(symbols, byCategory) {
    const cats = {};
    (symbols || []).forEach(sym => {
        for (const [cat, catInfo] of Object.entries(byCategory)) {
            if (catInfo.symbols && catInfo.symbols.includes(sym)) {
                cats[cat] = (cats[cat] || 0) + 1;
                break;
            }
        }
    });
    return cats;
}

function buildHierarchicalRanking(data) {
    /*
     * Returns array of groups: [{exe: {...}, libs: [{...}, ...]}, ...]
     * Each group = one ELF executable + its direct dependency libraries.
     * Libs that have OpenSSL symbols from deeper in the tree are surfaced.
     * Supports: single dependency_tree or dependency_trees[] (multi-process).
     * Groups are sorted by groupTotal (exe + lib syms) descending.
     */
    const trees = data.dependency_trees || (data.dependency_tree ? [data.dependency_tree] : []);
    if (trees.length === 0) return null;

    const byFile = data.openssl_symbols?.by_file || {};
    const byCategory = data.openssl_symbols?.by_category || {};

    /* Match by_file path to tree node path using basename prefix.
     * Handles: /usr/lib/.../libssh.so.4.9.6 vs /lib/.../libssh.so.4 */
    function matchByFile(nodePath) {
        if (byFile[nodePath]) return byFile[nodePath];
        const nodeBase = (nodePath || '').split('/').pop();
        if (!nodeBase) return null;
        for (const [filePath, info] of Object.entries(byFile)) {
            const fileBase = filePath.split('/').pop();
            if (fileBase.startsWith(nodeBase) || nodeBase.startsWith(fileBase.split('.so')[0] + '.so')) {
                return info;
            }
        }
        return null;
    }

    /* Walk tree recursively to find all nodes with OpenSSL symbols */
    function collectSymLibs(node, depth, seen) {
        const results = [];
        for (const child of (node.children || [])) {
            if (child.error === 'circular dependency') continue;
            const childKey = child.path || child.name;
            if (seen.has(childKey)) continue;
            seen.add(childKey);
            const fileInfo = matchByFile(child.path);
            if (fileInfo && fileInfo.count > 0) {
                results.push({ node: child, fileInfo, depth });
            }
            results.push(...collectSymLibs(child, depth + 1, seen));
        }
        return results;
    }

    function buildGroup(node) {
        const exePath = node.path || '';
        const exeName = node.name || exePath.split('/').pop();
        const exeFileInfo = matchByFile(exePath) || {};
        const exeSymCount = exeFileInfo.count || node.openssl_symbols_count || 0;
        const exeSymbols = exeFileInfo.symbols || [];
        const directChildren = (node.children || []);
        const nonCircularCount = directChildren.filter(c => c.error !== 'circular dependency').length;

        /* Build lib rows from direct children */
        const seen = new Set();
        seen.add(exePath || exeName);
        let groupTotal = exeSymCount;

        const libs = directChildren.map(child => {
            const childKey = child.path || child.name;
            seen.add(childKey);
            const childFileInfo = matchByFile(child.path) || {};
            const childSymCount = childFileInfo.count || child.openssl_symbols_count || 0;
            const childSymbols = childFileInfo.symbols || [];
            const childDeps = (child.children || []).filter(c => c.error !== 'circular dependency').length;
            groupTotal += childSymCount;
            return {
                lib: child.name,
                libPath: child.path || '',
                syms: childSymCount,
                categories: buildCategoryBreakdown(childSymbols, byCategory),
                isOpenSSL: child.is_openssl_lib || false,
                isCircular: child.error === 'circular dependency',
                childCount: childDeps,
                depth: 1
            };
        });

        /* Surface transitive deps that have OpenSSL symbols */
        const deepLibs = collectSymLibs(node, 1, new Set([exePath || exeName]));
        for (const dl of deepLibs) {
            const childKey = dl.node.path || dl.node.name;
            if (seen.has(childKey)) continue;
            seen.add(childKey);
            const childSymbols = dl.fileInfo.symbols || [];
            groupTotal += dl.fileInfo.count || 0;
            libs.push({
                lib: dl.node.name,
                libPath: dl.node.path || '',
                syms: dl.fileInfo.count || 0,
                categories: buildCategoryBreakdown(childSymbols, byCategory),
                isOpenSSL: dl.node.is_openssl_lib || false,
                isCircular: false,
                childCount: 0,
                depth: dl.depth
            });
        }

        const exe = {
            elf: exeName,
            elfPath: exePath,
            libCount: nonCircularCount,
            syms: exeSymCount,
            groupTotal: groupTotal,
            categories: buildCategoryBreakdown(exeSymbols, byCategory),
            isOpenSSL: node.is_openssl_lib || false
        };

        return { exe, libs };
    }

    const groups = trees.map(t => buildGroup(t));
    /* Default sort: groups by groupTotal (exe + lib syms) descending */
    groups.sort((a, b) => b.exe.groupTotal - a.exe.groupTotal);
    return groups;
}

function buildGlobalGraph(data) {
    const trees = data.dependency_trees || (data.dependency_tree ? [data.dependency_tree] : []);
    if (trees.length === 0) return null;

    const byFile = data.openssl_symbols?.by_file || {};
    const byCategory = data.openssl_symbols?.by_category || {};
    const nodes = {};
    const edgeMap = {};
    const chains = [];

    function baseName(p) { return (p || '').split('/').pop(); }

    function matchByFile(nodePath) {
        if (byFile[nodePath]) return byFile[nodePath];
        const nb = baseName(nodePath);
        if (!nb) return null;
        for (const [fp, info] of Object.entries(byFile)) {
            const fb = baseName(fp);
            if (fb.startsWith(nb) || nb.startsWith(fb.split('.so')[0] + '.so'))
                return info;
        }
        return null;
    }

    function getOrCreateNode(node) {
        const b = baseName(node.path || node.name);
        if (!nodes[b]) {
            const fi = matchByFile(node.path || node.name);
            nodes[b] = {
                name: b,
                paths: new Set(),
                processes: new Set(),
                osslSym: fi ? fi.count : 0,
                isOsslLib: node.is_openssl_lib || false,
                isRoot: false,
                depths: new Set(),
                fanIn: 0,
                fanOut: 0,
                tier: 'normal'
            };
        }
        nodes[b].paths.add(node.path || node.name);
        if (node.is_openssl_lib) nodes[b].isOsslLib = true;
        return b;
    }

    function walk(node, parentBase, procName, depth) {
        const b = getOrCreateNode(node);
        nodes[b].processes.add(procName);
        nodes[b].depths.add(depth);
        if (parentBase) {
            const ek = parentBase + '|' + b;
            if (!edgeMap[ek]) edgeMap[ek] = { from: parentBase, to: b, processes: new Set() };
            edgeMap[ek].processes.add(procName);
        }
        for (const ch of (node.children || [])) {
            if (ch.error === 'circular dependency') continue;
            walk(ch, b, procName, depth + 1);
        }
    }

    trees.forEach(t => {
        const rb = getOrCreateNode(t);
        nodes[rb].isRoot = true;
        walk(t, null, t.name, 0);
    });

    /* Compute fan-in / fan-out */
    const edges = Object.values(edgeMap);
    edges.forEach(e => {
        if (nodes[e.from]) nodes[e.from].fanOut++;
        if (nodes[e.to]) nodes[e.to].fanIn++;
    });

    /* Assign tiers */
    const systemLibs = new Set(['libc.so.6', 'ld-linux-aarch64.so.1', 'libm.so.6',
        'libpthread.so.0', 'libdl.so.2', 'librt.so.1']);
    for (const [b, n] of Object.entries(nodes)) {
        if (n.isRoot) n.tier = 'application';
        else if (n.isOsslLib) n.tier = 'crypto';
        else if (n.osslSym > 0) n.tier = 'middleware';
        else if (systemLibs.has(b)) n.tier = 'system';
    }

    /* Trace OpenSSL import chains (DFS from each root) */
    function traceChains(node, path, procName) {
        const b = baseName(node.path || node.name);
        const fi = matchByFile(node.path || node.name);
        const curPath = [...path, b];
        if (fi && fi.count > 0 && path.length > 0) {
            const cats = buildCategoryBreakdown(fi.symbols || [], byCategory);
            chains.push({ path: curPath, symbols: fi.count, categories: cats, process: procName });
        }
        for (const ch of (node.children || [])) {
            if (ch.error === 'circular dependency') continue;
            traceChains(ch, curPath, procName);
        }
    }
    trees.forEach(t => {
        const rb = baseName(t.path || t.name);
        const fi = matchByFile(t.path || t.name);
        if (fi && fi.count > 0) {
            const cats = buildCategoryBreakdown(fi.symbols || [], byCategory);
            chains.push({ path: [rb], symbols: fi.count, categories: cats, process: t.name });
        }
        for (const ch of (t.children || [])) {
            if (ch.error === 'circular dependency') continue;
            traceChains(ch, [rb], t.name);
        }
    });

    /* Deduplicate chains with same path (from different processes) */
    const chainMap = {};
    chains.forEach(c => {
        const key = c.path.join('|');
        if (!chainMap[key]) {
            chainMap[key] = { ...c, processes: [c.process] };
        } else {
            if (!chainMap[key].processes.includes(c.process))
                chainMap[key].processes.push(c.process);
        }
    });

    /* Convert sets to arrays for serialization */
    const nodeList = Object.values(nodes).map(n => ({
        ...n,
        paths: [...n.paths],
        processes: [...n.processes],
        depths: [...n.depths].sort((a, b) => a - b)
    }));

    return {
        nodes: nodeList,
        nodeMap: Object.fromEntries(nodeList.map(n => [n.name, n])),
        edges: edges.map(e => ({ ...e, processes: [...e.processes] })),
        chains: Object.values(chainMap).sort((a, b) => b.symbols - a.symbols),
        multiProc: trees.length > 1
    };
}

function renderDependencies(data) {
    const graph = data.globalGraph;
    if (!graph) {
        const depBtn = document.querySelector('button[onclick*="dependencies"]');
        if (depBtn) depBtn.style.display = 'none';
        return;
    }
    window.depGraph = graph;
    window.depFilter = 'ossl';

    /* Summary bar */
    const bar = document.getElementById('depSummaryBar');
    const shared = graph.nodes.filter(n => n.processes.length > 1).length;
    bar.innerHTML =
        '<div class="dep-stat"><span class="dep-stat-val">' + graph.nodes.length + '</span><span class="dep-stat-label">Libraries</span></div>' +
        '<div class="dep-stat"><span class="dep-stat-val">' + graph.edges.length + '</span><span class="dep-stat-label">Edges</span></div>' +
        (graph.multiProc ? '<div class="dep-stat"><span class="dep-stat-val">' + shared + '</span><span class="dep-stat-label">Shared</span></div>' : '') +
        '<div class="dep-stat"><span class="dep-stat-val">' + graph.chains.length + '</span><span class="dep-stat-label">OSSL Chains</span></div>';

    /* Import chains */
    renderDepChains(graph);

    /* Graph */
    renderDepGraph(graph, 'ossl');
    initDepGraphDrag();

    /* Statistics table */
    renderDepStatsTable(graph);
}

function renderDepChains(graph) {
    const container = document.getElementById('depChains');
    if (graph.chains.length === 0) {
        container.innerHTML = '<p style="color:#95a5a6">No OpenSSL import chains found.</p>';
        return;
    }
    let html = '<div class="dep-chains-title">OpenSSL Import Chains</div>';
    graph.chains.forEach(chain => {
        html += '<div class="dep-chain-row">';
        chain.path.forEach((node, i) => {
            if (i > 0) html += '<span class="dep-chain-arrow">&rarr;</span>';
            const isLast = i === chain.path.length - 1;
            const isFirst = i === 0;
            const cls = isFirst ? 'root' : (isLast ? 'target' : 'middle');
            const escaped = node.replace(/'/g, "\\\\'");
            html += '<span class="dep-chain-node ' + cls + '" onclick="showComponentDetail(\\'' + escaped + '\\')">' + node + '</span>';
        });
        html += '<span class="dep-chain-sym">';
        html += '<span class="sym-badge">' + chain.symbols + ' sym</span> ';
        html += renderCategoryDist(chain.categories);
        html += '</span>';
        if (graph.multiProc && chain.processes) {
            html += '<span style="color:#95a5a6;font-size:0.8em;margin-left:8px">(' + chain.processes.join(', ') + ')</span>';
        }
        html += '</div>';
    });
    container.innerHTML = html;
}

/* Layered DAG on Canvas */
function renderDepGraph(graph, filter) {
    const canvas = document.getElementById('depGraphCanvas');
    const ctx = canvas.getContext('2d');
    const dpr = window.devicePixelRatio || 1;

    /* Filter nodes */
    let visibleNodes, visibleEdges;
    if (filter === 'ossl') {
        const onPath = new Set();
        graph.chains.forEach(c => c.path.forEach(n => onPath.add(n)));
        visibleNodes = graph.nodes.filter(n => onPath.has(n.name));
        visibleEdges = graph.edges.filter(e => onPath.has(e.from) && onPath.has(e.to));
    } else if (filter === 'shared') {
        visibleNodes = graph.nodes.filter(n => n.processes.length > 1);
        const vset = new Set(visibleNodes.map(n => n.name));
        visibleEdges = graph.edges.filter(e => vset.has(e.from) && vset.has(e.to));
    } else {
        visibleNodes = [...graph.nodes];
        visibleEdges = [...graph.edges];
    }

    if (visibleNodes.length === 0) {
        const emptyW = canvas.offsetWidth || 800;
        canvas.style.width = emptyW + 'px';
        canvas.style.height = '60px';
        canvas.width = emptyW * dpr;
        canvas.height = 60 * dpr;
        ctx.scale(dpr, dpr);
        ctx.fillStyle = '#95a5a6';
        ctx.font = '14px -apple-system, sans-serif';
        ctx.fillText('No nodes to display for this filter.', 20, 35);
        return;
    }

    /* Topological layering */
    const nameMap = {};
    visibleNodes.forEach(n => nameMap[n.name] = n);
    const inDeg = {};
    visibleNodes.forEach(n => inDeg[n.name] = 0);
    visibleEdges.forEach(e => { if (inDeg[e.to] !== undefined) inDeg[e.to]++; });

    const layers = [];
    const assigned = {};
    const queue = visibleNodes.filter(n => (inDeg[n.name] || 0) === 0).map(n => n.name);
    if (queue.length === 0) queue.push(visibleNodes[0].name);
    queue.forEach(n => assigned[n] = 0);

    const adj = {};
    visibleEdges.forEach(e => {
        if (!adj[e.from]) adj[e.from] = [];
        adj[e.from].push(e.to);
    });

    let qi = 0;
    while (qi < queue.length) {
        const cur = queue[qi++];
        const layer = assigned[cur];
        (adj[cur] || []).forEach(child => {
            if (assigned[child] === undefined) {
                assigned[child] = layer + 1;
                queue.push(child);
            } else {
                assigned[child] = Math.max(assigned[child], layer + 1);
            }
        });
    }
    /* Assign unvisited */
    visibleNodes.forEach(n => { if (assigned[n.name] === undefined) assigned[n.name] = 0; });

    const maxLayer = Math.max(...Object.values(assigned), 0);
    for (let i = 0; i <= maxLayer; i++) layers[i] = [];
    visibleNodes.forEach(n => layers[assigned[n.name]].push(n));

    /* Barycenter ordering within layers */
    for (let i = 1; i <= maxLayer; i++) {
        layers[i].forEach(n => {
            const parents = visibleEdges.filter(e => e.to === n.name).map(e => e.from);
            if (parents.length > 0) {
                const prevLayer = layers[i - 1];
                const positions = parents.map(p => prevLayer.findIndex(x => x.name === p)).filter(x => x >= 0);
                n._bary = positions.length > 0 ? positions.reduce((a, b) => a + b, 0) / positions.length : 0;
            } else {
                n._bary = 0;
            }
        });
        layers[i].sort((a, b) => a._bary - b._bary);
    }

    /* Compute positions */
    const nodeW = 130;
    const nodeH = 32;
    const layerGap = 70;
    const nodeGap = 16;
    const padX = 30;
    const padY = 40;

    const maxNodesInLayer = Math.max(...layers.map(l => l.length), 1);
    const vp = document.getElementById('depGraphViewport');
    const containerW = (vp ? vp.clientWidth : canvas.parentElement.clientWidth) || 800;
    const canvasW = Math.max(maxNodesInLayer * (nodeW + nodeGap) + padX * 2, containerW);
    const canvasH = (maxLayer + 1) * (nodeH + layerGap) + padY * 2;

    canvas.style.width = canvasW + 'px';
    canvas.style.height = canvasH + 'px';
    canvas.width = canvasW * dpr;
    canvas.height = canvasH * dpr;
    ctx.setTransform(dpr, 0, 0, dpr, 0, 0);

    const pos = {};
    layers.forEach((layer, li) => {
        const totalW = layer.length * nodeW + (layer.length - 1) * nodeGap;
        const startX = (canvasW - totalW) / 2;
        layer.forEach((n, ni) => {
            pos[n.name] = {
                x: startX + ni * (nodeW + nodeGap) + nodeW / 2,
                y: padY + li * (nodeH + layerGap) + nodeH / 2
            };
        });
    });

    /* Build OSSL path edge set for highlighting */
    const osslPathEdges = new Set();
    const osslPathNodes = new Set();
    graph.chains.forEach(c => {
        c.path.forEach(n => osslPathNodes.add(n));
        for (let i = 0; i < c.path.length - 1; i++) {
            osslPathEdges.add(c.path[i] + '>' + c.path[i + 1]);
        }
    });

    /* Draw */
    ctx.clearRect(0, 0, canvasW, canvasH);

    /* Edges: normal first, then OSSL path on top */
    function drawEdge(e, color, width) {
        const from = pos[e.from];
        const to = pos[e.to];
        if (!from || !to) return;
        ctx.beginPath();
        ctx.moveTo(from.x, from.y + nodeH / 2);
        const midY = (from.y + to.y) / 2;
        ctx.bezierCurveTo(from.x, midY, to.x, midY, to.x, to.y - nodeH / 2);
        ctx.strokeStyle = color;
        ctx.lineWidth = width;
        ctx.stroke();
        const ax = to.x;
        const ay = to.y - nodeH / 2;
        ctx.beginPath();
        ctx.moveTo(ax - 4, ay - 6);
        ctx.lineTo(ax, ay);
        ctx.lineTo(ax + 4, ay - 6);
        ctx.fillStyle = color;
        ctx.fill();
    }

    const normalEdges = [];
    const highlightEdges = [];
    visibleEdges.forEach(e => {
        if (osslPathEdges.has(e.from + '>' + e.to)) {
            highlightEdges.push(e);
        } else {
            normalEdges.push(e);
        }
    });

    normalEdges.forEach(e => {
        const multi = e.processes && e.processes.length > 1;
        drawEdge(e, multi ? '#c8dff0' : '#e0e4e8', multi ? 1.5 : 1);
    });
    highlightEdges.forEach(e => {
        drawEdge(e, '#e67e22', 2.5);
    });

    /* Nodes */
    const tierColors = {
        application: { bg: '#eaf2f8', border: '#3498db', text: '#2c3e50' },
        crypto: { bg: '#fdedec', border: '#e74c3c', text: '#922b21' },
        middleware: { bg: '#fef9e7', border: '#f39c12', text: '#7d6608' },
        system: { bg: '#f2f3f4', border: '#bdc3c7', text: '#7f8c8d' },
        normal: { bg: '#fff', border: '#d5d8dc', text: '#555' }
    };

    visibleNodes.forEach(n => {
        const p = pos[n.name];
        if (!p) return;
        const tc = tierColors[n.tier] || tierColors.normal;
        const onPath = osslPathNodes.has(n.name);
        const hw = nodeW / 2;
        const hh = nodeH / 2;

        /* Glow for OSSL path nodes */
        if (onPath && filter !== 'ossl') {
            ctx.save();
            ctx.shadowColor = 'rgba(230, 126, 34, 0.5)';
            ctx.shadowBlur = 8;
            ctx.beginPath();
            ctx.roundRect(p.x - hw, p.y - hh, nodeW, nodeH, 5);
            ctx.fillStyle = tc.bg;
            ctx.fill();
            ctx.restore();
        }

        ctx.beginPath();
        ctx.roundRect(p.x - hw, p.y - hh, nodeW, nodeH, 5);
        ctx.fillStyle = tc.bg;
        ctx.fill();
        ctx.strokeStyle = onPath && filter !== 'ossl' ? '#e67e22' : tc.border;
        ctx.lineWidth = n.osslSym > 0 ? 2.5 : (onPath ? 2 : 1.5);
        ctx.stroke();

        /* Label */
        ctx.fillStyle = tc.text;
        ctx.font = (n.osslSym > 0 ? 'bold ' : '') + '11px -apple-system, BlinkMacSystemFont, sans-serif';
        ctx.textAlign = 'center';
        ctx.textBaseline = 'middle';
        let label = n.name;
        if (label.length > 18) label = label.substring(0, 16) + '..';
        ctx.fillText(label, p.x, p.y);

        /* Symbol count badge */
        if (n.osslSym > 0) {
            const badgeX = p.x + hw - 4;
            const badgeY = p.y - hh - 4;
            const badgeText = String(n.osslSym);
            ctx.font = 'bold 9px -apple-system, sans-serif';
            const bw = ctx.measureText(badgeText).width + 8;
            ctx.beginPath();
            ctx.roundRect(badgeX - bw, badgeY - 6, bw, 13, 3);
            ctx.fillStyle = '#e74c3c';
            ctx.fill();
            ctx.fillStyle = '#fff';
            ctx.textAlign = 'center';
            ctx.fillText(badgeText, badgeX - bw / 2, badgeY + 1);
        }

        /* Multi-process indicator */
        if (graph.multiProc && n.processes.length > 1) {
            ctx.font = '9px -apple-system, sans-serif';
            ctx.fillStyle = '#95a5a6';
            ctx.textAlign = 'center';
            ctx.fillText(n.processes.length + ' proc', p.x, p.y + hh + 12);
        }
    });

    /* Center viewport on content */
    if (vp && canvasW > vp.clientWidth) {
        vp.scrollLeft = (canvasW - vp.clientWidth) / 2;
    }
    if (vp) vp.scrollTop = 0;
}

function initDepGraphDrag() {
    const vp = document.getElementById('depGraphViewport');
    if (!vp || vp._dragInit) return;
    vp._dragInit = true;
    let dragging = false, sx, sy, sl, st;
    vp.addEventListener('mousedown', function(e) {
        dragging = true;
        vp.classList.add('dragging');
        sx = e.clientX; sy = e.clientY;
        sl = vp.scrollLeft; st = vp.scrollTop;
    });
    vp.addEventListener('mousemove', function(e) {
        if (!dragging) return;
        e.preventDefault();
        vp.scrollLeft = sl - (e.clientX - sx);
        vp.scrollTop = st - (e.clientY - sy);
    });
    vp.addEventListener('mouseup', function() {
        dragging = false; vp.classList.remove('dragging');
    });
    vp.addEventListener('mouseleave', function() {
        dragging = false; vp.classList.remove('dragging');
    });
}

function setDepFilter(filter, btn) {
    window.depFilter = filter;
    document.querySelectorAll('.dep-filter-btn').forEach(b => b.classList.remove('active'));
    if (btn) btn.classList.add('active');
    renderDepGraph(window.depGraph, filter);
}

let depSortCol = 'osslSym';
let depSortAsc = false;

function renderDepStatsTable(graph) {
    const table = document.getElementById('depStatsTable');
    const sortIcon = (col) => {
        if (depSortCol !== col) return '<span class="sort-icon">&#x2195;</span>';
        return depSortAsc ? '<span class="sort-icon active">&#x2191;</span>' : '<span class="sort-icon active">&#x2193;</span>';
    };

    const nodeList = [...graph.nodes];
    nodeList.sort((a, b) => {
        let cmp = 0;
        if (depSortCol === 'name') cmp = a.name.localeCompare(b.name);
        else if (depSortCol === 'procs') cmp = a.processes.length - b.processes.length;
        else if (depSortCol === 'fanIn') cmp = a.fanIn - b.fanIn;
        else if (depSortCol === 'fanOut') cmp = a.fanOut - b.fanOut;
        else if (depSortCol === 'osslSym') cmp = a.osslSym - b.osslSym;
        else if (depSortCol === 'tier') cmp = a.tier.localeCompare(b.tier);
        return depSortAsc ? cmp : -cmp;
    });

    const tierBadge = (t) => {
        const colors = { application: '#3498db', crypto: '#e74c3c', middleware: '#f39c12', system: '#bdc3c7', normal: '#ecf0f1' };
        const fg = t === 'normal' ? '#7f8c8d' : '#fff';
        return '<span style="display:inline-block;padding:1px 8px;border-radius:3px;font-size:0.8em;background:' + (colors[t] || '#ecf0f1') + ';color:' + fg + '">' + t + '</span>';
    };

    let tbody = '';
    nodeList.forEach(n => {
        const symStyle = n.osslSym > 0 ? 'font-weight:600;color:#c0392b' : 'color:#bdc3c7';
        const escaped = n.name.replace(/'/g, "\\\\'");
        const nameLink = n.osslSym > 0
            ? '<span class="component-link" onclick="showComponentDetail(\\'' + escaped + '\\')">' + n.name + '</span>'
            : n.name;
        tbody += '<tr>';
        tbody += '<td>' + nameLink + (n.isOsslLib ? ' <span class="openssl-marker">OpenSSL</span>' : '') + '</td>';
        if (graph.multiProc) tbody += '<td>' + n.processes.join(', ') + '</td>';
        tbody += '<td style="text-align:center">' + n.fanIn + '</td>';
        tbody += '<td style="text-align:center">' + n.fanOut + '</td>';
        tbody += '<td style="text-align:center;' + symStyle + '">' + n.osslSym + '</td>';
        tbody += '<td>' + tierBadge(n.tier) + '</td>';
        tbody += '</tr>';
    });

    const procsHeader = graph.multiProc
        ? '<th class="sortable" onclick="sortDepStats(\\'procs\\')">Procs ' + sortIcon('procs') + '</th>' : '';

    table.innerHTML = '<thead><tr>' +
        '<th class="sortable" onclick="sortDepStats(\\'name\\')">Library ' + sortIcon('name') + '</th>' +
        procsHeader +
        '<th class="sortable" onclick="sortDepStats(\\'fanIn\\')">Fan-in ' + sortIcon('fanIn') + '</th>' +
        '<th class="sortable" onclick="sortDepStats(\\'fanOut\\')">Fan-out ' + sortIcon('fanOut') + '</th>' +
        '<th class="sortable" onclick="sortDepStats(\\'osslSym\\')">OSSL Sym ' + sortIcon('osslSym') + '</th>' +
        '<th class="sortable" onclick="sortDepStats(\\'tier\\')">Tier ' + sortIcon('tier') + '</th>' +
        '</tr></thead><tbody>' + tbody + '</tbody>';
}

function sortDepStats(col) {
    if (depSortCol === col) depSortAsc = !depSortAsc;
    else { depSortCol = col; depSortAsc = false; }
    renderDepStatsTable(window.depGraph);
}

function normalizeData(data) {
    /* Normalize single/process and aggregated report formats to common structure */
    const reportType = data.meta?.report_type || 'single';
    const isSingleLike = (reportType === 'single' || reportType === 'process');

    if (isSingleLike) {
        /* Convert single report format to aggregated-like format */
        const byFile = data.openssl_symbols?.by_file || {};
        const byCategory = data.openssl_symbols?.by_category || {};

        /* Build ranking from by_file */
        const ranking = Object.entries(byFile)
            .map(([path, info]) => ({
                component: path.split('/').pop(),
                unique_symbols_count: info.count || 0,
                symbols: info.symbols || []
            }))
            .sort((a, b) => b.unique_symbols_count - a.unique_symbols_count)
            .map((item, i) => ({ ...item, rank: i + 1 }));

        /* Build components from by_file with category info */
        const components = {};
        Object.entries(byFile).forEach(([path, info]) => {
            const name = path.split('/').pop();
            components[name] = {
                executables: [name],
                unique_symbols_count: info.count || 0,
                unique_symbols: info.symbols || [],
                by_category: {}
            };
            /* Assign symbols to categories */
            (info.symbols || []).forEach(sym => {
                for (const [cat, catInfo] of Object.entries(byCategory)) {
                    if (catInfo.symbols && catInfo.symbols.includes(sym)) {
                        if (!components[name].by_category[cat]) {
                            components[name].by_category[cat] = { count: 0, symbols: [] };
                        }
                        components[name].by_category[cat].symbols.push(sym);
                        components[name].by_category[cat].count++;
                        break;
                    }
                }
            });
        });

        data.ranking = ranking;
        data.components = components;
        data._normalized = true;

        /* Build hierarchical ranking if dependency tree exists */
        data.hierarchicalRanking = buildHierarchicalRanking(data);

        /* Build global dependency graph */
        data.globalGraph = buildGlobalGraph(data);
    }

    return data;
}

function renderReport(data) {
    console.log('renderReport called, report_type:', data.meta?.report_type);
    console.log('ranking:', data.ranking?.length, 'components:', Object.keys(data.components || {}).length);

    try {
        currentData = normalizeData(data);
        console.log('After normalize - ranking:', currentData.ranking?.length);

        renderSummary(currentData);
        console.log('Summary rendered');

        renderRanking(currentData);
        console.log('Ranking rendered');

        renderDependencies(currentData);
        console.log('Dependencies rendered');

        renderCategories(currentData);
        console.log('Categories rendered');

        renderSymbols(currentData);
        console.log('Symbols rendered');

        populateFilters(currentData);
        console.log('Filters populated');
    } catch (e) {
        console.error('Error in renderReport:', e);
        document.body.innerHTML += '<div style="color:red;padding:20px;">Error: ' + e.message + '</div>';
    }
}

function renderSummary(data) {
    const container = document.getElementById('summaryCards');
    const summary = data.summary || {};

    const cards = [
        { label: 'Components', value: summary.total_components || summary.files_with_openssl_deps || 0 },
        { label: 'Total ELF Files', value: summary.total_executables || summary.total_elf_files || 0 },
        { label: 'Unique Symbols', value: summary.global_unique_symbols || summary.unique_openssl_symbols || 0 }
    ];

    container.innerHTML = cards.map(c => `
        <div class="card">
            <div class="value">${c.value}</div>
            <div class="label">${c.label}</div>
        </div>
    `).join('');

    /* Show process info bar if available */
    const proc = data.meta?.process;
    const piBar = document.getElementById('processInfoBar');
    if (proc && piBar) {
        const items = [
            { label: 'PID', value: proc.pid },
            { label: 'Process', value: proc.name },
            { label: 'Arch', value: data.meta?.target_arch || '' },
            { label: 'Libraries', value: proc.mapped_libraries_count || 0 },
            { label: 'RSS', value: proc.vm_rss_kb ? (proc.vm_rss_kb / 1024).toFixed(1) + ' MB' : '' }
        ];
        if (proc.runtime_loaded_count > 0) {
            items.push({ label: 'dlopen', value: proc.runtime_loaded_count });
        }
        piBar.innerHTML = items
            .filter(i => i.value !== '' && i.value !== undefined)
            .map(i => `<span class="pi-item"><span class="pi-label">${i.label}:</span>${i.value}</span>`)
            .join('');
        piBar.style.display = 'flex';
    }
}

/* Short display name for categories */
const CATEGORY_SHORT = {
    'ssl_core': 'SSL', 'ssl_tls': 'TLS',
    'crypto_evp': 'EVP', 'crypto_rsa': 'RSA',
    'crypto_ec': 'EC', 'crypto_bn': 'BN',
    'crypto_dsa': 'DSA', 'crypto_dh': 'DH',
    'crypto_bio': 'BIO', 'crypto_rand': 'RAND',
    'crypto_aes': 'AES', 'crypto_sm': 'SM',
    'crypto_x509': 'X509', 'crypto_pem': 'PEM',
    'crypto_err': 'ERR', 'crypto_engine': 'ENGINE',
    'crypto_hash': 'HASH', 'crypto_hmac': 'HMAC',
    'crypto_des': 'DES', 'crypto_chacha': 'CHACHA',
    'crypto_pkcs': 'PKCS', 'crypto_cms': 'CMS',
    'crypto_ocsp': 'OCSP', 'crypto_ts': 'TS',
    'crypto_obj': 'OBJ', 'crypto_kdf': 'KDF',
    'crypto_provider': 'PROV', 'crypto_asn1': 'ASN1',
    'openssl_util': 'UTIL', 'other': 'OTHER'
};

function renderCategoryDist(categories) {
    if (!categories || Object.keys(categories).length === 0) return '';
    const sorted = Object.entries(categories).sort((a, b) => b[1] - a[1]);
    const tags = sorted.map(([cat, count]) => {
        const color = CATEGORY_COLORS[cat] || CATEGORY_COLORS['other'];
        const label = CATEGORY_SHORT[cat] || cat;
        return '<span class="cat-tag" style="background:' + color + '">' + label + ':<span class="cat-count">' + count + '</span></span>';
    }).join('');
    return '<div class="cat-dist">' + tags + '</div>';
}

/*
 * Sort state:
 * - hierSort: for hierarchical table (process/single with dep tree)
 *   .col  = 'syms' | 'lib'  (library-level sort key)
 *   .asc  = bool
 *   .cycle = 0(desc) -> 1(asc) -> 2(original tree order)
 * - rankingSortCol/Asc: for flat table (aggregated/no-tree)
 */
let hierSort = { col: 'syms', asc: false, cycle: 0 };
let rankingSortCol = 'symbols';
let rankingSortAsc = false;

function renderRanking(data) {
    window.rankingData = data.ranking || [];
    window.hierGroups = data.hierarchicalRanking || null;
    /* Keep a deep copy of original lib order per group (tree order) */
    if (window.hierGroups) {
        window.hierGroupsOriginal = window.hierGroups.map(g => ({
            exe: g.exe,
            libs: [...g.libs]
        }));
    }
    renderRankingTable();

    /* Chart */
    const groups = window.hierGroups;
    if (groups && groups.length > 0) {
        /* Collect all files with symbols for stacked chart */
        const items = [];
        groups.forEach(g => {
            if (g.exe.syms > 0) items.push({ name: g.exe.elf, cats: g.exe.categories });
            g.libs.forEach(l => {
                if (l.syms > 0) items.push({ name: l.lib, cats: l.categories });
            });
        });
        if (items.length === 0) return;

        const labels = items.map(i => i.name);
        const allCats = new Set();
        items.forEach(i => Object.keys(i.cats).forEach(c => allCats.add(c)));

        const datasets = [...allCats].map(cat => ({
            label: cat,
            data: items.map(i => i.cats[cat] || 0),
            backgroundColor: CATEGORY_COLORS[cat] || CATEGORY_COLORS['other']
        }));

        const ctx = document.getElementById('rankingChart').getContext('2d');
        if (rankingChart) rankingChart.destroy();
        rankingChart = new Chart(ctx, {
            type: 'bar',
            data: { labels, datasets },
            options: {
                indexAxis: 'y',
                responsive: true,
                maintainAspectRatio: false,
                scales: { x: { stacked: true }, y: { stacked: true } },
                plugins: { legend: { position: 'bottom', labels: { boxWidth: 12, font: { size: 11 } } } }
            }
        });
    } else if (window.rankingData.length > 0) {
        const ctx = document.getElementById('rankingChart').getContext('2d');
        if (rankingChart) rankingChart.destroy();
        const ranking = window.rankingData;
        rankingChart = new Chart(ctx, {
            type: 'bar',
            data: {
                labels: ranking.slice(0, 15).map(r => r.component),
                datasets: [{
                    label: 'Symbols Used',
                    data: ranking.slice(0, 15).map(r => r.unique_symbols_count),
                    backgroundColor: '#3498db'
                }]
            },
            options: {
                indexAxis: 'y',
                responsive: true,
                maintainAspectRatio: false
            }
        });
    }
}

function sortHierarchical(col) {
    if (hierSort.col === col) {
        /* Same column: cycle desc(0) -> asc(1) -> original(2) -> desc(0) */
        hierSort.cycle = (hierSort.cycle + 1) % 3;
        hierSort.asc = hierSort.cycle === 1;
    } else {
        hierSort.col = col;
        hierSort.cycle = 0;
        hierSort.asc = false;
    }
    renderRankingTable();
}

function renderRankingTable() {
    const table = document.getElementById('rankingTable');
    const groups = window.hierGroups;

    if (groups && groups.length > 0) {
        /* --- Hierarchical table with two-level sort --- */
        const isOriginal = hierSort.cycle === 2;

        /* Level 1: sort EXE groups by groupTotal (exe + lib syms) */
        let sortedGroups = groups.map((g, i) => ({ group: g, origLibs: window.hierGroupsOriginal[i].libs }));
        if (!isOriginal) {
            sortedGroups.sort((a, b) => {
                const cmp = a.group.exe.groupTotal - b.group.exe.groupTotal;
                return hierSort.asc ? cmp : -cmp;
            });
        }

        /* Sort icon with 3 states */
        const hierIcon = (col) => {
            if (hierSort.col !== col) return '<span class="sort-icon">&#x2195;</span>';
            if (hierSort.cycle === 0) return '<span class="sort-icon active">&#x2193;</span>';
            if (hierSort.cycle === 1) return '<span class="sort-icon active">&#x2191;</span>';
            return '<span class="sort-icon active">&#x2261;</span>';
        };

        let tbody = '';
        sortedGroups.forEach(({ group, origLibs }) => {
            const exe = group.exe;

            /* Level 2: sort libs within group */
            let libs;
            if (isOriginal) {
                libs = [...origLibs];
            } else {
                libs = [...group.libs];
                libs.sort((a, b) => {
                    let cmp = 0;
                    if (hierSort.col === 'syms') {
                        cmp = a.syms - b.syms;
                    } else if (hierSort.col === 'lib') {
                        cmp = a.lib.localeCompare(b.lib);
                    }
                    return hierSort.asc ? cmp : -cmp;
                });
            }

            const rowSpan = libs.length + 1;
            const exeCats = renderCategoryDist(exe.categories);

            /* EXE row */
            const badgeText = exe.groupTotal !== exe.syms
                ? exe.groupTotal + ' sym (self: ' + exe.syms + ')'
                : exe.syms + ' sym';
            tbody += '<tr class="ranking-exe-row">';
            tbody += '<td rowspan="' + rowSpan + '" style="vertical-align:top">';
            tbody += '<span class="component-link" onclick="showComponentDetail(\\'';
            tbody += exe.elf.replace(/'/g, "\\\\'") + '\\')">';
            tbody += exe.elf + '</span>';
            tbody += '<span class="sym-badge">' + badgeText + '</span>';
            tbody += '</td>';
            tbody += '<td rowspan="' + rowSpan + '" style="vertical-align:top;text-align:center">';
            tbody += exe.libCount;
            tbody += '</td>';
            tbody += '<td style="color:#7f8c8d;font-style:italic">(executable)</td>';
            tbody += '<td>' + exe.syms + '</td>';
            tbody += '<td>' + exeCats + '</td>';
            tbody += '</tr>';

            /* Library rows */
            libs.forEach(r => {
                const libCats = renderCategoryDist(r.categories);
                let libDisplay = r.lib;
                if (r.syms > 0) {
                    const escaped = r.lib.replace(/'/g, "\\\\'");
                    libDisplay = '<span class="component-link" onclick="showComponentDetail(\\'' + escaped + '\\')">' + r.lib + '</span>';
                }
                if (r.isOpenSSL) libDisplay += '<span class="openssl-marker">OpenSSL</span>';
                if (r.isCircular) libDisplay += ' <span class="circular-tag">(circular)</span>';
                if (r.depth > 1) libDisplay += ' <span style="color:#e67e22;font-size:0.8em">(depth ' + r.depth + ')</span>';
                else if (r.childCount > 0) libDisplay += ' <span style="color:#95a5a6;font-size:0.85em">(' + r.childCount + ' deps)</span>';

                const symStyle = r.syms > 0 ? 'font-weight:600' : 'color:#bdc3c7';
                tbody += '<tr class="ranking-lib-row">';
                tbody += '<td>' + libDisplay + '</td>';
                tbody += '<td style="' + symStyle + '">' + r.syms + '</td>';
                tbody += '<td>' + (r.syms > 0 ? libCats : '') + '</td>';
                tbody += '</tr>';
            });
        });

        table.innerHTML = `
            <thead>
                <tr>
                    <th class="sortable" onclick="sortHierarchical('syms')">Process / ELF ${hierIcon('syms')}</th>
                    <th>Libs</th>
                    <th class="sortable" onclick="sortHierarchical('lib')">Library ${hierIcon('lib')}</th>
                    <th class="sortable" onclick="sortHierarchical('syms')">Symbols ${hierIcon('syms')}</th>
                    <th>Distribution</th>
                </tr>
            </thead>
            <tbody>${tbody}</tbody>
        `;
        return;
    }

    /* Fallback: flat ranking table for aggregated/directory scans */
    const ranking = [...window.rankingData];
    const total = currentData?.summary?.global_unique_symbols || currentData?.summary?.unique_openssl_symbols || 1;

    ranking.sort((a, b) => {
        let cmp = 0;
        if (rankingSortCol === 'component') {
            cmp = a.component.localeCompare(b.component);
        } else if (rankingSortCol === 'symbols') {
            cmp = a.unique_symbols_count - b.unique_symbols_count;
        }
        return rankingSortAsc ? cmp : -cmp;
    });

    const sortIcon = (col) => {
        if (rankingSortCol !== col) return '<span class="sort-icon">&#x2195;</span>';
        return rankingSortAsc ? '<span class="sort-icon active">&#x2191;</span>' : '<span class="sort-icon active">&#x2193;</span>';
    };

    table.innerHTML = `
        <thead>
            <tr>
                <th>#</th>
                <th class="sortable" onclick="sortRanking('component')">Component ${sortIcon('component')}</th>
                <th class="sortable" onclick="sortRanking('symbols')">Symbols ${sortIcon('symbols')}</th>
                <th>Distribution</th>
            </tr>
        </thead>
        <tbody>
            ${ranking.map((r, i) => {
                const pct = (r.unique_symbols_count / total * 100).toFixed(1);
                return `
                <tr>
                    <td>${i + 1}</td>
                    <td><span class="component-link" onclick="showComponentDetail('${r.component.replace(/'/g, "\\'")}')">${r.component}</span></td>
                    <td>${r.unique_symbols_count}</td>
                    <td>
                        <div class="bar-container" style="width:200px">
                            <div class="bar" style="width:${pct}%"></div>
                        </div>
                        ${pct}%
                    </td>
                </tr>`;
            }).join('')}
        </tbody>
    `;

    if (ranking.length === 0) {
        table.innerHTML += '<tbody><tr><td colspan="4">No data available</td></tr></tbody>';
    }
}

function sortRanking(col) {
    if (rankingSortCol === col) {
        rankingSortAsc = !rankingSortAsc;
    } else {
        rankingSortCol = col;
        rankingSortAsc = col === 'component';
    }
    renderRankingTable();
}

function renderCategories(data) {
    const components = data.components || {};
    const categoryTotals = {};

    /* Also check openssl_symbols.by_category for single reports */
    const directCategories = data.openssl_symbols?.by_category || {};

    if (Object.keys(directCategories).length > 0) {
        Object.entries(directCategories).forEach(([cat, catData]) => {
            categoryTotals[cat] = catData.count || 0;
        });
    } else {
        Object.values(components).forEach(comp => {
            Object.entries(comp.by_category || {}).forEach(([cat, catData]) => {
                const count = typeof catData === 'object' ? catData.count : 0;
                categoryTotals[cat] = (categoryTotals[cat] || 0) + count;
            });
        });
    }

    const sorted = Object.entries(categoryTotals).sort((a, b) => b[1] - a[1]);

    const table = document.getElementById('categoryTable');
    table.innerHTML = `
        <thead>
            <tr>
                <th>Category</th>
                <th>Total Symbols</th>
            </tr>
        </thead>
        <tbody>
            ${sorted.map(([cat, count]) => `
                <tr>
                    <td>${cat}</td>
                    <td>${count}</td>
                </tr>
            `).join('')}
        </tbody>
    `;

    const ctx = document.getElementById('categoryChart').getContext('2d');
    if (categoryChart) categoryChart.destroy();
    categoryChart = new Chart(ctx, {
        type: 'doughnut',
        data: {
            labels: sorted.slice(0, 10).map(s => s[0]),
            datasets: [{
                data: sorted.slice(0, 10).map(s => s[1]),
                backgroundColor: [
                    '#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6',
                    '#1abc9c', '#34495e', '#e67e22', '#95a5a6', '#d35400'
                ]
            }]
        },
        options: {
            responsive: true,
            maintainAspectRatio: false
        }
    });
}

function renderSymbols(data) {
    const components = data.components || {};
    const symbolMap = {};
    const isSingle = (data.meta?.report_type === 'single' || data.meta?.report_type === 'process');

    /* For single reports, use openssl_symbols.by_category directly */
    if (isSingle && data.openssl_symbols?.by_category) {
        const byCategory = data.openssl_symbols.by_category;
        const byFile = data.openssl_symbols.by_file || {};

        /* Build symbol -> files mapping */
        const symbolToFiles = {};
        Object.entries(byFile).forEach(([path, info]) => {
            const fileName = path.split('/').pop();
            (info.symbols || []).forEach(sym => {
                if (!symbolToFiles[sym]) symbolToFiles[sym] = [];
                if (!symbolToFiles[sym].includes(fileName)) {
                    symbolToFiles[sym].push(fileName);
                }
            });
        });

        /* Build symbolMap from categories */
        Object.entries(byCategory).forEach(([cat, catData]) => {
            (catData.symbols || []).forEach(sym => {
                if (!symbolMap[sym]) {
                    symbolMap[sym] = {
                        components: symbolToFiles[sym] || [],
                        category: cat
                    };
                }
            });
        });
    } else {
        /* Aggregated report format */
        Object.entries(components).forEach(([compName, comp]) => {
            Object.entries(comp.by_category || {}).forEach(([cat, catData]) => {
                const symbols = typeof catData === 'object' ? catData.symbols || [] : [];
                symbols.forEach(sym => {
                    if (!symbolMap[sym]) {
                        symbolMap[sym] = { components: [], category: cat };
                    }
                    if (!symbolMap[sym].components.includes(compName)) {
                        symbolMap[sym].components.push(compName);
                    }
                });
            });
        });
    }

    window.symbolData = Object.entries(symbolMap).map(([name, info]) => ({
        name,
        components: info.components,
        category: info.category
    })).sort((a, b) => a.name.localeCompare(b.name));

    renderSymbolTable(window.symbolData);
}

function renderSymbolTable(symbols) {
    const table = document.getElementById('symbolsTable');
    table.innerHTML = `
        <thead>
            <tr>
                <th>Symbol</th>
                <th>Components</th>
                <th>Category</th>
            </tr>
        </thead>
        <tbody>
            ${symbols.slice(0, 500).map(s => `
                <tr>
                    <td>${s.name}</td>
                    <td>${s.components.join(', ')}</td>
                    <td>${s.category}</td>
                </tr>
            `).join('')}
        </tbody>
    `;
    if (symbols.length > 500) {
        table.innerHTML += `<tfoot><tr><td colspan="3">Showing 500 of ${symbols.length} symbols. Use filters to narrow down.</td></tr></tfoot>`;
    }
}

function populateFilters(data) {
    const components = data.components || {};
    const categories = new Set();
    const isSingle = (data.meta?.report_type === 'single' || data.meta?.report_type === 'process');

    /* Get categories from openssl_symbols for single reports */
    if (isSingle && data.openssl_symbols?.by_category) {
        Object.keys(data.openssl_symbols.by_category).forEach(cat => categories.add(cat));
    } else {
        Object.values(components).forEach(comp => {
            Object.keys(comp.by_category || {}).forEach(cat => categories.add(cat));
        });
    }

    /* Get file/component names */
    let compNames = Object.keys(components);
    if (isSingle && data.openssl_symbols?.by_file) {
        compNames = Object.keys(data.openssl_symbols.by_file).map(p => p.split('/').pop());
    }

    const catFilter = document.getElementById('categoryFilter');
    catFilter.innerHTML = '<option value="">All Categories</option>' +
        [...categories].sort().map(c => `<option value="${c}">${c}</option>`).join('');

    const compFilter = document.getElementById('componentFilter');
    const filterLabel = isSingle ? 'All Files' : 'All Components';
    compFilter.innerHTML = `<option value="">${filterLabel}</option>` +
        compNames.sort().map(c => `<option value="${c}">${c}</option>`).join('');
}

function filterSymbols() {
    const search = document.getElementById('symbolSearch').value.toLowerCase();
    const category = document.getElementById('categoryFilter').value;
    const component = document.getElementById('componentFilter').value;

    const filtered = window.symbolData.filter(s => {
        if (search && !s.name.toLowerCase().includes(search)) return false;
        if (category && s.category !== category) return false;
        if (component && !s.components.includes(component)) return false;
        return true;
    });

    renderSymbolTable(filtered);
}

function showTab(tabId) {
    document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.getElementById(tabId).classList.add('active');
    event.target.classList.add('active');
}

function loadFile() {
    document.getElementById('fileInput').click();
}

function handleFileSelect(event) {
    const file = event.target.files[0];
    if (file) {
        const reader = new FileReader();
        reader.onload = function(e) {
            try {
                const data = JSON.parse(e.target.result);
                renderReport(data);
            } catch (err) {
                alert('Invalid JSON file');
            }
        };
        reader.readAsText(file);
    }
}

function exportExcel() {
    if (!currentData) {
        alert('No data to export');
        return;
    }

    /* Check if SheetJS loaded from CDN */
    if (window.XLSX_LOADED && typeof XLSX !== 'undefined' && XLSX.utils) {
        exportWithSheetJS();
    } else {
        /* Fallback to CSV */
        console.log('Using CSV fallback for export');
        exportAsCSV(currentData, 'openssl_report.xlsx');
    }
}

function exportWithSheetJS() {
    /*
     * Excel export with 8 sheets - matches CLI ExcelExporter exactly:
     * 1. Overview     - Metadata and summary statistics
     * 2. Files        - All scanned files with full attributes
     * 3. File-Symbol  - Flat table for pivot analysis
     * 4. Import Chains - Symbol import paths
     * 5. By Category  - Category-wise symbol statistics
     * 6. By Depth     - Depth-wise symbol statistics
     * 7. Dep Tree     - Flattened dependency tree
     * 8. Errors       - Scan errors and warnings
     */
    const wb = XLSX.utils.book_new();
    const meta = currentData.meta || {};
    const summary = currentData.summary || {};
    const reportType = meta.report_type || 'single';
    const isAggregated = reportType === 'aggregated';

    /* Build symbol-to-category mapping */
    const byCategory = currentData.openssl_symbols?.by_category || {};
    const symToCategory = {};
    Object.entries(byCategory).forEach(([cat, catData]) => {
        (catData.symbols || []).forEach(sym => { symToCategory[sym] = cat; });
    });

    /* ===== Sheet 1: Overview ===== */
    const overviewData = [
        ['REPORT METADATA', ''],
        ['Report Type', reportType],
        ['Tool Version', meta.tool_version || ''],
        ['Scan Time', meta.aggregation_time || meta.scan_time || ''],
        ['Scan Root', meta.scan_root || ''],
        ['Target Architecture', meta.target_arch || ''],
        ['', '']
    ];
    if (isAggregated) {
        overviewData.push(
            ['AGGREGATION INFO', ''],
            ['Source Reports', meta.source_reports_count || 0],
            ['Mapping File', meta.mapping_file || 'None'],
            ['', '']
        );
    }
    overviewData.push(
        ['SUMMARY STATISTICS', ''],
        ['Total Files Scanned', summary.total_files_scanned || summary.total_executables || 0],
        ['Total ELF Files', summary.total_elf_files || summary.total_executables || 0],
        ['Files with OpenSSL Deps', summary.files_with_openssl_deps || summary.total_components || 0],
        ['Total OpenSSL Symbols (refs)', summary.total_openssl_symbols || 0],
        ['Unique OpenSSL Symbols', summary.unique_openssl_symbols || summary.global_unique_symbols || 0],
        ['', '']
    );
    const openssl_libs = summary.openssl_libs_found || [];
    if (openssl_libs.length > 0) {
        overviewData.push(['DETECTED OPENSSL LIBRARIES', '']);
        openssl_libs.forEach(lib => overviewData.push(['', lib]));
    }
    const overviewWs = XLSX.utils.aoa_to_sheet(overviewData);
    overviewWs['!cols'] = [{ wch: 30 }, { wch: 60 }];
    XLSX.utils.book_append_sheet(wb, overviewWs, 'Overview');

    /* ===== Sheet 2: Files ===== */
    const filesData = [['File Path', 'File Name', 'Type', 'Arch', 'OpenSSL Direct', 'OpenSSL Transitive', 'OpenSSL Libs', 'Symbol Count', 'Direct Dependencies']];
    let filesDetail = currentData.files_detail || [];
    if (filesDetail.length === 0) {
        /* Fallback: build from by_file */
        const byFile = currentData.openssl_symbols?.by_file || {};
        Object.entries(byFile).forEach(([path, info]) => {
            filesDetail.push({
                path: path,
                type: path.includes('.so') ? 'shared_library' : 'executable',
                arch: meta.target_arch || '',
                openssl_deps: { direct: false, transitive: true, libs: [] },
                openssl_symbols_used: info.symbols || [],
                direct_deps: []
            });
        });
    }
    filesDetail.forEach(f => {
        const deps = f.openssl_deps || {};
        filesData.push([
            f.path || '',
            (f.path || '').split('/').pop(),
            f.type || '',
            f.arch || '',
            deps.direct ? 'Yes' : 'No',
            deps.transitive ? 'Yes' : 'No',
            (deps.libs || []).join(', '),
            (f.openssl_symbols_used || []).length,
            (f.direct_deps || []).join(', ')
        ]);
    });
    if (filesData.length > 1) {
        const filesWs = XLSX.utils.aoa_to_sheet(filesData);
        filesWs['!cols'] = [{ wch: 60 }, { wch: 25 }, { wch: 15 }, { wch: 10 }, { wch: 15 }, { wch: 18 }, { wch: 40 }, { wch: 12 }, { wch: 60 }];
        XLSX.utils.book_append_sheet(wb, filesWs, 'Files');
    }

    /* ===== Sheet 3: File-Symbol (core pivot table) ===== */
    const fileSymbolData = [['Component', 'Binary', 'Symbol', 'Category']];
    const byFile = currentData.openssl_symbols?.by_file || {};
    Object.entries(byFile).forEach(([path, info]) => {
        const fileName = path.split('/').pop();
        (info.symbols || []).sort().forEach(sym => {
            fileSymbolData.push([fileName, fileName, sym, symToCategory[sym] || 'other']);
        });
    });
    /* For aggregated reports: use executables_detail if available */
    if (fileSymbolData.length === 1 && currentData.components) {
        Object.entries(currentData.components).forEach(([compName, compData]) => {
            const execDetail = compData.executables_detail || {};
            if (Object.keys(execDetail).length > 0) {
                /* Use binary-level detail */
                Object.entries(execDetail).forEach(([binName, binData]) => {
                    Object.entries(binData.by_category || {}).forEach(([cat, catData]) => {
                        (catData.symbols || []).sort().forEach(sym => {
                            fileSymbolData.push([compName, binName, sym, cat]);
                        });
                    });
                });
            } else {
                /* Fallback to component-level aggregation */
                Object.entries(compData.by_category || {}).forEach(([cat, catData]) => {
                    (catData.symbols || []).sort().forEach(sym => {
                        fileSymbolData.push([compName, compName, sym, cat]);
                    });
                });
            }
        });
    }
    if (fileSymbolData.length > 1) {
        const fileSymbolWs = XLSX.utils.aoa_to_sheet(fileSymbolData);
        fileSymbolWs['!cols'] = [{ wch: 60 }, { wch: 25 }, { wch: 35 }, { wch: 18 }];
        XLSX.utils.book_append_sheet(wb, fileSymbolWs, 'File-Symbol');
    }

    /* ===== Sheet 4: Import Chains ===== */
    const importChainsData = [['Source File', 'File Name', 'Symbol', 'Category', 'Import Chain', 'Depth']];
    const importChains = currentData.openssl_symbols?.import_chains || {};
    Object.entries(importChains).sort().forEach(([sym, chains]) => {
        const cat = symToCategory[sym] || 'other';
        (chains || []).forEach(chainItem => {
            let sourceFile, chainStr, depth;
            if (typeof chainItem === 'object') {
                sourceFile = chainItem.source_file || '';
                if (!sourceFile && chainItem.component) {
                    sourceFile = chainItem.component + '/' + (chainItem.binary || '');
                }
                chainStr = chainItem.chain || '';
                depth = chainItem.depth || 0;
            } else {
                sourceFile = '';
                chainStr = chainItem;
                depth = (chainStr.match(/ -> /g) || []).length;
            }
            const fileName = sourceFile ? sourceFile.split('/').pop() : '';
            importChainsData.push([sourceFile, fileName, sym, cat, chainStr, depth]);
        });
    });
    if (importChainsData.length === 1) {
        importChainsData.push(['No import chain data available', '', '', '', '', '']);
    }
    const importChainsWs = XLSX.utils.aoa_to_sheet(importChainsData);
    importChainsWs['!cols'] = [{ wch: 50 }, { wch: 20 }, { wch: 35 }, { wch: 15 }, { wch: 80 }, { wch: 10 }];
    XLSX.utils.book_append_sheet(wb, importChainsWs, 'Import Chains');

    /* ===== Sheet 5: By Category ===== */
    const categoryData = [['Category', 'Symbol Count', 'Percentage', 'Symbols']];
    let catStats = {};
    Object.entries(byCategory).forEach(([cat, catData]) => {
        catStats[cat] = {
            count: catData.count || (catData.symbols || []).length,
            symbols: catData.symbols || []
        };
    });
    /* Fallback for aggregated reports */
    if (Object.keys(catStats).length === 0 && currentData.components) {
        Object.values(currentData.components).forEach(comp => {
            Object.entries(comp.by_category || {}).forEach(([cat, catData]) => {
                if (!catStats[cat]) catStats[cat] = { count: 0, symbols: new Set() };
                const symbols = catData.symbols || [];
                catStats[cat].count += symbols.length;
                symbols.forEach(s => catStats[cat].symbols.add(s));
            });
        });
        Object.keys(catStats).forEach(cat => {
            if (catStats[cat].symbols instanceof Set) {
                catStats[cat].symbols = Array.from(catStats[cat].symbols);
            }
        });
    }
    const totalCatSymbols = Object.values(catStats).reduce((sum, c) => sum + c.count, 0) || 1;
    Object.entries(catStats)
        .sort((a, b) => b[1].count - a[1].count)
        .forEach(([cat, data]) => {
            const pct = ((data.count / totalCatSymbols) * 100).toFixed(1) + '%';
            categoryData.push([cat, data.count, pct, data.symbols.sort().join(', ')]);
        });
    categoryData.push(['TOTAL', totalCatSymbols, '100%', '']);
    const categoryWs = XLSX.utils.aoa_to_sheet(categoryData);
    categoryWs['!cols'] = [{ wch: 20 }, { wch: 15 }, { wch: 12 }, { wch: 100 }];
    XLSX.utils.book_append_sheet(wb, categoryWs, 'By Category');

    /* ===== Sheet 6: By Depth ===== */
    const depthData = [['Depth', 'Description', 'Symbol Count', 'Symbols']];
    const byDepth = currentData.openssl_symbols?.by_depth || {};
    const depthDescs = {
        0: 'Root binary itself',
        1: 'Direct dependencies (depth 1)',
        2: 'Transitive dependencies (depth 2)'
    };
    if (Object.keys(byDepth).length === 0) {
        depthData.push(['No depth data available', '', '', '']);
        depthData.push(['(Depth analysis requires single-binary scan mode)', '', '', '']);
    } else {
        Object.keys(byDepth).sort().forEach(depthKey => {
            const depthInfo = byDepth[depthKey];
            const depthNum = parseInt(depthKey.replace('depth_', ''), 10) || 0;
            const desc = depthDescs[depthNum] || ('Transitive dependencies (depth ' + depthNum + ')');
            const count = depthInfo.count || (depthInfo.symbols || []).length;
            depthData.push([depthNum, desc, count, (depthInfo.symbols || []).sort().join(', ')]);
        });
    }
    const depthWs = XLSX.utils.aoa_to_sheet(depthData);
    depthWs['!cols'] = [{ wch: 10 }, { wch: 35 }, { wch: 15 }, { wch: 100 }];
    XLSX.utils.book_append_sheet(wb, depthWs, 'By Depth');

    /* ===== Sheet 7: Dep Tree (flattened) ===== */
    const depTreeData = [['Parent', 'Child', 'Depth', 'Is OpenSSL Lib', 'Symbol Count', 'Full Path']];
    const depTree = currentData.dependency_tree;
    if (!depTree) {
        depTreeData.push(['No dependency tree data available', '', '', '', '', '']);
        depTreeData.push(['(Dependency tree requires single-binary scan mode)', '', '', '', '', '']);
    } else {
        function flattenTree(node, parent, depth) {
            depTreeData.push([
                parent || '(root)',
                node.name || '',
                depth,
                node.is_openssl_lib ? 'Yes' : 'No',
                node.openssl_symbols_count || 0,
                node.path || ''
            ]);
            (node.children || []).forEach(child => flattenTree(child, node.name, depth + 1));
        }
        flattenTree(depTree, null, 0);
    }
    const depTreeWs = XLSX.utils.aoa_to_sheet(depTreeData);
    depTreeWs['!cols'] = [{ wch: 25 }, { wch: 25 }, { wch: 10 }, { wch: 15 }, { wch: 12 }, { wch: 60 }];
    XLSX.utils.book_append_sheet(wb, depTreeWs, 'Dep Tree');

    /* ===== Sheet 8: Errors ===== */
    const errorsData = [['Severity', 'File', 'Error Message']];
    const errors = currentData.errors || [];
    if (errors.length === 0) {
        errorsData.push(['No errors', '', '']);
    } else {
        errors.forEach(err => {
            if (typeof err === 'object') {
                errorsData.push([
                    (err.severity || 'warning').toUpperCase(),
                    err.file || '',
                    err.error || ''
                ]);
            } else {
                errorsData.push(['WARNING', '', String(err)]);
            }
        });
    }
    const errorsWs = XLSX.utils.aoa_to_sheet(errorsData);
    errorsWs['!cols'] = [{ wch: 12 }, { wch: 50 }, { wch: 60 }];
    XLSX.utils.book_append_sheet(wb, errorsWs, 'Errors');

    XLSX.writeFile(wb, 'openssl_report.xlsx');
}

/* Component Detail Modal Functions */
function showComponentDetail(componentName) {
    const modal = document.getElementById('componentModal');
    const title = document.getElementById('modalTitle');
    const summary = document.getElementById('modalSummary');
    const body = document.getElementById('modalBody');

    title.textContent = componentName;

    /* Get component data */
    const compData = currentData.components?.[componentName];
    const isSingle = (currentData.meta?.report_type === 'single' || currentData.meta?.report_type === 'process');

    if (!compData && isSingle) {
        /* For single/process reports, build from openssl_symbols */
        const byFile = currentData.openssl_symbols?.by_file || {};
        const byCategory = currentData.openssl_symbols?.by_category || {};

        /* Find matching file: exact path, endsWith, or basename prefix */
        let fileData = null;
        let filePath = '';
        const nameBase = (componentName || '').split('/').pop();
        for (const [path, info] of Object.entries(byFile)) {
            if (path === componentName || path.endsWith('/' + componentName)) {
                fileData = info;
                filePath = path;
                break;
            }
            const pathBase = path.split('/').pop();
            if (nameBase && (pathBase.startsWith(nameBase) ||
                nameBase.startsWith(pathBase.split('.so')[0] + '.so'))) {
                fileData = info;
                filePath = path;
                break;
            }
        }

        if (fileData) {
            /* Direct match: show this file's symbols */
            const symbolsByCategory = {};
            (fileData.symbols || []).forEach(sym => {
                let found = false;
                for (const [cat, catInfo] of Object.entries(byCategory)) {
                    if (catInfo.symbols && catInfo.symbols.includes(sym)) {
                        if (!symbolsByCategory[cat]) symbolsByCategory[cat] = [];
                        symbolsByCategory[cat].push(sym);
                        found = true;
                        break;
                    }
                }
                if (!found) {
                    if (!symbolsByCategory['other']) symbolsByCategory['other'] = [];
                    symbolsByCategory['other'].push(sym);
                }
            });

            renderComponentModal(summary, body, {
                totalSymbols: fileData.count || fileData.symbols?.length || 0,
                categoryCount: Object.keys(symbolsByCategory).length,
                binaryCount: 1,
                symbolsByCategory: symbolsByCategory,
                executables: null
            });
        } else if (window.hierGroups) {
            /* No direct match: look for EXE group and aggregate all lib symbols */
            const group = window.hierGroups.find(g => g.exe.elf === componentName);
            if (group) {
                const allSymbols = [];
                const binaryNames = [];
                group.libs.forEach(lib => {
                    if (lib.syms > 0) {
                        const libInfo = Object.entries(byFile).find(([p]) => {
                            const pb = p.split('/').pop();
                            return pb.startsWith(lib.lib) || lib.lib.startsWith(pb.split('.so')[0] + '.so');
                        });
                        if (libInfo) {
                            allSymbols.push(...(libInfo[1].symbols || []));
                            binaryNames.push(lib.lib + ' (' + lib.syms + ')');
                        }
                    }
                });
                const symbolsByCategory = {};
                allSymbols.forEach(sym => {
                    let found = false;
                    for (const [cat, catInfo] of Object.entries(byCategory)) {
                        if (catInfo.symbols && catInfo.symbols.includes(sym)) {
                            if (!symbolsByCategory[cat]) symbolsByCategory[cat] = [];
                            symbolsByCategory[cat].push(sym);
                            found = true;
                            break;
                        }
                    }
                    if (!found) {
                        if (!symbolsByCategory['other']) symbolsByCategory['other'] = [];
                        symbolsByCategory['other'].push(sym);
                    }
                });
                const unique = new Set(allSymbols);
                title.textContent = componentName + ' (via ' + binaryNames.length + ' libraries)';
                renderComponentModal(summary, body, {
                    totalSymbols: unique.size,
                    categoryCount: Object.keys(symbolsByCategory).length,
                    binaryCount: binaryNames.length,
                    symbolsByCategory: symbolsByCategory,
                    executables: binaryNames
                });
            } else {
                body.innerHTML = '<p>No OpenSSL symbol data for this component.</p>';
            }
        } else {
            body.innerHTML = '<p>No data found for this component.</p>';
        }
        modal.classList.add('active');
        return;
    } else if (compData) {
        /* Aggregated report format - check for executables_detail */
        const execDetail = compData.executables_detail || {};
        const hasDetail = Object.keys(execDetail).length > 0;

        if (hasDetail) {
            /* Three-level hierarchy: Component -> Binary -> Category */
            renderComponentModalWithBinaries(summary, body, compData);
        } else {
            /* Legacy format: Component -> Category */
            const symbolsByCategory = {};
            Object.entries(compData.by_category || {}).forEach(([cat, catData]) => {
                const symbols = typeof catData === 'object' ? catData.symbols || [] : [];
                if (symbols.length > 0) {
                    symbolsByCategory[cat] = symbols;
                }
            });

            renderComponentModal(summary, body, {
                totalSymbols: compData.unique_symbols_count || compData.unique_symbols?.length || 0,
                categoryCount: Object.keys(symbolsByCategory).length,
                binaryCount: compData.executables?.length || 0,
                symbolsByCategory: symbolsByCategory,
                executables: compData.executables
            });
        }
    } else {
        body.innerHTML = '<p>No data found for this component.</p>';
    }

    modal.classList.add('active');
    document.body.style.overflow = 'hidden';
}

function renderComponentModalWithBinaries(summary, body, compData) {
    /* Three-level hierarchy: Component -> Binary -> Category */
    const execDetail = compData.executables_detail || {};
    const binaryCount = Object.keys(execDetail).length;
    const totalSymbols = compData.unique_symbols_count || 0;
    const categoryCount = Object.keys(compData.by_category || {}).length;

    /* Render summary stats */
    summary.innerHTML = `
        <div class="stat">
            <div class="value">${binaryCount}</div>
            <div class="label">Binaries</div>
        </div>
        <div class="stat">
            <div class="value">${totalSymbols}</div>
            <div class="label">Total APIs</div>
        </div>
        <div class="stat">
            <div class="value">${categoryCount}</div>
            <div class="label">Categories</div>
        </div>
    `;

    /* Sort binaries by symbol count (descending) */
    const sortedBinaries = Object.entries(execDetail)
        .sort((a, b) => (b[1].unique_symbols_count || 0) - (a[1].unique_symbols_count || 0));

    /* Render binary sections */
    body.innerHTML = sortedBinaries.map(([binName, binData]) => {
        const binCategories = binData.by_category || {};
        const sortedCats = Object.entries(binCategories)
            .sort((a, b) => (b[1].count || b[1].symbols?.length || 0) - (a[1].count || a[1].symbols?.length || 0));

        return `
            <div class="binary-section">
                <div class="binary-header" onclick="toggleBinarySection(this)">
                    <span class="binary-toggle">&#9654;</span>
                    <span class="binary-name">${binName}</span>
                    <span class="binary-stats">${binData.unique_symbols_count || 0} APIs in ${sortedCats.length} categories</span>
                </div>
                <div class="binary-content" style="display:none;">
                    ${sortedCats.map(([cat, catData]) => {
                        const symbols = catData.symbols || [];
                        return `
                            <div class="category-section">
                                <h3>
                                    <span>${cat}</span>
                                    <span class="count">${symbols.length} APIs</span>
                                </h3>
                                <div class="symbol-list">
                                    ${symbols.sort().map(s => `<span class="symbol-tag">${s}</span>`).join('')}
                                </div>
                            </div>
                        `;
                    }).join('')}
                </div>
            </div>
        `;
    }).join('');
}

function toggleBinarySection(header) {
    const content = header.nextElementSibling;
    const toggle = header.querySelector('.binary-toggle');
    if (content.style.display === 'none') {
        content.style.display = 'block';
        toggle.innerHTML = '&#9660;';
        header.classList.add('expanded');
    } else {
        content.style.display = 'none';
        toggle.innerHTML = '&#9654;';
        header.classList.remove('expanded');
    }
}

function renderComponentModal(summary, body, data) {
    /* Render summary stats */
    summary.innerHTML = `
        <div class="stat">
            <div class="value">${data.totalSymbols}</div>
            <div class="label">Total APIs</div>
        </div>
        <div class="stat">
            <div class="value">${data.categoryCount}</div>
            <div class="label">Categories</div>
        </div>
    `;

    /* Sort categories by symbol count (descending) */
    const sortedCategories = Object.entries(data.symbolsByCategory)
        .sort((a, b) => b[1].length - a[1].length);

    /* Render category sections */
    body.innerHTML = sortedCategories.map(([cat, symbols]) => `
        <div class="category-section">
            <h3>
                <span>${cat}</span>
                <span class="count">${symbols.length} APIs</span>
            </h3>
            <div class="symbol-list">
                ${symbols.sort().map(s => `<span class="symbol-tag">${s}</span>`).join('')}
            </div>
        </div>
    `).join('');
}

function closeModal() {
    const modal = document.getElementById('componentModal');
    modal.classList.remove('active');
    document.body.style.overflow = '';
}

/* Close modal on outside click */
document.addEventListener('click', function(e) {
    const modal = document.getElementById('componentModal');
    if (e.target === modal) {
        closeModal();
    }
});

/* Close modal on Escape key */
document.addEventListener('keydown', function(e) {
    if (e.key === 'Escape') {
        closeModal();
    }
});
'''

    def _get_chartjs_minimal(self) -> str:
        """Get self-contained SVG chart implementation (no external deps)."""
        return '''
/* Self-contained SVG Chart Implementation - No External Dependencies */
const ChartRegistry = {
    instances: {},
    getChart: function(canvas) {
        return this.instances[canvas.id];
    }
};

function createChart(ctx, config) {
    const canvas = ctx.canvas;
    if (!canvas) {
        console.error('Chart: canvas not found');
        return { destroy: function() {} };
    }
    const container = canvas.parentElement;
    if (!container) {
        console.error('Chart: container not found');
        return { destroy: function() {} };
    }

    /* Get dimensions with fallbacks */
    let width = container.clientWidth || container.offsetWidth || 600;
    let height = container.clientHeight || container.offsetHeight || 300;

    /* Ensure minimum dimensions */
    if (width < 100) width = 600;
    if (height < 100) height = 300;

    /* Replace canvas with SVG */
    const svg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
    svg.setAttribute('width', '100%');
    svg.setAttribute('height', '100%');
    svg.setAttribute('viewBox', `0 0 ${width} ${height}`);
    svg.style.display = 'block';

    canvas.style.display = 'none';
    container.appendChild(svg);

    const colors = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6',
                    '#1abc9c', '#34495e', '#e67e22', '#95a5a6', '#d35400',
                    '#c0392b', '#27ae60', '#2980b9', '#8e44ad', '#f1c40f'];

    if (config.type === 'bar' && config.options?.indexAxis === 'y') {
        drawHorizontalBarChart(svg, config.data, width, height, colors);
    } else if (config.type === 'doughnut') {
        drawDoughnutChart(svg, config.data, width, height, colors);
    }

    const instance = {
        canvas: canvas,
        container: container,
        svg: svg,
        destroy: function() {
            if (this.svg && this.svg.parentElement) {
                this.svg.remove();
            }
            this.canvas.style.display = 'block';
            delete ChartRegistry.instances[this.canvas.id];
        }
    };

    ChartRegistry.instances[canvas.id] = instance;
    return instance;
}

function drawHorizontalBarChart(svg, data, width, height, colors) {
    const labels = data.labels || [];
    const values = data.datasets?.[0]?.data || [];
    const maxVal = Math.max(...values, 1);

    const margin = { top: 20, right: 20, bottom: 20, left: 120 };
    const chartW = width - margin.left - margin.right;
    const chartH = height - margin.top - margin.bottom;
    const barHeight = Math.min(25, (chartH / labels.length) * 0.7);
    const barGap = (chartH - barHeight * labels.length) / (labels.length + 1);

    labels.forEach((label, i) => {
        const y = margin.top + barGap * (i + 1) + barHeight * i;
        const barW = (values[i] / maxVal) * chartW;

        /* Label */
        const text = document.createElementNS('http://www.w3.org/2000/svg', 'text');
        text.setAttribute('x', margin.left - 5);
        text.setAttribute('y', y + barHeight / 2 + 4);
        text.setAttribute('text-anchor', 'end');
        text.setAttribute('font-size', '11');
        text.setAttribute('fill', '#333');
        text.textContent = label.length > 15 ? label.substring(0, 15) + '...' : label;
        svg.appendChild(text);

        /* Bar */
        const rect = document.createElementNS('http://www.w3.org/2000/svg', 'rect');
        rect.setAttribute('x', margin.left);
        rect.setAttribute('y', y);
        rect.setAttribute('width', Math.max(barW, 2));
        rect.setAttribute('height', barHeight);
        rect.setAttribute('fill', colors[i % colors.length]);
        rect.setAttribute('rx', '3');
        svg.appendChild(rect);

        /* Value */
        const valText = document.createElementNS('http://www.w3.org/2000/svg', 'text');
        valText.setAttribute('x', margin.left + barW + 5);
        valText.setAttribute('y', y + barHeight / 2 + 4);
        valText.setAttribute('font-size', '11');
        valText.setAttribute('fill', '#666');
        valText.textContent = values[i];
        svg.appendChild(valText);
    });
}

function drawDoughnutChart(svg, data, width, height, colors) {
    const labels = data.labels || [];
    const values = data.datasets?.[0]?.data || [];
    const total = values.reduce((a, b) => a + b, 0) || 1;

    const cx = width / 2;
    const cy = height / 2;
    const outerR = Math.min(width, height) / 2 - 40;
    const innerR = outerR * 0.5;

    let startAngle = -Math.PI / 2;

    values.forEach((val, i) => {
        const angle = (val / total) * Math.PI * 2;
        const endAngle = startAngle + angle;

        const x1 = cx + outerR * Math.cos(startAngle);
        const y1 = cy + outerR * Math.sin(startAngle);
        const x2 = cx + outerR * Math.cos(endAngle);
        const y2 = cy + outerR * Math.sin(endAngle);
        const x3 = cx + innerR * Math.cos(endAngle);
        const y3 = cy + innerR * Math.sin(endAngle);
        const x4 = cx + innerR * Math.cos(startAngle);
        const y4 = cy + innerR * Math.sin(startAngle);

        const largeArc = angle > Math.PI ? 1 : 0;

        const path = document.createElementNS('http://www.w3.org/2000/svg', 'path');
        path.setAttribute('d', `M ${x1} ${y1} A ${outerR} ${outerR} 0 ${largeArc} 1 ${x2} ${y2} L ${x3} ${y3} A ${innerR} ${innerR} 0 ${largeArc} 0 ${x4} ${y4} Z`);
        path.setAttribute('fill', colors[i % colors.length]);
        svg.appendChild(path);

        startAngle = endAngle;
    });

    /* Legend */
    const legendX = width - 120;
    let legendY = 20;
    labels.slice(0, 8).forEach((label, i) => {
        const rect = document.createElementNS('http://www.w3.org/2000/svg', 'rect');
        rect.setAttribute('x', legendX);
        rect.setAttribute('y', legendY);
        rect.setAttribute('width', '12');
        rect.setAttribute('height', '12');
        rect.setAttribute('fill', colors[i % colors.length]);
        svg.appendChild(rect);

        const text = document.createElementNS('http://www.w3.org/2000/svg', 'text');
        text.setAttribute('x', legendX + 16);
        text.setAttribute('y', legendY + 10);
        text.setAttribute('font-size', '10');
        text.setAttribute('fill', '#333');
        text.textContent = label.length > 12 ? label.substring(0, 12) + '..' : label;
        svg.appendChild(text);

        legendY += 18;
    });
}

/* Chart.js compatibility wrapper - must support 'new' keyword */
function ChartConstructor(ctx, config) {
    return createChart(ctx, config);
}
ChartConstructor.prototype.destroy = function() {};
window.Chart = ChartConstructor;
'''

    def _get_sheetjs_minimal(self) -> str:
        """Get SheetJS library embedded locally and CSV fallback."""
        try:
            sheetjs_code = resources.files('openssl_scanner.resources').joinpath(
                'xlsx.full.min.js'
            ).read_text(encoding='utf-8')
        except Exception as e:
            logger.warning(f"Failed to load SheetJS from resources: {e}")
            sheetjs_code = "/* SheetJS not available */"

        return '''
/* SheetJS Library (embedded locally) */
''' + sheetjs_code + '''
window.XLSX_LOADED = (typeof XLSX !== 'undefined');
console.log('SheetJS loaded:', window.XLSX_LOADED ? 'embedded' : 'unavailable');

/* CSV fallback export */
function exportAsCSV(data, filename) {
    const rows = [];

    /* Header */
    rows.push(['OpenSSL Dependency Analysis Report']);
    rows.push([]);

    /* Summary */
    rows.push(['Summary']);
    rows.push(['Components', data.summary?.total_components || data.summary?.files_with_openssl_deps || 0]);
    rows.push(['Total ELF Files', data.summary?.total_executables || data.summary?.total_elf_files || 0]);
    rows.push(['Unique Symbols', data.summary?.global_unique_symbols || data.summary?.unique_openssl_symbols || 0]);
    rows.push([]);

    /* Ranking */
    rows.push(['Ranking']);
    rows.push(['Rank', 'Component', 'Symbols', 'Percentage']);
    const total = data.summary?.global_unique_symbols || data.summary?.unique_openssl_symbols || 1;
    (data.ranking || []).forEach((r, i) => {
        const pct = ((r.unique_symbols_count / total) * 100).toFixed(1);
        rows.push([r.rank || i + 1, r.component, r.unique_symbols_count, pct + '%']);
    });
    rows.push([]);

    /* Symbols */
    rows.push(['All Symbols']);
    rows.push(['Symbol', 'Components', 'Category']);
    (window.symbolData || []).forEach(s => {
        rows.push([s.name, s.components.join('; '), s.category]);
    });

    /* Convert to CSV */
    const csv = rows.map(row =>
        row.map(cell => {
            const str = String(cell == null ? '' : cell);
            return str.includes(',') || str.includes('"') || str.includes('\\n')
                ? '"' + str.replace(/"/g, '""') + '"'
                : str;
        }).join(',')
    ).join('\\n');

    /* Add BOM for Excel UTF-8 compatibility */
    const bom = '\\uFEFF';
    const blob = new Blob([bom + csv], { type: 'text/csv;charset=utf-8' });
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = filename.replace('.xlsx', '.csv');
    a.click();
    URL.revokeObjectURL(url);
}
'''


class Exporter:
    """Main exporter class."""

    def __init__(self) -> None:
        self._excel_exporter = ExcelExporter()
        self._html_exporter = HTMLExporter()

    def export(self, report_path: str, output_path: str,
               format: Optional[str] = None) -> None:
        """
        Export report to specified format.

        Args:
            report_path: Path to JSON report file
            output_path: Output file path
            format: Output format ('xlsx' or 'html'), auto-detected if None
        """
        if not os.path.isfile(report_path):
            raise FileNotFoundError(f"Report file not found: {report_path}")

        if format is None:
            ext = os.path.splitext(output_path)[1].lower()
            format = ext[1:] if ext else 'xlsx'

        if format == 'xlsx':
            self._excel_exporter.export(report_path, output_path)
        elif format in ('html', 'htm'):
            self._html_exporter.export(report_path, output_path)
        else:
            raise ValueError(f"Unsupported format: {format}")
