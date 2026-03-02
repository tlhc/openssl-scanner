"""
HAP package scanning report, classification, and summary generation.

Extracted from __main__.py to separate business logic from CLI dispatch.
"""

import json
import logging
import os
import zipfile

logger = logging.getLogger(__name__)

_CONTAINER_EXTENSIONS = {'.zip', '.app'}
_INNER_PACKAGE_EXTENSIONS = {'.hap', '.har', '.hsp'}

_HAP_SUMMARY_COLUMNS = [
    ('pkg_name',        55, 'Package Name'),
    ('pkg_type',         8, 'Type'),
    ('version',         12, 'Version'),
    ('abi',             15, 'ABI'),
    ('so_files',        10, '.so Files'),
    ('openssl_usage',   22, 'OpenSSL Usage'),
    ('detection',       14, 'Detection'),
    ('static_syms',     14, 'Static Symbols'),
    ('dynamic_syms',    14, 'Dynamic Symbols'),
    ('dlopen_syms',     14, 'dlopen Symbols'),
    ('total_syms',      12, 'Total Symbols'),
    ('top_category',    18, 'Top Category'),
    ('ssl_core',        10, 'ssl_core'),
    ('crypto_evp',      10, 'crypto_evp'),
    ('crypto_x509',     10, 'crypto_x509'),
    ('crypto_ec',       10, 'crypto_ec'),
    ('crypto_hash',     10, 'crypto_hash'),
    ('crypto_sm',       10, 'crypto_sm'),
    ('crypto_bio',      10, 'crypto_bio'),
    ('other_cats',      10, 'Other Cats'),
    ('dlopen_libs',     30, 'dlopen Libs'),
    ('custom_match',    22, 'Custom Match'),
]

_HAP_HIGHLIGHT_CATS = [
    'ssl_core', 'crypto_evp', 'crypto_x509',
    'crypto_ec', 'crypto_hash', 'crypto_sm', 'crypto_bio',
]


class PkgEntry:
    """A scannable package reference -- standalone file or entry inside a container."""
    __slots__ = ('path', 'container', 'zip_entry', 'display_name')

    def __init__(self, path, container=None, zip_entry=None, display_name=None):
        self.path = path
        self.container = container
        self.zip_entry = zip_entry
        self.display_name = display_name or os.path.basename(path or '')


def plan_packages(packages, _logger=None):
    """Enumerate scannable packages without extracting containers.

    Peeks into ZIP/APP containers to list inner HAPs, but does NOT
    extract them. Returns a flat list of PkgEntry descriptors.
    Actual extraction happens one-at-a-time in the scan loop.
    """
    if _logger is None:
        _logger = logger
    plan = []
    for pkg in packages:
        ext = os.path.splitext(pkg)[1].lower()
        if ext not in _CONTAINER_EXTENSIONS:
            plan.append(PkgEntry(path=pkg))
            continue

        try:
            with zipfile.ZipFile(pkg, 'r') as zf:
                nested = [
                    e for e in zf.namelist()
                    if os.path.splitext(e)[1].lower() in _INNER_PACKAGE_EXTENSIONS
                    and not e.startswith('__MACOSX')
                ]
        except (zipfile.BadZipFile, OSError) as e:
            _logger.warning("Cannot open %s: %s, keeping as-is", pkg, e)
            plan.append(PkgEntry(path=pkg))
            continue

        if not nested:
            plan.append(PkgEntry(path=pkg))
            continue

        container_stem = os.path.splitext(os.path.basename(pkg))[0]
        seen_names = set()
        for entry in nested:
            inner_name = os.path.basename(entry)
            if not inner_name:
                continue
            inner_base, inner_ext = os.path.splitext(inner_name)
            safe_name = f"{container_stem}_{inner_base}{inner_ext}"
            if safe_name in seen_names:
                counter = 2
                while f"{container_stem}_{inner_base}_{counter}{inner_ext}" in seen_names:
                    counter += 1
                safe_name = f"{container_stem}_{inner_base}_{counter}{inner_ext}"
            seen_names.add(safe_name)
            plan.append(PkgEntry(
                path=None,
                container=pkg,
                zip_entry=entry,
                display_name=safe_name,
            ))
        _logger.info("Container %s -> %d inner packages",
                      os.path.basename(pkg), len(nested))

    return plan


def extract_pkg_entry(entry, _logger=None):
    """Extract a single PkgEntry just-in-time. Returns (actual_path, tmp_to_cleanup).

    For standalone packages, returns (path, None).
    For container entries, extracts to a temp file and returns (tmp_path, tmp_path).
    """
    if _logger is None:
        _logger = logger
    if entry.container is None:
        return entry.path, None

    import tempfile as _tempfile
    import shutil as _shutil
    fd, tmp_path = _tempfile.mkstemp(suffix=os.path.splitext(entry.display_name)[1],
                                      prefix='hap_')
    try:
        with zipfile.ZipFile(entry.container, 'r') as zf:
            with zf.open(entry.zip_entry) as src, os.fdopen(fd, 'wb') as dst:
                _shutil.copyfileobj(src, dst)
                fd = -1
    except Exception:
        if fd >= 0:
            os.close(fd)
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise
    _logger.debug("Extracted %s:%s -> %s", os.path.basename(entry.container),
                   entry.zip_entry, tmp_path)
    return tmp_path, tmp_path


def merge_hap_results(results):
    """Merge multiple ScanResult objects into a single batch result.

    Combines files_detail, symbols_by_file, symbols_by_category,
    all_unique_symbols, and all counter fields from every result.
    """
    from .scanner import ScanResult

    base = results[0]
    merged = ScanResult(
        target=base.target,
        scan_time=base.scan_time,
        tool_version=base.tool_version,
        arch=base.arch,
        report_type='package_batch',
    )

    all_files_detail = []
    merged_by_file = {}
    merged_by_category = {}
    unique_syms = set()
    total_scanned = 0
    total_elf = 0
    with_openssl = 0
    with_static = 0
    with_dlopen = 0
    merged_dlsym_by_file = {}
    dlsym_unique = set()
    dlopen_libs = set()
    ossl_libs = set()
    all_errors = []

    for r in results:
        all_files_detail.extend(r.files_detail)
        merged_by_file.update(r.symbols_by_file)
        for cat, syms in r.symbols_by_category.items():
            if cat not in merged_by_category:
                merged_by_category[cat] = set()
            merged_by_category[cat].update(syms)
        unique_syms.update(r.all_unique_symbols)
        total_scanned += r.total_files_scanned
        total_elf += r.total_elf_files
        with_openssl += r.files_with_openssl
        with_static += r.files_with_static_openssl
        with_dlopen += r.files_with_dlopen
        merged_dlsym_by_file.update(r.dlsym_symbols_by_file)
        dlsym_unique.update(r.all_dlsym_symbols)
        dlopen_libs.update(r.dlopen_libs_detected)
        ossl_libs.update(r.openssl_libs_found)
        all_errors.extend(r.errors)

    merged.files_detail = all_files_detail
    merged.symbols_by_file = merged_by_file
    merged.symbols_by_category = {
        cat: sorted(syms) for cat, syms in merged_by_category.items()
    }
    merged.all_unique_symbols = sorted(unique_syms)
    merged.total_files_scanned = total_scanned
    merged.total_elf_files = total_elf
    merged.files_with_openssl = with_openssl
    merged.files_with_static_openssl = with_static
    merged.files_with_dlopen = with_dlopen
    merged.dlsym_symbols_by_file = merged_dlsym_by_file
    merged.all_dlsym_symbols = sorted(dlsym_unique)
    merged.dlopen_libs_detected = sorted(dlopen_libs)
    merged.openssl_libs_found = sorted(ossl_libs)
    merged.errors = all_errors

    merged.package_info = results[0].package_info
    merged.package_info['batch'] = [
        r.package_info for r in results[1:]
    ]

    return merged


def resolve_hap_output_names(names, output_dir, ext):
    """Map package display names to unique output file paths."""
    used = set()
    result = {}
    for pkg in names:
        base = os.path.splitext(os.path.basename(pkg))[0]
        candidate = base
        counter = 2
        while candidate in used:
            candidate = f"{base}_{counter}"
            counter += 1
        used.add(candidate)
        result[pkg] = os.path.join(output_dir, candidate + ext)
    return result


def hap_write_single_report(result, pkg_path, out_path, reporter, json_only):
    """Write a single package report immediately after scanning."""
    json_report = reporter.generate_json(result)

    json_path = os.path.splitext(out_path)[0] + '.json'
    with open(json_path, 'w', encoding='utf-8') as f:
        f.write(json_report)

    if not json_only:
        from .exporter import Exporter
        Exporter().export(json_path, out_path)


def collect_bundled_names(removed_libs, extract_result):
    """Collect bundled OpenSSL library basenames from multiple sources."""
    names = list(removed_libs)
    if extract_result.openssl_lib:
        bn = os.path.basename(extract_result.openssl_lib)
        if bn not in names:
            names.append(bn)
    if extract_result.openssl_ssl:
        bn = os.path.basename(extract_result.openssl_ssl)
        if bn not in names:
            names.append(bn)
    return names


def detect_static_providers(scan_result):
    """Identify .so files with statically linked OpenSSL and their consumers.

    Multi-ABI packages may contain the same library under arm64-v8a/ and
    x86_64/ etc.  We deduplicate by basename, keeping the entry with the
    highest symbol count.

    Returns (bundled_str, providers_list):
        bundled_str:    'Yes (static, shared)' | 'Yes (static)' | None
        providers_list: [{file, confidence, symbols, consumers}, ...]
    """
    best = {}
    for fr in scan_result.files_detail:
        if not fr.static_openssl:
            continue
        if fr.static_openssl_confidence not in ('high', 'medium'):
            continue
        basename = os.path.basename(fr.path)
        sym_count = len(fr.openssl_symbols)
        prev = best.get(basename)
        if prev is None or sym_count > prev[1]:
            best[basename] = (fr, sym_count)

    if not best:
        return None, []

    providers = []
    for basename, (fr, sym_count) in best.items():
        consumers = sorted({os.path.basename(g.path)
                            for g in scan_result.files_detail
                            if basename in g.direct_deps
                            and os.path.basename(g.path) != basename})
        providers.append({
            'file': basename,
            'confidence': fr.static_openssl_confidence,
            'symbols': sym_count,
            'consumers': consumers,
        })

    has_shared = any(p['consumers'] for p in providers)
    bundled_str = 'Yes (static, shared)' if has_shared else 'Yes (static)'
    return bundled_str, providers


def _lib_stem(name):
    """Extract library stem: 'libcrypto.so.3' -> 'libcrypto'."""
    bn = os.path.basename(name)
    return bn.split('.so')[0] if '.so' in bn else bn


def _dlopen_targets_resolved(dlopen_libs, bundled_basenames, patterns):
    """Check if ALL OpenSSL dlopen targets are resolved by bundled libs.

    Returns False (unresolved) when dlopen_libs is empty (unknown target)
    or any OpenSSL target lacks a matching bundled lib.
    """
    if not dlopen_libs or not bundled_basenames:
        return False
    ossl_targets = [lib for lib in dlopen_libs
                    if any(os.path.basename(lib).lower().startswith(p)
                           for p in patterns)]
    if not ossl_targets:
        return False
    bundled_stems = {_lib_stem(b).lower() for b in bundled_basenames}
    return all(_lib_stem(t).lower() in bundled_stems for t in ossl_targets)


def _dt_needed_resolved(openssl_libs, bundled_basenames, patterns):
    """Check if DT_NEEDED OpenSSL libs are all resolved by bundled libs.

    Returns False when no OpenSSL lib in DT_NEEDED or when any needed
    OpenSSL lib lacks a matching bundled lib.
    """
    needed = [lib for lib in openssl_libs
              if any(os.path.basename(lib).lower().startswith(p)
                     for p in patterns)]
    if not needed or not bundled_basenames:
        return False
    bundled_stems = {_lib_stem(b).lower() for b in bundled_basenames}
    return all(_lib_stem(lib).lower() in bundled_stems for lib in needed)


def classify_hap_detection(result):
    """Classify OpenSSL usage and return deduped symbol sets.

    Returns ``(method, static_syms, dynamic_syms, dlopen_syms, ossl_type)``
    where the middle three are **sets** of unique symbol names.

    Per-library resolution: each library's OpenSSL dependency must be
    satisfied within the HAP (static linking or bundled .so).  If any
    library has an unresolved external dependency the HAP is System-Link.

    Static providers (high/medium confidence) are treated as bundled libs
    for dependency resolution purposes.
    """
    from .constants import OPENSSL_LIBRARY_PATTERNS

    pi = result.package_info or {}
    bundled_basenames = set(pi.get('bundled_openssl_files', []))

    for fr in result.files_detail:
        if (fr.static_openssl
                and fr.static_openssl_confidence in ('high', 'medium')):
            bundled_basenames.add(os.path.basename(fr.path))

    static_syms = set()
    dynamic_syms = set()
    dlopen_syms = set()
    has_dynamic = has_static = has_dlopen = False
    has_unresolved_external = False

    def _has_ossl_dlopen_evidence(fr):
        return fr.dlsym_symbols or any(
            any(os.path.basename(lib).lower().startswith(p)
                for p in OPENSSL_LIBRARY_PATTERNS)
            for lib in fr.dlopen_libs)

    for fr in result.files_detail:
        if fr.static_openssl:
            has_static = True
            if fr.uses_dlopen and _has_ossl_dlopen_evidence(fr):
                has_dlopen = True
                dlopen_syms.update(fr.dlsym_symbols)
                static_only = set(fr.openssl_symbols) - set(fr.dlsym_symbols)
                static_syms.update(static_only)
                if not _dlopen_targets_resolved(
                        fr.dlopen_libs, bundled_basenames,
                        OPENSSL_LIBRARY_PATTERNS):
                    has_unresolved_external = True
            else:
                static_syms.update(fr.openssl_symbols)
                if fr.openssl_libs and not _dt_needed_resolved(
                        fr.openssl_libs, bundled_basenames,
                        OPENSSL_LIBRARY_PATTERNS):
                    has_unresolved_external = True
        elif fr.uses_dlopen:
            if not fr.dlsym_symbols and not fr.openssl_symbols \
                    and not _has_ossl_dlopen_evidence(fr):
                continue
            has_dlopen = True
            dlopen_syms.update(fr.dlsym_symbols)
            non_dlsym = set(fr.openssl_symbols) - set(fr.dlsym_symbols)
            if non_dlsym:
                has_dynamic = True
                dynamic_syms.update(non_dlsym)
            if not _dlopen_targets_resolved(
                    fr.dlopen_libs, bundled_basenames,
                    OPENSSL_LIBRARY_PATTERNS):
                has_unresolved_external = True
        elif fr.openssl_symbols:
            has_dynamic = True
            dynamic_syms.update(fr.openssl_symbols)
            if not _dt_needed_resolved(
                    fr.openssl_libs, bundled_basenames,
                    OPENSSL_LIBRARY_PATTERNS):
                has_unresolved_external = True

    methods = []
    if has_dynamic:
        methods.append('Dynamic')
    if has_static:
        methods.append('Static')
    if has_dlopen:
        methods.append('dlopen')

    if not methods:
        method = 'None'
    elif len(methods) == 1:
        method = methods[0]
    else:
        method = 'Mixed'

    total = static_syms | dynamic_syms | dlopen_syms
    if not total and not has_dlopen:
        ossl_type = 'No-OpenSSL'
    elif not has_unresolved_external:
        ossl_type = 'Self-Contained'
    else:
        ossl_type = 'System-Link'

    return method, static_syms, dynamic_syms, dlopen_syms, ossl_type


def build_hap_summary_row(result, pkg_path, method, s_syms, d_syms, dl_syms,
                           ossl_type, custom_result=None):
    """Build a dict of column values for one package.

    All symbol counts derive from the deduped sets returned by
    ``classify_hap_detection`` so that Static + Dynamic + dlopen
    decomposition is internally consistent.
    """
    pi = result.package_info or {}
    abi = pi.get('scanned_abi', '')
    if isinstance(abi, list):
        abi = ', '.join(abi)

    cat_counts = {}
    for cat, syms in result.symbols_by_category.items():
        cat_counts[cat] = len(syms)

    highlight_sum = sum(cat_counts.get(c, 0) for c in _HAP_HIGHLIGHT_CATS)
    other_cats = sum(cat_counts.values()) - highlight_sum

    top_cat = ''
    if cat_counts:
        top_cat = max(cat_counts, key=cat_counts.get)

    bundled_raw = pi.get('bundled_openssl', False)
    if isinstance(bundled_raw, str) and bundled_raw.startswith('Yes'):
        detail = bundled_raw.removeprefix('Yes').strip()
        openssl_usage = f'Bundled {detail}' if detail else 'Bundled'
    elif bundled_raw is True:
        openssl_usage = 'Bundled'
    elif ossl_type == 'System-Link':
        openssl_usage = 'System-Link'
    else:
        openssl_usage = 'None'

    bundle = pi.get('bundle_name', '')
    module = pi.get('module_name', '')
    source = os.path.splitext(os.path.basename(pkg_path))[0]
    if bundle and module:
        pkg_name = f'{bundle}/{module} ({source})'
    elif bundle:
        pkg_name = f'{bundle} ({source})'
    else:
        pkg_name = source

    return {
        'pkg_name': pkg_name,
        'pkg_type': pi.get('package_type', ''),
        'version': pi.get('version_name', ''),
        'abi': abi,
        'so_files': pi.get('native_libs_count', 0),
        'openssl_usage': openssl_usage,
        'detection': method,
        'static_syms': len(s_syms),
        'dynamic_syms': len(d_syms),
        'dlopen_syms': len(dl_syms),
        'total_syms': len(s_syms | d_syms | dl_syms),
        'top_category': top_cat,
        'other_cats': other_cats,
        'dlopen_libs': ', '.join(result.dlopen_libs_detected),
        'custom_match': (custom_result.summary_text() if custom_result
                         else pi.get('custom_match', '')),
    }


def generate_hap_summary(all_results, scanned_packages, output_dir,
                          custom_results=None):
    """Generate Package Summary XLSX for HAP batch scan.

    Returns the output path on success, None on failure.
    """
    from . import _vendor  # noqa: F401
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Package Summary"

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="E8F4FC", end_color="E8F4FC", fill_type="solid"
    )
    total_font = Font(bold=True)
    total_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    col_keys = [c[0] for c in _HAP_SUMMARY_COLUMNS]
    for col_idx, (_, width, title) in enumerate(_HAP_SUMMARY_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    rows = []
    all_static = set()
    all_dynamic = set()
    all_dlopen = set()
    for idx, (result, pkg_path) in enumerate(zip(all_results, scanned_packages)):
        method, s_syms, d_syms, dl_syms, ossl_type = \
            classify_hap_detection(result)
        cr = custom_results[idx] if custom_results and idx < len(custom_results) else None
        row = build_hap_summary_row(
            result, pkg_path, method, s_syms, d_syms, dl_syms, ossl_type,
            custom_result=cr)
        for cat in _HAP_HIGHLIGHT_CATS:
            row[cat] = len(result.symbols_by_category.get(cat, []))
        rows.append(row)
        all_static |= s_syms
        all_dynamic |= d_syms
        all_dlopen |= dl_syms

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(col_keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ''))

    total_row = len(rows) + 2

    cat_union = {}
    for r in all_results:
        for cat, syms in r.symbols_by_category.items():
            cat_union.setdefault(cat, set()).update(syms)

    highlight_union = sum(len(cat_union.get(c, set())) for c in _HAP_HIGHLIGHT_CATS)
    other_union = sum(len(v) for v in cat_union.values()) - highlight_union

    top_cat_total = ''
    if cat_union:
        top_cat_total = max(cat_union, key=lambda c: len(cat_union[c]))

    usage_counts = {}
    for r in rows:
        u = r.get('openssl_usage', '')
        if u:
            usage_counts[u] = usage_counts.get(u, 0) + 1
    usage_summary = ', '.join(f'{v} {k}' for k, v in sorted(usage_counts.items()))

    total_data = {
        'pkg_name': 'TOTAL',
        'pkg_type': '',
        'version': '',
        'abi': '',
        'so_files': sum(r['so_files'] for r in rows),
        'openssl_usage': usage_summary,
        'detection': '',
        'static_syms': sum(r.get('static_syms', 0) for r in rows),
        'dynamic_syms': sum(r.get('dynamic_syms', 0) for r in rows),
        'dlopen_syms': sum(r.get('dlopen_syms', 0) for r in rows),
        'total_syms': sum(r.get('total_syms', 0) for r in rows),
        'top_category': top_cat_total,
        'other_cats': sum(r.get('other_cats', 0) for r in rows),
        'dlopen_libs': '',
        'custom_match': '',
    }
    for cat in _HAP_HIGHLIGHT_CATS:
        total_data[cat] = sum(r.get(cat, 0) for r in rows)

    for col_idx, key in enumerate(col_keys, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=total_data.get(key, ''))
        cell.font = total_font
        cell.fill = total_fill

    last_col = get_column_letter(len(_HAP_SUMMARY_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{total_row - 1}"

    summary_path = os.path.join(output_dir, 'summary.xlsx')
    wb.save(summary_path)
    logger.info("HAP summary saved to: %s", summary_path)
    return summary_path


def load_scan_result_from_json(json_path):
    """Reconstruct a ScanResult from a per-package JSON report.

    Returns (ScanResult, pkg_path) or (None, None) on error.
    """
    from .scanner import ScanResult, FileResult

    try:
        with open(json_path, 'r') as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError, UnicodeDecodeError) as e:
        logger.warning("Skipping %s: %s", json_path, e)
        return None, None

    meta = data.get('meta', {})
    if meta.get('report_type') != 'package':
        return None, None

    summary = data.get('summary', {})
    ossl_syms = data.get('openssl_symbols', {})

    files_detail = []
    for fd in data.get('files_detail', []):
        dlopen_det = fd.get('dlopen_detection', {})
        fr = FileResult(
            path=fd.get('path', ''),
            file_type=fd.get('type', ''),
            arch=fd.get('arch', ''),
            direct_deps=fd.get('direct_deps', []),
            openssl_direct=fd.get('openssl_deps', {}).get('direct', False),
            openssl_transitive=fd.get('openssl_deps', {}).get('transitive', False),
            openssl_libs=fd.get('openssl_deps', {}).get('libs', []),
            openssl_symbols=fd.get('openssl_symbols_used', []),
            static_openssl=fd.get('static_openssl', False),
            static_openssl_confidence=str(
                fd.get('static_openssl_confidence') or ''),
            static_openssl_confidence_reason=fd.get(
                'static_openssl_confidence_reason', ''),
            static_ssl_library=fd.get('static_ssl_library', ''),
            uses_dlopen=dlopen_det.get('uses_dlopen', False),
            dlsym_symbols=dlopen_det.get('dlopen_symbols', []),
            dlopen_libs=dlopen_det.get('dlopen_libs', []),
            dlopen_confidence=dlopen_det.get('confidence', 'high'),
        )
        files_detail.append(fr)

    by_cat = {}
    for cat, info in ossl_syms.get('by_category', {}).items():
        by_cat[cat] = info.get('symbols', []) if isinstance(info, dict) else info

    by_file = {}
    for path, info in ossl_syms.get('by_file', {}).items():
        by_file[path] = info.get('symbols', []) if isinstance(info, dict) else info

    dlopen_analysis = data.get('dlopen_analysis', {})

    result = ScanResult(
        target=meta.get('scan_root', ''),
        scan_time=meta.get('scan_time', ''),
        tool_version=meta.get('tool_version', ''),
        arch=meta.get('target_arch', ''),
        report_type=meta.get('report_type', 'package'),
        total_files_scanned=summary.get('total_files_scanned', 0),
        total_elf_files=summary.get('total_elf_files', 0),
        files_with_openssl=summary.get('files_with_openssl_deps', 0),
        openssl_libs_found=summary.get('openssl_libs_found', []),
        files_detail=files_detail,
        symbols_by_file=by_file,
        symbols_by_category=by_cat,
        all_unique_symbols=ossl_syms.get('all_unique', []),
        files_with_static_openssl=summary.get('files_with_static_openssl', 0),
        files_with_dlopen=summary.get('files_with_dlopen', 0),
        all_dlsym_symbols=dlopen_analysis.get('all_dlopen_symbols', []),
        dlopen_libs_detected=summary.get('dlopen_libs_detected', []),
    )

    result.package_info = meta.get('package')

    pkg_path = ''
    if result.package_info:
        pkg_path = result.package_info.get('package_path', '')
    if not pkg_path:
        pkg_path = json_path

    return result, pkg_path
