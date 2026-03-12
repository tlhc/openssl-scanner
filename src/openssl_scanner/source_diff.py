"""
Diff engine for comparing two source scan JSON reports.

Compares call sites by identity key (file_path, caller_function, ossl_symbol)
and produces structured deltas at call-site, symbol, file, and metric layers.
"""

import json
import logging
import os
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional, Set, Tuple, TYPE_CHECKING

if TYPE_CHECKING:
    from .hitls_compat import HiTLSCompat

logger = logging.getLogger(__name__)

XLSX_MAX_ROW = 1048576


class DiffStatus(str, Enum):
    ADDED = "added"
    REMOVED = "removed"
    CHANGED = "changed"
    MOVED = "moved"
    UNCHANGED = "unchanged"


STATUS_SORT_ORDER: Dict[DiffStatus, int] = {
    DiffStatus.ADDED: 0,
    DiffStatus.REMOVED: 1,
    DiffStatus.CHANGED: 2,
    DiffStatus.MOVED: 3,
    DiffStatus.UNCHANGED: 4,
}


@dataclass
class MetricDelta:
    name: str
    old_value: int
    new_value: int
    delta: int


@dataclass
class ArgDelta:
    status: DiffStatus
    old_line: Optional[int] = None
    new_line: Optional[int] = None
    old_args: str = ""
    new_args: str = ""


@dataclass
class CallSiteDelta:
    status: DiffStatus
    identity_key: Tuple[str, str, str]
    old_count: int
    new_count: int
    old_lines: List[int] = field(default_factory=list)
    new_lines: List[int] = field(default_factory=list)
    category: str = ""
    arg_deltas: List[ArgDelta] = field(default_factory=list)


@dataclass
class SymbolDelta:
    status: DiffStatus
    symbol: str
    category: str
    old_count: int
    new_count: int


@dataclass
class FileDelta:
    status: DiffStatus
    file_path: str
    old_call_count: int
    new_call_count: int
    old_symbols: int = 0
    new_symbols: int = 0
    added_symbols: List[str] = field(default_factory=list)
    removed_symbols: List[str] = field(default_factory=list)


@dataclass
class ProjectDelta:
    project: str
    metrics: List[MetricDelta] = field(default_factory=list)
    call_site_delta: List[CallSiteDelta] = field(default_factory=list)
    symbol_delta: List[SymbolDelta] = field(default_factory=list)
    file_delta: List[FileDelta] = field(default_factory=list)
    has_call_site_changes: bool = False


@dataclass
class DiffResult:
    old_label: str
    new_label: str
    projects: List[ProjectDelta] = field(default_factory=list)
    old_scan_time: str = ""
    new_scan_time: str = ""
    is_combo: bool = False

    def is_empty(self) -> bool:
        for proj in self.projects:
            if proj.has_call_site_changes:
                return False
            for m in proj.metrics:
                if m.delta != 0:
                    return False
            for csd in proj.call_site_delta:
                if csd.status != DiffStatus.UNCHANGED:
                    return False
            for sd in proj.symbol_delta:
                if sd.status != DiffStatus.UNCHANGED:
                    return False
            for fd in proj.file_delta:
                if fd.status != DiffStatus.UNCHANGED:
                    return False
        return True


def _strip_prefix(file_path: str, prefix: str) -> str:
    """Strip a leading prefix from a file path."""
    if not prefix:
        return file_path
    prefix = prefix.rstrip("/") + "/"
    if file_path.startswith(prefix):
        return file_path[len(prefix):]
    return file_path


def _normalize_path(file_path: str, target: str) -> str:
    """Make file_path relative to target directory."""
    if not target:
        return file_path
    target = target.rstrip("/") + "/"
    if file_path.startswith(target):
        return file_path[len(target):]
    return file_path


def _normalize_call_sites(
    call_sites: List[Dict[str, Any]],
    target: str,
    prefix: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Normalize file_path in each call site dict.

    When prefix is given, strip it from both file_path and target before
    making paths target-relative.
    """
    effective_target = target
    if prefix:
        effective_target = _strip_prefix(target, prefix)

    result = []
    for cs in call_sites:
        cs = dict(cs)
        fp = cs.get("file_path", "")
        if prefix:
            fp = _strip_prefix(fp, prefix)
        fp = _normalize_path(fp, effective_target)
        cs["file_path"] = fp
        result.append(cs)
    return result


def load_report(
    path: str,
    prefix: Optional[str] = None,
) -> Dict[str, Any]:
    """Load a source scan or combo scan JSON report.

    Returns a normalized dict with keys:
        report_type: 'source_scan' or 'combo_scan'
        projects: list of dicts, each with 'project', 'call_sites', 'summary'

    For source_scan reports, a single-element projects list is returned
    with project name derived from the target directory basename.
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    meta = data.get("meta", {})
    report_type = meta.get("report_type", "source_scan")

    if report_type not in ("source_scan", "combo_scan"):
        raise ValueError(
            f"Unsupported report_type '{report_type}' in {path}; "
            "expected 'source_scan' or 'combo_scan'"
        )

    scan_time = meta.get("scan_time", meta.get("merge_time", ""))

    if report_type == "combo_scan":
        projects = []
        for proj in data.get("projects", []):
            target = proj.get("target", "")
            call_sites = _normalize_call_sites(
                proj.get("call_sites", []), target, prefix
            )
            projects.append({
                "project": proj.get("project", ""),
                "call_sites": call_sites,
                "summary": {
                    "total_files_scanned": proj.get("total_files_scanned", 0),
                    "files_with_calls": proj.get("files_with_calls", 0),
                    "total_call_sites": proj.get("total_call_sites", 0),
                },
            })
        return {
            "report_type": report_type,
            "projects": projects,
            "scan_time": scan_time,
        }

    target = meta.get("target", "")
    summary = data.get("summary", {})
    call_sites = _normalize_call_sites(
        data.get("call_sites", []), target, prefix
    )
    project_name = os.path.basename(target.rstrip("/")) if target else "unknown"
    return {
        "report_type": "source_scan",
        "projects": [{
            "project": project_name,
            "call_sites": call_sites,
            "summary": {
                "total_files_scanned": summary.get("total_files_scanned", 0),
                "files_with_calls": summary.get("files_with_calls", 0),
                "total_call_sites": summary.get("total_call_sites", 0),
            },
        }],
        "scan_time": scan_time,
    }


IdentityKey = Tuple[str, str, str]


def _build_identity_map(
    call_sites: List[Dict[str, Any]],
) -> Dict[IdentityKey, List[Dict[str, Any]]]:
    """Group call sites by identity key (file_path, caller_function, ossl_symbol)."""
    result: Dict[IdentityKey, List[Dict[str, Any]]] = {}
    for cs in call_sites:
        key = (
            cs.get("file_path", ""),
            cs.get("caller_function", ""),
            cs.get("ossl_symbol", ""),
        )
        result.setdefault(key, []).append(cs)
    return result


def _match_call_sites_within_group(
    old_entries: List[Dict[str, Any]],
    new_entries: List[Dict[str, Any]],
) -> List[ArgDelta]:
    """Match individual call sites within an identity group using 4-phase greedy.

    Phase 1: Exact match   (line == line AND args == args)  -> UNCHANGED
    Phase 2: Same args     (args == args, closest line)     -> MOVED
    Phase 3: Same line     (line == line AND args != args)  -> CHANGED
    Phase 4: Nearest line  (remaining, closest |line| pair) -> CHANGED
    Residual: unmatched old -> REMOVED, unmatched new -> ADDED

    Phase 2 runs before Phase 3 to prefer content identity (same args = MOVED)
    over location identity (same line = CHANGED). This correctly classifies
    code shifts (all lines move by N, args unchanged) as MOVED rather than
    CHANGED.
    """
    old_raw_lines = [e.get("line_number") for e in old_entries]
    new_raw_lines = [e.get("line_number") for e in new_entries]
    old_items = [
        (ln or 0, e.get("call_args") or "")
        for ln, e in zip(old_raw_lines, old_entries)
    ]
    new_items = [
        (ln or 0, e.get("call_args") or "")
        for ln, e in zip(new_raw_lines, new_entries)
    ]

    matched_old: set = set()
    matched_new: set = set()
    result: List[ArgDelta] = []

    for oi, (ol, oa) in enumerate(old_items):
        for ni, (nl, na) in enumerate(new_items):
            if ni in matched_new:
                continue
            if ol == nl and oa == na:
                matched_old.add(oi)
                matched_new.add(ni)
                result.append(ArgDelta(
                    status=DiffStatus.UNCHANGED,
                    old_line=old_raw_lines[oi], new_line=new_raw_lines[ni],
                    old_args=oa, new_args=na,
                ))
                break

    for oi, (ol, oa) in enumerate(old_items):
        if oi in matched_old:
            continue
        best_ni = None
        best_dist = None
        for ni, (nl, na) in enumerate(new_items):
            if ni in matched_new:
                continue
            if oa == na:
                dist = abs(ol - nl)
                if best_dist is None or dist < best_dist:
                    best_ni = ni
                    best_dist = dist
        if best_ni is not None:
            nl, na = new_items[best_ni]
            matched_old.add(oi)
            matched_new.add(best_ni)
            result.append(ArgDelta(
                status=DiffStatus.MOVED,
                old_line=old_raw_lines[oi], new_line=new_raw_lines[best_ni],
                old_args=oa, new_args=na,
            ))

    for oi, (ol, oa) in enumerate(old_items):
        if oi in matched_old:
            continue
        for ni, (nl, na) in enumerate(new_items):
            if ni in matched_new:
                continue
            if ol == nl:
                matched_old.add(oi)
                matched_new.add(ni)
                result.append(ArgDelta(
                    status=DiffStatus.CHANGED,
                    old_line=old_raw_lines[oi], new_line=new_raw_lines[ni],
                    old_args=oa, new_args=na,
                ))
                break

    for oi, (ol, oa) in enumerate(old_items):
        if oi in matched_old:
            continue
        best_ni = None
        best_dist = None
        for ni, (nl, na) in enumerate(new_items):
            if ni in matched_new:
                continue
            dist = abs(ol - nl)
            if best_dist is None or dist < best_dist:
                best_ni = ni
                best_dist = dist
        if best_ni is not None:
            nl, na = new_items[best_ni]
            matched_old.add(oi)
            matched_new.add(best_ni)
            result.append(ArgDelta(
                status=DiffStatus.CHANGED,
                old_line=old_raw_lines[oi], new_line=new_raw_lines[best_ni],
                old_args=oa, new_args=na,
            ))

    for oi, (ol, oa) in enumerate(old_items):
        if oi in matched_old:
            continue
        result.append(ArgDelta(
            status=DiffStatus.REMOVED,
            old_line=old_raw_lines[oi], old_args=oa,
        ))

    for ni, (nl, na) in enumerate(new_items):
        if ni in matched_new:
            continue
        result.append(ArgDelta(
            status=DiffStatus.ADDED,
            new_line=new_raw_lines[ni], new_args=na,
        ))

    result.sort(key=lambda ad: (
        ad.new_line if ad.new_line is not None else (ad.old_line or 0),
        ad.old_line if ad.old_line is not None else 0,
    ))
    return result


def _derive_group_status(arg_deltas: List[ArgDelta]) -> DiffStatus:
    """Derive group-level status from individual arg deltas."""
    if not arg_deltas:
        return DiffStatus.UNCHANGED
    statuses = {ad.status for ad in arg_deltas}
    if statuses == {DiffStatus.UNCHANGED}:
        return DiffStatus.UNCHANGED
    if statuses <= {DiffStatus.UNCHANGED, DiffStatus.MOVED}:
        return DiffStatus.MOVED
    if DiffStatus.ADDED in statuses and len(statuses) == 1:
        return DiffStatus.ADDED
    if DiffStatus.REMOVED in statuses and len(statuses) == 1:
        return DiffStatus.REMOVED
    return DiffStatus.CHANGED


def _compute_call_site_delta(
    old_map: Dict[IdentityKey, List[Dict[str, Any]]],
    new_map: Dict[IdentityKey, List[Dict[str, Any]]],
    include_unchanged: bool = False,
) -> List[CallSiteDelta]:
    """Compute per-identity-key call site deltas."""
    all_keys = set(old_map.keys()) | set(new_map.keys())
    deltas: List[CallSiteDelta] = []

    for key in sorted(all_keys):
        old_entries = old_map.get(key, [])
        new_entries = new_map.get(key, [])
        old_count = len(old_entries)
        new_count = len(new_entries)
        old_lines = sorted(e.get("line_number") or 0 for e in old_entries)
        new_lines = sorted(e.get("line_number") or 0 for e in new_entries)
        category = ""
        if new_entries:
            category = new_entries[0].get("category", "")
        elif old_entries:
            category = old_entries[0].get("category", "")

        arg_deltas = _match_call_sites_within_group(old_entries, new_entries)
        status = _derive_group_status(arg_deltas)

        if status == DiffStatus.UNCHANGED and not include_unchanged:
            continue

        deltas.append(CallSiteDelta(
            status=status,
            identity_key=key,
            old_count=old_count,
            new_count=new_count,
            old_lines=old_lines,
            new_lines=new_lines,
            category=category,
            arg_deltas=arg_deltas,
        ))

    deltas.sort(key=lambda d: (STATUS_SORT_ORDER.get(d.status, 99), d.identity_key))
    return deltas


def _aggregate_symbol_delta(
    old_cs: List[Dict[str, Any]],
    new_cs: List[Dict[str, Any]],
) -> List[SymbolDelta]:
    """Derive symbol-level deltas from raw call sites."""
    old_sym_count: Counter = Counter()
    old_sym_cat: Dict[str, str] = {}
    for cs in old_cs:
        sym = cs.get("ossl_symbol", "")
        old_sym_count[sym] += 1
        if sym not in old_sym_cat:
            old_sym_cat[sym] = cs.get("category", "")

    new_sym_count: Counter = Counter()
    new_sym_cat: Dict[str, str] = {}
    for cs in new_cs:
        sym = cs.get("ossl_symbol", "")
        new_sym_count[sym] += 1
        if sym not in new_sym_cat:
            new_sym_cat[sym] = cs.get("category", "")

    all_syms = set(old_sym_count.keys()) | set(new_sym_count.keys())
    result: List[SymbolDelta] = []

    for sym in sorted(all_syms):
        oc = old_sym_count.get(sym, 0)
        nc = new_sym_count.get(sym, 0)
        cat = new_sym_cat.get(sym, old_sym_cat.get(sym, ""))

        if oc == 0:
            status = DiffStatus.ADDED
        elif nc == 0:
            status = DiffStatus.REMOVED
        elif oc != nc:
            status = DiffStatus.CHANGED
        else:
            status = DiffStatus.UNCHANGED

        result.append(SymbolDelta(
            status=status,
            symbol=sym,
            category=cat,
            old_count=oc,
            new_count=nc,
        ))

    result.sort(key=lambda d: (STATUS_SORT_ORDER.get(d.status, 99), d.symbol))
    return result


def _aggregate_file_delta(
    old_cs: List[Dict[str, Any]],
    new_cs: List[Dict[str, Any]],
) -> List[FileDelta]:
    """Derive file-level deltas from raw call sites."""
    old_file_count: Counter = Counter()
    old_file_syms: Dict[str, Set[str]] = {}
    for cs in old_cs:
        fp = cs.get("file_path", "")
        old_file_count[fp] += 1
        old_file_syms.setdefault(fp, set()).add(cs.get("ossl_symbol", ""))

    new_file_count: Counter = Counter()
    new_file_syms: Dict[str, Set[str]] = {}
    for cs in new_cs:
        fp = cs.get("file_path", "")
        new_file_count[fp] += 1
        new_file_syms.setdefault(fp, set()).add(cs.get("ossl_symbol", ""))

    all_files = set(old_file_count.keys()) | set(new_file_count.keys())
    result: List[FileDelta] = []

    for fp in sorted(all_files):
        oc = old_file_count.get(fp, 0)
        nc = new_file_count.get(fp, 0)
        os_set = old_file_syms.get(fp, set())
        ns_set = new_file_syms.get(fp, set())

        if oc == 0:
            status = DiffStatus.ADDED
        elif nc == 0:
            status = DiffStatus.REMOVED
        elif oc != nc or os_set != ns_set:
            status = DiffStatus.CHANGED
        else:
            status = DiffStatus.UNCHANGED

        result.append(FileDelta(
            status=status,
            file_path=fp,
            old_call_count=oc,
            new_call_count=nc,
            old_symbols=len(os_set),
            new_symbols=len(ns_set),
            added_symbols=sorted(ns_set - os_set),
            removed_symbols=sorted(os_set - ns_set),
        ))

    result.sort(key=lambda d: (STATUS_SORT_ORDER.get(d.status, 99), d.file_path))
    return result


def _compute_metrics(
    old_summary: Dict[str, Any],
    new_summary: Dict[str, Any],
    old_cs: List[Dict[str, Any]],
    new_cs: List[Dict[str, Any]],
) -> List[MetricDelta]:
    """Compute summary metric deltas.

    total_files_scanned comes from pre-computed summary (not affected by
    category filtering).  files_with_calls, total_call_sites, and
    unique_symbols are recomputed from the (possibly filtered) call-site
    lists so that --ignore-categories produces consistent numbers.
    """
    metrics = []

    ov = old_summary.get("total_files_scanned", 0)
    nv = new_summary.get("total_files_scanned", 0)
    metrics.append(MetricDelta(
        name="total_files_scanned", old_value=ov, new_value=nv, delta=nv - ov))

    old_files = set(cs.get("file_path", "") for cs in old_cs)
    new_files = set(cs.get("file_path", "") for cs in new_cs)
    old_files.discard("")
    new_files.discard("")
    metrics.append(MetricDelta(
        name="files_with_calls",
        old_value=len(old_files), new_value=len(new_files),
        delta=len(new_files) - len(old_files)))

    metrics.append(MetricDelta(
        name="total_call_sites",
        old_value=len(old_cs), new_value=len(new_cs),
        delta=len(new_cs) - len(old_cs)))

    old_syms = set(cs.get("ossl_symbol", "") for cs in old_cs)
    new_syms = set(cs.get("ossl_symbol", "") for cs in new_cs)
    old_syms.discard("")
    new_syms.discard("")
    metrics.append(MetricDelta(
        name="unique_symbols",
        old_value=len(old_syms),
        new_value=len(new_syms),
        delta=len(new_syms) - len(old_syms),
    ))

    return metrics


def diff_single(
    old_data: Dict[str, Any],
    new_data: Dict[str, Any],
    ignore_categories: Optional[Set[str]] = None,
    include_unchanged: bool = False,
    summary_only: bool = False,
) -> ProjectDelta:
    """Diff two single-project datasets.

    Parameters:
        old_data: dict with 'call_sites' list and 'summary' dict
        new_data: dict with 'call_sites' list and 'summary' dict
        ignore_categories: set of category names to exclude before diffing
        include_unchanged: if True, include UNCHANGED entries in call_site_delta
        summary_only: if True, return only metrics and symbol/file deltas (no call sites)

    Returns:
        ProjectDelta with metrics, call_site_delta, symbol_delta, file_delta
    """
    project_name = new_data.get("project", old_data.get("project", ""))

    old_cs = list(old_data.get("call_sites", []))
    new_cs = list(new_data.get("call_sites", []))

    if ignore_categories:
        old_cs = [cs for cs in old_cs if cs.get("category", "") not in ignore_categories]
        new_cs = [cs for cs in new_cs if cs.get("category", "") not in ignore_categories]

    old_map = _build_identity_map(old_cs)
    new_map = _build_identity_map(new_cs)

    call_deltas = _compute_call_site_delta(old_map, new_map, include_unchanged)
    symbol_deltas = _aggregate_symbol_delta(old_cs, new_cs)
    file_deltas = _aggregate_file_delta(old_cs, new_cs)

    old_summary = old_data.get("summary", {})
    new_summary = new_data.get("summary", {})
    metrics = _compute_metrics(old_summary, new_summary, old_cs, new_cs)

    has_cs_changes = any(
        d.status != DiffStatus.UNCHANGED for d in call_deltas
    )

    if summary_only:
        call_deltas = []

    return ProjectDelta(
        project=project_name,
        metrics=metrics,
        call_site_delta=call_deltas,
        symbol_delta=symbol_deltas,
        file_delta=file_deltas,
        has_call_site_changes=has_cs_changes,
    )


def diff_combo(
    old_report: Dict[str, Any],
    new_report: Dict[str, Any],
    ignore_categories: Optional[Set[str]] = None,
    include_unchanged: bool = False,
    summary_only: bool = False,
) -> DiffResult:
    """Diff two combo scan reports across all projects.

    Matches projects by name and produces per-project ProjectDelta entries.
    Projects only in new_report appear as all-ADDED; projects only in
    old_report appear as all-REMOVED.
    """
    old_map: Dict[str, Dict[str, Any]] = {
        p["project"]: p for p in old_report.get("projects", [])
    }
    new_map: Dict[str, Dict[str, Any]] = {
        p["project"]: p for p in new_report.get("projects", [])
    }

    all_projects = sorted(set(old_map.keys()) | set(new_map.keys()))
    empty_proj = {"project": "", "call_sites": [], "summary": {}}

    project_deltas: List[ProjectDelta] = []
    for name in all_projects:
        old_proj = old_map.get(name, dict(empty_proj, project=name))
        new_proj = new_map.get(name, dict(empty_proj, project=name))
        pd = diff_single(
            old_proj, new_proj,
            ignore_categories=ignore_categories,
            include_unchanged=include_unchanged,
            summary_only=summary_only,
        )
        project_deltas.append(pd)

    old_label = old_report.get("_path", "old")
    new_label = new_report.get("_path", "new")

    return DiffResult(
        old_label=old_label,
        new_label=new_label,
        projects=project_deltas,
        old_scan_time=old_report.get("scan_time", ""),
        new_scan_time=new_report.get("scan_time", ""),
        is_combo=True,
    )


_METRIC_LABELS = {
    "total_files_scanned": "Files Scanned",
    "files_with_calls": "Files with Calls",
    "total_call_sites": "Call Sites",
    "unique_symbols": "Unique Symbols",
}


class SourceDiffJsonExporter:
    """Export a DiffResult to JSON."""

    def __init__(self, hitls_compat: Optional['HiTLSCompat'] = None):
        self._hitls_compat = hitls_compat

    def export(self, result: DiffResult, output_path: str) -> None:
        parent = os.path.dirname(output_path)
        if parent:
            os.makedirs(parent, exist_ok=True)

        if result.is_combo:
            doc = self._build_combo(result)
        else:
            doc = self._build_single(result)

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(doc, f, indent=2, ensure_ascii=False)

    def _build_meta(self, result: DiffResult, report_type: str) -> Dict[str, Any]:
        meta: Dict[str, Any] = {
            "report_type": report_type,
            "diff_time": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "tool_version": "1.0.0",
            "old_report": result.old_label,
            "new_report": result.new_label,
        }
        if result.old_scan_time:
            meta["old_scan_time"] = result.old_scan_time
        if result.new_scan_time:
            meta["new_scan_time"] = result.new_scan_time
        return meta

    def _serialize_project(self, pd: ProjectDelta) -> Dict[str, Any]:
        summary_delta: Dict[str, Any] = {}
        for m in pd.metrics:
            summary_delta[m.name] = {
                "old": m.old_value,
                "new": m.new_value,
                "delta": m.delta,
            }

        symbol_delta = []
        for sd in pd.symbol_delta:
            sd_entry = {
                "status": sd.status.value,
                "symbol": sd.symbol,
                "category": sd.category,
                "old_count": sd.old_count,
                "new_count": sd.new_count,
            }
            if self._hitls_compat is not None:
                h_status, h_equiv = self._hitls_compat.lookup(sd.symbol)
                sd_entry["hitls_status"] = h_status
                sd_entry["hitls_equiv"] = h_equiv
            symbol_delta.append(sd_entry)

        file_delta = [
            {
                "status": fd.status.value,
                "file_path": fd.file_path,
                "old_call_count": fd.old_call_count,
                "new_call_count": fd.new_call_count,
                "old_symbols": fd.old_symbols,
                "new_symbols": fd.new_symbols,
                "added_symbols": fd.added_symbols,
                "removed_symbols": fd.removed_symbols,
            }
            for fd in pd.file_delta
        ]

        call_site_delta = []
        for csd in pd.call_site_delta:
            entry = {
                "status": csd.status.value,
                "file_path": csd.identity_key[0],
                "caller_function": csd.identity_key[1],
                "ossl_symbol": csd.identity_key[2],
                "category": csd.category,
                "old_count": csd.old_count,
                "new_count": csd.new_count,
                "old_lines": csd.old_lines,
                "new_lines": csd.new_lines,
            }
            if csd.arg_deltas:
                entry["arg_deltas"] = [
                    {
                        "status": ad.status.value,
                        "old_line": ad.old_line,
                        "new_line": ad.new_line,
                        "old_args": ad.old_args,
                        "new_args": ad.new_args,
                    }
                    for ad in csd.arg_deltas
                ]
            call_site_delta.append(entry)

        return {
            "summary_delta": summary_delta,
            "symbol_delta": symbol_delta,
            "file_delta": file_delta,
            "call_site_delta": call_site_delta,
        }

    def _build_single(self, result: DiffResult) -> Dict[str, Any]:
        pd = result.projects[0] if result.projects else ProjectDelta(project="")
        doc = {"meta": self._build_meta(result, "source_diff")}
        doc.update(self._serialize_project(pd))
        return doc

    def _build_combo(self, result: DiffResult) -> Dict[str, Any]:
        projects = []
        for pd in result.projects:
            entry = {"project": pd.project}
            entry.update(self._serialize_project(pd))
            projects.append(entry)
        return {
            "meta": self._build_meta(result, "combo_diff"),
            "projects": projects,
        }


def format_console(result: DiffResult) -> str:
    """Format a DiffResult as human-readable console text."""
    if result.is_empty():
        return "  No changes detected."

    lines: List[str] = []
    lines.append(f"  Source Diff: {result.old_label} -> {result.new_label}")
    lines.append("  " + "=" * 44)

    for pd in result.projects:
        lines.append("")
        lines.append(f"  Project: {pd.project}")

        if pd.metrics:
            lines.append("  Summary:")
            for m in pd.metrics:
                label = _METRIC_LABELS.get(m.name, m.name)
                sign = "+" if m.delta > 0 else ""
                lines.append(
                    f"    {label + ':':22s}{m.old_value:>6} -> {m.new_value:<6}"
                    f"  ({sign}{m.delta})"
                )

        added_syms = [s for s in pd.symbol_delta
                      if s.status == DiffStatus.ADDED]
        if added_syms:
            lines.append(f"  Symbols Added ({len(added_syms)}):")
            for s in added_syms:
                lines.append(f"    + {s.symbol:30s} [{s.category}]")

        removed_syms = [s for s in pd.symbol_delta
                        if s.status == DiffStatus.REMOVED]
        if removed_syms:
            lines.append(f"  Symbols Removed ({len(removed_syms)}):")
            for s in removed_syms:
                lines.append(f"    - {s.symbol:30s} [{s.category}]")

        changed_syms = [s for s in pd.symbol_delta
                        if s.status == DiffStatus.CHANGED]
        if changed_syms:
            lines.append(f"  Symbols Changed ({len(changed_syms)}):")
            for s in changed_syms:
                sign = "+" if s.new_count > s.old_count else ""
                delta = s.new_count - s.old_count
                lines.append(
                    f"    ~ {s.symbol:30s} [{s.category}]"
                    f"  {s.old_count} -> {s.new_count} ({sign}{delta})"
                )

        moved_syms = [s for s in pd.symbol_delta
                      if s.status == DiffStatus.MOVED]
        if moved_syms:
            lines.append(f"  Symbols Moved ({len(moved_syms)}):")
            for s in moved_syms:
                lines.append(f"    > {s.symbol:30s} [{s.category}]")

        added_files = [fd for fd in pd.file_delta
                       if fd.status == DiffStatus.ADDED]
        if added_files:
            lines.append(f"  Files Added ({len(added_files)}):")
            for fd in added_files:
                lines.append(f"    + {fd.file_path}")

        removed_files = [fd for fd in pd.file_delta
                         if fd.status == DiffStatus.REMOVED]
        if removed_files:
            lines.append(f"  Files Removed ({len(removed_files)}):")
            for fd in removed_files:
                lines.append(f"    - {fd.file_path}")

        changed_files = [fd for fd in pd.file_delta
                         if fd.status == DiffStatus.CHANGED]
        if changed_files:
            lines.append(f"  Files Changed ({len(changed_files)}):")
            for fd in changed_files:
                sign = "+" if fd.new_call_count > fd.old_call_count else ""
                delta = fd.new_call_count - fd.old_call_count
                lines.append(
                    f"    ~ {fd.file_path}"
                    f"  {fd.old_call_count} -> {fd.new_call_count} calls"
                    f" ({sign}{delta})"
                )

        cs_added = [csd for csd in pd.call_site_delta
                    if csd.status == DiffStatus.ADDED]
        cs_removed = [csd for csd in pd.call_site_delta
                      if csd.status == DiffStatus.REMOVED]
        if cs_added:
            lines.append(f"  Call Sites Added ({len(cs_added)}):")
            for csd in cs_added:
                fp, caller, sym = csd.identity_key
                lines.append(f"    + {fp} :: {caller} :: {sym}")
        if cs_removed:
            lines.append(f"  Call Sites Removed ({len(cs_removed)}):")
            for csd in cs_removed:
                fp, caller, sym = csd.identity_key
                lines.append(f"    - {fp} :: {caller} :: {sym}")

        args_changed = []
        for csd in pd.call_site_delta:
            for ad in csd.arg_deltas:
                if ad.status == DiffStatus.CHANGED and ad.old_args != ad.new_args:
                    args_changed.append((csd.identity_key, ad))
        if args_changed:
            lines.append(f"  Args Changed ({len(args_changed)}):")
            for key, ad in args_changed:
                file_path, caller, symbol = key
                line_label = f"L{ad.old_line}" if ad.old_line is not None else "L?"
                lines.append(
                    f"    {file_path} :: {caller} :: {symbol}  {line_label}")
                lines.append(f"      old: {ad.old_args}")
                lines.append(f"      new: {ad.new_args}")

    return "\n".join(lines)


DIFF_COLORS = {
    "header_blue": "E8F4FC",
    "header_green": "F0F8E8",
    "added": "C6EFCE",
    "removed": "FFC7CE",
    "changed": "FFEB9C",
    "moved": "D9D9D9",
    "delta_pos": "006100",
    "delta_neg": "9C0006",
    "delta_zero": "808080",
}


_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value: str) -> str:
    """Neutralize Excel formula injection by prefixing with single quote."""
    if value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _get_status_fill(status_str):
    """Return a PatternFill for a DiffStatus value string."""
    from . import _vendor  # noqa: F401
    from openpyxl.styles import PatternFill

    color_map = {
        "added": DIFF_COLORS["added"],
        "removed": DIFF_COLORS["removed"],
        "changed": DIFF_COLORS["changed"],
        "moved": DIFF_COLORS["moved"],
    }
    color = color_map.get(status_str)
    if color:
        return PatternFill(start_color=color, end_color=color, fill_type="solid")
    return None


def _delta_font(delta):
    """Return a Font with color based on delta sign."""
    from . import _vendor  # noqa: F401
    from openpyxl.styles import Font

    if delta > 0:
        return Font(color=DIFF_COLORS["delta_pos"])
    elif delta < 0:
        return Font(color=DIFF_COLORS["delta_neg"])
    return Font(color=DIFF_COLORS["delta_zero"])


def _write_header(ws, columns, header_font, header_fill):
    """Write header row and set column widths.  Returns column count."""
    from . import _vendor  # noqa: F401
    from openpyxl.utils import get_column_letter

    for col_idx, (width, title) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return len(columns)


def _set_auto_filter(ws, row, num_cols):
    """Set auto_filter if sheet has data rows beyond header."""
    if row > 2:
        from . import _vendor  # noqa: F401
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{row - 1}"


class SourceDiffExcelExporter:
    """Export a DiffResult to XLSX with conditional coloring."""

    def __init__(self, include_unchanged: bool = False,
                 hitls_compat: Optional['HiTLSCompat'] = None):
        self._include_unchanged = include_unchanged
        self._hitls_compat = hitls_compat

    def export(self, result: DiffResult, output_path: str) -> None:
        from . import _vendor  # noqa: F401
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        parent = os.path.dirname(output_path)
        if parent:
            os.makedirs(parent, exist_ok=True)

        wb = Workbook()

        green_font = Font(bold=True)
        green_fill = PatternFill(
            start_color=DIFF_COLORS["header_green"],
            end_color=DIFF_COLORS["header_green"],
            fill_type="solid",
        )
        blue_font = Font(bold=True)
        blue_fill = PatternFill(
            start_color=DIFF_COLORS["header_blue"],
            end_color=DIFF_COLORS["header_blue"],
            fill_type="solid",
        )

        self._write_summary_sheet(wb, result, green_font, green_fill)
        self._write_symbol_sheet(wb, result, blue_font, blue_fill)
        self._write_file_sheet(wb, result, blue_font, blue_fill)
        self._write_callsite_sheet(wb, result, blue_font, blue_fill)

        if result.is_combo:
            self._write_project_sheet(wb, result, green_font, green_fill)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_path)

    def _write_summary_sheet(self, wb, result, header_font, header_fill):
        ws = wb.create_sheet("Summary Delta")
        columns = [(30, "Metric"), (12, "Old"), (12, "New"), (12, "Delta")]
        num_cols = _write_header(ws, columns, header_font, header_fill)

        if len(result.projects) == 1:
            metrics = result.projects[0].metrics
        else:
            agg: Dict[str, List[int]] = {}
            old_all_syms: Set[str] = set()
            new_all_syms: Set[str] = set()
            for proj in result.projects:
                for m in proj.metrics:
                    if m.name == "unique_symbols":
                        continue
                    if m.name not in agg:
                        agg[m.name] = [0, 0]
                    agg[m.name][0] += m.old_value
                    agg[m.name][1] += m.new_value
                for sd in proj.symbol_delta:
                    if sd.old_count > 0:
                        old_all_syms.add(sd.symbol)
                    if sd.new_count > 0:
                        new_all_syms.add(sd.symbol)
            metrics = []
            for name in ("total_files_scanned", "files_with_calls",
                         "total_call_sites"):
                if name in agg:
                    ov, nv = agg[name]
                    metrics.append(MetricDelta(name=name, old_value=ov,
                                              new_value=nv, delta=nv - ov))
            metrics.append(MetricDelta(
                name="unique_symbols",
                old_value=len(old_all_syms),
                new_value=len(new_all_syms),
                delta=len(new_all_syms) - len(old_all_syms)))

        row = 2
        for m in metrics:
            label = _METRIC_LABELS.get(m.name, m.name)
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=m.old_value)
            ws.cell(row=row, column=3, value=m.new_value)
            delta_cell = ws.cell(row=row, column=4, value=m.delta)
            delta_cell.font = _delta_font(m.delta)
            row += 1

        cat_old: Dict[str, Set[str]] = {}
        cat_new: Dict[str, Set[str]] = {}
        for proj in result.projects:
            for sd in proj.symbol_delta:
                if sd.old_count > 0:
                    cat_old.setdefault(sd.category, set()).add(sd.symbol)
                if sd.new_count > 0:
                    cat_new.setdefault(sd.category, set()).add(sd.symbol)

        all_cats = sorted(set(cat_old.keys()) | set(cat_new.keys()))
        if all_cats:
            row += 1
        for cat in all_cats:
            ov = len(cat_old.get(cat, set()))
            nv = len(cat_new.get(cat, set()))
            delta = nv - ov
            ws.cell(row=row, column=1, value=f"[{cat}]")
            ws.cell(row=row, column=2, value=ov)
            ws.cell(row=row, column=3, value=nv)
            delta_cell = ws.cell(row=row, column=4, value=delta)
            delta_cell.font = _delta_font(delta)
            row += 1

        _set_auto_filter(ws, row, num_cols)

    def _write_symbol_sheet(self, wb, result, header_font, header_fill):
        ws = wb.create_sheet("Symbol Delta")
        columns = [
            (35, "OpenSSL Symbol"), (20, "Category"), (12, "Status"),
            (12, "Old Calls"), (12, "New Calls"), (10, "Delta"),
        ]
        if self._hitls_compat is not None:
            columns.append((15, "HiTLS Status"))
            columns.append((30, "HiTLS Equivalent"))
        num_cols = _write_header(ws, columns, header_font, header_fill)

        all_symbols: Dict[str, SymbolDelta] = {}
        for proj in result.projects:
            for sd in proj.symbol_delta:
                if sd.symbol in all_symbols:
                    existing = all_symbols[sd.symbol]
                    merged_old = existing.old_count + sd.old_count
                    merged_new = existing.new_count + sd.new_count
                    if merged_old == 0:
                        status = DiffStatus.ADDED
                    elif merged_new == 0:
                        status = DiffStatus.REMOVED
                    elif merged_old != merged_new:
                        status = DiffStatus.CHANGED
                    else:
                        status = DiffStatus.UNCHANGED
                    all_symbols[sd.symbol] = SymbolDelta(
                        status=status, symbol=sd.symbol, category=sd.category,
                        old_count=merged_old, new_count=merged_new,
                    )
                else:
                    all_symbols[sd.symbol] = sd

        entries = sorted(
            all_symbols.values(),
            key=lambda d: (STATUS_SORT_ORDER.get(d.status, 99), d.category, d.symbol),
        )

        row = 2
        for sd in entries:
            if sd.status == DiffStatus.UNCHANGED and not self._include_unchanged:
                continue
            ws.cell(row=row, column=1, value=sd.symbol)
            ws.cell(row=row, column=2, value=sd.category)
            status_cell = ws.cell(row=row, column=3, value=sd.status.value)
            fill = _get_status_fill(sd.status.value)
            if fill:
                status_cell.fill = fill
            ws.cell(row=row, column=4, value=sd.old_count)
            ws.cell(row=row, column=5, value=sd.new_count)
            delta = sd.new_count - sd.old_count
            delta_cell = ws.cell(row=row, column=6, value=delta)
            delta_cell.font = _delta_font(delta)
            if self._hitls_compat is not None:
                h_status, h_equiv = self._hitls_compat.lookup(sd.symbol)
                ws.cell(row=row, column=7, value=h_status)
                ws.cell(row=row, column=8, value=h_equiv or '')
            row += 1

        _set_auto_filter(ws, row, num_cols)

    def _write_file_sheet(self, wb, result, header_font, header_fill):
        ws = wb.create_sheet("File Delta")
        columns = [
            (60, "File Path"), (12, "Status"), (12, "Old Calls"),
            (12, "New Calls"), (10, "Delta"), (12, "Old Symbols"),
            (12, "New Symbols"), (40, "Added Symbols"), (40, "Removed Symbols"),
        ]
        num_cols = _write_header(ws, columns, header_font, header_fill)

        is_combo = len(result.projects) > 1
        all_files: List[Tuple[str, FileDelta]] = []
        for proj in result.projects:
            for fd in proj.file_delta:
                if fd.status == DiffStatus.UNCHANGED and not self._include_unchanged:
                    continue
                prefix = f"{proj.project}/" if is_combo else ""
                all_files.append((prefix + fd.file_path, fd))

        all_files.sort(
            key=lambda x: (
                STATUS_SORT_ORDER.get(x[1].status, 99),
                -abs(x[1].new_call_count - x[1].old_call_count),
            )
        )

        row = 2
        for display_path, fd in all_files:
            ws.cell(row=row, column=1, value=_safe_cell(display_path))
            status_cell = ws.cell(row=row, column=2, value=fd.status.value)
            fill = _get_status_fill(fd.status.value)
            if fill:
                status_cell.fill = fill
            ws.cell(row=row, column=3, value=fd.old_call_count)
            ws.cell(row=row, column=4, value=fd.new_call_count)
            delta = fd.new_call_count - fd.old_call_count
            delta_cell = ws.cell(row=row, column=5, value=delta)
            delta_cell.font = _delta_font(delta)
            ws.cell(row=row, column=6, value=fd.old_symbols)
            ws.cell(row=row, column=7, value=fd.new_symbols)
            ws.cell(row=row, column=8, value=", ".join(fd.added_symbols))
            ws.cell(row=row, column=9, value=", ".join(fd.removed_symbols))
            row += 1

        _set_auto_filter(ws, row, num_cols)

    def _write_callsite_sheet(self, wb, result, header_font, header_fill):
        base_title = "Call Site Delta"
        is_combo = result.is_combo
        if is_combo:
            columns = [
                (20, "Project"), (12, "Status"), (50, "File Path"),
                (30, "Caller Function"), (35, "OpenSSL Symbol"),
                (20, "Category"), (10, "Old Line"), (10, "New Line"),
                (30, "Old Args"), (30, "New Args"),
            ]
        else:
            columns = [
                (12, "Status"), (50, "File Path"), (30, "Caller Function"),
                (35, "OpenSSL Symbol"), (20, "Category"), (10, "Old Line"),
                (10, "New Line"), (30, "Old Args"), (30, "New Args"),
            ]

        ws_first = wb.create_sheet(base_title)
        num_cols = _write_header(ws_first, columns, header_font, header_fill)
        ws = ws_first
        sheet_num = 1

        all_cs: List[Tuple[str, CallSiteDelta]] = []
        for proj in result.projects:
            for csd in proj.call_site_delta:
                all_cs.append((proj.project, csd))

        all_cs.sort(
            key=lambda x: (
                STATUS_SORT_ORDER.get(x[1].status, 99),
                x[0],
                x[1].identity_key[0],
                x[1].identity_key[2],
            )
        )

        row = 2
        for proj_name, csd in all_cs:
            file_path, caller, symbol = csd.identity_key
            if not csd.arg_deltas:
                continue
            for ad in csd.arg_deltas:
                if row > XLSX_MAX_ROW:
                    sheet_num += 1
                    ws = wb.create_sheet(f"{base_title} ({sheet_num})")
                    _write_header(ws, columns, header_font, header_fill)
                    row = 2
                col = 1
                if is_combo:
                    ws.cell(row=row, column=col, value=_safe_cell(proj_name))
                    col += 1
                status_cell = ws.cell(row=row, column=col, value=ad.status.value)
                fill = _get_status_fill(ad.status.value)
                if fill:
                    status_cell.fill = fill
                col += 1
                ws.cell(row=row, column=col, value=_safe_cell(file_path))
                col += 1
                ws.cell(row=row, column=col, value=_safe_cell(caller))
                col += 1
                ws.cell(row=row, column=col, value=symbol)
                col += 1
                ws.cell(row=row, column=col, value=csd.category)
                col += 1
                ws.cell(row=row, column=col,
                        value=ad.old_line if ad.old_line is not None else "")
                col += 1
                ws.cell(row=row, column=col,
                        value=ad.new_line if ad.new_line is not None else "")
                col += 1
                ws.cell(row=row, column=col, value=_safe_cell(ad.old_args))
                col += 1
                ws.cell(row=row, column=col, value=_safe_cell(ad.new_args))
                row += 1

        if sheet_num == 1:
            _set_auto_filter(ws_first, row, num_cols)
        else:
            _set_auto_filter(ws_first, XLSX_MAX_ROW + 1, num_cols)
            _set_auto_filter(ws, row, num_cols)
            logger.info("Call Site Delta split across %d sheets", sheet_num)

    def _write_project_sheet(self, wb, result, header_font, header_fill):
        ws = wb.create_sheet("Project Delta")
        columns = [
            (30, "Project"), (12, "Status"), (12, "Old Calls"),
            (12, "New Calls"), (10, "Delta"), (12, "Old Symbols"),
            (12, "New Symbols"), (40, "Added Symbols"), (40, "Removed Symbols"),
        ]
        num_cols = _write_header(ws, columns, header_font, header_fill)

        row = 2
        for proj in result.projects:
            sd_statuses = {s.status for s in proj.symbol_delta}
            if sd_statuses == {DiffStatus.ADDED}:
                proj_status = DiffStatus.ADDED
            elif sd_statuses == {DiffStatus.REMOVED}:
                proj_status = DiffStatus.REMOVED
            elif sd_statuses - {DiffStatus.UNCHANGED}:
                proj_status = DiffStatus.CHANGED
            else:
                proj_status = DiffStatus.UNCHANGED

            if proj_status == DiffStatus.UNCHANGED and not self._include_unchanged:
                continue

            old_calls = sum(s.old_count for s in proj.symbol_delta)
            new_calls = sum(s.new_count for s in proj.symbol_delta)
            old_syms = sum(1 for s in proj.symbol_delta if s.old_count > 0)
            new_syms = sum(1 for s in proj.symbol_delta if s.new_count > 0)
            added = sorted(s.symbol for s in proj.symbol_delta
                           if s.status == DiffStatus.ADDED)
            removed = sorted(s.symbol for s in proj.symbol_delta
                             if s.status == DiffStatus.REMOVED)

            ws.cell(row=row, column=1, value=_safe_cell(proj.project))
            status_cell = ws.cell(row=row, column=2, value=proj_status.value)
            fill = _get_status_fill(proj_status.value)
            if fill:
                status_cell.fill = fill
            ws.cell(row=row, column=3, value=old_calls)
            ws.cell(row=row, column=4, value=new_calls)
            delta = new_calls - old_calls
            delta_cell = ws.cell(row=row, column=5, value=delta)
            delta_cell.font = _delta_font(delta)
            ws.cell(row=row, column=6, value=old_syms)
            ws.cell(row=row, column=7, value=new_syms)
            ws.cell(row=row, column=8, value=", ".join(added))
            ws.cell(row=row, column=9, value=", ".join(removed))
            row += 1

        _set_auto_filter(ws, row, num_cols)
