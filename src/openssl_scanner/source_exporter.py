"""
Export source scan results to XLSX or JSON.

Single-sheet XLSX with call site details, matching existing exporter styling.
Multi-sheet merge: combine multiple XLSX reports into one workbook.
"""

# ruff: noqa: I001

import json
import logging
import os
import re
from typing import TYPE_CHECKING, Any, Dict, List, Optional, Tuple

from .source_analyzer import SourceScanResult

if TYPE_CHECKING:
    from .hitls_compat import HiTLSCompat

logger = logging.getLogger(__name__)

XLSX_MAX_ROW = 1048576

COLUMNS = [
    ('file_path',         60, 'File Path'),
    ('file_name',         25, 'File Name'),
    ('caller_function',   30, 'Caller Function'),
    ('line_number',       12, 'Line'),
    ('ossl_symbol',       35, 'OpenSSL Symbol'),
    ('category',          20, 'Category'),
    ('call_args',         60, 'Call Arguments'),
]

LAST_COL_LETTER = chr(64 + len(COLUMNS))


SUMMARY_SHEET_COLUMNS = [
    ('ossl_symbol',  35, 'OpenSSL Symbol'),
    ('category',     20, 'Category'),
    ('call_count',   12, 'Calls'),
    ('file_count',   12, 'Files'),
    ('file_list',    80, 'File List'),
]

MERGE_SUMMARY_SHEET_COLUMNS = [
    ('ossl_symbol',    35, 'OpenSSL Symbol'),
    ('category',       20, 'Category'),
    ('call_count',     12, 'Calls'),
    ('file_count',     12, 'Files'),
    ('file_list',      80, 'File List'),
    ('project_count',  12, 'Projects'),
    ('project_list',   40, 'Project List'),
]

HITLS_COVERAGE_ROWS = [
    ('total_symbols', 'Unique OpenSSL Symbols'),
    ('available', 'Available'),
    ('partial', 'Partial'),
    ('not_available', 'Not Available'),
    ('unknown', 'Unknown'),
    ('direct_replace_ratio', 'Direct Replace Ratio (%)'),
    ('direct_or_partial_replace_ratio', 'Direct+Partial Replace Ratio (%)'),
]

RECOVERY_COLUMNS = [
    ('extraction_source', 24, 'Extraction Source'),
    ('confidence', 14, 'Confidence'),
    ('parser_diagnostic_class', 28, 'Parser Diagnostic Class'),
]


def _has_fallback_sites(result: SourceScanResult) -> bool:
    return any(cs.extraction_source != 'ast' for cs in result.call_sites)


def _row_has_recovery(row: List) -> bool:
    return len(row) >= len(COLUMNS) + len(RECOVERY_COLUMNS)


def _row_columns(rows: List[List]) -> List[Tuple[str, int, str]]:
    if any(_row_has_recovery(row) for row in rows):
        return COLUMNS + RECOVERY_COLUMNS
    return COLUMNS


def _call_site_row(cs: Any, include_recovery: bool = False) -> List:
    row = [
        cs.file_path, cs.file_name, cs.caller_function,
        cs.line_number, cs.ossl_symbol, cs.category, cs.call_args,
    ]
    if include_recovery:
        row.extend([
            cs.extraction_source if cs.extraction_source != 'ast' else '',
            cs.confidence if cs.extraction_source != 'ast' else '',
            cs.parser_diagnostic_class if cs.extraction_source != 'ast' else '',
        ])
    return row


def _write_hitls_coverage_sheet(wb, symbols, header_font, hitls_compat) -> None:
    """Write HiTLS overall replacement coverage sheet."""
    from openpyxl.styles import PatternFill

    if hitls_compat is None or not hitls_compat.is_loaded():
        return

    ws = wb.create_sheet(title="HiTLS Coverage")
    summary_fill = PatternFill(
        start_color="F0F8E8", end_color="F0F8E8", fill_type="solid"
    )
    ws.cell(row=1, column=1, value='Metric').font = header_font
    ws.cell(row=1, column=2, value='Value').font = header_font
    ws.cell(row=1, column=1).fill = summary_fill
    ws.cell(row=1, column=2).fill = summary_fill
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18

    coverage = hitls_compat.get_coverage_summary(symbols)
    for row_idx, (key, title) in enumerate(HITLS_COVERAGE_ROWS, 2):
        ws.cell(row=row_idx, column=1, value=title)
        ws.cell(row=row_idx, column=2, value=coverage[key])


class SourceExcelExporter:
    """Export SourceScanResult to XLSX with call sites and symbol summary."""

    def export(self, result: SourceScanResult, output_path: str,
               hitls_compat: Optional['HiTLSCompat'] = None) -> None:
        from . import _vendor  # noqa: F401
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        include_recovery = _has_fallback_sites(result)
        cols = list(COLUMNS)
        if hitls_compat is not None:
            cols.insert(6, ('hitls_status', 15, 'HiTLS Status'))
            cols.insert(7, ('hitls_equiv', 30, 'HiTLS Replacement'))
        if include_recovery:
            cols.extend(RECOVERY_COLUMNS)
        last_col = chr(64 + len(cols)) if len(cols) <= 26 else 'A' + chr(64 + len(cols) - 26)

        base_title = "OpenSSL Call Sites"
        wb = Workbook()
        ws_first = wb.active
        ws_first.title = base_title

        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="E8F4FC", end_color="E8F4FC", fill_type="solid"
        )

        def _init_sheet(ws_target):
            for col_idx, (_, width, title) in enumerate(cols, 1):
                cell = ws_target.cell(row=1, column=col_idx, value=title)
                cell.font = header_font
                cell.fill = header_fill
                ws_target.column_dimensions[
                    chr(64 + col_idx) if col_idx <= 26
                    else 'A' + chr(64 + col_idx - 26)
                ].width = width

        _init_sheet(ws_first)
        ws = ws_first

        row_idx = 2
        sheet_num = 1
        for cs in result.call_sites:
            if row_idx > XLSX_MAX_ROW:
                sheet_num += 1
                ws = wb.create_sheet(f"{base_title} ({sheet_num})")
                _init_sheet(ws)
                row_idx = 2
            ws.cell(row=row_idx, column=1, value=cs.file_path)
            ws.cell(row=row_idx, column=2, value=cs.file_name)
            ws.cell(row=row_idx, column=3, value=cs.caller_function)
            ws.cell(row=row_idx, column=4, value=cs.line_number)
            ws.cell(row=row_idx, column=5, value=cs.ossl_symbol)
            ws.cell(row=row_idx, column=6, value=cs.category)
            if hitls_compat is not None:
                h_status, h_equiv = hitls_compat.lookup(cs.ossl_symbol)
                ws.cell(row=row_idx, column=7, value=h_status)
                ws.cell(row=row_idx, column=8, value=h_equiv or '')
                ws.cell(row=row_idx, column=9, value=cs.call_args)
                next_col = 10
            else:
                ws.cell(row=row_idx, column=7, value=cs.call_args)
                next_col = 8
            if include_recovery:
                ws.cell(row=row_idx, column=next_col, value=cs.extraction_source)
                ws.cell(row=row_idx, column=next_col + 1, value=cs.confidence)
                ws.cell(
                    row=row_idx, column=next_col + 2,
                    value=cs.parser_diagnostic_class,
                )
            row_idx += 1

        if result.call_sites:
            last_row = min(len(result.call_sites) + 1, XLSX_MAX_ROW)
            ws_first.auto_filter.ref = f"A1:{last_col}{last_row}"

        if sheet_num > 1:
            logger.info("Call Sites split across %d sheets (%d rows)",
                        sheet_num, len(result.call_sites))

        self._write_symbol_summary(wb, result, header_font, header_fill,
                                   hitls_compat=hitls_compat)
        if hitls_compat is not None:
            _write_hitls_coverage_sheet(
                wb, set(result.unique_symbols), header_font, hitls_compat,
            )

        wb.save(output_path)
        logger.info("XLSX report saved to: %s", output_path)

    def _write_symbol_summary(self, wb, result: SourceScanResult,
                              header_font, header_fill,
                              hitls_compat: Optional['HiTLSCompat'] = None) -> None:
        """Write Symbol Summary sheet: one row per unique symbol."""
        ws = wb.create_sheet(title="Symbol Summary")

        cols = list(SUMMARY_SHEET_COLUMNS)
        if hitls_compat is not None:
            cols.append(('hitls_status', 15, 'HiTLS Status'))
            cols.append(('hitls_equiv', 30, 'HiTLS Replacement'))

        for col_idx, (_, width, title) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[
                chr(64 + col_idx) if col_idx <= 26
                else 'A' + chr(64 + col_idx - 26)
            ].width = width

        sym_data: Dict[str, Dict] = {}
        for cs in result.call_sites:
            if cs.ossl_symbol not in sym_data:
                sym_data[cs.ossl_symbol] = {
                    'category': cs.category,
                    'calls': 0,
                    'files': set(),
                }
            sym_data[cs.ossl_symbol]['calls'] += 1
            sym_data[cs.ossl_symbol]['files'].add(cs.file_name)

        rows = sorted(sym_data.items(),
                       key=lambda x: (x[1]['category'], x[0]))

        for row_idx, (symbol, info) in enumerate(rows, 2):
            file_list = ', '.join(sorted(info['files']))
            ws.cell(row=row_idx, column=1, value=symbol)
            ws.cell(row=row_idx, column=2, value=info['category'])
            ws.cell(row=row_idx, column=3, value=info['calls'])
            ws.cell(row=row_idx, column=4, value=len(info['files']))
            ws.cell(row=row_idx, column=5, value=file_list)
            if hitls_compat is not None:
                h_status, h_equiv = hitls_compat.lookup(symbol)
                ws.cell(row=row_idx, column=6, value=h_status)
                ws.cell(row=row_idx, column=7, value=h_equiv or '')

        last_letter = chr(64 + len(cols))
        if rows:
            ws.auto_filter.ref = f"A1:{last_letter}{len(rows) + 1}"


class SourceJsonExporter:
    """Export SourceScanResult to JSON."""

    def export(self, result: SourceScanResult,
               output_path: Optional[str] = None,
               hitls_compat: Optional['HiTLSCompat'] = None) -> str:
        call_sites = []
        for cs in result.call_sites:
            entry = {
                'file_path': cs.file_path,
                'file_name': cs.file_name,
                'caller_function': cs.caller_function,
                'line_number': cs.line_number,
                'column': cs.column,
                'ossl_symbol': cs.ossl_symbol,
                'category': cs.category,
                'call_args': cs.call_args,
                'language': cs.language,
            }
            if cs.extraction_source != 'ast':
                entry['extraction_source'] = cs.extraction_source
                entry['confidence'] = cs.confidence
                entry['parser_diagnostic_class'] = cs.parser_diagnostic_class
            if hitls_compat is not None:
                h_status, h_equiv = hitls_compat.lookup(cs.ossl_symbol)
                entry['hitls_status'] = h_status
                entry['hitls_equiv'] = h_equiv
                entry['hitls_replacement'] = h_equiv
            call_sites.append(entry)

        summary = {
            'total_files_scanned': result.total_files_scanned,
            'files_with_calls': result.files_with_calls,
            'total_call_sites': result.total_call_sites,
            'unique_symbols_count': len(result.unique_symbols),
            'unique_symbols': result.unique_symbols,
            'symbols_by_category': result.symbols_by_category,
        }
        fallback_count = sum(
            1 for cs in result.call_sites if cs.extraction_source != 'ast'
        )
        if fallback_count:
            summary['fallback_call_sites'] = fallback_count
            summary['files_with_fallback_call_sites'] = len({
                cs.file_path
                for cs in result.call_sites
                if cs.extraction_source != 'ast'
            })
        if hitls_compat is not None:
            coverage = hitls_compat.get_coverage_summary(
                set(result.unique_symbols))
            summary['hitls_coverage'] = {
                key: coverage[key]
                for key in ('available', 'partial', 'not_available', 'unknown')
            }
            summary['hitls_direct_replace_ratio'] = (
                coverage['direct_replace_ratio']
            )
            summary['hitls_direct_or_partial_replace_ratio'] = (
                coverage['direct_or_partial_replace_ratio']
            )

        data = {
            'meta': {
                'tool_version': result.tool_version,
                'report_type': 'source_scan',
                'scan_time': result.scan_time,
                'target': result.target,
            },
            'summary': summary,
            'call_sites': call_sites,
            'errors': result.errors,
        }

        json_str = json.dumps(data, indent=2, ensure_ascii=False)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(json_str)
            logger.info("JSON report saved to: %s", output_path)

        return json_str


SUMMARY_COLUMNS = [
    ('Project',         30),
    ('Files Scanned',   15),
    ('Files with Calls', 16),
    ('Call Sites',      12),
    ('Unique Symbols',  15),
    ('Top Category',    20),
    ('Top Cat Symbols', 15),
]


class SourceMergeExporter:
    """Merge multiple source scan XLSX reports into one multi-sheet workbook."""

    def merge(self, input_paths: List[str], output_path: str,
              hitls_compat: Optional['HiTLSCompat'] = None) -> Dict:
        """Read XLSX reports and write a merged workbook.

        Returns:
            Dict with per-project stats for console summary.
        """
        input_paths = list(input_paths)
        project_data = self._load_projects(input_paths, self._read_xlsx)
        for pdata, path in zip(project_data, input_paths):
            pdata['source'] = path
        return self._merge_to_workbook(project_data, output_path,
                                       hitls_compat=hitls_compat)

    def merge_from_json(self, input_paths: List[str],
                        output_path: str,
                        hitls_compat: Optional['HiTLSCompat'] = None) -> Dict:
        """Read JSON reports and write a merged XLSX (or JSON) workbook.

        Unlike merge() which reads XLSX, this reads the faster JSON format
        and preserves total_files_scanned (not '--').
        """
        project_data = self._load_projects(input_paths, self._read_json)
        return self._merge_to_workbook(project_data, output_path,
                                       hitls_compat=hitls_compat)

    def _load_projects(self, input_paths, reader):
        """Load project data from input files using the given reader.

        Args:
            input_paths: List of file paths to read.
            reader: Callable returning (name, files_scanned, rows).
        """
        project_data = []
        for path in input_paths:
            name, files_scanned, rows = reader(path)
            project_data.append({
                'name': name, 'files_scanned': files_scanned, 'rows': rows,
            })
        return project_data

    def _merge_to_workbook(self, project_data: List[Dict],
                           output_path: str,
                           hitls_compat: Optional['HiTLSCompat'] = None) -> Dict:
        """Shared implementation for all XLSX merge paths.

        Args:
            project_data: List of dicts with keys:
                name (str), files_scanned (int or '--'), rows (List[List])
                Optional: source (str) - preserved in stats for merge() callers
        """
        from . import _vendor  # noqa: F401
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="E8F4FC", end_color="E8F4FC", fill_type="solid"
        )
        summary_fill = PatternFill(
            start_color="F0F8E8", end_color="F0F8E8", fill_type="solid"
        )

        names = self._resolve_sheet_names([p['name'] for p in project_data])

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        for col_idx, (title, width) in enumerate(SUMMARY_COLUMNS, 1):
            cell = ws_summary.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = summary_fill
            ws_summary.column_dimensions[chr(64 + col_idx)].width = width

        stats = []
        all_symbols = set()
        all_rows: List[List] = []
        has_files_scanned = any(
            p['files_scanned'] != '--' for p in project_data)

        def _init_project_sheet(target_ws, columns):
            for col_idx, (_, width, title) in enumerate(columns, 1):
                cell = target_ws.cell(row=1, column=col_idx, value=title)
                cell.font = header_font
                cell.fill = header_fill
                target_ws.column_dimensions[
                    chr(64 + col_idx) if col_idx <= 26
                    else 'A' + chr(64 + col_idx - 26)
                ].width = width

        for idx, pdata in enumerate(project_data):
            sheet_name = names[idx]
            rows = pdata['rows']
            files_scanned = pdata['files_scanned']
            project_columns = _row_columns(rows)

            ws_first = wb.create_sheet(title=sheet_name)
            _init_project_sheet(ws_first, project_columns)
            ws = ws_first

            row_idx = 2
            sheet_num = 1
            for row_data in rows:
                if row_idx > XLSX_MAX_ROW:
                    sheet_num += 1
                    overflow_name = f"{sheet_name[:25]} ({sheet_num})"
                    ws = wb.create_sheet(title=overflow_name)
                    _init_project_sheet(ws, project_columns)
                    row_idx = 2
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                row_idx += 1

            if rows:
                last = min(len(rows) + 1, XLSX_MAX_ROW)
                last_col = chr(64 + len(project_columns))
                ws_first.auto_filter.ref = f"A1:{last_col}{last}"

            if sheet_num > 1:
                logger.info("Project '%s' split across %d sheets (%d rows)",
                            sheet_name, sheet_num, len(rows))

            cat_counts = self._count_categories(rows)
            top_cat, top_count = '', 0
            if cat_counts:
                top_cat = max(cat_counts, key=lambda cat: cat_counts[cat])
                top_count = cat_counts[top_cat]

            symbols = set()
            files_with_calls = set()
            for row in rows:
                symbols.add(row[4])
                files_with_calls.add(row[0])
                all_rows.append(list(row) + [sheet_name])
                all_symbols.add(row[4])

            info = {
                'project': sheet_name,
                'files_scanned': files_scanned,
                'call_sites': len(rows),
                'unique_symbols': len(symbols),
                'files_with_calls': len(files_with_calls),
                'top_category': top_cat,
                'top_cat_symbols': top_count,
            }
            if 'source' in pdata:
                info['source'] = pdata['source']
            stats.append(info)

            row_num = idx + 2
            ws_summary.cell(row=row_num, column=1, value=sheet_name)
            ws_summary.cell(row=row_num, column=2, value=files_scanned)
            ws_summary.cell(row=row_num, column=3, value=len(files_with_calls))
            ws_summary.cell(row=row_num, column=4, value=len(rows))
            ws_summary.cell(row=row_num, column=5, value=len(symbols))
            ws_summary.cell(row=row_num, column=6, value=top_cat)
            ws_summary.cell(row=row_num, column=7, value=top_count)

        total_row = len(project_data) + 2
        total_font = Font(bold=True)
        ws_summary.cell(row=total_row, column=1, value="TOTAL").font = total_font
        if has_files_scanned:
            ws_summary.cell(row=total_row, column=2,
                            value=sum(s['files_scanned'] for s in stats
                                      if s['files_scanned'] != '--')).font = total_font
        ws_summary.cell(row=total_row, column=3,
                        value=sum(s['files_with_calls'] for s in stats)).font = total_font
        ws_summary.cell(row=total_row, column=4,
                        value=sum(s['call_sites'] for s in stats)).font = total_font
        ws_summary.cell(row=total_row, column=5,
                        value=len(all_symbols)).font = total_font

        self._write_symbol_summary_from_rows(
            wb, "Symbol Summary", all_rows, header_font, header_fill,
            hitls_compat=hitls_compat)
        if hitls_compat is not None:
            _write_hitls_coverage_sheet(
                wb, all_symbols, header_font, hitls_compat,
            )

        wb.save(output_path)
        logger.info("Merged XLSX report saved to: %s", output_path)
        return {'sheets': stats, 'total_symbols': len(all_symbols)}

    def _read_xlsx(self, path: str) -> Tuple[str, str, List[List]]:
        """Read call site rows from a source scan XLSX.

        Reads all sheets whose title starts with "OpenSSL Call Sites"
        to handle overflow continuation sheets.

        Returns:
            (project_name, files_scanned_placeholder, list_of_row_values)
        """
        from . import _vendor  # noqa: F401
        from openpyxl import load_workbook

        base_title = "OpenSSL Call Sites"
        wb = load_workbook(path, read_only=True, data_only=True)
        rows = []
        base_titles = [title for _, _, title in COLUMNS]
        recovery_titles = [title for _, _, title in RECOVERY_COLUMNS]
        for sn in wb.sheetnames:
            if sn == base_title or sn.startswith(base_title + " ("):
                ws = wb[sn]
                col_map = []
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx == 0:
                        headers = [str(v) if v else '' for v in row]
                        titles = list(base_titles)
                        if 'Extraction Source' in headers:
                            titles.extend(recovery_titles)
                        col_map = [
                            headers.index(title) if title in headers else None
                            for title in titles
                        ]
                        continue
                    if any(v is not None for v in row):
                        vals = [
                            row[i] if i is not None and i < len(row) else ''
                            for i in col_map
                        ]
                        rows.append(vals)
        wb.close()

        name = os.path.splitext(os.path.basename(path))[0]
        return name, '--', rows

    def _read_json(self, path: str) -> Tuple[str, int, List[List]]:
        """Read call site rows from a source scan JSON.

        Returns:
            (project_name, files_scanned, list_of_row_values)
        Row format matches XLSX: [file_path, file_name, caller_function,
            line_number, ossl_symbol, category, call_args]
        """
        with open(path, encoding='utf-8') as f:
            data = json.load(f)

        files_scanned = data.get('summary', {}).get('total_files_scanned', 0)
        rows = []
        for cs in data.get('call_sites', []):
            rows.append([
                cs.get('file_path', ''),
                cs.get('file_name', ''),
                cs.get('caller_function', ''),
                cs.get('line_number', 0),
                cs.get('ossl_symbol', ''),
                cs.get('category', ''),
                cs.get('call_args', ''),
            ])
            if cs.get('extraction_source') == 'parser-diagnostic-text':
                rows[-1].extend([
                    cs.get('extraction_source', ''),
                    cs.get('confidence', ''),
                    cs.get('parser_diagnostic_class', ''),
                ])

        name = os.path.splitext(os.path.basename(path))[0]
        return name, files_scanned, rows

    def _resolve_sheet_names(self, names: List[str]) -> List[str]:
        """Ensure unique sheet names within Excel's 31-char limit."""
        invalid_re = re.compile(r'[\[\]:*?/\\]')
        result = []
        used = set()
        counter: Dict[str, int] = {}
        for name in names:
            short = invalid_re.sub('_', name)[:31]
            if short in used:
                base = short
                counter.setdefault(base, 0)
                while short in used:
                    counter[base] += 1
                    suffix = f"_{counter[base]}"
                    short = base[:31 - len(suffix)] + suffix
            used.add(short)
            result.append(short)
        return result

    def _count_categories(self, rows: List[List]) -> Dict[str, int]:
        """Count unique symbols per category from raw rows."""
        cat_syms: Dict[str, set] = {}
        for row in rows:
            cat = row[5] if len(row) > 5 else ''
            sym = row[4] if len(row) > 4 else ''
            if cat:
                if cat not in cat_syms:
                    cat_syms[cat] = set()
                cat_syms[cat].add(sym)
        return {cat: len(syms) for cat, syms in cat_syms.items()}

    def _write_symbol_summary_from_rows(self, wb, sheet_name: str,
                                        rows: List[List],
                                        header_font, header_fill,
                                        hitls_compat: Optional['HiTLSCompat'] = None) -> None:
        """Write Symbol Summary sheet from tagged call site rows.

        Each row is expected to have a project name appended as the last
        element by the merge() caller.
        """
        ws = wb.create_sheet(title=sheet_name)

        columns = list(MERGE_SUMMARY_SHEET_COLUMNS)
        if hitls_compat is not None:
            columns.append(('hitls_status', 15, 'HiTLS Status'))
            columns.append(('hitls_equiv', 30, 'HiTLS Replacement'))

        for col_idx, (_, width, title) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[
                chr(64 + col_idx) if col_idx <= 26
                else 'A' + chr(64 + col_idx - 26)
            ].width = width

        sym_data: Dict[str, Dict] = {}
        for row in rows:
            if len(row) < 6:
                continue
            sym = row[4]
            cat = row[5]
            fname = row[0] if len(row) > 0 else ''
            project = row[-1] if len(row) > len(COLUMNS) else ''
            if sym not in sym_data:
                sym_data[sym] = {
                    'category': cat, 'calls': 0,
                    'files': set(), 'projects': set(),
                }
            sym_data[sym]['calls'] += 1
            if fname:
                sym_data[sym]['files'].add(fname)
            if project:
                sym_data[sym]['projects'].add(project)

        sorted_rows = sorted(sym_data.items(),
                              key=lambda x: (x[1]['category'], x[0]))

        for row_idx, (symbol, info) in enumerate(sorted_rows, 2):
            file_list = ', '.join(sorted(info['files']))
            project_list = ', '.join(sorted(info['projects']))
            ws.cell(row=row_idx, column=1, value=symbol)
            ws.cell(row=row_idx, column=2, value=info['category'])
            ws.cell(row=row_idx, column=3, value=info['calls'])
            ws.cell(row=row_idx, column=4, value=len(info['files']))
            ws.cell(row=row_idx, column=5, value=file_list)
            ws.cell(row=row_idx, column=6, value=len(info['projects']))
            ws.cell(row=row_idx, column=7, value=project_list)
            if hitls_compat is not None:
                h_status, h_equiv = hitls_compat.lookup(symbol)
                ws.cell(row=row_idx, column=8, value=h_status)
                ws.cell(row=row_idx, column=9, value=h_equiv or '')

        last_letter = chr(64 + len(columns))
        if sorted_rows:
            ws.auto_filter.ref = f"A1:{last_letter}{len(sorted_rows) + 1}"

    def merge_from_results(self, named_results: List[Tuple[str, 'SourceScanResult']],
                           output_path: str,
                           hitls_compat: Optional['HiTLSCompat'] = None) -> Dict:
        """Merge from in-memory SourceScanResult objects.

        Args:
            named_results: List of (project_name, SourceScanResult) tuples
            output_path: Path to write merged XLSX (or JSON if .json)

        Returns:
            Dict with per-project stats for console summary.
        """
        ext = os.path.splitext(output_path)[1].lower()
        if ext == '.json':
            return self._merge_to_json(named_results, output_path,
                                       hitls_compat=hitls_compat)

        project_data = []
        for name, result in named_results:
            rows = self._result_to_rows(result)
            project_data.append({
                'name': name,
                'files_scanned': result.total_files_scanned,
                'rows': rows,
            })
        return self._merge_to_workbook(project_data, output_path,
                                       hitls_compat=hitls_compat)

    def _result_to_rows(self, result: 'SourceScanResult') -> List[List]:
        """Convert SourceScanResult.call_sites to row lists."""
        return [
            _call_site_row(cs, include_recovery=_has_fallback_sites(result))
            for cs in result.call_sites
        ]

    def _merge_to_json(self, named_results, output_path,
                       hitls_compat: Optional['HiTLSCompat'] = None):
        """Merge results to a single JSON file."""
        projects = []
        all_symbols = set()
        stats = []

        for name, result in named_results:
            call_sites = []
            for cs in result.call_sites:
                cs_entry = {
                    'file_path': cs.file_path,
                    'file_name': cs.file_name,
                    'caller_function': cs.caller_function,
                    'line_number': cs.line_number,
                    'ossl_symbol': cs.ossl_symbol,
                    'category': cs.category,
                    'call_args': cs.call_args,
                }
                if cs.extraction_source != 'ast':
                    cs_entry['extraction_source'] = cs.extraction_source
                    cs_entry['confidence'] = cs.confidence
                    cs_entry['parser_diagnostic_class'] = (
                        cs.parser_diagnostic_class
                    )
                if hitls_compat is not None:
                    h_status, h_equiv = hitls_compat.lookup(cs.ossl_symbol)
                    cs_entry['hitls_status'] = h_status
                    cs_entry['hitls_equiv'] = h_equiv
                    cs_entry['hitls_replacement'] = h_equiv
                call_sites.append(cs_entry)
            entry = {
                'project': name,
                'target': result.target,
                'total_files_scanned': result.total_files_scanned,
                'files_with_calls': result.files_with_calls,
                'total_call_sites': result.total_call_sites,
                'unique_symbols': result.unique_symbols,
                'symbols_by_category': result.symbols_by_category,
                'call_sites': call_sites,
            }
            projects.append(entry)
            all_symbols.update(result.unique_symbols)

            cat_syms: Dict[str, set] = {}
            for cs in result.call_sites:
                if cs.category:
                    cat_syms.setdefault(cs.category, set()).add(cs.ossl_symbol)
            top_cat, top_count = '', 0
            if cat_syms:
                top_cat = max(cat_syms, key=lambda c: len(cat_syms[c]))
                top_count = len(cat_syms[top_cat])

            stats.append({
                'project': name,
                'files_scanned': result.total_files_scanned,
                'call_sites': result.total_call_sites,
                'unique_symbols': len(result.unique_symbols),
                'files_with_calls': result.files_with_calls,
                'top_category': top_cat,
                'top_cat_symbols': top_count,
            })

        meta = {
            'report_type': 'combo_scan',
            'total_projects': len(projects),
            'total_call_sites': sum(p['total_call_sites'] for p in projects),
            'total_unique_symbols': len(all_symbols),
        }
        if hitls_compat is not None:
            coverage = hitls_compat.get_coverage_summary(all_symbols)
            meta['hitls_coverage'] = {
                key: coverage[key]
                for key in ('available', 'partial', 'not_available', 'unknown')
            }
            meta['hitls_direct_replace_ratio'] = (
                coverage['direct_replace_ratio']
            )
            meta['hitls_direct_or_partial_replace_ratio'] = (
                coverage['direct_or_partial_replace_ratio']
            )
        merged = {'meta': meta, 'projects': projects}

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(merged, f, indent=2, ensure_ascii=False)
        logger.info("Merged JSON report saved to: %s", output_path)
        return {'sheets': stats, 'total_symbols': len(all_symbols)}
