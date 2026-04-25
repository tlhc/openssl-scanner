"""Summary workbook generation for source-scan JSON reports."""

from __future__ import annotations

import json
import logging
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Sequence

logger = logging.getLogger(__name__)

README_NAMES = (
    "README.md",
    "README_zh.md",
    "README.en.md",
    "README_CN.md",
    "Readme.md",
    "readme.md",
)
METADATA_NAMES = (
    "package.json",
    "Cargo.toml",
    "bundle.json",
    "bundle.json5",
    "oh-package.json5",
    "module.json",
    "module.json5",
)

SUMMARY_HEADERS = [
    "仓库名称",
    "仓库路径",
    "用途描述",
    "地址",
    "HM标签",
    "标签依据",
    "扫描文件数",
    "命中文件数",
    "OpenSSL调用点数量",
    "使用OSSL接口数量",
    "可直接替换接口数",
    "可部分替换接口数",
    "不可替换接口数",
    "未知接口数",
    "直接替换率(符号)",
    "直接+部分替换率(符号)",
    "可直接替换调用点数",
    "可部分替换调用点数",
    "不可替换调用点数",
    "未知调用点数",
    "直接替换率(调用点)",
    "直接+部分替换率(调用点)",
    "Top阻塞接口",
    "JSON报告",
    "XLSX报告",
]

BLOCKER_HEADERS = [
    "仓库名称",
    "仓库路径",
    "OpenSSL接口",
    "HiTLS状态",
    "调用点数",
    "类别",
    "命中文件数",
    "命中文件列表",
    "JSON报告",
]

NONZERO_INDEX_FIELDS = [
    "使用OSSL接口数量",
    "OpenSSL调用点数量",
    "仓库名称",
    "仓库路径",
    "HM标签",
    "地址",
    "直接替换率(符号)",
    "直接+部分替换率(符号)",
    "直接替换率(调用点)",
    "直接+部分替换率(调用点)",
]


def _round_ratio(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round((numerator / denominator) * 100, 2)


def _normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text.replace("\r\n", "\n").replace("\r", "\n")).strip()


def _read_text_file(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        try:
            return path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            return path.read_text(encoding="latin-1", errors="ignore")


def _extract_json_description(file_path: Path) -> str | None:
    try:
        data = json.loads(_read_text_file(file_path))
    except json.JSONDecodeError:
        return None
    for key in ("description", "summary", "name"):
        value = data.get(key)
        if isinstance(value, str) and value.strip():
            return _normalize_text(value)[:300]
    package = data.get("package")
    if isinstance(package, dict):
        value = package.get("description")
        if isinstance(value, str) and value.strip():
            return _normalize_text(value)[:300]
    return None


def _extract_toml_description(file_path: Path) -> str | None:
    match = re.search(
        r'^\s*description\s*=\s*"(.*)"\s*$',
        _read_text_file(file_path),
        re.MULTILINE,
    )
    if match:
        return _normalize_text(match.group(1))[:300]
    return None


def _extract_json5_description(file_path: Path) -> str | None:
    content = _read_text_file(file_path)
    for pattern in (
        r'description\s*:\s*"([^"]+)"',
        r'summary\s*:\s*"([^"]+)"',
        r"description\s*:\s*'([^']+)'",
        r"summary\s*:\s*'([^']+)'",
    ):
        match = re.search(pattern, content)
        if match:
            return _normalize_text(match.group(1))[:300]
    return None


class SourceSummaryBuilder:
    """Build repo-level workbook summaries from source-scan JSON reports."""

    def __init__(
        self,
        *,
        source_root: str | None = None,
        case_sensitive_source_root: str | None = None,
        manifest_path: str | None = None,
    ) -> None:
        self.source_root: Path | None = (
            Path(source_root).resolve() if source_root else None
        )
        if case_sensitive_source_root:
            self.case_sensitive_source_root: Path | None = Path(
                case_sensitive_source_root
            ).resolve()
        else:
            self.case_sensitive_source_root = None
        self.manifest_path: Path | None = (
            Path(manifest_path).resolve() if manifest_path else None
        )
        self.manifest_root: Path | None = (
            self.manifest_path.parent if self.manifest_path else None
        )
        self.manifest_projects = (
            self._load_manifest_projects(self.manifest_path)
            if self.manifest_path and self.manifest_path.is_file()
            else {}
        )

    def build_from_inputs(
        self,
        inputs: Sequence[str],
    ) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
        report_paths = self._discover_report_paths(inputs)
        rows: list[dict[str, Any]] = []
        blockers: list[dict[str, Any]] = []
        for report_path in report_paths:
            report = self._load_source_scan(report_path)
            if report is None:
                continue
            row, blocker_rows = self._build_summary_row(report_path, report)
            rows.append(row)
            blockers.extend(blocker_rows)

        rows.sort(
            key=lambda item: (
                -int(item["使用OSSL接口数量"]),
                -int(item["OpenSSL调用点数量"]),
                str(item["仓库名称"]),
            )
        )
        blockers.sort(
            key=lambda item: (
                str(item["仓库名称"]),
                -int(item["调用点数"]),
                str(item["OpenSSL接口"]),
                str(item["HiTLS状态"]),
            )
        )
        return rows, blockers

    def write_workbook(
        self,
        rows: Sequence[dict[str, Any]],
        blockers: Sequence[dict[str, Any]],
        output_path: str,
    ) -> None:
        from . import _vendor  # noqa: F401,I001
        from openpyxl import Workbook  # type: ignore[import-untyped]
        from openpyxl.styles import Font  # type: ignore[import-untyped]
        from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

        workbook = Workbook()
        all_sheet = workbook.active
        all_sheet.title = "AllRepos"
        used_sheet = workbook.create_sheet("OpenSSLUsed")
        blocker_sheet = workbook.create_sheet("TopBlockers")

        self._write_sheet(all_sheet, list(rows), SUMMARY_HEADERS, Font(bold=True), get_column_letter)
        nonzero_rows = [row for row in rows if int(row["使用OSSL接口数量"]) > 0]
        self._write_sheet(used_sheet, nonzero_rows, SUMMARY_HEADERS, Font(bold=True), get_column_letter)
        self._write_sheet(blocker_sheet, list(blockers), BLOCKER_HEADERS, Font(bold=True), get_column_letter)

        out_path = Path(output_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(out_path))

    def write_nonzero_index(
        self,
        rows: Sequence[dict[str, Any]],
        output_path: str,
    ) -> None:
        lines: list[str] = []
        for row in rows:
            if self._int_value(row.get("使用OSSL接口数量")) <= 0:
                continue
            lines.append("\t".join(str(row[field]) for field in NONZERO_INDEX_FIELDS))
        Path(output_path).write_text(
            "\n".join(lines) + ("\n" if lines else ""),
            encoding="utf-8",
        )

    def _discover_report_paths(self, inputs: Sequence[str]) -> list[Path]:
        report_paths: list[Path] = []
        seen: set[Path] = set()
        for raw in inputs:
            path = Path(raw).resolve()
            if path.is_dir():
                candidates = sorted(path.glob("*.json"))
            elif path.is_file():
                candidates = [path]
            else:
                raise FileNotFoundError(f"Input not found: {path}")
            for candidate in candidates:
                if candidate not in seen:
                    seen.add(candidate)
                    report_paths.append(candidate)
        return sorted(report_paths)

    def _load_source_scan(self, report_path: Path) -> dict[str, Any] | None:
        try:
            payload = json.loads(report_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            logger.warning("Skipping unreadable report %s: %s", report_path, exc)
            return None
        if not isinstance(payload, dict):
            return None
        meta = payload.get("meta")
        summary = payload.get("summary")
        if not isinstance(meta, dict) or not isinstance(summary, dict):
            return None
        if meta.get("report_type") != "source_scan":
            return None
        return payload

    def _build_summary_row(
        self,
        report_path: Path,
        payload: dict[str, Any],
    ) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        meta = self._dict_value(payload.get("meta"))
        summary = self._dict_value(payload.get("summary"))
        call_sites = payload.get("call_sites", [])
        target_path = str(meta.get("target", ""))
        target = Path(target_path)
        relative_path = self._resolve_relative_path(target_path)
        repo_path = target if target.exists() else None
        manifest_info = self.manifest_projects.get(relative_path, {})
        repo_name = str(manifest_info.get("repo_name") or Path(relative_path).name or report_path.stem)
        repo_url = str(manifest_info.get("repo_url") or "")
        purpose = self._extract_purpose_description(repo_path) if repo_path else ""
        hm_tag, hm_tag_basis = self._classify_hm_tag(relative_path)

        symbol_coverage = summary.get("hitls_coverage")
        if not isinstance(symbol_coverage, dict):
            symbol_coverage = {}
        symbol_available = int(symbol_coverage.get("available", 0))
        symbol_partial = int(symbol_coverage.get("partial", 0))
        symbol_not_available = int(symbol_coverage.get("not_available", 0))
        symbol_unknown = int(symbol_coverage.get("unknown", 0))
        symbol_total = int(summary.get("unique_symbols_count", 0))

        call_counts = {"available": 0, "partial": 0, "not_available": 0, "unknown": 0}
        blocker_map: dict[tuple[str, str], dict[str, Any]] = {}
        if isinstance(call_sites, list):
            for item in call_sites:
                if not isinstance(item, dict):
                    continue
                status = str(item.get("hitls_status") or "unknown")
                if status not in call_counts:
                    status = "unknown"
                call_counts[status] += 1
                if status not in {"not_available", "unknown"}:
                    continue
                symbol = str(item.get("ossl_symbol") or "")
                if not symbol:
                    continue
                key = (symbol, status)
                entry = blocker_map.setdefault(
                    key,
                    {
                        "symbol": symbol,
                        "status": status,
                        "count": 0,
                        "categories": set(),
                        "files": set(),
                    },
                )
                entry["count"] = self._int_value(entry.get("count")) + 1
                category = str(item.get("category") or "")
                if category:
                    entry["categories"].add(category)
                file_name = str(item.get("file_name") or item.get("file_path") or "")
                if file_name:
                    entry["files"].add(file_name)

        call_total = int(summary.get("total_call_sites", 0))
        row = {
            "仓库名称": repo_name,
            "仓库路径": relative_path,
            "用途描述": purpose,
            "地址": repo_url,
            "HM标签": hm_tag,
            "标签依据": hm_tag_basis,
            "扫描文件数": int(summary.get("total_files_scanned", 0)),
            "命中文件数": int(summary.get("files_with_calls", 0)),
            "OpenSSL调用点数量": call_total,
            "使用OSSL接口数量": symbol_total,
            "可直接替换接口数": symbol_available,
            "可部分替换接口数": symbol_partial,
            "不可替换接口数": symbol_not_available,
            "未知接口数": symbol_unknown,
            "直接替换率(符号)": float(summary.get("hitls_direct_replace_ratio", _round_ratio(symbol_available, symbol_total))),
            "直接+部分替换率(符号)": float(summary.get("hitls_direct_or_partial_replace_ratio", _round_ratio(symbol_available + symbol_partial, symbol_total))),
            "可直接替换调用点数": call_counts["available"],
            "可部分替换调用点数": call_counts["partial"],
            "不可替换调用点数": call_counts["not_available"],
            "未知调用点数": call_counts["unknown"],
            "直接替换率(调用点)": _round_ratio(call_counts["available"], call_total),
            "直接+部分替换率(调用点)": _round_ratio(call_counts["available"] + call_counts["partial"], call_total),
            "Top阻塞接口": self._format_top_blockers(blocker_map),
            "JSON报告": str(report_path),
            "XLSX报告": str(report_path.with_suffix(".xlsx")),
        }

        blocker_rows: list[dict[str, Any]] = []
        for entry in sorted(
            blocker_map.values(),
            key=lambda item: (
                -self._int_value(item.get("count")),
                str(item.get("symbol", "")),
                str(item.get("status", "")),
            ),
        ):
            files = sorted(str(v) for v in entry.get("files", set()))
            categories = sorted(str(v) for v in entry.get("categories", set()))
            blocker_rows.append(
                {
                    "仓库名称": repo_name,
                    "仓库路径": relative_path,
                    "OpenSSL接口": str(entry.get("symbol", "")),
                    "HiTLS状态": str(entry.get("status", "")),
                    "调用点数": self._int_value(entry.get("count")),
                    "类别": ", ".join(categories),
                    "命中文件数": len(files),
                    "命中文件列表": ", ".join(files),
                    "JSON报告": str(report_path),
                }
            )

        return row, blocker_rows

    def _write_sheet(
        self,
        worksheet: Any,
        rows: Sequence[dict[str, Any]],
        headers: Sequence[str],
        header_font: Any,
        get_column_letter: Any,
    ) -> None:
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = header_font
        for row in rows:
            worksheet.append([row.get(header, "") for header in headers])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 80)

    def _resolve_relative_path(self, target_path: str) -> str:
        target = Path(target_path).resolve()
        candidates = [root for root in (self.source_root, self.case_sensitive_source_root) if root]
        for root in candidates:
            try:
                return str(target.relative_to(root))
            except ValueError:
                marker = f"/{root.name}/"
                target_text = str(target)
                if marker in target_text:
                    return target_text.split(marker, 1)[1]
        return target.name

    def _load_manifest_projects(
        self,
        manifest_path: Path,
    ) -> dict[str, dict[str, str]]:
        project_by_path: dict[str, dict[str, str]] = {}
        self._collect_manifest_projects(
            ET.parse(manifest_path).getroot(),
            self.manifest_root or manifest_path.parent,
            {},
            "",
            project_by_path,
        )
        return project_by_path

    def _collect_manifest_projects(
        self,
        manifest_node: ET.Element,
        base_dir: Path,
        remote_by_name: dict[str, str],
        default_remote: str,
        project_by_path: dict[str, dict[str, str]],
    ) -> None:
        current_remotes = dict(remote_by_name)
        current_default = default_remote
        for remote_node in manifest_node.findall("remote"):
            remote_name = remote_node.attrib.get("name", "")
            remote_fetch = remote_node.attrib.get("fetch", "").rstrip("/")
            if remote_name and remote_fetch:
                current_remotes[remote_name] = remote_fetch
        default_node = manifest_node.find("default")
        if default_node is not None:
            current_default = default_node.attrib.get("remote", current_default)
        for project_node in manifest_node.findall("project"):
            project_path = project_node.attrib.get("path", "")
            project_name = project_node.attrib.get("name", "")
            if not project_path or not project_name:
                continue
            remote_name = project_node.attrib.get("remote", current_default)
            remote_fetch = current_remotes.get(remote_name, "")
            repo_url = f"{remote_fetch}/{project_name}" if remote_fetch else ""
            project_by_path[project_path] = {
                "repo_name": project_name,
                "repo_url": repo_url,
            }
        for include_node in manifest_node.findall("include"):
            include_name = include_node.attrib.get("name", "")
            if not include_name:
                continue
            include_path = base_dir / include_name
            if include_path.is_file():
                self._collect_manifest_projects(
                    ET.parse(include_path).getroot(),
                    base_dir,
                    current_remotes,
                    current_default,
                    project_by_path,
                )

    def _classify_hm_tag(self, relative_path: str) -> tuple[str, str]:
        if not relative_path or relative_path == Path(relative_path).name:
            return "", ""
        if relative_path == "third_party" or relative_path.startswith("third_party/"):
            return "第三方", "路径前缀 third_party/"
        if relative_path in ("vendor", "device") or relative_path.startswith(("vendor/", "device/")):
            return "厂商/芯片", "路径前缀 vendor/ 或 device/"
        if relative_path == "productdefine" or relative_path.startswith("productdefine/"):
            return "设备预装(产品相关)", "路径前缀 productdefine/"
        if relative_path == "kernel" or relative_path.startswith("kernel/") or relative_path == "base/tee/tee_os_kernel":
            return "内核/上游移植", "路径前缀 kernel/ 或精确路径 base/tee/tee_os_kernel"
        if relative_path == "applications/sample" or relative_path.startswith("applications/sample/"):
            return "HM自研组件", "路径前缀 applications/sample/"
        if relative_path == "applications/standard" or relative_path.startswith("applications/standard/"):
            app_leaf = relative_path.split("/", 2)[-1]
            if app_leaf == "app_samples" or app_leaf.endswith("_wrapper"):
                return "HM自研组件", "applications/standard 下的 samples/wrapper 仓"
            return "HM预置应用", "路径前缀 applications/standard/ 且非 samples/wrapper"
        prefixes = (
            "arkcompiler/",
            "base/",
            "build/",
            "commonlibrary/",
            "developtools/",
            "docs/",
            "docs_cangjie/",
            "domains/",
            "drivers/",
            "foundation/",
            "ide/",
            "interface/",
            "napi_generator/",
            "test/",
        )
        if relative_path in tuple(prefix.rstrip("/") for prefix in prefixes) or relative_path.startswith(prefixes):
            return "HM自研组件", "OpenHarmony 自有子系统路径前缀"
        return "待人工确认", "未命中当前规则，需人工判定"

    def _extract_purpose_description(self, repo_path: Path) -> str:
        for readme_name in README_NAMES:
            readme_path = repo_path / readme_name
            if readme_path.exists():
                desc = self._extract_readme_description(readme_path)
                if desc:
                    return desc
        for metadata_name in METADATA_NAMES:
            metadata_path = repo_path / metadata_name
            if not metadata_path.exists():
                continue
            if metadata_path.name == "package.json" or metadata_path.suffix == ".json":
                desc = _extract_json_description(metadata_path)
            elif metadata_path.suffix == ".toml":
                desc = _extract_toml_description(metadata_path)
            else:
                desc = _extract_json5_description(metadata_path)
            if desc:
                return desc
        return ""

    def _extract_readme_description(self, readme_path: Path) -> str | None:
        lines = _read_text_file(readme_path).splitlines()
        paragraph: list[str] = []
        in_code_block = False
        for raw_line in lines[:200]:
            line = raw_line.strip()
            if line.startswith("```"):
                in_code_block = not in_code_block
                continue
            if in_code_block:
                continue
            if not line:
                if paragraph:
                    break
                continue
            if line.startswith("#"):
                continue
            if line.startswith("![](") or line.startswith("[!["):
                continue
            if line.startswith("|") and line.endswith("|"):
                continue
            if re.fullmatch(r"[-=]{3,}", line):
                continue
            paragraph.append(line)
        if paragraph:
            return _normalize_text(" ".join(paragraph))[:300]
        return None

    def _format_top_blockers(
        self,
        blocker_map: dict[tuple[str, str], dict[str, Any]],
    ) -> str:
        parts: list[str] = []
        for entry in sorted(
            blocker_map.values(),
            key=lambda item: (
                -self._int_value(item.get("count")),
                str(item.get("symbol", "")),
                str(item.get("status", "")),
            ),
        )[:5]:
            parts.append(
                f"{entry.get('symbol', '')}[{entry.get('status', '')}]"
                f"({self._int_value(entry.get('count'))})"
            )
        return ", ".join(parts)

    def _dict_value(self, value: Any) -> dict[str, Any]:
        return value if isinstance(value, dict) else {}

    def _int_value(self, value: Any) -> int:
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str):
            try:
                return int(value)
            except ValueError:
                return 0
        return 0
