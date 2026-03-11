"""
R1: auto_filter correctness tests for all 5 XLSX overflow points.

Verifies:
- auto_filter is applied to the FIRST sheet (not continuation sheets)
- auto_filter row range stays within XLSX_MAX_ROW
- auto_filter ref is correct in both overflow and no-overflow cases
- exporter.py File-Symbol and Import Chains sheets have no auto_filter

Overflow locations tested:
1. source_exporter.SourceExcelExporter.export()          (Call Sites)
2. source_exporter.SourceMergeExporter._merge_to_workbook() (per-project)
3. source_diff.SourceDiffExcelExporter._write_callsite_sheet() (Call Site Delta)
4. exporter.ExcelExporter._create_file_symbol_sheet()     (File-Symbol)
5. exporter.ExcelExporter._create_import_chains_sheet()   (Import Chains)
"""

import json
import os
import sys
from unittest import mock

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openssl_scanner import _vendor  # noqa: F401
from openpyxl import load_workbook

from openssl_scanner.source_analyzer import CallSite, SourceScanResult
from openssl_scanner.source_exporter import (
    SourceExcelExporter, SourceMergeExporter, COLUMNS, LAST_COL_LETTER,
)
from openssl_scanner.exporter import ExcelExporter
from openssl_scanner.source_diff import (
    ArgDelta, CallSiteDelta, DiffResult, DiffStatus, FileDelta,
    MetricDelta, ProjectDelta, SourceDiffExcelExporter, SymbolDelta,
)

MAX_ROW_PATCH = 50


def _make_call_site(idx=0, symbol="SSL_connect", category="ssl_core"):
    return CallSite(
        file_path=f"/tmp/test/file{idx}.c",
        file_name=f"file{idx}.c",
        caller_function=f"func{idx}",
        line_number=idx + 1,
        column=4,
        ossl_symbol=symbol,
        category=category,
        call_args=f"(arg{idx})",
        language="c",
    )


def _make_result(call_sites=None, target="/tmp/test"):
    if call_sites is None:
        call_sites = []
    unique = sorted(set(cs.ossl_symbol for cs in call_sites))
    cat_map = {}
    for cs in call_sites:
        cat_map.setdefault(cs.category, [])
        if cs.ossl_symbol not in cat_map[cs.category]:
            cat_map[cs.category].append(cs.ossl_symbol)
    return SourceScanResult(
        target=target,
        scan_time="2026-01-01T00:00:00",
        tool_version="1.0.0",
        total_files_scanned=1,
        files_with_calls=1 if call_sites else 0,
        total_call_sites=len(call_sites),
        unique_symbols=unique,
        symbols_by_category=cat_map,
        call_sites=call_sites,
        errors=[],
    )


def _make_merge_rows(n):
    return [
        [f"/f{i}.c", f"f{i}.c", f"fn{i}", i + 1,
         "SSL_connect", "ssl_core", "()", "dynamic-link"]
        for i in range(n)
    ]


def _make_elf_report(file_symbol_pairs=None, import_chains=None):
    by_file = {}
    by_category = {}
    if file_symbol_pairs:
        for fpath, syms in file_symbol_pairs:
            by_file[fpath] = {'count': len(syms), 'symbols': syms}
            for sym in syms:
                cat = 'ssl_core' if sym.startswith('SSL_') else 'crypto_evp'
                by_category.setdefault(cat, {'count': 0, 'symbols': []})
                if sym not in by_category[cat]['symbols']:
                    by_category[cat]['symbols'].append(sym)
                    by_category[cat]['count'] += 1

    chains = import_chains or {}

    return {
        'meta': {
            'tool_version': '1.0.0',
            'report_type': 'single',
            'scan_time': '2026-01-01T00:00:00',
            'scan_root': '/usr/bin/app',
            'target_arch': 'aarch64',
        },
        'summary': {
            'total_files_scanned': 1,
            'total_elf_files': 1,
            'files_with_openssl_deps': 1,
            'total_openssl_symbols': 0,
            'unique_openssl_symbols': 0,
            'openssl_libs_found': [],
            'files_with_static_openssl': 0,
            'files_with_dlopen': 0,
            'dlopen_unique_symbols': 0,
            'dlopen_libs_detected': [],
        },
        'openssl_symbols': {
            'by_file': by_file,
            'by_category': by_category,
            'by_depth': {},
            'import_chains': chains,
            'all_unique': [],
        },
        'files_detail': [],
        'dependency_tree': {},
    }


def _make_diff_result(n_callsites, is_combo=False):
    csd_list = []
    for i in range(n_callsites):
        ad = ArgDelta(
            status=DiffStatus.ADDED,
            old_line=None,
            new_line=i + 1,
            old_args="",
            new_args=f"(arg{i})",
        )
        csd = CallSiteDelta(
            status=DiffStatus.ADDED,
            identity_key=(f"/tmp/f{i}.c", f"func{i}", "SSL_connect"),
            old_count=0,
            new_count=1,
            old_lines=[],
            new_lines=[i + 1],
            category="ssl_core",
            arg_deltas=[ad],
        )
        csd_list.append(csd)

    proj = ProjectDelta(
        project="test_proj",
        metrics=[MetricDelta("total_call_sites", 0, n_callsites, n_callsites)],
        call_site_delta=csd_list,
        symbol_delta=[
            SymbolDelta(DiffStatus.ADDED, "SSL_connect", "ssl_core", 0, n_callsites)
        ],
        file_delta=[
            FileDelta(DiffStatus.ADDED, f"/tmp/f{i}.c", 0, 1)
            for i in range(min(n_callsites, 10))
        ],
    )
    return DiffResult(
        old_label="old",
        new_label="new",
        projects=[proj],
        old_scan_time="2026-01-01T00:00:00",
        new_scan_time="2026-01-02T00:00:00",
        is_combo=is_combo,
    )


def _auto_filter_ref(ws):
    """Extract auto_filter.ref, normalizing empty/None to None."""
    ref = ws.auto_filter.ref
    if ref is None or ref == "" or ref == "A1:A1":
        return None
    return ref


class TestAutoFilterOverflow:
    """Overflow case (120 items, MAX_ROW=50): verify auto_filter on first sheet only,
    row range capped at MAX_ROW, continuation sheets have no auto_filter."""

    def test_source_exporter_overflow(self, tmp_path):
        """Location 1: SourceExcelExporter -- 120 call sites, MAX_ROW=50."""
        path = str(tmp_path / "overflow.xlsx")
        sites = [_make_call_site(i) for i in range(120)]
        result = _make_result(sites)

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceExcelExporter().export(result, path)

        wb = load_workbook(path)

        ws1 = wb["OpenSSL Call Sites"]
        ref = _auto_filter_ref(ws1)
        assert ref is not None, "First sheet must have auto_filter"
        assert ref == f"A1:{LAST_COL_LETTER}{MAX_ROW_PATCH}", (
            f"First sheet auto_filter should be A1:{LAST_COL_LETTER}{MAX_ROW_PATCH}, "
            f"got {ref}"
        )

        end_row = int(ref.split(LAST_COL_LETTER)[1])
        assert end_row <= MAX_ROW_PATCH, (
            f"auto_filter end row {end_row} exceeds XLSX_MAX_ROW {MAX_ROW_PATCH}"
        )

        for sn in wb.sheetnames:
            if sn.startswith("OpenSSL Call Sites ("):
                ws_cont = wb[sn]
                cont_ref = _auto_filter_ref(ws_cont)
                assert cont_ref is None, (
                    f"Continuation sheet '{sn}' should not have auto_filter, got {cont_ref}"
                )

    def test_merge_exporter_overflow(self, tmp_path):
        """Location 2: SourceMergeExporter -- 120 rows, MAX_ROW=50."""
        path = str(tmp_path / "merge_overflow.xlsx")
        rows = _make_merge_rows(120)
        project_data = [{'name': 'proj', 'files_scanned': 10, 'rows': rows}]

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)

        ws1 = wb["proj"]
        ref = _auto_filter_ref(ws1)
        assert ref is not None, "First sheet must have auto_filter"
        assert ref == f"A1:{LAST_COL_LETTER}{MAX_ROW_PATCH}", (
            f"First sheet auto_filter should be A1:{LAST_COL_LETTER}{MAX_ROW_PATCH}, "
            f"got {ref}"
        )

        end_row = int(ref.split(LAST_COL_LETTER)[1])
        assert end_row <= MAX_ROW_PATCH, (
            f"auto_filter end row {end_row} exceeds XLSX_MAX_ROW {MAX_ROW_PATCH}"
        )

        for sn in wb.sheetnames:
            if sn.startswith("proj") and "(" in sn:
                ws_cont = wb[sn]
                cont_ref = _auto_filter_ref(ws_cont)
                assert cont_ref is None, (
                    f"Continuation sheet '{sn}' should not have auto_filter, got {cont_ref}"
                )

    def test_source_diff_overflow_first_sheet(self, tmp_path):
        """Location 3: SourceDiffExcelExporter -- 120 items, MAX_ROW=50.
        Verify first sheet auto_filter ends at row MAX_ROW_PATCH."""
        path = str(tmp_path / "diff_overflow.xlsx")
        result = _make_diff_result(120, is_combo=False)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)

        ws1 = wb["Call Site Delta"]
        ref = _auto_filter_ref(ws1)
        assert ref is not None, "First sheet must have auto_filter"

        parts = ref.replace("A1:", "")
        end_row = int(''.join(c for c in parts if c.isdigit()))
        assert end_row == MAX_ROW_PATCH, (
            f"First sheet auto_filter end row should be {MAX_ROW_PATCH}, got {end_row} "
            f"(ref={ref})"
        )
        assert end_row <= MAX_ROW_PATCH, (
            f"auto_filter end row {end_row} exceeds XLSX_MAX_ROW {MAX_ROW_PATCH}"
        )

    def test_source_diff_overflow_last_sheet_has_auto_filter(self, tmp_path):
        """Location 3: SourceDiffExcelExporter -- current behavior sets auto_filter
        on the LAST continuation sheet as well. Document this behavior."""
        path = str(tmp_path / "diff_overflow2.xlsx")
        result = _make_diff_result(120, is_combo=False)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)

        callsite_sheets = [
            sn for sn in wb.sheetnames if sn.startswith("Call Site Delta")
        ]
        assert len(callsite_sheets) == 3, (
            f"Expected 3 Call Site Delta sheets, got {callsite_sheets}"
        )

        last_sheet_name = callsite_sheets[-1]
        ws_last = wb[last_sheet_name]
        last_ref = _auto_filter_ref(ws_last)
        assert last_ref is not None, (
            "source_diff sets auto_filter on last continuation sheet "
            "(current behavior, differs from source_exporter)"
        )

        parts = last_ref.replace("A1:", "")
        end_row = int(''.join(c for c in parts if c.isdigit()))
        assert end_row < MAX_ROW_PATCH, (
            f"Last sheet auto_filter end row {end_row} should be less than "
            f"{MAX_ROW_PATCH} since it's a partial sheet"
        )

    def test_file_symbol_no_auto_filter(self, tmp_path):
        """Location 4: exporter.py File-Symbol -- no auto_filter set (expected)."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        pairs = [(f"/bin/app{i}", [f"SSL_{i}"]) for i in range(120)]
        data = _make_elf_report(file_symbol_pairs=pairs)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        assert "File-Symbol" in wb.sheetnames

        ws = wb["File-Symbol"]
        ref = _auto_filter_ref(ws)
        assert ref is None, (
            f"File-Symbol sheet should have no auto_filter (current code), got {ref}"
        )

    def test_import_chains_no_auto_filter(self, tmp_path):
        """Location 5: exporter.py Import Chains -- no auto_filter set (expected)."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        chains = {}
        for i in range(120):
            sym = f"SSL_{i}"
            chains[sym] = [
                {"source_file": f"/bin/app{i}",
                 "chain": f"app{i} -> libssl.so", "depth": 1}
            ]
        data = _make_elf_report(import_chains=chains)
        by_cat = {'ssl_core': {'count': 120, 'symbols': list(chains.keys())}}
        data['openssl_symbols']['by_category'] = by_cat
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        assert "Import Chains" in wb.sheetnames

        ws = wb["Import Chains"]
        ref = _auto_filter_ref(ws)
        assert ref is None, (
            f"Import Chains sheet should have no auto_filter (current code), got {ref}"
        )


class TestAutoFilterNoOverflow:
    """No-overflow case (30 items, MAX_ROW=50): verify auto_filter covers all rows."""

    def test_source_exporter_no_overflow(self, tmp_path):
        """Location 1: 30 call sites, MAX_ROW=50 -> 1 sheet, auto_filter to row 31."""
        path = str(tmp_path / "no_overflow.xlsx")
        sites = [_make_call_site(i) for i in range(30)]
        result = _make_result(sites)

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceExcelExporter().export(result, path)

        wb = load_workbook(path)
        ws = wb["OpenSSL Call Sites"]
        ref = _auto_filter_ref(ws)
        assert ref == f"A1:{LAST_COL_LETTER}31", (
            f"auto_filter should be A1:{LAST_COL_LETTER}31 (header + 30 data), got {ref}"
        )
        assert "OpenSSL Call Sites (2)" not in wb.sheetnames

    def test_merge_exporter_no_overflow(self, tmp_path):
        """Location 2: 30 rows, MAX_ROW=50 -> 1 sheet, auto_filter to row 31."""
        path = str(tmp_path / "merge_no_overflow.xlsx")
        rows = _make_merge_rows(30)
        project_data = [{'name': 'proj', 'files_scanned': 5, 'rows': rows}]

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        ws = wb["proj"]
        ref = _auto_filter_ref(ws)
        assert ref == f"A1:{LAST_COL_LETTER}31", (
            f"auto_filter should be A1:{LAST_COL_LETTER}31, got {ref}"
        )

    def test_source_diff_no_overflow(self, tmp_path):
        """Location 3: 30 call site deltas, MAX_ROW=50 -> 1 sheet."""
        path = str(tmp_path / "diff_no_overflow.xlsx")
        result = _make_diff_result(30, is_combo=False)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        ws = wb["Call Site Delta"]
        ref = _auto_filter_ref(ws)
        assert ref is not None, "auto_filter must be set for non-empty data"

        parts = ref.replace("A1:", "")
        end_row = int(''.join(c for c in parts if c.isdigit()))
        assert end_row == 31, (
            f"auto_filter end row should be 31 (header + 30 data), got {end_row}"
        )
        assert "Call Site Delta (2)" not in wb.sheetnames

    def test_file_symbol_no_overflow_no_auto_filter(self, tmp_path):
        """Location 4: File-Symbol with 30 items, no overflow, still no auto_filter."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        pairs = [(f"/bin/app{i}", [f"SSL_{i}"]) for i in range(30)]
        data = _make_elf_report(file_symbol_pairs=pairs)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["File-Symbol"]
        ref = _auto_filter_ref(ws)
        assert ref is None, (
            f"File-Symbol has no auto_filter in current code, got {ref}"
        )
        assert "File-Symbol (2)" not in wb.sheetnames

    def test_import_chains_no_overflow_no_auto_filter(self, tmp_path):
        """Location 5: Import Chains with 30 items, no overflow, still no auto_filter."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        chains = {}
        for i in range(30):
            sym = f"SSL_{i}"
            chains[sym] = [
                {"source_file": f"/bin/app{i}",
                 "chain": f"app{i} -> libssl.so", "depth": 1}
            ]
        data = _make_elf_report(import_chains=chains)
        by_cat = {'ssl_core': {'count': 30, 'symbols': list(chains.keys())}}
        data['openssl_symbols']['by_category'] = by_cat
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Import Chains"]
        ref = _auto_filter_ref(ws)
        assert ref is None, (
            f"Import Chains has no auto_filter in current code, got {ref}"
        )


class TestSourceDiffAutoFilterArithmetic:
    """Task 3: Verify _set_auto_filter(ws_first, XLSX_MAX_ROW + 1, num_cols)
    produces the correct ref = A1:{col}XLSX_MAX_ROW."""

    def test_overflow_first_sheet_ref_exactly_max_row(self, tmp_path):
        """XLSX_MAX_ROW+1 passed to _set_auto_filter -> ref ends at XLSX_MAX_ROW.
        With MAX_ROW=50: _set_auto_filter(ws, 51, 9) -> A1:I50."""
        path = str(tmp_path / "arith.xlsx")
        result = _make_diff_result(120, is_combo=False)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        ws1 = wb["Call Site Delta"]
        ref = _auto_filter_ref(ws1)
        assert ref is not None

        assert ref == "A1:I50", (
            f"Expected A1:I50 (9 columns, end row = MAX_ROW), got {ref}"
        )

    def test_overflow_last_sheet_ref_less_than_max_row(self, tmp_path):
        """Last continuation sheet should have auto_filter ending at actual data row,
        which is less than MAX_ROW for a partial sheet."""
        path = str(tmp_path / "arith2.xlsx")
        result = _make_diff_result(120, is_combo=False)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        callsite_sheets = [
            sn for sn in wb.sheetnames if sn.startswith("Call Site Delta")
        ]
        ws_last = wb[callsite_sheets[-1]]
        ref = _auto_filter_ref(ws_last)
        assert ref is not None

        parts = ref.replace("A1:", "")
        end_row = int(''.join(c for c in parts if c.isdigit()))
        assert end_row < MAX_ROW_PATCH, (
            f"Last sheet end row {end_row} should be < {MAX_ROW_PATCH}"
        )

        expected_last_sheet_rows = 120 - 49 * 2
        expected_end = expected_last_sheet_rows + 1
        assert end_row == expected_end, (
            f"Last sheet should have {expected_last_sheet_rows} data rows, "
            f"auto_filter end={end_row}, expected {expected_end}"
        )

    def test_no_overflow_single_sheet_ref(self, tmp_path):
        """No overflow: _set_auto_filter(ws_first, row, 9) where row=31.
        ref should be A1:I30."""
        path = str(tmp_path / "arith3.xlsx")
        result = _make_diff_result(30, is_combo=False)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        ws = wb["Call Site Delta"]
        ref = _auto_filter_ref(ws)
        assert ref == "A1:I31", (
            f"Expected A1:I31 (header + 30 data rows), got {ref}"
        )

    def test_combo_overflow_first_sheet(self, tmp_path):
        """Combo mode overflow: first sheet auto_filter has 10 columns (includes Project)."""
        path = str(tmp_path / "combo_arith.xlsx")
        result = _make_diff_result(120, is_combo=True)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        ws1 = wb["Call Site Delta"]
        ref = _auto_filter_ref(ws1)
        assert ref is not None
        assert ref == "A1:J50", (
            f"Combo mode: expected A1:J50 (10 columns, end=MAX_ROW), got {ref}"
        )


class TestAutoFilterBoundary:
    """Boundary case: exactly MAX_ROW-1 data rows (fills exactly, no overflow)."""

    def test_source_exporter_exact_fill(self, tmp_path):
        """49 call sites at MAX_ROW=50: rows 2-50, auto_filter A1:H50."""
        path = str(tmp_path / "exact.xlsx")
        sites = [_make_call_site(i) for i in range(49)]
        result = _make_result(sites)

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceExcelExporter().export(result, path)

        wb = load_workbook(path)
        ws = wb["OpenSSL Call Sites"]
        ref = _auto_filter_ref(ws)
        assert ref == f"A1:{LAST_COL_LETTER}50", (
            f"49 data rows -> auto_filter to row 50, got {ref}"
        )
        assert "OpenSSL Call Sites (2)" not in wb.sheetnames

    def test_source_exporter_one_over_exact(self, tmp_path):
        """50 call sites at MAX_ROW=50: first sheet gets 49 data rows,
        second sheet gets 1 data row. auto_filter on first = A1:H50."""
        path = str(tmp_path / "one_over.xlsx")
        sites = [_make_call_site(i) for i in range(50)]
        result = _make_result(sites)

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceExcelExporter().export(result, path)

        wb = load_workbook(path)
        ws1 = wb["OpenSSL Call Sites"]
        ref = _auto_filter_ref(ws1)
        assert ref == f"A1:{LAST_COL_LETTER}50", (
            f"50 data rows overflow -> first sheet auto_filter capped at 50, got {ref}"
        )
        assert "OpenSSL Call Sites (2)" in wb.sheetnames

        ws2 = wb["OpenSSL Call Sites (2)"]
        cont_ref = _auto_filter_ref(ws2)
        assert cont_ref is None, (
            f"Continuation sheet should not have auto_filter, got {cont_ref}"
        )

    def test_merge_exporter_exact_fill(self, tmp_path):
        """49 rows at MAX_ROW=50: auto_filter A1:H50."""
        path = str(tmp_path / "merge_exact.xlsx")
        rows = _make_merge_rows(49)
        project_data = [{'name': 'proj', 'files_scanned': 5, 'rows': rows}]

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        ws = wb["proj"]
        ref = _auto_filter_ref(ws)
        assert ref == f"A1:{LAST_COL_LETTER}50", (
            f"49 data rows -> auto_filter to row 50, got {ref}"
        )

    def test_source_diff_exact_fill(self, tmp_path):
        """49 arg deltas at MAX_ROW=50: auto_filter A1:I50."""
        path = str(tmp_path / "diff_exact.xlsx")
        result = _make_diff_result(49, is_combo=False)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", MAX_ROW_PATCH):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        ws = wb["Call Site Delta"]
        ref = _auto_filter_ref(ws)
        assert ref == "A1:I50", (
            f"49 data rows -> auto_filter to row 50, got {ref}"
        )
        assert "Call Site Delta (2)" not in wb.sheetnames
