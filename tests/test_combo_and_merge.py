"""Tests for combo-scan helpers, SourceMergeExporter, and multi-target source mode."""

import json
import os
import tempfile

import pytest

tree_sitter = pytest.importorskip("tree_sitter")

from openssl_scanner.source_analyzer import CallSite, SourceScanResult
from openssl_scanner.source_exporter import (
    SourceExcelExporter,
    SourceJsonExporter,
    SourceMergeExporter,
)


def _make_call_site(symbol="SSL_connect", category="ssl_core",
                    file_path="/tmp/src/a.c", file_name="a.c",
                    caller="main", line=10, args="(ctx)"):
    return CallSite(
        file_path=file_path,
        file_name=file_name,
        caller_function=caller,
        line_number=line,
        column=4,
        ossl_symbol=symbol,
        category=category,
        call_args=args,
        language="c",
    )


def _make_result(target="/tmp/proj", call_sites=None):
    if call_sites is None:
        call_sites = []
    unique = sorted(set(cs.ossl_symbol for cs in call_sites))
    cats = {}
    for cs in call_sites:
        if cs.category:
            cats.setdefault(cs.category, set()).add(cs.ossl_symbol)
    symbols_by_category = {k: len(v) for k, v in cats.items()}
    return SourceScanResult(
        target=target,
        scan_time="2026-01-01T00:00:00",
        tool_version="1.0.0",
        total_files_scanned=3,
        files_with_calls=1 if call_sites else 0,
        total_call_sites=len(call_sites),
        unique_symbols=unique,
        symbols_by_category=symbols_by_category,
        call_sites=call_sites,
        errors=[],
    )


def _write_json_report(path, target="/tmp/proj", call_sites=None):
    """Write a per-project JSON report file (as source command does)."""
    if call_sites is None:
        call_sites = [_make_call_site()]
    result = _make_result(target=target, call_sites=call_sites)
    exporter = SourceJsonExporter()
    exporter.export(result, path)
    return result


def _write_xlsx_report(path, call_sites=None):
    """Write a per-project XLSX report file (as source command does)."""
    if call_sites is None:
        call_sites = [_make_call_site()]
    result = _make_result(call_sites=call_sites)
    exporter = SourceExcelExporter()
    exporter.export(result, path)
    return result


class TestResolveComboNames:
    """Test _resolve_combo_names from __main__.py."""

    def _call(self, project_dirs, root):
        from openssl_scanner.__main__ import _resolve_combo_names
        return _resolve_combo_names(project_dirs, root)

    def test_single_project(self, tmp_path):
        proj = str(tmp_path / "mylib")
        os.makedirs(proj)
        names = self._call([proj], str(tmp_path))
        assert names == ["mylib"]

    def test_two_distinct_projects(self, tmp_path):
        a = str(tmp_path / "alpha")
        b = str(tmp_path / "beta")
        os.makedirs(a)
        os.makedirs(b)
        names = self._call([a, b], str(tmp_path))
        assert names == ["alpha", "beta"]

    def test_nested_project(self, tmp_path):
        proj = str(tmp_path / "vendor" / "openssl")
        os.makedirs(proj)
        names = self._call([proj], str(tmp_path))
        assert names == ["vendor_openssl"]

    def test_duplicate_basenames_get_index_suffix(self, tmp_path):
        a = str(tmp_path / "x" / "src")
        b = str(tmp_path / "y" / "src")
        os.makedirs(a)
        os.makedirs(b)
        names = self._call([a, b], str(tmp_path))
        assert len(names) == 2
        assert len(set(names)) == 2

    def test_root_itself(self, tmp_path):
        names = self._call([str(tmp_path)], str(tmp_path))
        assert len(names) == 1
        assert names[0] == tmp_path.name

    def test_dedup_collision(self, tmp_path):
        """When parent-prefix dedup creates another collision, the set-based
        fallback should guarantee uniqueness."""
        d1 = str(tmp_path / "a" / "lib")
        d2 = str(tmp_path / "b" / "lib")
        d3 = str(tmp_path / "c" / "lib")
        for d in [d1, d2, d3]:
            os.makedirs(d)
        names = self._call([d1, d2, d3], str(tmp_path))
        assert len(names) == 3
        assert len(set(names)) == 3


class TestResolveOutputNames:
    """Test _resolve_output_names from __main__.py."""

    def _call(self, targets, output_arg, ext):
        from openssl_scanner.__main__ import _resolve_output_names
        return _resolve_output_names(targets, output_arg, ext)

    def test_single_target(self, tmp_path):
        out_dir = str(tmp_path / "out")
        os.makedirs(out_dir)
        result = self._call(["/src/myproj"], out_dir, ".xlsx")
        assert len(result) == 1
        path = result["/src/myproj"]
        assert path.endswith("myproj.xlsx")
        assert out_dir in path

    def test_two_unique_targets(self, tmp_path):
        out_dir = str(tmp_path / "out")
        os.makedirs(out_dir)
        result = self._call(["/a/alpha", "/b/beta"], out_dir, ".json")
        assert "alpha.json" in result["/a/alpha"]
        assert "beta.json" in result["/b/beta"]

    def test_collision_adds_parent(self, tmp_path):
        out_dir = str(tmp_path / "out")
        os.makedirs(out_dir)
        result = self._call(["/x/src", "/y/src"], out_dir, ".xlsx")
        paths = list(result.values())
        assert len(set(paths)) == 2
        basenames = [os.path.basename(p) for p in paths]
        assert "x_src.xlsx" in basenames
        assert "y_src.xlsx" in basenames

    def test_triple_collision_dedup(self, tmp_path):
        out_dir = str(tmp_path / "out")
        os.makedirs(out_dir)
        result = self._call(
            ["/a/src", "/b/src", "/c/src"], out_dir, ".xlsx"
        )
        paths = list(result.values())
        assert len(set(paths)) == 3


class TestDuplicatePathDedup:
    """Test that duplicate paths in -f list are deduplicated."""

    def test_exact_duplicate_removed(self, tmp_path):
        d = tmp_path / "proj"
        d.mkdir()
        raw = [str(d), str(d), str(d)]
        targets = list(dict.fromkeys(os.path.abspath(t) for t in raw))
        assert len(targets) == 1
        assert targets[0] == os.path.abspath(str(d))

    def test_relative_and_absolute_duplicate(self, tmp_path, monkeypatch):
        d = tmp_path / "proj"
        d.mkdir()
        monkeypatch.chdir(tmp_path)
        raw = ["proj", str(d)]
        targets = list(dict.fromkeys(os.path.abspath(t) for t in raw))
        assert len(targets) == 1

    def test_distinct_paths_preserved(self, tmp_path):
        a = tmp_path / "alpha"
        b = tmp_path / "beta"
        a.mkdir()
        b.mkdir()
        raw = [str(a), str(b)]
        targets = list(dict.fromkeys(os.path.abspath(t) for t in raw))
        assert len(targets) == 2

    def test_order_preserved(self, tmp_path):
        a = tmp_path / "alpha"
        b = tmp_path / "beta"
        c = tmp_path / "gamma"
        a.mkdir()
        b.mkdir()
        c.mkdir()
        raw = [str(b), str(a), str(b), str(c), str(a)]
        targets = list(dict.fromkeys(os.path.abspath(t) for t in raw))
        assert len(targets) == 3
        assert targets == [str(b), str(a), str(c)]


class TestCmdSourceDuplicatePaths:
    """Integration test: cmd_source with duplicate paths in -f list.

    Before the fix, duplicate paths caused TypeError because
    results_ordered had None slots from dict key collision.
    """

    def test_from_file_with_duplicates(self, tmp_path):
        from argparse import Namespace
        from openssl_scanner.__main__ import cmd_source

        src_dir = tmp_path / "proj"
        src_dir.mkdir()
        c_file = src_dir / "hello.c"
        c_file.write_text('#include <stdio.h>\nint main() { return 0; }\n')

        list_file = tmp_path / "list.txt"
        list_file.write_text(f"{src_dir}\n{src_dir}\n{src_dir}\n")

        out_dir = tmp_path / "out"

        args = Namespace(
            target=[],
            from_file=str(list_file),
            output=str(out_dir),
            jobs=1,
            no_recursive=False,
            json_only=True,
            verbose=0,
            log_file=None,
        )

        rc = cmd_source(args)
        assert rc == 0

    def test_positional_duplicates(self, tmp_path):
        from argparse import Namespace
        from openssl_scanner.__main__ import cmd_source

        a = tmp_path / "alpha"
        b = tmp_path / "beta"
        a.mkdir()
        b.mkdir()
        (a / "a.c").write_text('int f() { return 0; }\n')
        (b / "b.c").write_text('int g() { return 0; }\n')

        out_dir = tmp_path / "out"

        args = Namespace(
            target=[str(a), str(b), str(a)],
            from_file=None,
            output=str(out_dir),
            jobs=1,
            no_recursive=False,
            json_only=True,
            verbose=0,
            log_file=None,
        )

        rc = cmd_source(args)
        assert rc == 0
        reports = list(out_dir.glob("*.json"))
        assert len(reports) == 2


class TestComboMergeJson:
    """Test _combo_merge_json from __main__.py."""

    def _call(self, json_files, output_path):
        from openssl_scanner.__main__ import _combo_merge_json
        return _combo_merge_json(json_files, output_path)

    def test_merge_two_projects(self, tmp_path):
        j1 = str(tmp_path / "projA.json")
        j2 = str(tmp_path / "projB.json")
        _write_json_report(j1, target="/src/projA", call_sites=[
            _make_call_site("SSL_connect", "ssl_core"),
            _make_call_site("EVP_DigestInit", "crypto_evp"),
        ])
        _write_json_report(j2, target="/src/projB", call_sites=[
            _make_call_site("SSL_read", "ssl_core"),
        ])

        out = str(tmp_path / "merged.json")
        stats = self._call([j1, j2], out)

        assert os.path.isfile(out)
        with open(out, 'r') as f:
            data = json.load(f)

        assert data['meta']['report_type'] == 'combo_scan'
        assert data['meta']['total_projects'] == 2
        assert data['meta']['total_call_sites'] == 3
        assert data['meta']['total_unique_symbols'] == 3
        assert len(data['projects']) == 2
        assert data['projects'][0]['project'] == 'projA'
        assert data['projects'][1]['project'] == 'projB'
        assert len(data['projects'][0]['call_sites']) == 2
        assert len(data['projects'][1]['call_sites']) == 1

        assert stats['total_symbols'] == 3
        assert len(stats['sheets']) == 2
        assert stats['sheets'][0]['call_sites'] == 2
        assert stats['sheets'][1]['call_sites'] == 1

    def test_merge_empty_project(self, tmp_path):
        j1 = str(tmp_path / "empty.json")
        _write_json_report(j1, call_sites=[])

        out = str(tmp_path / "merged.json")
        stats = self._call([j1], out)

        with open(out, 'r') as f:
            data = json.load(f)
        assert data['meta']['total_call_sites'] == 0
        assert data['meta']['total_unique_symbols'] == 0
        assert len(data['projects']) == 1

    def test_project_entry_schema(self, tmp_path):
        j1 = str(tmp_path / "proj.json")
        _write_json_report(j1, call_sites=[
            _make_call_site("EVP_EncryptInit", "crypto_evp"),
        ])
        out = str(tmp_path / "merged.json")
        self._call([j1], out)

        with open(out, 'r') as f:
            data = json.load(f)

        proj = data['projects'][0]
        expected_keys = {
            'project', 'target', 'total_files_scanned', 'files_with_calls',
            'total_call_sites', 'unique_symbols', 'symbols_by_category',
            'call_sites'
        }
        assert set(proj.keys()) == expected_keys

    def test_stats_top_category(self, tmp_path):
        j1 = str(tmp_path / "proj.json")
        _write_json_report(j1, call_sites=[
            _make_call_site("EVP_DigestInit", "crypto_evp"),
            _make_call_site("EVP_EncryptInit", "crypto_evp"),
            _make_call_site("SSL_connect", "ssl_core"),
        ])
        out = str(tmp_path / "merged.json")
        stats = self._call([j1], out)

        assert stats['sheets'][0]['top_category'] == 'crypto_evp'
        assert stats['sheets'][0]['top_cat_symbols'] == 2


class TestMergeExporterFromJson:
    """Test SourceMergeExporter.merge_from_json."""

    def test_merge_two_json_to_xlsx(self, tmp_path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        j1 = str(tmp_path / "alpha.json")
        j2 = str(tmp_path / "beta.json")
        _write_json_report(j1, call_sites=[
            _make_call_site("SSL_connect", "ssl_core"),
        ])
        _write_json_report(j2, call_sites=[
            _make_call_site("EVP_DigestInit", "crypto_evp"),
            _make_call_site("EVP_EncryptInit", "crypto_evp"),
        ])

        out = str(tmp_path / "merged.xlsx")
        merger = SourceMergeExporter()
        stats = merger.merge_from_json([j1, j2], out)

        assert os.path.isfile(out)
        wb = load_workbook(out, read_only=True)
        sheets = wb.sheetnames
        assert sheets[0] == "Summary"
        assert "alpha" in sheets
        assert "beta" in sheets
        assert sheets[-1] == "Symbol Summary"

        ws_sum = wb["Summary"]
        rows = list(ws_sum.iter_rows(values_only=True))
        assert rows[0][0] == "Project"
        assert rows[1][0] == "alpha"
        assert rows[2][0] == "beta"
        assert rows[3][0] == "TOTAL"
        assert rows[3][3] == 3

        ws_alpha = wb["alpha"]
        data_rows = list(ws_alpha.iter_rows(values_only=True))
        assert data_rows[0][4] == "OpenSSL Symbol"
        assert data_rows[1][4] == "SSL_connect"

        ws_sym = wb["Symbol Summary"]
        sym_rows = list(ws_sym.iter_rows(values_only=True))
        assert len(sym_rows) >= 4
        symbols_found = {r[0] for r in sym_rows[1:]}
        assert "SSL_connect" in symbols_found
        assert "EVP_DigestInit" in symbols_found
        wb.close()

        assert stats['total_symbols'] == 3
        assert len(stats['sheets']) == 2

    def test_merge_preserves_files_scanned(self, tmp_path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        j1 = str(tmp_path / "proj.json")
        _write_json_report(j1, call_sites=[_make_call_site()])
        out = str(tmp_path / "merged.xlsx")

        merger = SourceMergeExporter()
        merger.merge_from_json([j1], out)

        wb = load_workbook(out, read_only=True)
        ws_sum = wb["Summary"]
        rows = list(ws_sum.iter_rows(values_only=True))
        files_col = rows[1][1]
        assert files_col == 3
        wb.close()

    def test_merge_single_project(self, tmp_path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        j1 = str(tmp_path / "solo.json")
        _write_json_report(j1, call_sites=[
            _make_call_site("SSL_write", "ssl_core"),
        ])
        out = str(tmp_path / "merged.xlsx")
        merger = SourceMergeExporter()
        stats = merger.merge_from_json([j1], out)

        wb = load_workbook(out, read_only=True)
        assert "Summary" in wb.sheetnames
        assert "solo" in wb.sheetnames
        assert "Symbol Summary" in wb.sheetnames
        wb.close()

        assert stats['total_symbols'] == 1


class TestMergeExporterFromResults:
    """Test SourceMergeExporter.merge_from_results."""

    def test_merge_to_xlsx(self, tmp_path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        r1 = _make_result("/src/alpha", [
            _make_call_site("SSL_connect", "ssl_core"),
        ])
        r2 = _make_result("/src/beta", [
            _make_call_site("EVP_DigestInit", "crypto_evp"),
        ])

        out = str(tmp_path / "merged.xlsx")
        merger = SourceMergeExporter()
        stats = merger.merge_from_results(
            [("alpha", r1), ("beta", r2)], out
        )

        wb = load_workbook(out, read_only=True)
        assert wb.sheetnames[0] == "Summary"
        assert "alpha" in wb.sheetnames
        assert "beta" in wb.sheetnames
        assert wb.sheetnames[-1] == "Symbol Summary"

        ws_sum = wb["Summary"]
        rows = list(ws_sum.iter_rows(values_only=True))
        assert rows[3][0] == "TOTAL"
        assert rows[3][3] == 2
        wb.close()

        assert stats['total_symbols'] == 2

    def test_merge_to_json(self, tmp_path):
        r1 = _make_result("/src/alpha", [
            _make_call_site("SSL_connect", "ssl_core"),
        ])

        out = str(tmp_path / "merged.json")
        merger = SourceMergeExporter()
        stats = merger.merge_from_results([("alpha", r1)], out)

        with open(out, 'r') as f:
            data = json.load(f)
        assert data['meta']['report_type'] == 'combo_scan'
        assert data['meta']['total_projects'] == 1
        assert len(data['projects']) == 1
        assert data['projects'][0]['project'] == 'alpha'
        assert stats['total_symbols'] == 1

    def test_merge_empty_results(self, tmp_path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        r1 = _make_result("/src/empty", [])

        out = str(tmp_path / "merged.xlsx")
        merger = SourceMergeExporter()
        stats = merger.merge_from_results([("empty", r1)], out)

        wb = load_workbook(out, read_only=True)
        ws_empty = wb["empty"]
        data_rows = list(ws_empty.iter_rows(values_only=True))
        assert len(data_rows) == 1
        wb.close()

        assert stats['sheets'][0]['call_sites'] == 0

    def test_preserves_total_files_scanned(self, tmp_path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        r1 = _make_result("/src/proj", [_make_call_site()])

        out = str(tmp_path / "merged.xlsx")
        merger = SourceMergeExporter()
        merger.merge_from_results([("proj", r1)], out)

        wb = load_workbook(out, read_only=True)
        ws_sum = wb["Summary"]
        rows = list(ws_sum.iter_rows(values_only=True))
        assert rows[1][1] == 3
        wb.close()


class TestMergeExporterFromXlsx:
    """Test SourceMergeExporter.merge (XLSX roundtrip)."""

    def test_merge_two_xlsx(self, tmp_path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        x1 = str(tmp_path / "alpha.xlsx")
        x2 = str(tmp_path / "beta.xlsx")
        _write_xlsx_report(x1, [
            _make_call_site("SSL_connect", "ssl_core"),
        ])
        _write_xlsx_report(x2, [
            _make_call_site("EVP_DigestInit", "crypto_evp"),
        ])

        out = str(tmp_path / "merged.xlsx")
        merger = SourceMergeExporter()
        stats = merger.merge([x1, x2], out)

        assert os.path.isfile(out)
        wb = load_workbook(out, read_only=True)
        assert wb.sheetnames[0] == "Summary"
        assert wb.sheetnames[-1] == "Symbol Summary"
        assert len(wb.sheetnames) == 4

        ws_sum = wb["Summary"]
        rows = list(ws_sum.iter_rows(values_only=True))
        assert rows[3][0] == "TOTAL"
        wb.close()

        assert stats['total_symbols'] == 2


class TestResolveSheetNames:
    """Test SourceMergeExporter._resolve_sheet_names."""

    def test_no_duplicates(self):
        m = SourceMergeExporter()
        result = m._resolve_sheet_names(["alpha", "beta", "gamma"])
        assert result == ["alpha", "beta", "gamma"]

    def test_duplicates_get_suffix(self):
        m = SourceMergeExporter()
        result = m._resolve_sheet_names(["lib", "lib", "lib"])
        assert len(set(result)) == 3
        assert result[0] == "lib"
        assert "_1" in result[1]
        assert "_2" in result[2]

    def test_truncation_at_31_chars(self):
        m = SourceMergeExporter()
        long_name = "a" * 40
        result = m._resolve_sheet_names([long_name])
        assert len(result[0]) <= 31

    def test_long_duplicate_truncation(self):
        m = SourceMergeExporter()
        long_name = "a" * 40
        result = m._resolve_sheet_names([long_name, long_name])
        assert len(set(result)) == 2
        for name in result:
            assert len(name) <= 31


class TestExportResultParentDir:
    """Test _export_result creates parent dirs."""

    def _call(self, result, output_path):
        from openssl_scanner.__main__ import _export_result
        return _export_result(result, output_path)

    def test_creates_parent_dir(self, tmp_path):
        result = _make_result(call_sites=[_make_call_site()])
        out = str(tmp_path / "nested" / "dir" / "report.json")
        self._call(result, out)
        assert os.path.isfile(out)

    def test_xlsx_export(self, tmp_path):
        result = _make_result(call_sites=[_make_call_site()])
        out = str(tmp_path / "report.xlsx")
        self._call(result, out)
        assert os.path.isfile(out)

    def test_json_export(self, tmp_path):
        result = _make_result(call_sites=[_make_call_site()])
        out = str(tmp_path / "report.json")
        self._call(result, out)
        assert os.path.isfile(out)
        with open(out) as f:
            data = json.load(f)
        assert data['meta']['report_type'] == 'source_scan'
