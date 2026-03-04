
import json
import os
import sys
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openssl_scanner.exporter import ExcelExporter, HTMLExporter, Exporter
from openssl_scanner import _vendor  # noqa: F401


def _make_single_report():
    """Build a representative single scan report dict."""
    return {
        'meta': {
            'tool_version': '1.0.0',
            'report_type': 'single',
            'scan_time': '2026-02-27T12:00:00',
            'scan_root': '/usr/bin/test_app',
            'target_arch': 'aarch64',
        },
        'summary': {
            'total_files_scanned': 5,
            'total_elf_files': 3,
            'files_with_openssl_deps': 2,
            'total_openssl_symbols': 4,
            'unique_openssl_symbols': 3,
            'openssl_libs_found': ['libcrypto.so.3'],
            'files_with_static_openssl': 0,
            'files_with_dlopen': 0,
            'dlopen_unique_symbols': 0,
            'dlopen_libs_detected': [],
        },
        'openssl_symbols': {
            'by_file': {
                '/usr/bin/test_app': {
                    'count': 2,
                    'symbols': ['SSL_connect', 'EVP_sha256'],
                },
                '/usr/lib/libhelper.so': {
                    'count': 2,
                    'symbols': ['EVP_sha256', 'BIO_new'],
                },
            },
            'by_category': {
                'ssl_core': {'count': 1, 'symbols': ['SSL_connect']},
                'crypto_evp': {'count': 1, 'symbols': ['EVP_sha256']},
                'crypto_bio': {'count': 1, 'symbols': ['BIO_new']},
            },
            'by_depth': {
                'depth_1': {
                    'count': 2,
                    'symbols': ['SSL_connect', 'EVP_sha256'],
                    'files': ['/usr/bin/test_app'],
                },
            },
            'import_chains': {
                'SSL_connect': [
                    {
                        'source_file': '/usr/bin/test_app',
                        'chain': 'test_app -> libssl.so -> libcrypto.so',
                        'depth': 2,
                    },
                ],
            },
            'all_unique': ['SSL_connect', 'EVP_sha256', 'BIO_new'],
        },
        'files_detail': [
            {
                'path': '/usr/bin/test_app',
                'type': 'executable',
                'arch': 'aarch64',
                'direct_deps': ['libssl.so.3', 'libcrypto.so.3', 'libc.so'],
                'openssl_deps': {
                    'direct': True,
                    'transitive': False,
                    'libs': ['libcrypto.so.3'],
                },
                'openssl_symbols_used': ['SSL_connect', 'EVP_sha256'],
                'static_openssl': False,
                'error': None,
            },
            {
                'path': '/usr/lib/libhelper.so',
                'type': 'shared_library',
                'arch': 'aarch64',
                'direct_deps': ['libc.so'],
                'openssl_deps': {
                    'direct': False,
                    'transitive': True,
                    'libs': [],
                },
                'openssl_symbols_used': ['EVP_sha256', 'BIO_new'],
                'static_openssl': False,
                'error': None,
            },
        ],
        'dependency_tree': {
            'name': 'test_app',
            'path': '/usr/bin/test_app',
            'is_openssl_lib': False,
            'openssl_symbols_count': 2,
            'children': [
                {
                    'name': 'libcrypto.so.3',
                    'path': '/usr/lib/libcrypto.so.3',
                    'is_openssl_lib': True,
                    'openssl_symbols_count': 0,
                },
            ],
        },
        'errors': [
            {'severity': 'warning', 'file': '/usr/lib/bad.so', 'error': 'parse error'},
        ],
    }


def _make_empty_report():
    """Report with no symbols or files."""
    return {
        'meta': {
            'tool_version': '1.0.0',
            'report_type': 'single',
            'scan_time': '2026-02-27T12:00:00',
            'scan_root': '/empty',
            'target_arch': 'aarch64',
        },
        'summary': {
            'total_files_scanned': 0,
            'total_elf_files': 0,
            'files_with_openssl_deps': 0,
            'total_openssl_symbols': 0,
            'unique_openssl_symbols': 0,
            'openssl_libs_found': [],
        },
        'openssl_symbols': {
            'by_file': {},
            'by_category': {},
            'by_depth': {},
            'import_chains': {},
            'all_unique': [],
        },
        'files_detail': [],
        'errors': [],
    }


def _write_json(tmp_path, filename, data):
    path = os.path.join(str(tmp_path), filename)
    with open(path, 'w') as f:
        json.dump(data, f)
    return path


class TestExcelExporterSheetNames:
    """Verify the 8-sheet structure of Excel output."""

    def test_sheet_names(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')

        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        expected = [
            'Overview', 'Files', 'File-Symbol', 'Import Chains',
            'By Category', 'By Depth', 'Dep Tree', 'Errors',
        ]
        assert wb.sheetnames == expected
        wb.close()


class TestExcelExporterOverview:
    """Test Overview sheet content."""

    def test_overview_metadata(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Overview']

        values = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0]:
                values[row[0]] = row[1]

        assert values.get('Report Type') == 'single'
        assert values.get('Tool Version') == '1.0.0'
        assert values.get('Target Architecture') == 'aarch64'
        wb.close()

    def test_overview_summary_stats(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Overview']

        values = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0]:
                values[row[0]] = row[1]

        assert values.get('Total Files Scanned') == '5'
        assert values.get('Unique OpenSSL Symbols') == '3'
        wb.close()


class TestExcelExporterFiles:
    """Test Files sheet content."""

    def test_files_headers(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']

        headers = [cell.value for cell in ws[1]]
        assert 'File Path' in headers
        assert 'File Name' in headers
        assert 'Type' in headers
        assert 'Arch' in headers
        assert 'Link Type' in headers
        assert 'Symbol Count' in headers
        wb.close()

    def test_files_data_rows(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']

        paths = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                paths.append(row[0])
        assert '/usr/bin/test_app' in paths
        assert '/usr/lib/libhelper.so' in paths
        wb.close()

    def test_files_link_type_dynamic(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == '/usr/bin/test_app':
                assert row[4] == 'dynamic-link'
                break
        wb.close()

    def test_files_empty_report(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_empty_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']

        assert ws.cell(row=2, column=1).value == 'No file data available'
        wb.close()


class TestExcelExporterFileSymbol:
    """Test File-Symbol (flat pivot) sheet."""

    def test_file_symbol_headers(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['File-Symbol']

        headers = [cell.value for cell in ws[1]]
        assert headers == ['Component', 'Binary', 'Symbol', 'Category', 'Detection']
        wb.close()

    def test_file_symbol_rows_present(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['File-Symbol']

        symbols = set()
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
            if row[0]:
                symbols.add(row[0])

        assert 'SSL_connect' in symbols
        assert 'EVP_sha256' in symbols
        assert 'BIO_new' in symbols
        wb.close()


class TestExcelExporterImportChains:
    """Test Import Chains sheet."""

    def test_import_chains_headers(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Import Chains']

        headers = [cell.value for cell in ws[1]]
        assert 'Symbol' in headers
        assert 'Import Chain' in headers
        assert 'Depth' in headers
        wb.close()

    def test_import_chains_data(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Import Chains']

        row2 = [cell.value for cell in ws[2]]
        assert row2[0] == '/usr/bin/test_app'
        assert row2[2] == 'SSL_connect'
        assert row2[4] == 'test_app -> libssl.so -> libcrypto.so'
        assert row2[5] == 2
        wb.close()

    def test_import_chains_empty(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_empty_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Import Chains']

        assert ws.cell(row=2, column=1).value == 'No import chain data available'
        wb.close()


class TestExcelExporterCategory:
    """Test By Category sheet."""

    def test_category_data(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['By Category']

        categories = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and row[0] != 'TOTAL':
                categories.add(row[0])

        assert 'ssl_core' in categories
        assert 'crypto_evp' in categories
        assert 'crypto_bio' in categories
        wb.close()

    def test_category_total_row(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['By Category']

        last_row = ws.max_row
        assert ws.cell(row=last_row, column=1).value == 'TOTAL'
        assert ws.cell(row=last_row, column=2).value == 3
        wb.close()


class TestExcelExporterDepth:
    """Test By Depth sheet."""

    def test_depth_data(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['By Depth']

        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=3).value == 2
        wb.close()

    def test_depth_empty(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_empty_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['By Depth']

        assert ws.cell(row=2, column=1).value == 'No depth data available'
        wb.close()


class TestExcelExporterDepTree:
    """Test Dep Tree sheet."""

    def test_dep_tree_data(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Dep Tree']

        assert ws.cell(row=2, column=1).value == '(root)'
        assert ws.cell(row=2, column=2).value == 'test_app'
        assert ws.cell(row=3, column=1).value == 'test_app'
        assert ws.cell(row=3, column=2).value == 'libcrypto.so.3'
        assert ws.cell(row=3, column=4).value == 'Yes'
        wb.close()

    def test_dep_tree_empty(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_empty_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Dep Tree']

        assert ws.cell(row=2, column=1).value == 'No dependency tree data available'
        wb.close()


class TestExcelExporterErrors:
    """Test Errors sheet."""

    def test_errors_data(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Errors']

        assert ws.cell(row=2, column=1).value == 'WARNING'
        assert ws.cell(row=2, column=2).value == '/usr/lib/bad.so'
        assert ws.cell(row=2, column=3).value == 'parse error'
        wb.close()

    def test_errors_empty(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_empty_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Errors']

        assert ws.cell(row=2, column=1).value == 'No errors'
        wb.close()


class TestExcelExporterDlopenReport:
    """Test Excel export with dlopen detection data."""

    def test_dlopen_link_type(self, tmp_path):
        report = _make_single_report()
        report['files_detail'][0]['dlopen_detection'] = {
            'uses_dlopen': True,
            'dlopen_symbols': ['AES_encrypt'],
            'confidence': 'high',
        }
        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == '/usr/bin/test_app':
                assert row[4] == 'dlopen'
                break
        wb.close()

    def test_dlopen_inferred_link_type(self, tmp_path):
        report = _make_single_report()
        report['files_detail'][0]['dlopen_detection'] = {
            'uses_dlopen': True,
            'dlopen_symbols': ['AES_encrypt'],
            'confidence': 'inferred',
        }
        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == '/usr/bin/test_app':
                assert row[4] == 'dlopen-infer'
                break
        wb.close()


class TestExcelExporterStaticReport:
    """Test Excel export with static OpenSSL detection data."""

    def test_static_link_type(self, tmp_path):
        report = _make_single_report()
        report['files_detail'][0]['static_openssl'] = True
        report['files_detail'][0]['openssl_deps']['direct'] = False
        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == '/usr/bin/test_app':
                assert row[4] == 'static-link'
                break
        wb.close()


class TestHTMLExporter:
    """Tests for HTMLExporter."""

    def test_produces_valid_html(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.html')
        HTMLExporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            content = f.read()

        assert content.startswith('<!DOCTYPE html>')
        assert '</html>' in content

    def test_html_contains_report_data(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.html')
        HTMLExporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            content = f.read()

        assert 'REPORT_DATA' in content
        assert 'SSL_connect' in content
        assert 'EVP_sha256' in content

    def test_html_contains_chartjs(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.html')
        HTMLExporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            content = f.read()

        assert '<canvas' in content
        assert 'renderReport' in content

    def test_html_self_contained(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.html')
        HTMLExporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            content = f.read()

        assert '<style>' in content
        assert '<script>' in content

    def test_html_empty_report(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_empty_report())
        output_path = os.path.join(str(tmp_path), 'output.html')
        HTMLExporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            content = f.read()

        assert '<!DOCTYPE html>' in content


class TestExporterFacade:
    """Tests for the Exporter dispatch facade."""

    def test_xlsx_dispatch(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')

        Exporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        assert 'Overview' in wb.sheetnames
        wb.close()

    def test_html_dispatch(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.html')

        Exporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            assert '<!DOCTYPE html>' in f.read()

    def test_htm_dispatch(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.htm')

        Exporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            assert '<!DOCTYPE html>' in f.read()

    def test_explicit_format_override(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.dat')

        Exporter().export(report_path, output_path, format='html')

        with open(output_path, 'r') as f:
            assert '<!DOCTYPE html>' in f.read()

    def test_unsupported_format_raises(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.pdf')

        with pytest.raises(ValueError, match='Unsupported format'):
            Exporter().export(report_path, output_path)

    def test_missing_report_file_raises(self, tmp_path):
        output_path = os.path.join(str(tmp_path), 'output.xlsx')

        with pytest.raises(FileNotFoundError):
            Exporter().export('/nonexistent/report.json', output_path)

    def test_default_format_is_xlsx(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output_no_ext.xlsx')

        Exporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        assert 'Overview' in wb.sheetnames
        wb.close()


class TestExcelExporterHeaderStyling:
    """Verify header rows have the expected bold font styling."""

    def test_header_bold(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        ws = wb['Files']
        assert ws.cell(row=1, column=1).font.bold is True

        ws = wb['By Category']
        assert ws.cell(row=1, column=1).font.bold is True

        wb.close()


class TestExcelExporterLegacyImportChains:
    """Test import chains with legacy string format."""

    def test_legacy_string_chains(self, tmp_path):
        report = _make_single_report()
        report['openssl_symbols']['import_chains'] = {
            'SSL_connect': ['test_app -> libssl.so -> libcrypto.so'],
        }

        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Import Chains']

        assert ws.cell(row=2, column=5).value == 'test_app -> libssl.so -> libcrypto.so'
        assert ws.cell(row=2, column=6).value == 2
        wb.close()


def _make_aggregated_report():
    """Build a representative aggregated scan report dict."""
    return {
        'meta': {
            'tool_version': '1.0.0',
            'report_type': 'aggregated',
            'aggregation_time': '2026-03-01T00:00:00',
            'source_reports_count': 2,
            'mapping_file': 'mapping.json',
        },
        'summary': {
            'total_components': 2,
            'total_executables': 3,
            'global_unique_symbols': 4,
        },
        'ranking': [
            {'rank': 1, 'component': 'http_framework', 'unique_symbols_count': 3},
            {'rank': 2, 'component': 'crypto_lib', 'unique_symbols_count': 2},
        ],
        'components': {
            'http_framework': {
                'executables': ['nginx', 'httpd'],
                'executables_detail': {
                    'nginx': {
                        'name': 'nginx',
                        'path': '/usr/sbin/nginx',
                        'unique_symbols_count': 2,
                        'unique_symbols': ['SSL_connect', 'EVP_sha256'],
                        'by_category': {
                            'ssl_core': {'count': 1, 'symbols': ['SSL_connect']},
                            'crypto_evp': {'count': 1, 'symbols': ['EVP_sha256']},
                        },
                    },
                    'httpd': {
                        'name': 'httpd',
                        'path': '/usr/sbin/httpd',
                        'unique_symbols_count': 1,
                        'unique_symbols': ['BIO_new'],
                        'by_category': {
                            'crypto_bio': {'count': 1, 'symbols': ['BIO_new']},
                        },
                    },
                },
                'unique_symbols_count': 3,
                'unique_symbols': ['SSL_connect', 'EVP_sha256', 'BIO_new'],
                'by_category': {
                    'ssl_core': {'count': 1, 'symbols': ['SSL_connect']},
                    'crypto_evp': {'count': 1, 'symbols': ['EVP_sha256']},
                    'crypto_bio': {'count': 1, 'symbols': ['BIO_new']},
                },
            },
            'crypto_lib': {
                'executables': ['libcurl.so'],
                'executables_detail': {
                    'libcurl.so': {
                        'name': 'libcurl.so',
                        'path': '/usr/lib/libcurl.so',
                        'unique_symbols_count': 2,
                        'unique_symbols': ['RAND_bytes', 'SSL_connect'],
                        'by_category': {
                            'crypto_rand': {'count': 1, 'symbols': ['RAND_bytes']},
                            'ssl_core': {'count': 1, 'symbols': ['SSL_connect']},
                        },
                    },
                },
                'unique_symbols_count': 2,
                'unique_symbols': ['RAND_bytes', 'SSL_connect'],
                'by_category': {
                    'crypto_rand': {'count': 1, 'symbols': ['RAND_bytes']},
                    'ssl_core': {'count': 1, 'symbols': ['SSL_connect']},
                },
            },
        },
        'unclassified': None,
        'global_unique_symbols': ['BIO_new', 'EVP_sha256', 'RAND_bytes', 'SSL_connect'],
        'openssl_symbols': {
            'import_chains': {
                'SSL_connect': [
                    {
                        'component': 'http_framework',
                        'binary': 'nginx',
                        'chain': 'nginx -> libssl.so -> libcrypto.so',
                        'depth': 2,
                    },
                ],
            },
            'by_depth': {
                'depth_1': {
                    'count': 2,
                    'symbols': ['SSL_connect', 'EVP_sha256'],
                    'files': ['nginx'],
                    'components': ['http_framework'],
                },
            },
            'by_category': {
                'ssl_core': {'count': 1, 'symbols': ['SSL_connect']},
                'crypto_evp': {'count': 1, 'symbols': ['EVP_sha256']},
                'crypto_bio': {'count': 1, 'symbols': ['BIO_new']},
                'crypto_rand': {'count': 1, 'symbols': ['RAND_bytes']},
            },
        },
        'errors': [],
    }


class TestExcelExporterAggregated:
    """Test Excel export with aggregated report data."""

    def test_aggregated_sheet_names(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_aggregated_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        expected = [
            'Overview', 'Files', 'File-Symbol', 'Import Chains',
            'By Category', 'By Depth', 'Dep Tree', 'Errors',
        ]
        assert wb.sheetnames == expected
        wb.close()

    def test_aggregated_overview_report_type(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_aggregated_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Overview']

        values = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0]:
                values[row[0]] = row[1]

        assert values.get('Report Type') == 'aggregated'
        wb.close()

    def test_aggregated_overview_aggregation_time(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_aggregated_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Overview']

        values = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0]:
                values[row[0]] = row[1]

        assert values.get('Scan Time') == '2026-03-01T00:00:00'
        wb.close()

    def test_aggregated_overview_source_reports(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_aggregated_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Overview']

        values = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0]:
                values[row[0]] = row[1]

        assert values.get('Source Reports') == '2'
        wb.close()

    def test_aggregated_file_symbol_has_component(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_aggregated_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['File-Symbol']

        components = set()
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            if row[0]:
                components.add(row[0])

        assert 'http_framework' in components
        assert 'crypto_lib' in components
        wb.close()

    def test_aggregated_file_symbol_has_symbols(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_aggregated_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['File-Symbol']

        symbols = set()
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
            if row[0]:
                symbols.add(row[0])

        assert 'SSL_connect' in symbols
        assert 'EVP_sha256' in symbols
        assert 'BIO_new' in symbols
        assert 'RAND_bytes' in symbols
        wb.close()

    def test_aggregated_category_sheet(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_aggregated_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['By Category']

        categories = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and row[0] != 'TOTAL':
                categories.add(row[0])

        assert 'ssl_core' in categories
        assert 'crypto_evp' in categories
        assert 'crypto_bio' in categories
        assert 'crypto_rand' in categories
        wb.close()

    def test_aggregated_files_sheet_with_components(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_aggregated_report())
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']

        paths = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                paths.append(row[0])

        assert 'nginx' in paths
        assert 'httpd' in paths
        assert 'libcurl.so' in paths
        wb.close()


class TestExcelExporterEdgeCases:
    """Test edge cases and unusual report data."""

    def test_report_no_openssl_symbols_key(self, tmp_path):
        report = _make_single_report()
        del report['openssl_symbols']

        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        assert wb.sheetnames == [
            'Overview', 'Files', 'File-Symbol', 'Import Chains',
            'By Category', 'By Depth', 'Dep Tree', 'Errors',
        ]
        ws = wb['By Category']
        assert ws.cell(row=ws.max_row, column=1).value == 'TOTAL'
        assert ws.cell(row=ws.max_row, column=2).value == 1
        assert ws.max_row == 2
        ws = wb['File-Symbol']
        assert ws.max_row >= 2
        wb.close()

    def test_unicode_file_paths(self, tmp_path):
        report = _make_single_report()
        report['files_detail'][0]['path'] = '/usr/lib/\u5f00\u6e90/libtest.so'
        report['openssl_symbols']['by_file'] = {
            '/usr/lib/\u5f00\u6e90/libtest.so': {
                'count': 1,
                'symbols': ['SSL_connect'],
            },
        }

        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']

        paths = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                paths.append(row[0])
        assert '/usr/lib/\u5f00\u6e90/libtest.so' in paths
        wb.close()

    def test_report_only_errors(self, tmp_path):
        report = _make_empty_report()
        report['errors'] = [
            {'severity': 'error', 'file': '/bad.so', 'error': 'corrupt ELF'},
            {'severity': 'warning', 'file': '/warn.so', 'error': 'missing section'},
        ]

        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Errors']

        errors = []
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            if row[0] and row[0] != 'No errors':
                errors.append(row[2])

        assert 'corrupt ELF' in errors
        assert 'missing section' in errors
        wb.close()

    def test_null_values_in_fields(self, tmp_path):
        report = _make_single_report()
        report['files_detail'][0]['arch'] = None
        report['files_detail'][0]['type'] = None

        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Files']
        headers = [cell.value for cell in ws[1]]
        arch_col = headers.index('Arch') + 1
        type_col = headers.index('Type') + 1
        assert ws.cell(row=2, column=arch_col).value is None
        assert ws.cell(row=2, column=type_col).value is None
        assert ws.cell(row=2, column=1).value == '/usr/bin/test_app'
        wb.close()

    def test_empty_by_category(self, tmp_path):
        report = _make_single_report()
        report['openssl_symbols']['by_category'] = {}

        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['By Category']

        last_row = ws.max_row
        assert ws.cell(row=last_row, column=1).value == 'TOTAL'
        wb.close()


class TestExporterErrorHandling:
    """Test error handling and malformed inputs."""

    def test_malformed_json_raises(self, tmp_path):
        path = os.path.join(str(tmp_path), 'bad.json')
        with open(path, 'w') as f:
            f.write('{not valid json}')

        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        with pytest.raises(json.JSONDecodeError):
            ExcelExporter().export(path, output_path)

    def test_json_missing_meta_no_crash(self, tmp_path):
        report = {'summary': {}, 'openssl_symbols': {}, 'files_detail': [], 'errors': []}

        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        assert wb.sheetnames == [
            'Overview', 'Files', 'File-Symbol', 'Import Chains',
            'By Category', 'By Depth', 'Dep Tree', 'Errors',
        ]
        ws = wb['Overview']
        values = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0]:
                values[row[0]] = row[1]
        assert values.get('Report Type') == 'single'
        assert values.get('Tool Version') is None
        wb.close()

    def test_report_with_empty_strings(self, tmp_path):
        report = _make_single_report()
        report['meta']['scan_root'] = ''
        report['meta']['target_arch'] = ''
        report['files_detail'][0]['path'] = ''

        report_path = _write_json(tmp_path, 'report.json', report)
        output_path = os.path.join(str(tmp_path), 'output.xlsx')
        ExcelExporter().export(report_path, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Overview']
        values = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0]:
                values[row[0]] = row[1]
        assert 'Scan Root' in values
        assert values['Scan Root'] is None
        assert 'Target Architecture' in values
        assert values['Target Architecture'] is None
        ws = wb['Files']
        assert ws.cell(row=2, column=1).value is None
        wb.close()

    def test_output_missing_parent_dirs_raises(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        nested_dir = os.path.join(str(tmp_path), 'a', 'b', 'c')
        output_path = os.path.join(nested_dir, 'output.xlsx')

        with pytest.raises((FileNotFoundError, OSError)):
            ExcelExporter().export(report_path, output_path)


class TestHTMLExporterDetailed:
    """Detailed tests for HTML output structure."""

    def test_html_contains_tab_elements(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.html')
        HTMLExporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            content = f.read()

        assert 'id="ranking"' in content
        assert 'id="dependencies"' in content
        assert 'id="categories"' in content
        assert 'id="symbols"' in content

    def test_html_contains_chart_canvases(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.html')
        HTMLExporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            content = f.read()

        assert 'id="rankingChart"' in content
        assert 'id="categoryChart"' in content

    def test_html_no_external_resources(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.html')
        HTMLExporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            content = f.read()

        assert '<script src=' not in content
        assert '<link rel="stylesheet" href="http' not in content

    def test_html_contains_export_buttons(self, tmp_path):
        report_path = _write_json(tmp_path, 'report.json', _make_single_report())
        output_path = os.path.join(str(tmp_path), 'output.html')
        HTMLExporter().export(report_path, output_path)

        with open(output_path, 'r') as f:
            content = f.read()

        assert 'Export Excel' in content
        assert 'Load JSON' in content
