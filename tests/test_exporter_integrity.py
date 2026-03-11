"""R2: ExcelExporter data integrity tests.

Verifies that XLSX output from ExcelExporter matches the source JSON exactly:
- File-Symbol sheet row count, unique symbols, (file,symbol) pair fidelity
- Overview sheet totals match JSON summary
- By Category sum matches unique symbol count
- Files sheet row count matches JSON files_detail
- total_rows == 0 fallback logic for three data source paths
- Batch scan cross-validation (per-package JSON vs XLSX)
"""

import json
import os
import subprocess
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openssl_scanner import _vendor  # noqa: F401
from openpyxl import load_workbook
from openssl_scanner.exporter import ExcelExporter


HAP_DIR = os.path.join(os.path.dirname(__file__), '..', 'hap_test_packages')
HARFLIX_HAP = os.path.join(HAP_DIR, 'Harflix-V1.0.5-unsigned.hap')


def _make_elf_report(**overrides):
    """Minimal valid ELF scan report dict."""
    report = {
        'meta': {
            'tool_version': '1.0.0',
            'report_type': 'single',
            'scan_time': '2026-01-01T00:00:00',
            'scan_root': '/usr/bin/app',
            'target_arch': 'aarch64',
        },
        'summary': {
            'total_files_scanned': 0,
            'total_elf_files': 0,
            'files_with_openssl_deps': 0,
            'total_openssl_symbols': 0,
            'unique_openssl_symbols': 0,
            'openssl_libs_found': [],
            'files_with_static_openssl': 0,
            'files_with_dlopen': 0,
            'dlopen_unique_symbols': 0,
            'dlopen_libs_detected': [],
        },
        'openssl_symbols': {
            'by_file': {},
            'by_category': {},
            'by_depth': {},
            'import_chains': {},
            'all_unique': [],
        },
        'files_detail': [],
        'dependency_tree': {},
    }
    report.update(overrides)
    return report


def _export_json_to_xlsx(data, tmp_path, name="report"):
    """Write data as JSON and export to XLSX, return (json_path, xlsx_path)."""
    json_path = os.path.join(str(tmp_path), f"{name}.json")
    xlsx_path = os.path.join(str(tmp_path), f"{name}.xlsx")
    with open(json_path, 'w') as f:
        json.dump(data, f)
    ExcelExporter().export(json_path, xlsx_path)
    return json_path, xlsx_path


def _read_file_symbol_rows(xlsx_path):
    """Read all data rows from File-Symbol sheet(s), return list of tuples."""
    wb = load_workbook(xlsx_path, read_only=True)
    fs_sheets = [s for s in wb.sheetnames if s.startswith('File-Symbol')]
    rows = []
    for sn in fs_sheets:
        ws = wb[sn]
        for idx, r in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                continue
            if len(r) >= 3 and r[0] and r[0] != 'No file-symbol data available':
                rows.append(r)
    wb.close()
    return rows


def _read_overview_value(xlsx_path, label):
    """Read a specific label's value from the Overview sheet."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb['Overview']
    for r in ws.iter_rows(values_only=True):
        if len(r) >= 2 and r[0] == label:
            wb.close()
            return r[1]
    wb.close()
    return None


def _read_category_total(xlsx_path):
    """Read the TOTAL from By Category sheet."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb['By Category']
    for r in ws.iter_rows(values_only=True):
        if len(r) >= 2 and r[0] == 'TOTAL':
            wb.close()
            return r[1]
    wb.close()
    return None


def _read_files_count(xlsx_path):
    """Count data rows in Files sheet."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb['Files']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return len(rows) - 1


class TestHarflixIntegrity:
    """Task 1: Real HAP scan data integrity against Harflix package."""

    @pytest.fixture(scope="class")
    def harflix_data(self, tmp_path_factory):
        """Run HAP scan once and return (json_data, xlsx_path)."""
        if not os.path.exists(HARFLIX_HAP):
            pytest.skip("Harflix HAP not available")
        out_dir = tmp_path_factory.mktemp("harflix")
        xlsx_path = os.path.join(str(out_dir), "harflix.xlsx")
        result = subprocess.run(
            [sys.executable, '-m', 'openssl_scanner', 'hap',
             HARFLIX_HAP, '-o', xlsx_path],
            capture_output=True, text=True, timeout=60,
            cwd=os.path.join(os.path.dirname(__file__), '..'),
        )
        assert result.returncode == 0, f"HAP scan failed: {result.stderr}"

        json_path = xlsx_path.replace('.xlsx', '.json')
        assert os.path.exists(json_path), "JSON report not generated"
        with open(json_path) as f:
            data = json.load(f)
        return data, xlsx_path

    def test_file_symbol_unique_symbols_match(self, harflix_data):
        """Unique symbols in File-Symbol sheet = JSON unique_openssl_symbols."""
        data, xlsx_path = harflix_data
        rows = _read_file_symbol_rows(xlsx_path)
        xlsx_unique = set(r[2] for r in rows if len(r) >= 3)
        json_unique = data['summary']['unique_openssl_symbols']
        assert len(xlsx_unique) == json_unique

    def test_file_symbol_unique_files_match(self, harflix_data):
        """Unique files in File-Symbol = files in by_file with symbols."""
        data, xlsx_path = harflix_data
        rows = _read_file_symbol_rows(xlsx_path)
        xlsx_files = set(r[0] for r in rows)

        by_file = data['openssl_symbols']['by_file']
        json_files = set(by_file.keys())
        assert xlsx_files == json_files

    def test_file_symbol_pairs_exact_match(self, harflix_data):
        """Every (file, symbol) pair in XLSX exists in JSON and vice versa."""
        data, xlsx_path = harflix_data
        rows = _read_file_symbol_rows(xlsx_path)
        xlsx_pairs = set((r[0], r[2]) for r in rows if len(r) >= 3)

        json_pairs = set()
        for path, info in data['openssl_symbols']['by_file'].items():
            syms = info.get('symbols', []) if isinstance(info, dict) else []
            for sym in syms:
                json_pairs.add((path, sym))

        assert xlsx_pairs == json_pairs

    def test_file_symbol_no_duplicates(self, harflix_data):
        """No duplicate rows in File-Symbol sheet."""
        _, xlsx_path = harflix_data
        rows = _read_file_symbol_rows(xlsx_path)
        pair_list = [(r[0], r[2]) for r in rows if len(r) >= 3]
        assert len(pair_list) == len(set(pair_list))

    def test_overview_total_symbols(self, harflix_data):
        """Overview 'Total OpenSSL Symbols (refs)' matches JSON."""
        data, xlsx_path = harflix_data
        val = _read_overview_value(xlsx_path, 'Total OpenSSL Symbols (refs)')
        assert int(val) == data['summary']['total_openssl_symbols']

    def test_overview_unique_symbols(self, harflix_data):
        """Overview 'Unique OpenSSL Symbols' matches JSON."""
        data, xlsx_path = harflix_data
        val = _read_overview_value(xlsx_path, 'Unique OpenSSL Symbols')
        assert int(val) == data['summary']['unique_openssl_symbols']

    def test_category_sum_equals_unique(self, harflix_data):
        """Sum of all category counts = unique symbol count."""
        data, xlsx_path = harflix_data
        cat_total = _read_category_total(xlsx_path)
        assert cat_total == data['summary']['unique_openssl_symbols']

    def test_files_count(self, harflix_data):
        """Files sheet row count = total_files_scanned."""
        data, xlsx_path = harflix_data
        count = _read_files_count(xlsx_path)
        assert count == data['summary']['total_files_scanned']


class TestFallbackByFile:
    """Task 2a: files_detail fallback when by_file is empty."""

    def test_files_detail_only(self, tmp_path):
        """With empty by_file, File-Symbol falls back to files_detail."""
        data = _make_elf_report()
        data['summary']['total_files_scanned'] = 2
        data['summary']['files_with_openssl_deps'] = 1
        data['openssl_symbols']['by_category'] = {
            'ssl_core': {'count': 2, 'symbols': ['SSL_connect', 'SSL_read']},
            'crypto_evp': {'count': 1, 'symbols': ['EVP_sha256']},
        }
        data['files_detail'] = [
            {
                'path': '/usr/bin/app',
                'type': 'executable', 'arch': 'aarch64',
                'direct_deps': [],
                'openssl_deps': {'direct': True, 'transitive': True, 'libs': ['libcrypto.so.3']},
                'openssl_symbols_used': ['SSL_connect', 'SSL_read', 'EVP_sha256'],
            },
            {
                'path': '/usr/bin/noossl',
                'type': 'executable', 'arch': 'aarch64',
                'direct_deps': [],
                'openssl_deps': {'direct': False, 'transitive': False, 'libs': []},
                'openssl_symbols_used': [],
            },
        ]

        _, xlsx_path = _export_json_to_xlsx(data, tmp_path, "fallback_a")
        rows = _read_file_symbol_rows(xlsx_path)

        assert len(rows) == 3
        symbols = set(r[2] for r in rows)
        assert symbols == {'SSL_connect', 'SSL_read', 'EVP_sha256'}
        files = set(r[0] for r in rows)
        assert files == {'/usr/bin/app'}


class TestFallbackComponents:
    """Task 2b: components fallback when both by_file and files_detail are empty."""

    def test_components_only(self, tmp_path):
        """With empty by_file and files_detail, falls back to components."""
        data = _make_elf_report()
        data['meta']['report_type'] = 'aggregated'
        data['components'] = {
            'my_component': {
                'by_category': {
                    'ssl_core': {'count': 1, 'symbols': ['SSL_connect']},
                    'crypto_evp': {'count': 1, 'symbols': ['EVP_sha256']},
                },
                'executables_detail': {
                    'my_binary': {
                        'by_category': {
                            'ssl_core': {'count': 1, 'symbols': ['SSL_connect']},
                            'crypto_evp': {'count': 1, 'symbols': ['EVP_sha256']},
                        }
                    }
                }
            }
        }

        _, xlsx_path = _export_json_to_xlsx(data, tmp_path, "fallback_b")
        rows = _read_file_symbol_rows(xlsx_path)

        assert len(rows) == 2
        symbols = set(r[2] for r in rows)
        assert symbols == {'SSL_connect', 'EVP_sha256'}


class TestByFilePriority:
    """Task 2c: by_file takes priority when both by_file and files_detail exist."""

    def test_by_file_priority_over_files_detail(self, tmp_path):
        """When by_file has data, files_detail is NOT used for File-Symbol."""
        data = _make_elf_report()
        data['openssl_symbols']['by_file'] = {
            '/usr/bin/app': {'count': 1, 'symbols': ['SSL_connect']}
        }
        data['openssl_symbols']['by_category'] = {
            'ssl_core': {'count': 1, 'symbols': ['SSL_connect']}
        }
        data['files_detail'] = [
            {
                'path': '/usr/bin/app',
                'openssl_symbols_used': ['SSL_connect', 'SSL_read', 'EVP_sha256'],
                'openssl_deps': {},
            }
        ]

        _, xlsx_path = _export_json_to_xlsx(data, tmp_path, "priority")
        rows = _read_file_symbol_rows(xlsx_path)

        assert len(rows) == 1
        assert rows[0][2] == 'SSL_connect'


class TestBatchIntegrity:
    """Task 3: Batch scan data integrity across multiple packages."""

    @pytest.fixture(scope="class")
    def batch_results(self, tmp_path_factory):
        """Run batch HAP scan and return output directory path."""
        if not os.path.isdir(HAP_DIR):
            pytest.skip("HAP test packages not available")
        out_dir = str(tmp_path_factory.mktemp("batch"))
        result = subprocess.run(
            [sys.executable, '-m', 'openssl_scanner', 'hap',
             HAP_DIR, '-o', out_dir, '-j', '1'],
            capture_output=True, text=True, timeout=300,
            cwd=os.path.join(os.path.dirname(__file__), '..'),
        )
        assert result.returncode == 0, f"Batch scan failed: {result.stderr}"
        return out_dir

    def _top_packages(self, batch_dir, n=3):
        """Return top N packages by symbol count as (json_path, xlsx_path) pairs."""
        import glob
        pkg_stats = []
        for jf in glob.glob(os.path.join(batch_dir, '*.json')):
            if os.path.basename(jf) == 'summary.json':
                continue
            with open(jf) as f:
                data = json.load(f)
            unique = data.get('summary', {}).get('unique_openssl_symbols', 0)
            xlsx_path = jf.replace('.json', '.xlsx')
            if os.path.exists(xlsx_path):
                pkg_stats.append((unique, jf, xlsx_path))
        pkg_stats.sort(reverse=True)
        return [(jf, xp) for _, jf, xp in pkg_stats[:n]]

    def test_top3_file_symbol_row_count(self, batch_results):
        """File-Symbol row count = sum of symbols across all files in JSON."""
        for json_path, xlsx_path in self._top_packages(batch_results, 3):
            with open(json_path) as f:
                data = json.load(f)
            by_file = data.get('openssl_symbols', {}).get('by_file', {})
            json_total = sum(
                len(info.get('symbols', []) if isinstance(info, dict) else [])
                for info in by_file.values()
            )
            xlsx_rows = _read_file_symbol_rows(xlsx_path)
            pkg_name = os.path.basename(json_path)
            assert len(xlsx_rows) == json_total, (
                f"{pkg_name}: XLSX rows={len(xlsx_rows)}, JSON={json_total}"
            )

    def test_top3_overview_matches(self, batch_results):
        """Overview sheet totals match JSON summary."""
        for json_path, xlsx_path in self._top_packages(batch_results, 3):
            with open(json_path) as f:
                data = json.load(f)
            summary = data['summary']
            pkg_name = os.path.basename(json_path)

            ov_total = _read_overview_value(xlsx_path, 'Total OpenSSL Symbols (refs)')
            assert int(ov_total) == summary['total_openssl_symbols'], (
                f"{pkg_name}: Overview total={ov_total}, JSON={summary['total_openssl_symbols']}"
            )

            ov_unique = _read_overview_value(xlsx_path, 'Unique OpenSSL Symbols')
            assert int(ov_unique) == summary['unique_openssl_symbols'], (
                f"{pkg_name}: Overview unique={ov_unique}, JSON={summary['unique_openssl_symbols']}"
            )

    def test_top3_no_duplicates(self, batch_results):
        """No duplicate (file, symbol) rows in any top package."""
        for json_path, xlsx_path in self._top_packages(batch_results, 3):
            rows = _read_file_symbol_rows(xlsx_path)
            pair_list = [(r[0], r[2]) for r in rows if len(r) >= 3]
            pkg_name = os.path.basename(json_path)
            assert len(pair_list) == len(set(pair_list)), (
                f"{pkg_name}: {len(pair_list) - len(set(pair_list))} duplicate pairs"
            )

    def test_top3_category_sum(self, batch_results):
        """By Category TOTAL = unique symbol count for each top package."""
        for json_path, xlsx_path in self._top_packages(batch_results, 3):
            with open(json_path) as f:
                data = json.load(f)
            cat_total = _read_category_total(xlsx_path)
            expected = data['summary']['unique_openssl_symbols']
            pkg_name = os.path.basename(json_path)
            assert cat_total == expected, (
                f"{pkg_name}: category total={cat_total}, expected={expected}"
            )
