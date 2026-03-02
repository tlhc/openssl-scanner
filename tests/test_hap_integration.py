"""Integration tests for the hap subcommand end-to-end pipeline."""

import io
import json
import os
import shutil
import struct
import sys
import tempfile
import zipfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openssl_scanner.__main__ import main
from openssl_scanner.hap_report import (
    classify_hap_detection as _classify_hap_detection,
    build_hap_summary_row as _build_hap_summary_row,
    _HAP_SUMMARY_COLUMNS,
    load_scan_result_from_json as _load_scan_result_from_json,
    detect_static_providers as _detect_static_providers,
)
from openssl_scanner.scanner import ScanResult, FileResult


def _minimal_elf64():
    """Minimal valid ELF64 shared library header (aarch64)."""
    e_ident = b'\x7fELF'
    e_ident += b'\x02\x01\x01'
    e_ident += b'\x00' * 9
    header = e_ident
    header += struct.pack('<H', 3)    # ET_DYN
    header += struct.pack('<H', 183)  # EM_AARCH64
    header += struct.pack('<I', 1)
    header += struct.pack('<Q', 0)    # e_entry
    header += struct.pack('<Q', 0)    # e_phoff
    header += struct.pack('<Q', 0)    # e_shoff
    header += struct.pack('<I', 0)
    header += struct.pack('<H', 64)
    header += struct.pack('<H', 0) * 5
    return header


def _create_test_hap(path, bundle_name="com.test.scanner",
                     module_name="entry", include_openssl=False):
    """Create a test HAP for integration testing."""
    with zipfile.ZipFile(path, 'w') as zf:
        zf.writestr("module.json", json.dumps({
            "module": {
                "name": module_name,
                "type": "entry",
                "deviceTypes": ["default", "tablet"]
            },
            "app": {
                "bundleName": bundle_name,
                "versionCode": 1000000,
                "versionName": "1.0.0",
                "minAPIVersion": 11
            }
        }))
        zf.writestr("libs/arm64-v8a/libentry.so", _minimal_elf64())
        if include_openssl:
            zf.writestr("libs/arm64-v8a/libcrypto.so.3", _minimal_elf64())
    return path


def _create_test_app(path, hap_count=2):
    """Create a test APP containing multiple HAPs."""
    with zipfile.ZipFile(path, 'w') as app_zf:
        for i in range(hap_count):
            name = f"module{i}.hap"
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w') as hap_zf:
                hap_zf.writestr("module.json", json.dumps({
                    "module": {"name": f"module{i}", "type": "entry",
                               "deviceTypes": ["default"]},
                    "app": {"bundleName": "com.test.app",
                            "versionCode": 1, "versionName": "1.0.0",
                            "minAPIVersion": 11}
                }))
                hap_zf.writestr(f"libs/arm64-v8a/libmod{i}.so", _minimal_elf64())
            app_zf.writestr(name, buf.getvalue())
    return path


class TestHapSubcommandCLI:
    """Test hap subcommand CLI behavior."""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_hap_subcommand_help(self, capsys):
        """hap --help should show usage and return 0."""
        with pytest.raises(SystemExit) as exc_info:
            main(['hap', '--help'])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()
        assert 'HAP/HAR/HSP/APP' in captured.out

    def test_hap_subcommand_missing_file(self):
        """Non-existent target should return error code 1."""
        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', '/nonexistent/path.hap', '-o', output, '--json-only'])
        assert ret == 1

    def test_hap_subcommand_invalid_file(self):
        """Invalid ZIP file should return error code but not crash."""
        invalid = os.path.join(self.tmpdir, "bad.hap")
        with open(invalid, 'wb') as f:
            f.write(b"not a zip at all")

        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', invalid, '-o', output, '--json-only'])
        assert ret == 1

    def test_hap_no_native_libs(self):
        """HAP with no native libs should warn and return 1."""
        hap_path = os.path.join(self.tmpdir, "no_native.hap")
        with zipfile.ZipFile(hap_path, 'w') as zf:
            zf.writestr("module.json", json.dumps({
                "module": {"name": "entry", "type": "entry"},
                "app": {"bundleName": "com.test.empty"}
            }))

        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', hap_path, '-o', output, '--json-only'])
        assert ret == 1

    def test_hap_no_openssl_scans_with_builtin(self):
        """HAP with native libs but no bundled OpenSSL should scan using built-in symbols."""
        hap_path = os.path.join(self.tmpdir, "no_ssl.hap")
        _create_test_hap(hap_path, include_openssl=False)

        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', hap_path, '-o', output, '--json-only'])
        assert ret == 0
        assert os.path.isfile(output)

    def test_hap_directory_no_packages(self):
        """Directory with no packages should return 1."""
        empty_dir = os.path.join(self.tmpdir, "empty")
        os.makedirs(empty_dir)

        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', empty_dir, '-o', output, '--json-only'])
        assert ret == 1

    def test_hap_directory_finds_packages(self):
        """Directory scan should find and scan packages using built-in symbols."""
        pkg_dir = os.path.join(self.tmpdir, "packages")
        os.makedirs(pkg_dir)
        _create_test_hap(os.path.join(pkg_dir, "a.hap"))
        _create_test_hap(os.path.join(pkg_dir, "b.hap"))

        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', pkg_dir, '-o', output, '--json-only'])
        assert ret == 0
        assert os.path.isfile(output)


class TestHapExtractPipeline:
    """Test the extraction + metadata pipeline without requiring real OpenSSL."""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_keep_extracted_preserves_files(self):
        """--keep-extracted should leave temp files in place."""
        hap_path = os.path.join(self.tmpdir, "keep.hap")
        _create_test_hap(hap_path)

        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', hap_path, '-o', output, '--json-only',
                     '--keep-extracted'])
        assert ret == 0
        assert os.path.isfile(output)

    def test_hap_recognized_in_command_list(self):
        """'hap' should be recognized as a valid subcommand (not falling through to scan)."""
        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', '/nonexistent.hap', '-o', output, '--json-only'])
        assert ret == 1


class TestHapReportMetadata:
    """Test report metadata structure from hap_extractor + reporter integration."""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_extractor_metadata_structure(self):
        """Verify HapExtractor produces correct metadata that cmd_hap would use."""
        from openssl_scanner.hap_extractor import HapExtractor

        hap_path = os.path.join(self.tmpdir, "meta_test.hap")
        _create_test_hap(hap_path, bundle_name="com.test.meta",
                         module_name="feature")

        extractor = HapExtractor()
        result = extractor.extract(hap_path)

        assert result.metadata.bundle_name == "com.test.meta"
        assert result.metadata.module_name == "feature"
        assert result.metadata.package_type == "hap"
        assert result.metadata.version_name == "1.0.0"
        assert result.metadata.version_code == 1000000
        assert result.metadata.min_api_version == 11
        assert "default" in result.metadata.device_types
        assert "arm64-v8a" in result.metadata.abis_found
        assert len(result.so_files) >= 1

        extractor.cleanup(result)

    def test_app_file_extracts_sub_packages(self):
        """APP file should produce sub_packages with their own metadata."""
        from openssl_scanner.hap_extractor import HapExtractor

        app_path = os.path.join(self.tmpdir, "test.app")
        _create_test_app(app_path, hap_count=3)

        extractor = HapExtractor()
        result = extractor.extract(app_path)

        assert len(result.sub_packages) == 3
        module_names = [sub.metadata.module_name for sub in result.sub_packages]
        assert "module0" in module_names
        assert "module1" in module_names
        assert "module2" in module_names

        total_so = sum(len(sub.so_files) for sub in result.sub_packages)
        assert total_so == 3

        extractor.cleanup(result)

    def test_reporter_package_info_in_json(self):
        """Verify Reporter correctly serializes package_info to JSON."""
        from openssl_scanner.reporter import Reporter

        result = ScanResult(
            target="/tmp/test.hap",
            scan_time="2026-02-09T00:00:00",
            tool_version="0.1.0",
            arch="aarch64",
            report_type="package",
        )
        result.package_info = {
            'package_type': 'hap',
            'bundle_name': 'com.test.reporter',
            'module_name': 'entry',
            'module_type': 'entry',
            'version_name': '1.0.0',
            'version_code': 100,
            'scanned_abi': 'arm64-v8a',
            'native_libs_count': 3,
            'bundled_openssl': False,
        }

        reporter = Reporter()
        json_str = reporter.generate_json(result)
        data = json.loads(json_str)

        assert data['meta']['report_type'] == 'package'
        assert 'package' in data['meta']
        pkg = data['meta']['package']
        assert pkg['bundle_name'] == 'com.test.reporter'
        assert pkg['scanned_abi'] == 'arm64-v8a'
        assert pkg['native_libs_count'] == 3

    def test_reporter_summary_includes_package_info(self):
        """Verify console summary includes package metadata."""
        from openssl_scanner.reporter import Reporter

        result = ScanResult(
            target="/tmp/test.hap",
            scan_time="2026-02-09T00:00:00",
            tool_version="0.1.0",
            arch="aarch64",
            report_type="package",
        )
        result.package_info = {
            'package_type': 'hap',
            'bundle_name': 'com.test.summary',
            'module_name': 'entry',
            'module_type': 'entry',
            'version_name': '2.0.0',
            'version_code': 200,
            'scanned_abi': 'arm64-v8a',
            'native_libs_count': 5,
        }

        reporter = Reporter()
        summary = reporter.generate_summary(result)

        assert 'HAP' in summary
        assert 'com.test.summary' in summary
        assert '2.0.0' in summary
        assert 'arm64-v8a' in summary


class TestHapSummaryReport:
    """Test Package Summary XLSX generated during batch HAP scan."""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _batch_scan(self, hap_count=3):
        """Create HAPs in a dir, batch scan to output dir, return output dir."""
        pkg_dir = os.path.join(self.tmpdir, "packages")
        os.makedirs(pkg_dir)
        for i in range(hap_count):
            _create_test_hap(
                os.path.join(pkg_dir, f"app{i}.hap"),
                bundle_name=f"com.test.app{i}",
                module_name=f"mod{i}",
            )
        out_dir = os.path.join(self.tmpdir, "output")
        ret = main(['hap', pkg_dir, '-o', out_dir, '--json-only'])
        return ret, out_dir

    def test_summary_xlsx_generated(self):
        """Batch scan should produce summary.xlsx in output dir."""
        ret, out_dir = self._batch_scan(2)
        assert ret == 0
        summary = os.path.join(out_dir, "summary.xlsx")
        assert os.path.isfile(summary)

    def test_summary_has_correct_columns(self):
        """Header row should have all 22 columns."""
        ret, out_dir = self._batch_scan(2)
        assert ret == 0
        summary = os.path.join(out_dir, "summary.xlsx")

        from openpyxl import load_workbook
        wb = load_workbook(summary)
        ws = wb.active
        assert ws.title == "Package Summary"
        headers = [ws.cell(row=1, column=c).value for c in range(1, 23)]
        assert headers[0] == "Package Name"
        assert headers[5] == "OpenSSL Usage"
        assert headers[6] == "Detection"
        assert headers[7] == "Static Symbols"
        assert headers[8] == "Dynamic Symbols"
        assert headers[9] == "dlopen Symbols"
        assert headers[10] == "Total Symbols"
        assert headers[11] == "Top Category"
        assert headers[20] == "dlopen Libs"
        assert headers[21] == "Custom Match"
        assert len(headers) == 22

    def test_summary_row_count(self):
        """Should have header + N data rows + TOTAL row."""
        ret, out_dir = self._batch_scan(3)
        assert ret == 0
        summary = os.path.join(out_dir, "summary.xlsx")

        from openpyxl import load_workbook
        wb = load_workbook(summary)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Package Name"
        assert ws.cell(row=5, column=1).value == "TOTAL"
        for r in range(2, 5):
            val = ws.cell(row=r, column=1).value
            assert val is not None and val != "TOTAL"

    def test_summary_total_row_so_count(self):
        """TOTAL row .so Files should be sum of all packages."""
        ret, out_dir = self._batch_scan(3)
        assert ret == 0

        from openpyxl import load_workbook
        wb = load_workbook(os.path.join(out_dir, "summary.xlsx"))
        ws = wb.active
        total_row = ws.max_row
        assert ws.cell(row=total_row, column=1).value == "TOTAL"
        pkg_so = sum(ws.cell(row=r, column=5).value or 0
                     for r in range(2, total_row))
        assert ws.cell(row=total_row, column=5).value == pkg_so

    def test_summary_openssl_usage_for_minimal_elf(self):
        """Minimal ELFs have no OpenSSL -> OpenSSL Usage should be None."""
        ret, out_dir = self._batch_scan(2)
        assert ret == 0

        from openpyxl import load_workbook
        wb = load_workbook(os.path.join(out_dir, "summary.xlsx"))
        ws = wb.active
        for r in range(2, ws.max_row):
            assert ws.cell(row=r, column=6).value == 'None'
        total_row = ws.max_row
        val = ws.cell(row=total_row, column=6).value
        assert 'None' in str(val)

    def test_summary_detection_none_for_minimal_elf(self):
        """Minimal ELFs have no OpenSSL -> Detection should be None."""
        ret, out_dir = self._batch_scan(2)
        assert ret == 0

        from openpyxl import load_workbook
        wb = load_workbook(os.path.join(out_dir, "summary.xlsx"))
        ws = wb.active
        for r in range(2, ws.max_row):
            assert ws.cell(row=r, column=7).value == 'None'


class TestSummaryTotalRow:
    """Tests for TOTAL row aggregation in Package Summary."""

    def test_total_row_mixed_usage_aggregation(self):
        """TOTAL row should aggregate different openssl_usage values."""
        results = []
        pkg_paths = []

        none_result = ScanResult(
            target="/t", scan_time="", tool_version="0.1", arch="aarch64")
        none_result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=[]),
        ]
        none_result.symbols_by_category = {}
        none_result.package_info = {
            'bundle_name': 'com.test.none', 'module_name': 'entry',
            'package_type': 'hap', 'version_name': '1.0',
            'scanned_abi': ['arm64-v8a'], 'native_libs_count': 1,
            'bundled_openssl': False,
        }
        results.append(none_result)
        pkg_paths.append('/none.hap')

        bundled_result = ScanResult(
            target="/t", scan_time="", tool_version="0.1", arch="aarch64")
        bundled_result.files_detail = [
            FileResult(path="/b.so", file_type="shared_library",
                       arch="aarch64", direct_deps=["libcrypto.so"],
                       openssl_direct=True, openssl_transitive=False,
                       openssl_libs=["libcrypto.so"],
                       openssl_symbols=["SSL_connect"]),
        ]
        bundled_result.symbols_by_category = {'ssl_core': ['SSL_connect']}
        bundled_result.package_info = {
            'bundle_name': 'com.test.bundled', 'module_name': 'entry',
            'package_type': 'hap', 'version_name': '1.0',
            'scanned_abi': ['arm64-v8a'], 'native_libs_count': 2,
            'bundled_openssl': True,
            'bundled_openssl_files': ['libcrypto.so'],
        }
        results.append(bundled_result)
        pkg_paths.append('/bundled.hap')

        rows = []
        for result, pkg_path in zip(results, pkg_paths):
            method, s, d, dl, ossl_type = _classify_hap_detection(result)
            row = _build_hap_summary_row(
                result, pkg_path, method, s, d, dl, ossl_type)
            rows.append(row)

        assert rows[0]['openssl_usage'] == 'None'
        assert rows[1]['openssl_usage'] == 'Bundled'

        usage_counts = {}
        for r in rows:
            u = r.get('openssl_usage', '')
            if u:
                usage_counts[u] = usage_counts.get(u, 0) + 1
        usage_summary = ', '.join(
            f'{v} {k}' for k, v in sorted(usage_counts.items()))

        assert '1 Bundled' in usage_summary
        assert '1 None' in usage_summary


class TestClassifyHapDetection:
    """Unit tests for _classify_hap_detection helper."""

    def test_no_openssl(self):
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=[]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'None'
        assert len(s_set) == 0 and len(d_set) == 0 and len(dl_set) == 0
        assert ossl_type == 'No-OpenSSL'

    def test_dynamic_only(self):
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_connect"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'Dynamic'
        assert ossl_type == 'System-Link'
        assert d_set == {"SSL_connect"} and len(s_set) == 0

    def test_static_only(self):
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256"],
                       static_openssl=True),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'Static'
        assert ossl_type == 'Self-Contained'
        assert s_set == {"EVP_sha256"}

    def test_dlopen_only(self):
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_read"],
                       uses_dlopen=True, dlsym_symbols=["SSL_read"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'dlopen'
        assert ossl_type == 'System-Link'
        assert dl_set == {"SSL_read"}

    def test_mixed(self):
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_connect"]),
            FileResult(path="/b.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256"],
                       static_openssl=True),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'Mixed'
        assert s_set == {"EVP_sha256"} and d_set == {"SSL_connect"}
        assert ossl_type == 'System-Link'

    def test_static_majority_but_unresolved_dynamic(self):
        """Static lib self-sufficient but dynamic lib unresolved -> System-Link.

        Per-library resolution: libcrypto.so is self-sufficient (static),
        but app.so has dynamic OpenSSL symbols with no bundled .so to satisfy
        them.  The HAP is NOT Self-Contained despite static count > dynamic.
        """
        static_syms = [f"EVP_sym_{i}" for i in range(50)]
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/libcrypto.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=static_syms, static_openssl=True),
            FileResult(path="/app.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_connect", "SSL_read"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'Mixed'
        assert len(s_set) == 50 and len(d_set) == 2
        assert ossl_type == 'System-Link'

    def test_system_link_dynamic_majority(self):
        """Package with many dynamic symbols and few static -> System-Link."""
        dynamic_syms = [f"SSL_sym_{i}" for i in range(30)]
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=dynamic_syms),
            FileResult(path="/b.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha1", "EVP_sha256"],
                       static_openssl=True),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert len(s_set) == 2 and len(d_set) == 30
        assert ossl_type == 'System-Link'

    def test_equal_static_dynamic_is_system_link(self):
        """Equal static and dynamic counts -> System-Link (not >)."""
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_connect", "SSL_read", "SSL_write"]),
            FileResult(path="/b.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha1", "EVP_sha256", "EVP_md5"],
                       static_openssl=True),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert len(s_set) == 3 and len(d_set) == 3
        assert ossl_type == 'System-Link'

    def test_dlopen_plus_dynamic_system_link(self):
        """Package with both dlopen and dynamic (no static) -> System-Link."""
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_connect"]),
            FileResult(path="/b.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_read"],
                       uses_dlopen=True, dlsym_symbols=["SSL_read"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert len(d_set) == 1 and len(dl_set) == 1 and len(s_set) == 0
        assert ossl_type == 'System-Link'

    def test_hybrid_dlopen_file_with_direct_und(self):
        """File with both dlopen symbols and direct UND -> split counts."""
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_connect", "SSL_read", "EVP_sha256"],
                       uses_dlopen=True, dlsym_symbols=["EVP_sha256"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert dl_set == {"EVP_sha256"}
        assert d_set == {"SSL_connect", "SSL_read"}
        assert ossl_type == 'System-Link'

    def test_static_dlopen_overlap_not_double_counted(self):
        """Static file with dlopen: dlsym symbols should NOT inflate static count.

        scanner.py appends filtered dlsym symbols to openssl_symbols, so
        fr.openssl_symbols = UND + dlsym-only.  Without the fix, all of
        openssl_symbols would go into static_syms, inflating the count.

        Setup: static file with 1 true UND symbol + 2 dlsym-appended symbols.
        Correct: static=1, dlopen=2 -> 1 < 2 -> System-Link.
        Bug:     static=3, dlopen=2 -> 3 > 2 -> Self-Contained (wrong).
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/libcrypto.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256", "SSL_read", "SSL_write"],
                       static_openssl=True, uses_dlopen=True,
                       dlsym_symbols=["SSL_read", "SSL_write"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert s_set == {"EVP_sha256"}
        assert dl_set == {"SSL_read", "SSL_write"}
        assert len(s_set) == 1 and len(dl_set) == 2
        assert ossl_type == 'System-Link'

    def test_static_dlopen_all_symbols_are_dlsym(self):
        """Static file where ALL symbols are dlsym -> empty static_only.

        static_only = openssl_symbols - dlsym_symbols = empty.
        method='Mixed' (has_static + has_dlopen), ossl_type='System-Link'.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_read", "SSL_write"],
                       static_openssl=True, uses_dlopen=True,
                       dlsym_symbols=["SSL_read", "SSL_write"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert s_set == set()
        assert dl_set == {"SSL_read", "SSL_write"}
        assert method == 'Mixed'
        assert ossl_type == 'System-Link'

    def test_dedup_shared_symbols_across_static_files(self):
        """Duplicate symbols across static files count once after dedup.

        Two static .so files share the same 2 symbols -> s_set has 2, not 4.
        Three dynamic symbols -> external=3 > static=2 -> System-Link.
        Without dedup this would be 4 > 3 -> Self-Contained (wrong).
        """
        shared_static = ["EVP_sha256", "EVP_md5"]
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=shared_static, static_openssl=True),
            FileResult(path="/b.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=shared_static, static_openssl=True),
            FileResult(path="/app.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_connect", "SSL_read", "SSL_write"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert len(s_set) == 2
        assert len(d_set) == 3
        assert ossl_type == 'System-Link'
        assert len(s_set | d_set | dl_set) == 5

    def test_bundled_resolves_dynamic_dt_needed(self):
        """HAP with bundled libcrypto.so.3 satisfies DT_NEEDED -> Self-Contained.

        Production flow: cmd_hap removes bundled OpenSSL .so before scanning,
        storing filenames in package_info['bundled_openssl_files'].  app.so
        DT_NEEDs libcrypto.so.3 which matches the bundled lib stem.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.package_info = {
            'bundled_openssl': True,
            'bundled_openssl_files': ['libcrypto.so.3'],
        }
        result.files_detail = [
            FileResult(path="/libs/arm64-v8a/app.so",
                       file_type="shared_library", arch="aarch64",
                       direct_deps=["libcrypto.so.3"], openssl_direct=True,
                       openssl_transitive=False,
                       openssl_libs=["libcrypto.so.3"],
                       openssl_symbols=["SSL_connect", "SSL_read"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert ossl_type == 'Self-Contained'

    def test_bundled_resolves_dlopen(self):
        """HAP with bundled libcrypto.so.3 satisfies dlopen -> Self-Contained.

        lib.so does dlopen("libcrypto.so"), stem matches bundled libcrypto.so.3.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.package_info = {
            'bundled_openssl': True,
            'bundled_openssl_files': ['libcrypto.so.3'],
        }
        result.files_detail = [
            FileResult(path="/libs/arm64-v8a/lib.so",
                       file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_read"],
                       uses_dlopen=True, dlsym_symbols=["SSL_read"],
                       dlopen_libs=["libcrypto.so"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert ossl_type == 'Self-Contained'

    def test_static_plus_dlopen_no_bundle(self):
        """Static lib with dlopen symbols, no bundled .so -> System-Link.

        One lib has static OpenSSL for its own needs but also dlopen-loads
        additional symbols.  Without bundled .so, the dlopen part is unresolved.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256", "SSL_read"],
                       static_openssl=True, uses_dlopen=True,
                       dlsym_symbols=["SSL_read"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert s_set == {"EVP_sha256"}
        assert dl_set == {"SSL_read"}
        assert ossl_type == 'System-Link'

    def test_static_plus_dlopen_with_bundle(self):
        """Static lib with dlopen symbols, bundled .so present -> Self-Contained.

        The static portion self-resolves; the dlopen("libcrypto.so") resolves
        via bundled libcrypto.so.3 (stem match).
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.package_info = {
            'bundled_openssl': True,
            'bundled_openssl_files': ['libcrypto.so.3'],
        }
        result.files_detail = [
            FileResult(path="/libs/arm64-v8a/lib.so",
                       file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256", "SSL_read"],
                       static_openssl=True, uses_dlopen=True,
                       dlsym_symbols=["SSL_read"],
                       dlopen_libs=["libcrypto.so"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert ossl_type == 'Self-Contained'

    def test_dlopen_partial_bundle_system_link(self):
        """dlopen needs both libcrypto + libssl but only libcrypto bundled.

        Per-target resolution: ALL OpenSSL dlopen targets must resolve.
        libssl.so is missing -> System-Link.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.package_info = {
            'bundled_openssl': True,
            'bundled_openssl_files': ['libcrypto.so.3'],
        }
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_read", "EVP_sha256"],
                       uses_dlopen=True,
                       dlsym_symbols=["SSL_read", "EVP_sha256"],
                       dlopen_libs=["libcrypto.so", "libssl.so"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert ossl_type == 'System-Link'

    def test_dlopen_absolute_path_resolved(self):
        """dlopen with absolute path '/system/lib64/libcrypto.so' is normalized.

        Basename extraction ensures path prefix does not break matching.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.package_info = {
            'bundled_openssl': True,
            'bundled_openssl_files': ['libcrypto.so.3'],
        }
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_read"],
                       uses_dlopen=True, dlsym_symbols=["SSL_read"],
                       dlopen_libs=["/system/lib64/libcrypto.so"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert ossl_type == 'Self-Contained'

    def test_hybrid_static_plus_dt_needed_unresolved(self):
        """Static OpenSSL lib that also DT_NEEDs external libssl -> System-Link.

        File implements some OpenSSL symbols (static) but also imports from
        libssl.so.3 via DT_NEEDED.  Without bundled libssl, the dynamic part
        is unresolved.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=["libssl.so.3"],
                       openssl_direct=True,
                       openssl_transitive=False,
                       openssl_libs=["libssl.so.3"],
                       openssl_symbols=["EVP_sha256", "SSL_connect"],
                       static_openssl=True),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert ossl_type == 'System-Link'

    def test_hybrid_static_plus_dt_needed_resolved(self):
        """Static OpenSSL lib that DT_NEEDs libssl, bundled -> Self-Contained."""
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.package_info = {
            'bundled_openssl': True,
            'bundled_openssl_files': ['libssl.so.3'],
        }
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=["libssl.so.3"],
                       openssl_direct=True,
                       openssl_transitive=False,
                       openssl_libs=["libssl.so.3"],
                       openssl_symbols=["EVP_sha256", "SSL_connect"],
                       static_openssl=True),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert ossl_type == 'Self-Contained'

    def test_bundled_openssl_files_absent_fallback(self):
        """bundled_openssl=True but bundled_openssl_files missing -> System-Link.

        Tests backward compat: old scan results without bundled_openssl_files
        key produce conservative System-Link (not crash).
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.package_info = {'bundled_openssl': True}
        result.files_detail = [
            FileResult(path="/app.so", file_type="shared_library",
                       arch="aarch64", direct_deps=["libcrypto.so.3"],
                       openssl_direct=True, openssl_transitive=False,
                       openssl_libs=["libcrypto.so.3"],
                       openssl_symbols=["SSL_connect"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert ossl_type == 'System-Link'

    def test_harflix_scenario_system_link(self):
        """Harflix-like: static ijkplayer + dlopen reddownload -> System-Link.

        libijkplayer.so: 5215 static OpenSSL symbols, self-sufficient.
        libreddownload.so: 756 dlopen-detected symbols, no bundled .so.
        The HAP is NOT Self-Contained because libreddownload.so needs system
        OpenSSL to resolve its dlopen("libcrypto.so") call.
        """
        static_syms = [f"OSSL_{i}" for i in range(100)]
        dlopen_syms = [f"EVP_{i}" for i in range(50)]
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/libijkplayer.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=static_syms, static_openssl=True),
            FileResult(path="/libreddownload.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=dlopen_syms,
                       uses_dlopen=True, dlsym_symbols=dlopen_syms),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert len(s_set) == 100
        assert len(dl_set) == 50
        assert ossl_type == 'System-Link'

    def test_multiple_static_all_self_sufficient(self):
        """Multiple static-only libs with no external deps -> Self-Contained."""
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/a.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256"], static_openssl=True),
            FileResult(path="/b.so", file_type="shared_library", arch="aarch64",
                       direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_connect"], static_openssl=True),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert s_set == {"EVP_sha256", "SSL_connect"}
        assert ossl_type == 'Self-Contained'

    def test_dlopen_no_lib_name_no_bundle(self):
        """dlopen with empty dlopen_libs (unknown target) -> System-Link.

        When dlopen_libs is empty, the scanner couldn't determine which .so
        the library is trying to load.  Without a bundled OpenSSL, this is
        conservatively treated as unresolved.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_read"],
                       uses_dlopen=True, dlsym_symbols=["SSL_read"],
                       dlopen_libs=[]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert ossl_type == 'System-Link'

    def test_static_dlopen_empty_evidence_ignored(self):
        """D1: Static file with uses_dlopen=True but zero OpenSSL dlopen evidence.

        The file imports dlopen/dlsym for non-OpenSSL purposes.  Both
        dlsym_symbols and dlopen_libs are empty -- no OpenSSL signal at all.
        The dlopen branch should be skipped; detection stays Static-only.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256", "EVP_md5"],
                       static_openssl=True, uses_dlopen=True,
                       dlsym_symbols=[], dlopen_libs=[]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'Static'
        assert s_set == {"EVP_sha256", "EVP_md5"}
        assert dl_set == set()
        assert ossl_type == 'Self-Contained'

    def test_static_dlopen_non_openssl_libs_ignored(self):
        """D2: Static file with dlopen_libs targeting non-OpenSSL library.

        The file dlopen-loads libsqlite.so which doesn't match any
        OPENSSL_LIBRARY_PATTERNS. The dlopen branch should be skipped
        despite dlopen_libs being non-empty.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256"],
                       static_openssl=True, uses_dlopen=True,
                       dlsym_symbols=[],
                       dlopen_libs=["libsqlite.so"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'Static'
        assert s_set == {"EVP_sha256"}
        assert dl_set == set()
        assert ossl_type == 'Self-Contained'

    def test_static_dlopen_libs_only_preserved(self):
        """D3: Static file with dlopen_libs but empty dlsym_symbols.

        dlopen_analyzer found library name (e.g., libcrypto.so in .rodata)
        but disassembly xref failed to resolve individual symbols.
        The dlopen_libs signal should still activate the dlopen branch.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256"],
                       static_openssl=True, uses_dlopen=True,
                       dlsym_symbols=[],
                       dlopen_libs=["libcrypto.so"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'Mixed'
        assert s_set == {"EVP_sha256"}
        assert dl_set == set()
        assert ossl_type == 'System-Link'

    def test_nonstat_dlopen_empty_evidence_skipped(self):
        """D4: Non-static file with uses_dlopen but zero evidence -> skip entirely.

        This tests the existing guard at L397 (non-static branch).
        File has uses_dlopen=True but no dlsym_symbols and no openssl_symbols.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=[],
                       uses_dlopen=True, dlsym_symbols=[], dlopen_libs=[]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'None'
        assert ossl_type == 'No-OpenSSL'

    def test_multifile_static_empty_dlopen_plus_dynamic(self):
        """D7: Multi-file package with one static+empty-dlopen + one dynamic.

        File A: static_openssl + uses_dlopen but empty evidence -> Static only
        File B: dynamic UND symbols -> Dynamic

        Package should be Mixed(Static+Dynamic), NOT Mixed(Static+Dynamic+dlopen).
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/static.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256"],
                       static_openssl=True, uses_dlopen=True,
                       dlsym_symbols=[], dlopen_libs=[]),
            FileResult(path="/app.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["SSL_connect"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'Mixed'
        assert s_set == {"EVP_sha256"}
        assert d_set == {"SSL_connect"}
        assert dl_set == set()
        assert ossl_type == 'System-Link'

    def test_nonstat_dlopen_libs_only_system_link(self):
        """D5: Non-static file with dlopen_libs but no resolved symbols.

        dlopen_analyzer found 'libcrypto.so' in .rodata but xref failed.
        The file should NOT be skipped; method=dlopen, type=System-Link.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=[],
                       uses_dlopen=True, dlsym_symbols=[],
                       dlopen_libs=["libcrypto.so"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'dlopen'
        assert dl_set == set()
        assert ossl_type == 'System-Link'

    def test_nonstat_dlopen_libs_only_bundled_self_contained(self):
        """D6: Non-static dlopen_libs only, bundled lib resolves target.

        Same as D5 but with bundled libcrypto.so.3 -> Self-Contained.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.package_info = {
            'bundled_openssl_files': ['libcrypto.so.3'],
        }
        result.files_detail = [
            FileResult(path="/lib.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=[],
                       uses_dlopen=True, dlsym_symbols=[],
                       dlopen_libs=["libcrypto.so"]),
        ]
        method, s_set, d_set, dl_set, ossl_type = _classify_hap_detection(result)
        assert method == 'dlopen'
        assert ossl_type == 'Self-Contained'


class TestBundledOpenSSLDetection:
    """Tests for bundled_openssl field correctness."""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_bundled_openssl_yes_for_separate_so(self):
        """HAP with libcrypto.so.3 should have bundled_openssl=True in JSON."""
        hap_path = os.path.join(self.tmpdir, "bundled.hap")
        _create_test_hap(hap_path, include_openssl=True)

        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', hap_path, '-o', output, '--json-only'])
        assert ret == 0

        with open(output) as f:
            data = json.load(f)
        assert data['meta']['package']['bundled_openssl'] is True

    def test_bundled_openssl_no_without_openssl(self):
        """HAP without OpenSSL libs should have bundled_openssl=False."""
        hap_path = os.path.join(self.tmpdir, "plain.hap")
        _create_test_hap(hap_path, include_openssl=False)

        output = os.path.join(self.tmpdir, "report.json")
        ret = main(['hap', hap_path, '-o', output, '--json-only'])
        assert ret == 0

        with open(output) as f:
            data = json.load(f)
        assert data['meta']['package']['bundled_openssl'] is False

    def test_summary_row_static_no_so_file(self):
        """Static OpenSSL without bundled .so -> openssl_usage='None'.

        No standalone .so, no high/medium confidence providers ->
        bundled_openssl=False.  Low-confidence static detection is
        insufficient evidence; maps to None in OpenSSL Usage column.
        """
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/libapp.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=True,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=["EVP_sha256", "SSL_connect"],
                       static_openssl=True),
        ]
        result.files_with_static_openssl = 1
        result.package_info = {
            'bundle_name': 'com.test.static',
            'package_type': 'hap',
            'version_name': '1.0.0',
            'scanned_abi': ['arm64-v8a'],
            'native_libs_count': 1,
            'bundled_openssl': False,
        }
        method, s, d, dl, ossl_type = _classify_hap_detection(result)
        row = _build_hap_summary_row(result, "/t.hap", method, s, d, dl,
                                     ossl_type)
        assert row['openssl_usage'] == 'None'

    def test_summary_row_no_openssl(self):
        """No OpenSSL .so and no static -> openssl_usage='None'."""
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(path="/libapp.so", file_type="shared_library",
                       arch="aarch64", direct_deps=[], openssl_direct=False,
                       openssl_transitive=False, openssl_libs=[],
                       openssl_symbols=[]),
        ]
        result.package_info = {
            'bundle_name': 'com.test.plain',
            'package_type': 'hap',
            'version_name': '1.0.0',
            'scanned_abi': ['arm64-v8a'],
            'native_libs_count': 1,
            'bundled_openssl': False,
        }
        method, s, d, dl, ossl_type = _classify_hap_detection(result)
        row = _build_hap_summary_row(result, "/t.hap", method, s, d, dl,
                                     ossl_type)
        assert row['openssl_usage'] == 'None'

    def test_libopenssl_pattern_detected(self):
        """libopenssl.so should be recognized as an OpenSSL library."""
        from openssl_scanner.openssl_matcher import OpenSSLMatcher

        matcher = OpenSSLMatcher()
        assert matcher.is_openssl_library("libopenssl.so") is True
        assert matcher.is_openssl_library("libopenssl.so.1.1") is True

    def test_openssl_usage_column_with_bundled_hap(self):
        """Batch scan with one bundled HAP -> OpenSSL Usage shows 'Bundled'."""
        pkg_dir = os.path.join(self.tmpdir, "packages")
        os.makedirs(pkg_dir)
        _create_test_hap(
            os.path.join(pkg_dir, "bundled.hap"),
            bundle_name="com.test.bundled",
            include_openssl=True,
        )
        _create_test_hap(
            os.path.join(pkg_dir, "plain.hap"),
            bundle_name="com.test.plain",
            include_openssl=False,
        )
        out_dir = os.path.join(self.tmpdir, "output")
        ret = main(['hap', pkg_dir, '-o', out_dir, '--json-only'])
        assert ret == 0

        summary = os.path.join(out_dir, "summary.xlsx")
        assert os.path.isfile(summary)

        from openpyxl import load_workbook
        wb = load_workbook(summary)
        ws = wb.active
        usage_vals = {}
        for r in range(2, ws.max_row):
            name = ws.cell(row=r, column=1).value
            usage = ws.cell(row=r, column=6).value
            if name and name != "TOTAL":
                usage_vals[name] = usage
        assert usage_vals.get("com.test.bundled/entry (bundled)") == "Bundled"
        assert usage_vals.get("com.test.plain/entry (plain)") == "None"


    def test_app_summary_distinguishes_modules(self):
        """APP with multiple HAPs -> summary rows use bundle_name/module_name."""
        pkg_dir = os.path.join(self.tmpdir, "packages")
        os.makedirs(pkg_dir)
        _create_test_app(os.path.join(pkg_dir, "multi.app"), hap_count=3)

        out_dir = os.path.join(self.tmpdir, "output")
        ret = main(['hap', pkg_dir, '-o', out_dir, '--json-only'])
        assert ret == 0

        summary = os.path.join(out_dir, "summary.xlsx")
        assert os.path.isfile(summary)

        from openpyxl import load_workbook
        wb = load_workbook(summary)
        ws = wb.active
        pkg_names = []
        for r in range(2, ws.max_row):
            name = ws.cell(row=r, column=1).value
            if name and name != "TOTAL":
                pkg_names.append(name)

        assert len(pkg_names) == 3
        assert "com.test.app/module0 (multi_module0)" in pkg_names
        assert "com.test.app/module1 (multi_module1)" in pkg_names
        assert "com.test.app/module2 (multi_module2)" in pkg_names


class TestLoadScanResultFromJson:
    """Tests for _load_scan_result_from_json round-trip fidelity."""

    def test_static_confidence_survives_round_trip(self):
        """static_openssl_confidence must survive JSON serialize -> deserialize.

        Fix #1 from agent team review: _load_scan_result_from_json was not
        reading back static_openssl_confidence, causing hap-summary to
        misclassify Bundled (static) as None.
        """
        report_json = {
            "meta": {
                "report_type": "package",
                "scan_root": "/t",
                "scan_time": "",
                "tool_version": "0.1",
                "target_arch": "aarch64",
                "package": {
                    "bundle_name": "com.test.roundtrip",
                    "package_type": "hap",
                    "version_name": "1.0",
                    "scanned_abi": ["arm64-v8a"],
                    "native_libs_count": 1,
                    "bundled_openssl": "Yes (static)",
                    "bundled_openssl_files": [],
                    "static_openssl_providers": [{
                        "file": "libapp.so",
                        "confidence": "high",
                        "symbols": 500,
                        "consumers": [],
                    }],
                },
            },
            "summary": {
                "total_files_scanned": 1,
                "total_elf_files": 1,
                "files_with_openssl_deps": 1,
                "total_openssl_symbols": 500,
                "unique_openssl_symbols": 500,
            },
            "openssl_symbols": {
                "by_category": {"crypto_evp": ["EVP_sha256"]},
                "by_file": {},
            },
            "files_detail": [{
                "path": "/libapp.so",
                "type": "shared_library",
                "arch": "aarch64",
                "direct_deps": [],
                "openssl_deps": {"direct": True, "transitive": False,
                                 "libs": []},
                "openssl_symbols_used": ["EVP_sha256", "SSL_connect"],
                "static_openssl": True,
                "static_openssl_confidence": "high",
                "static_openssl_confidence_reason": "500 syms exported",
                "static_ssl_library": "OpenSSL",
            }],
        }

        with tempfile.NamedTemporaryFile(
                mode='w', suffix='.json', delete=False) as f:
            json.dump(report_json, f)
            json_path = f.name

        try:
            result, pkg_path = _load_scan_result_from_json(json_path)
            assert result is not None

            fr = result.files_detail[0]
            assert fr.static_openssl is True
            assert fr.static_openssl_confidence == 'high'
            assert fr.static_openssl_confidence_reason == '500 syms exported'
            assert fr.static_ssl_library == 'OpenSSL'

            bundled_str, providers = _detect_static_providers(result)
            assert bundled_str is not None
            assert len(providers) == 1
            assert providers[0]['confidence'] == 'high'

        finally:
            os.unlink(json_path)

    def test_missing_confidence_defaults_empty(self):
        """JSON without static_openssl_confidence -> defaults to ''."""
        report_json = {
            "meta": {"report_type": "package", "scan_root": "/t",
                     "scan_time": "", "tool_version": "0.1",
                     "target_arch": "aarch64", "package": {}},
            "summary": {},
            "openssl_symbols": {"by_category": {}, "by_file": {}},
            "files_detail": [{
                "path": "/a.so", "type": "shared_library",
                "arch": "aarch64", "direct_deps": [],
                "openssl_deps": {"direct": False, "transitive": False,
                                 "libs": []},
                "openssl_symbols_used": [],
                "static_openssl": True,
            }],
        }

        with tempfile.NamedTemporaryFile(
                mode='w', suffix='.json', delete=False) as f:
            json.dump(report_json, f)
            json_path = f.name

        try:
            result, _ = _load_scan_result_from_json(json_path)
            assert result is not None
            fr = result.files_detail[0]
            assert fr.static_openssl is True
            assert fr.static_openssl_confidence == ''
            assert fr.static_ssl_library == ''
        finally:
            os.unlink(json_path)


class TestOpenSSLUsageScenarios:
    """Tests for openssl_usage classification across all detection scenarios."""

    def _make_result(self, files_detail, package_info, symbols_by_category=None):
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = files_detail
        result.symbols_by_category = symbols_by_category or {}
        result.package_info = package_info
        return result

    def _make_pkg_info(self, bundled=False, bundled_files=None,
                       static_providers=None, native_count=1):
        pi = {
            'bundle_name': 'com.test.pkg',
            'module_name': 'entry',
            'package_type': 'hap',
            'version_name': '1.0.0',
            'scanned_abi': ['arm64-v8a'],
            'native_libs_count': native_count,
            'bundled_openssl': bundled,
        }
        if bundled_files is not None:
            pi['bundled_openssl_files'] = bundled_files
        if static_providers is not None:
            pi['static_openssl_providers'] = static_providers
        return pi

    def _classify_and_build(self, result):
        method, s, d, dl, ossl_type = _classify_hap_detection(result)
        row = _build_hap_summary_row(
            result, "/test.hap", method, s, d, dl, ossl_type
        )
        return row, method, s, d, dl, ossl_type

    def test_s1_no_openssl(self):
        """No .so uses OpenSSL -> usage='None', detection='None'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=[]),
            ],
            package_info=self._make_pkg_info(bundled=False),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'None'
        assert row['detection'] == 'None'

    def test_s2_system_link_dynamic(self):
        """DT_NEED libcrypto.so (not bundled) -> usage='System-Link', detection='Dynamic'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=["libcrypto.so.3"],
                    openssl_direct=True, openssl_transitive=False,
                    openssl_libs=["libcrypto.so.3"],
                    openssl_symbols=["SSL_connect"]),
            ],
            package_info=self._make_pkg_info(bundled=False),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'System-Link'
        assert row['detection'] == 'Dynamic'

    def test_s3_bundled_standalone_so(self):
        """Bundled libcrypto.so + DT_NEED -> usage='Bundled', detection='Dynamic'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libcrypto.so.3", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=[]),
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=["libcrypto.so.3"],
                    openssl_direct=True, openssl_transitive=False,
                    openssl_libs=["libcrypto.so.3"],
                    openssl_symbols=["SSL_connect", "EVP_sha256"]),
            ],
            package_info=self._make_pkg_info(
                bundled=True, bundled_files=['libcrypto.so.3']),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled'
        assert row['detection'] == 'Dynamic'

    def test_s4_bundled_static_no_consumers(self):
        """Static high confidence, no consumers -> usage='Bundled (static)', detection='Static'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=True,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["EVP_sha256", "SSL_connect"],
                    static_openssl=True, static_openssl_confidence='high'),
            ],
            package_info=self._make_pkg_info(bundled='Yes (static)'),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled (static)'
        assert row['detection'] == 'Static'

    def test_s5_bundled_static_shared(self):
        """Static high conf + DT_NEED consumer -> usage='Bundled (static, shared)', detection='Static'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libprovider.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=True,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["EVP_sha256", "SSL_connect"],
                    static_openssl=True, static_openssl_confidence='high'),
                FileResult(
                    path="/libconsumer.so", file_type="shared_library",
                    arch="aarch64", direct_deps=["libprovider.so"],
                    openssl_direct=False, openssl_transitive=True,
                    openssl_libs=[], openssl_symbols=[]),
            ],
            package_info=self._make_pkg_info(
                bundled='Yes (static, shared)', native_count=2),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled (static, shared)'
        assert row['detection'] == 'Static'

    def test_s6_boringssl_static(self):
        """BoringSSL detected, 0 exported, 2 imported -> usage='Bundled (static)', detection='Static'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=True,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["BIO_new_bio_pair", "HMAC"],
                    static_openssl=True, static_openssl_confidence='medium',
                    static_ssl_library='BoringSSL'),
            ],
            package_info=self._make_pkg_info(bundled='Yes (static)'),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled (static)'
        assert row['detection'] == 'Static'

    def test_s7_low_confidence_filtered(self):
        """Low confidence static (confidence='none') -> usage='None', detection='None'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=[],
                    static_openssl=False),
            ],
            package_info=self._make_pkg_info(bundled=False),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'None'
        assert row['detection'] == 'None'

    def test_s8_static_plus_dynamic_consumer(self):
        """Static provider + dynamic consumer -> usage='Bundled (static, shared)', detection='Mixed'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libprovider.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=True,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["EVP_sha256", "SSL_CTX_new"],
                    static_openssl=True, static_openssl_confidence='high'),
                FileResult(
                    path="/libconsumer.so", file_type="shared_library",
                    arch="aarch64", direct_deps=["libprovider.so"],
                    openssl_direct=True, openssl_transitive=False,
                    openssl_libs=["libprovider.so"],
                    openssl_symbols=["SSL_connect"]),
            ],
            package_info=self._make_pkg_info(
                bundled='Yes (static, shared)', native_count=2),
        )
        row, method, s, d, dl, ossl_type = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled (static, shared)'
        assert row['detection'] == 'Mixed'

    def test_s9_static_plus_dlopen(self):
        """Static + dlopen in same package -> bundled_raw takes priority -> usage='Bundled (static)'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libstatic.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=True,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["EVP_sha256", "SSL_CTX_new"],
                    static_openssl=True, static_openssl_confidence='high'),
                FileResult(
                    path="/libdlopen.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["SSL_read"],
                    uses_dlopen=True,
                    dlsym_symbols=["SSL_read"],
                    dlopen_libs=["libcrypto.so"]),
            ],
            package_info=self._make_pkg_info(
                bundled='Yes (static)', native_count=2),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled (static)'

    def test_s10_app_multi_hap_mixed_usage(self):
        """Multiple results: one with OpenSSL, one without -> per-HAP usage correct."""
        result_ossl = ScanResult(
            target="/tmp/a", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result_ossl.files_detail = [
            FileResult(
                path="/libapp.so", file_type="shared_library",
                arch="aarch64", direct_deps=["libcrypto.so.3"],
                openssl_direct=True, openssl_transitive=False,
                openssl_libs=["libcrypto.so.3"],
                openssl_symbols=["SSL_connect"]),
        ]
        result_ossl.symbols_by_category = {'ssl_core': ['SSL_connect']}
        result_ossl.package_info = {
            'bundle_name': 'com.test.ossl', 'module_name': 'entry',
            'package_type': 'hap', 'version_name': '1.0.0',
            'scanned_abi': ['arm64-v8a'], 'native_libs_count': 1,
            'bundled_openssl': False,
        }

        result_plain = ScanResult(
            target="/tmp/b", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result_plain.files_detail = [
            FileResult(
                path="/libplain.so", file_type="shared_library",
                arch="aarch64", direct_deps=[], openssl_direct=False,
                openssl_transitive=False, openssl_libs=[],
                openssl_symbols=[]),
        ]
        result_plain.symbols_by_category = {}
        result_plain.package_info = {
            'bundle_name': 'com.test.plain', 'module_name': 'entry',
            'package_type': 'hap', 'version_name': '1.0.0',
            'scanned_abi': ['arm64-v8a'], 'native_libs_count': 1,
            'bundled_openssl': False,
        }

        m1, s1, d1, dl1, t1 = _classify_hap_detection(result_ossl)
        row1 = _build_hap_summary_row(result_ossl, "/a.hap", m1, s1, d1, dl1, t1)

        m2, s2, d2, dl2, t2 = _classify_hap_detection(result_plain)
        row2 = _build_hap_summary_row(result_plain, "/b.hap", m2, s2, d2, dl2, t2)

        assert row1['openssl_usage'] == 'System-Link'
        assert row1['detection'] == 'Dynamic'
        assert row2['openssl_usage'] == 'None'
        assert row2['detection'] == 'None'

    def test_s11_multi_abi_dedup(self):
        """Multi-ABI: same basename under arm64-v8a/ and x86_64/ deduped in providers."""
        result = ScanResult(
            target="/tmp/t", scan_time="", tool_version="0.1", arch="aarch64"
        )
        result.files_detail = [
            FileResult(
                path="/arm64-v8a/libapp.so", file_type="shared_library",
                arch="aarch64", direct_deps=[], openssl_direct=True,
                openssl_transitive=False, openssl_libs=[],
                openssl_symbols=["EVP_sha256", "SSL_connect"],
                static_openssl=True, static_openssl_confidence='high'),
            FileResult(
                path="/x86_64/libapp.so", file_type="shared_library",
                arch="x86_64", direct_deps=[], openssl_direct=True,
                openssl_transitive=False, openssl_libs=[],
                openssl_symbols=["EVP_sha256"],
                static_openssl=True, static_openssl_confidence='high'),
        ]
        bundled_str, providers = _detect_static_providers(result)
        assert bundled_str == 'Yes (static)'
        assert len(providers) == 1
        assert providers[0]['file'] == 'libapp.so'
        assert providers[0]['symbols'] == 2

    def test_s12_bundled_so_plus_static(self):
        """Bundled .so + static in same package -> usage='Bundled', detection='Mixed'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libcrypto.so.3", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=[]),
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=["libcrypto.so.3"],
                    openssl_direct=True, openssl_transitive=False,
                    openssl_libs=["libcrypto.so.3"],
                    openssl_symbols=["SSL_connect"]),
                FileResult(
                    path="/libstatic.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=True,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["EVP_sha256"],
                    static_openssl=True, static_openssl_confidence='high'),
            ],
            package_info=self._make_pkg_info(
                bundled=True, bundled_files=['libcrypto.so.3'],
                native_count=3),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled'
        assert row['detection'] == 'Mixed'

    def test_s13_dlopen_system(self):
        """dlopen libcrypto.so, not bundled -> usage='System-Link', detection='dlopen'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["SSL_read"],
                    uses_dlopen=True,
                    dlsym_symbols=["SSL_read"],
                    dlopen_libs=["libcrypto.so"]),
            ],
            package_info=self._make_pkg_info(bundled=False),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'System-Link'
        assert row['detection'] == 'dlopen'

    def test_s14_dlopen_bundled(self):
        """dlopen libcrypto.so, bundled in package -> usage='Bundled', detection='dlopen'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libcrypto.so.3", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=[]),
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["SSL_read"],
                    uses_dlopen=True,
                    dlsym_symbols=["SSL_read"],
                    dlopen_libs=["libcrypto.so"]),
            ],
            package_info=self._make_pkg_info(
                bundled=True, bundled_files=['libcrypto.so.3'],
                native_count=2),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled'
        assert row['detection'] == 'dlopen'

    def test_s15_dlsym_only_no_lib(self):
        """dlsym resolves OpenSSL symbols but no lib filename -> usage='System-Link', detection='dlopen'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["SSL_read"],
                    uses_dlopen=True,
                    dlsym_symbols=["SSL_read"],
                    dlopen_libs=[]),
            ],
            package_info=self._make_pkg_info(bundled=False),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'System-Link'
        assert row['detection'] == 'dlopen'

    def test_s16_dlopen_no_openssl_symbols(self):
        """dlopen exists but 0 OpenSSL symbols -> usage='None', detection='None'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=[],
                    uses_dlopen=True,
                    dlsym_symbols=[],
                    dlopen_libs=[]),
            ],
            package_info=self._make_pkg_info(bundled=False),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'None'
        assert row['detection'] == 'None'

    def test_s17_dlopen_non_openssl_filtered(self):
        """dlopen targets non-OpenSSL libs, 0 OpenSSL symbols -> usage='None', detection='None'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=[],
                    uses_dlopen=True,
                    dlsym_symbols=[],
                    dlopen_libs=["libsqlite.so"]),
            ],
            package_info=self._make_pkg_info(bundled=False),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'None'
        assert row['detection'] == 'None'

    def test_s18_dynamic_plus_dlopen(self):
        """Dynamic + dlopen in same package -> usage='System-Link', detection='Mixed'."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libdynamic.so", file_type="shared_library",
                    arch="aarch64", direct_deps=["libcrypto.so.3"],
                    openssl_direct=True, openssl_transitive=False,
                    openssl_libs=["libcrypto.so.3"],
                    openssl_symbols=["SSL_connect"]),
                FileResult(
                    path="/libdlopen.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["SSL_read"],
                    uses_dlopen=True,
                    dlsym_symbols=["SSL_read"],
                    dlopen_libs=["libcrypto.so"]),
            ],
            package_info=self._make_pkg_info(bundled=False, native_count=2),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'System-Link'
        assert row['detection'] == 'Mixed'

    def test_s19_static_dynamic_dlopen_all(self):
        """All three mechanisms -> bundled_raw='Yes (static)' takes priority."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libstatic.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=True,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["EVP_sha256", "SSL_CTX_new"],
                    static_openssl=True, static_openssl_confidence='high'),
                FileResult(
                    path="/libdynamic.so", file_type="shared_library",
                    arch="aarch64", direct_deps=["libcrypto.so.3"],
                    openssl_direct=True, openssl_transitive=False,
                    openssl_libs=["libcrypto.so.3"],
                    openssl_symbols=["SSL_connect"]),
                FileResult(
                    path="/libdlopen.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["SSL_read"],
                    uses_dlopen=True,
                    dlsym_symbols=["SSL_read"],
                    dlopen_libs=["libcrypto.so"]),
            ],
            package_info=self._make_pkg_info(
                bundled='Yes (static)', native_count=3),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled (static)'

    def test_s20_low_confidence_static_is_none(self):
        """Static only, low confidence filtered -> usage='None'.

        Library has static_openssl=True but confidence is empty (below
        high/medium threshold), so _detect_static_providers() skips it.
        Low-confidence detection is insufficient for any Bundled label.
        """
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=True,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=["EVP_sha256", "SSL_connect"],
                    static_openssl=True),
            ],
            package_info=self._make_pkg_info(bundled=False),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'None'
        assert row['detection'] == 'Static'

    def test_s21_empty_files_detail(self):
        """HAP with no .so files -> usage='None', detection='None'."""
        result = self._make_result(
            files_detail=[],
            package_info=self._make_pkg_info(bundled=False, native_count=0),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'None'
        assert row['detection'] == 'None'
        assert row['total_syms'] == 0

    def test_s22_bundled_yes_bare_no_trailing_space(self):
        """bundled_openssl='Yes' (bare, no parenthetical) -> 'Bundled' not 'Bundled '.

        Guards against trailing space when _detect_static_providers returns
        a bare 'Yes' without detail parenthetical.
        """
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=["libcrypto.so"],
                    openssl_direct=True, openssl_transitive=False,
                    openssl_libs=["libcrypto.so"],
                    openssl_symbols=["SSL_connect"]),
            ],
            package_info=self._make_pkg_info(bundled='Yes'),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'Bundled'
        assert not row['openssl_usage'].endswith(' ')

    def test_s23_bundled_none_value(self):
        """bundled_openssl=None (from JSON null) -> falls through to ossl_type."""
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=[], openssl_direct=False,
                    openssl_transitive=False, openssl_libs=[],
                    openssl_symbols=[]),
            ],
            package_info=self._make_pkg_info(bundled=None),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'None'

    def test_s24_system_link_via_static_unresolved(self):
        """Static OpenSSL lib that DT_NEEDs libcrypto.so (not bundled) -> System-Link.

        When static_openssl=True but confidence is empty, _detect_static_providers
        skips it. The DT_NEEDED libcrypto.so is unresolved -> System-Link.
        Detection is 'Static' because static_openssl=True routes symbols to
        the static bucket in _classify_hap_detection.
        """
        result = self._make_result(
            files_detail=[
                FileResult(
                    path="/libapp.so", file_type="shared_library",
                    arch="aarch64", direct_deps=["libcrypto.so"],
                    openssl_direct=True, openssl_transitive=False,
                    openssl_libs=["libcrypto.so"],
                    openssl_symbols=["SSL_connect", "EVP_sha256"],
                    static_openssl=True),
            ],
            package_info=self._make_pkg_info(bundled=False),
        )
        row, *_ = self._classify_and_build(result)
        assert row['openssl_usage'] == 'System-Link'
        assert row['detection'] == 'Static'


HAP_TEST_DIR = os.path.join(
    os.path.dirname(__file__), '..', 'hap_test_packages')


def _hap_path(name):
    """Return absolute path to a HAP in hap_test_packages/."""
    return os.path.join(HAP_TEST_DIR, name)


def _scan_real_hap(hap_file):
    """Scan a real HAP file and return (json_report_dict, summary_row)."""
    with tempfile.TemporaryDirectory() as td:
        out_json = os.path.join(td, 'report.json')
        rc = main(['hap', hap_file, '-o', out_json])
        assert rc is None or rc == 0, f'hap command failed: rc={rc}'
        with open(out_json) as f:
            report = json.load(f)

        result = ScanResult(
            target=hap_file, scan_time="", tool_version="test",
            arch="aarch64")
        result.report_type = 'package'
        result.package_info = report['meta']['package']
        result.symbols_by_category = report['openssl_symbols'].get(
            'by_category', {})

        frs = []
        for fd in report['files_detail']:
            fr = FileResult(
                path=fd['path'],
                file_type=fd.get('type', 'shared_library'),
                arch=fd.get('arch', 'aarch64'),
                direct_deps=fd.get('direct_deps', []),
                openssl_direct=fd['openssl_deps'].get('direct', False),
                openssl_transitive=fd['openssl_deps'].get(
                    'transitive', False),
                openssl_libs=fd['openssl_deps'].get('libs', []),
                openssl_symbols=fd.get('openssl_symbols_used', []),
            )
            fr.static_openssl = fd.get('static_openssl', False)
            fr.static_openssl_version = fd.get('static_openssl_version')
            fr.static_ssl_library = fd.get('static_ssl_library', '')
            fr.static_openssl_confidence = fd.get(
                'static_openssl_confidence', '')
            fr.openssl_exported = fd.get('openssl_exported', [])
            dl = fd.get('dlopen_detection', {})
            if dl:
                fr.uses_dlopen = dl.get('uses_dlopen', False)
                fr.dlsym_symbols = dl.get('dlopen_symbols', [])
                fr.dlopen_libs = dl.get('dlopen_libs', [])
            frs.append(fr)
        result.files_detail = frs

        method, s, d, dl, ossl_type = _classify_hap_detection(result)
        row = _build_hap_summary_row(
            result, hap_file, method, s, d, dl, ossl_type)
        return report, row


@pytest.mark.skipif(
    not os.path.isdir(HAP_TEST_DIR),
    reason="hap_test_packages directory not found")
class TestRealHAPOpenSSLUsage:
    """Integration tests using real HAP packages with genuine ELF binaries.

    These tests verify the full pipeline: extraction, symbol analysis,
    static detection, and openssl_usage classification using:
    - Real prebuilt OpenSSL .so files from ohos-rs/ohos-openssl
    - Real OpenSSL engine binaries (loader_attic.so)
    - Real third-party app HAP packages
    """

    @pytest.mark.skipif(
        not os.path.isfile(_hap_path('bundled_openssl_test.hap')),
        reason="bundled_openssl_test.hap not available")
    def test_bundled_usage_with_real_libcrypto(self):
        """HAP bundling real libcrypto.so (ohos-openssl) -> usage='Bundled'.

        The test HAP contains:
          - libcrypto.so (6.5MB, real OpenSSL 3.x for aarch64)
          - libmytls.so  (renamed libssl.so, DT_NEEDs libcrypto.so)
          - libc++_shared.so
        libcrypto.so matches OPENSSL_LIBRARY_PATTERNS so has_standalone=True.
        """
        _, row = _scan_real_hap(_hap_path('bundled_openssl_test.hap'))
        assert row['openssl_usage'] == 'Bundled', (
            f"Expected 'Bundled', got '{row['openssl_usage']}'")
        assert row['detection'] in ('Static', 'Dynamic', 'Mixed')
        assert row['total_syms'] > 0

    @pytest.mark.skipif(
        not os.path.isfile(_hap_path('syslink_openssl_test.hap')),
        reason="syslink_openssl_test.hap not available")
    def test_system_link_usage_with_real_engine(self):
        """HAP with real OpenSSL engine that DT_NEEDs libcrypto.so (not bundled).

        The test HAP contains:
          - libengine_attic.so (renamed loader_attic.so from ohos-openssl)
            DT_NEEDED: libcrypto.so, libc.so
            Imports 130+ OpenSSL symbols, exports 0 OpenSSL symbols
          - libc++_shared.so
        libcrypto.so is NOT bundled -> unresolved external -> System-Link.
        """
        _, row = _scan_real_hap(_hap_path('syslink_openssl_test.hap'))
        assert row['openssl_usage'] == 'System-Link', (
            f"Expected 'System-Link', got '{row['openssl_usage']}'")
        assert row['detection'] == 'Dynamic'
        assert row['total_syms'] > 100

    @pytest.mark.skipif(
        not os.path.isfile(_hap_path('aloesend-1.17.2-sideloaded.hap')),
        reason="aloesend HAP not available")
    def test_bundled_static_with_real_flutter_boringssl(self):
        """AloeSend (Flutter app) with statically linked BoringSSL.

        libflutter.so has BoringSSL compiled in (medium confidence).
        No standalone libcrypto.so/libssl.so -> Bundled (static).
        """
        _, row = _scan_real_hap(
            _hap_path('aloesend-1.17.2-sideloaded.hap'))
        assert row['openssl_usage'] == 'Bundled (static)', (
            f"Expected 'Bundled (static)', got '{row['openssl_usage']}'")

    @pytest.mark.skipif(
        not os.path.isfile(_hap_path('FinVideo-1.0.0.hap')),
        reason="FinVideo HAP not available")
    def test_bundled_static_shared_with_real_ffmpeg(self):
        """FinVideo (Jellyfin player) with FFmpeg statically linking OpenSSL.

        libwlffmpeg.so has OpenSSL compiled in (high confidence, 6000+ syms).
        Other libs (libwlmediautil, libwlplayer) consume it via DT_NEEDED
        -> Bundled (static, shared).
        """
        _, row = _scan_real_hap(_hap_path('FinVideo-1.0.0.hap'))
        assert row['openssl_usage'] == 'Bundled (static, shared)', (
            f"Expected 'Bundled (static, shared)', got "
            f"'{row['openssl_usage']}'")
        assert row['total_syms'] > 5000

    @pytest.mark.skipif(
        not os.path.isfile(_hap_path('ppsspp.hap')),
        reason="ppsspp HAP not available")
    def test_none_usage_with_real_gameemu(self):
        """PPSSPP (game emulator) has no OpenSSL usage -> usage='None'.

        libppsspp_jni.so exports AES_encrypt/AES_decrypt (own impl),
        but imports zero OpenSSL symbols.
        """
        _, row = _scan_real_hap(_hap_path('ppsspp.hap'))
        assert row['openssl_usage'] == 'None', (
            f"Expected 'None', got '{row['openssl_usage']}'")
        assert row['detection'] == 'None'
        assert row['total_syms'] == 0

    @pytest.mark.skipif(
        not os.path.isfile(_hap_path('ClashForHarmonyOS.hap')),
        reason="Clash HAP not available")
    def test_none_usage_with_real_vpn_app(self):
        """Clash (VPN proxy) uses Go crypto, no OpenSSL -> usage='None'."""
        _, row = _scan_real_hap(_hap_path('ClashForHarmonyOS.hap'))
        assert row['openssl_usage'] == 'None'
        assert row['total_syms'] == 0


HITLS_SYMS = ['HITLS_Connect', 'HITLS_Read', 'HITLS_Write', 'HITLS_Close',
              'HITLS_Accept', 'BSL_ERR_GetError', 'CRYPT_EAL_CipherInit']
WOLF_SYMS = ['wolfSSL_Init', 'wolfSSL_CTX_new', 'wolfSSL_new']


def _create_hap_with_elfs(path, bundle_name, module_name, elf_entries,
                          include_openssl=False):
    """Build a HAP ZIP with module.json and custom ELF .so files.

    Args:
        path: Output HAP file path.
        bundle_name: App bundle name for module.json.
        module_name: Module name for module.json.
        elf_entries: List of (filename, elf_bytes) tuples to place in
            libs/arm64-v8a/.
        include_openssl: If True, also include a minimal libcrypto.so.3.
    """
    with zipfile.ZipFile(path, 'w') as zf:
        zf.writestr("module.json", json.dumps({
            "module": {
                "name": module_name,
                "type": "entry",
                "deviceTypes": ["default"]
            },
            "app": {
                "bundleName": bundle_name,
                "versionCode": 1000000,
                "versionName": "1.0.0",
                "minAPIVersion": 11
            }
        }))
        for fname, data in elf_entries:
            zf.writestr(f"libs/arm64-v8a/{fname}", data)
        if include_openssl:
            zf.writestr("libs/arm64-v8a/libcrypto.so.3", _minimal_elf64())
    return path


def _scan_hap_json(hap_path, output_dir):
    """Run hap subcommand and return parsed JSON report dict."""
    os.makedirs(output_dir, exist_ok=True)
    ret = main(['hap', hap_path, '-o', output_dir, '--json-only'])
    assert ret is None or ret == 0, f'hap command failed: rc={ret}'
    json_files = [f for f in os.listdir(output_dir) if f.endswith('.json')]
    assert len(json_files) >= 1, f'No JSON output in {output_dir}'
    with open(os.path.join(output_dir, json_files[0])) as f:
        return json.load(f)


class TestCustomPatternHapIntegration:
    """End-to-end tests for CustomMatcher integration in the HAP pipeline.

    Verifies the full path: HAP zip -> extract .so -> CustomMatcher ->
    package_info['custom_match'] -> JSON report + summary.xlsx column 22.
    """

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _build_rodata(self, symbols):
        """Pack symbols as null-terminated strings into a .rodata blob."""
        return b'\x00'.join(s.encode() for s in symbols) + b'\x00'

    def test_b1_hitls_und_only(self):
        """B1: HiTLS UND symbols -> custom_match='openHiTLS (3)'."""
        from tests.fixtures.elf_builder import ELFBuilder
        elf = ELFBuilder()
        for sym in HITLS_SYMS[:3]:
            elf.add_dynsym(sym, defined=False)
        elf_bytes = elf.build()

        hap_path = os.path.join(self.tmpdir, 'hitls_und.hap')
        out_dir = os.path.join(self.tmpdir, 'out_b1')
        _create_hap_with_elfs(hap_path, 'com.test.hitls.und', 'entry',
                              [('libhitls.so', elf_bytes)])
        data = _scan_hap_json(hap_path, out_dir)

        pi = data['meta']['package']
        assert pi['custom_match'] == 'openHiTLS (3)'
        assert 'openHiTLS' in pi['custom_match_groups']
        assert len(pi['custom_match_groups']['openHiTLS']) == 3

    def test_b2_hitls_def_only(self):
        """B2: HiTLS DEF symbols (static link) -> custom_match='openHiTLS (3)'."""
        from tests.fixtures.elf_builder import ELFBuilder
        elf = ELFBuilder()
        for sym in HITLS_SYMS[:3]:
            elf.add_dynsym(sym, defined=True)
        elf_bytes = elf.build()

        hap_path = os.path.join(self.tmpdir, 'hitls_def.hap')
        out_dir = os.path.join(self.tmpdir, 'out_b2')
        _create_hap_with_elfs(hap_path, 'com.test.hitls.def', 'entry',
                              [('libhitls.so', elf_bytes)])
        data = _scan_hap_json(hap_path, out_dir)

        pi = data['meta']['package']
        assert pi['custom_match'] == 'openHiTLS (3)'
        assert 'openHiTLS' in pi['custom_match_groups']

    def test_b3_hitls_rodata_only(self):
        """B3: HiTLS strings in .rodata only -> custom_match='openHiTLS (3)'."""
        from tests.fixtures.elf_builder import ELFBuilder
        rodata = self._build_rodata(HITLS_SYMS[:3])
        elf = ELFBuilder().set_rodata(rodata)
        elf_bytes = elf.build()

        hap_path = os.path.join(self.tmpdir, 'hitls_rodata.hap')
        out_dir = os.path.join(self.tmpdir, 'out_b3')
        _create_hap_with_elfs(hap_path, 'com.test.hitls.rodata', 'entry',
                              [('libhitls.so', elf_bytes)])
        data = _scan_hap_json(hap_path, out_dir)

        pi = data['meta']['package']
        assert pi['custom_match'] == 'openHiTLS (3)'
        assert 'openHiTLS' in pi['custom_match_groups']
        assert len(pi['custom_match_groups']['openHiTLS']) == 3

    def test_b4_hitls_plus_bundled_openssl(self):
        """B4: HiTLS UND + bundled libcrypto.so.3 -> openssl_usage=Bundled."""
        from tests.fixtures.elf_builder import ELFBuilder
        elf = ELFBuilder()
        for sym in HITLS_SYMS[:2]:
            elf.add_dynsym(sym, defined=False)
        elf_bytes = elf.build()

        hap_path = os.path.join(self.tmpdir, 'hitls_bundled.hap')
        out_dir = os.path.join(self.tmpdir, 'out_b4')
        _create_hap_with_elfs(hap_path, 'com.test.hitls.bundled', 'entry',
                              [('libhitls.so', elf_bytes)],
                              include_openssl=True)
        data = _scan_hap_json(hap_path, out_dir)

        pi = data['meta']['package']
        assert 'openHiTLS' in pi['custom_match']
        assert pi.get('bundled_openssl') is not False

    def test_b5_hitls_plus_system_openssl(self):
        """B5: HiTLS UND + SSL_connect UND -> openssl_usage=System-Link."""
        from tests.fixtures.elf_builder import ELFBuilder
        elf = ELFBuilder()
        for sym in HITLS_SYMS[:3]:
            elf.add_dynsym(sym, defined=False)
        elf.add_dynsym('SSL_connect', defined=False)
        elf.add_dynsym('SSL_read', defined=False)
        elf_bytes = elf.build()

        hap_path = os.path.join(self.tmpdir, 'hitls_syslink.hap')
        out_dir = os.path.join(self.tmpdir, 'out_b5')
        _create_hap_with_elfs(hap_path, 'com.test.hitls.syslink', 'entry',
                              [('libhitls.so', elf_bytes)])
        data = _scan_hap_json(hap_path, out_dir)

        pi = data['meta']['package']
        assert 'openHiTLS' in pi['custom_match']

        result = ScanResult(
            target=hap_path, scan_time="", tool_version="test",
            arch="aarch64")
        result.report_type = 'package'
        result.package_info = pi
        result.symbols_by_category = data['openssl_symbols'].get(
            'by_category', {})
        frs = []
        for fd in data['files_detail']:
            fr = FileResult(
                path=fd['path'],
                file_type=fd.get('type', 'shared_library'),
                arch=fd.get('arch', 'aarch64'),
                direct_deps=fd.get('direct_deps', []),
                openssl_direct=fd['openssl_deps'].get('direct', False),
                openssl_transitive=fd['openssl_deps'].get('transitive', False),
                openssl_libs=fd['openssl_deps'].get('libs', []),
                openssl_symbols=fd.get('openssl_symbols_used', []),
            )
            frs.append(fr)
        result.files_detail = frs
        method, s, d, dl, ossl_type = _classify_hap_detection(result)
        row = _build_hap_summary_row(result, hap_path, method, s, d, dl,
                                     ossl_type)
        assert row['openssl_usage'] == 'System-Link'

    def test_b6_multi_lib(self):
        """B6: HiTLS UND + wolfSSL UND -> both groups in custom_match."""
        from tests.fixtures.elf_builder import ELFBuilder
        elf = ELFBuilder()
        elf.add_dynsym('HITLS_Connect', defined=False)
        elf.add_dynsym('wolfSSL_Init', defined=False)
        elf_bytes = elf.build()

        hap_path = os.path.join(self.tmpdir, 'multi_lib.hap')
        out_dir = os.path.join(self.tmpdir, 'out_b6')
        _create_hap_with_elfs(hap_path, 'com.test.multilib', 'entry',
                              [('libmulti.so', elf_bytes)])
        data = _scan_hap_json(hap_path, out_dir)

        pi = data['meta']['package']
        assert 'openHiTLS (1)' in pi['custom_match']
        assert 'wolfSSL (1)' in pi['custom_match']
        assert 'openHiTLS' in pi['custom_match_groups']
        assert 'wolfSSL' in pi['custom_match_groups']

    def test_b7_dedup_und_over_rodata(self):
        """B7: Same symbol in UND + rodata -> deduplicated, no double count."""
        from tests.fixtures.elf_builder import ELFBuilder
        rodata = self._build_rodata(['HITLS_Connect', 'HITLS_Read'])
        elf = (ELFBuilder()
               .add_dynsym('HITLS_Connect', defined=False)
               .set_rodata(rodata))
        elf_bytes = elf.build()

        hap_path = os.path.join(self.tmpdir, 'dedup.hap')
        out_dir = os.path.join(self.tmpdir, 'out_b7')
        _create_hap_with_elfs(hap_path, 'com.test.dedup', 'entry',
                              [('libdedup.so', elf_bytes)])
        data = _scan_hap_json(hap_path, out_dir)

        pi = data['meta']['package']
        assert pi['custom_match'] == 'openHiTLS (2)'
        assert len(pi['custom_match_groups']['openHiTLS']) == 2

    def test_b8_clean_hap(self):
        """B8: No custom pattern matches -> custom_match=''."""
        from tests.fixtures.elf_builder import ELFBuilder
        elf = (ELFBuilder()
               .add_dynsym('printf', defined=False)
               .add_dynsym('malloc', defined=False)
               .build())

        hap_path = os.path.join(self.tmpdir, 'clean.hap')
        out_dir = os.path.join(self.tmpdir, 'out_b8')
        _create_hap_with_elfs(hap_path, 'com.test.clean', 'entry',
                              [('libclean.so', elf)])
        data = _scan_hap_json(hap_path, out_dir)

        pi = data['meta']['package']
        assert pi['custom_match'] == ''
        assert pi.get('custom_match_groups', {}) == {}

    def test_b9_batch_scan_summary(self):
        """B9: Batch scan 3 HAPs -> per-package JSON + summary.xlsx correct."""
        from tests.fixtures.elf_builder import ELFBuilder

        pkg_dir = os.path.join(self.tmpdir, 'pkgs')
        os.makedirs(pkg_dir)

        elf_hitls = (ELFBuilder()
                     .add_dynsym('HITLS_Accept', defined=False)
                     .add_dynsym('HITLS_Connect', defined=False)
                     .build())
        _create_hap_with_elfs(
            os.path.join(pkg_dir, 'hitls.hap'),
            'com.test.batch.hitls', 'entry',
            [('libhitls.so', elf_hitls)])

        elf_wolf = (ELFBuilder()
                    .add_dynsym('wolfSSL_Init', defined=False)
                    .build())
        _create_hap_with_elfs(
            os.path.join(pkg_dir, 'wolf.hap'),
            'com.test.batch.wolf', 'entry',
            [('libwolf.so', elf_wolf)])

        elf_clean = (ELFBuilder()
                     .add_dynsym('printf', defined=False)
                     .build())
        _create_hap_with_elfs(
            os.path.join(pkg_dir, 'clean.hap'),
            'com.test.batch.clean', 'entry',
            [('libclean.so', elf_clean)])

        out_dir = os.path.join(self.tmpdir, 'out_b9')
        ret = main(['hap', pkg_dir, '-o', out_dir, '--json-only'])
        assert ret == 0

        summary_path = os.path.join(out_dir, 'summary.xlsx')
        assert os.path.isfile(summary_path)

        from openpyxl import load_workbook
        wb = load_workbook(summary_path)
        ws = wb.active

        col_keys = [c[0] for c in _HAP_SUMMARY_COLUMNS]
        cm_col = col_keys.index('custom_match') + 1

        header = ws.cell(row=1, column=cm_col).value
        assert header == 'Custom Match'

        cm_values = {}
        for r in range(2, ws.max_row + 1):
            pkg = ws.cell(row=r, column=1).value
            if pkg and pkg != 'TOTAL':
                cm_values[pkg] = ws.cell(row=r, column=cm_col).value or ''

        hitls_pkg = [k for k in cm_values if 'hitls' in k.lower()]
        wolf_pkg = [k for k in cm_values if 'wolf' in k.lower()]
        clean_pkg = [k for k in cm_values if 'clean' in k.lower()]

        assert len(hitls_pkg) == 1
        assert 'openHiTLS (2)' in cm_values[hitls_pkg[0]]

        assert len(wolf_pkg) == 1
        assert 'wolfSSL (1)' in cm_values[wolf_pkg[0]]

        assert len(clean_pkg) == 1
        assert cm_values[clean_pkg[0]] == ''

    def test_b10_json_roundtrip_hap_summary(self):
        """B10: hap -> JSON -> hap-summary -> custom_match survives."""
        from tests.fixtures.elf_builder import ELFBuilder

        elf = (ELFBuilder()
               .add_dynsym('HITLS_Accept', defined=False)
               .add_dynsym('HITLS_Connect', defined=False)
               .add_dynsym('HITLS_Read', defined=False)
               .build())

        hap_path = os.path.join(self.tmpdir, 'roundtrip.hap')
        json_dir = os.path.join(self.tmpdir, 'json_out')
        _create_hap_with_elfs(hap_path, 'com.test.roundtrip', 'entry',
                              [('libhitls.so', elf)])

        ret1 = main(['hap', hap_path, '-o', json_dir, '--json-only'])
        assert ret1 is None or ret1 == 0

        json_files = [f for f in os.listdir(json_dir) if f.endswith('.json')]
        assert len(json_files) >= 1

        with open(os.path.join(json_dir, json_files[0])) as f:
            original = json.load(f)
        original_cm = original['meta']['package']['custom_match']
        assert original_cm == 'openHiTLS (3)'

        summary_dir = os.path.join(self.tmpdir, 'summary_out')
        os.makedirs(summary_dir, exist_ok=True)
        summary_xlsx = os.path.join(summary_dir, 'summary.xlsx')
        ret2 = main(['hap-summary', json_dir, '-o', summary_xlsx])
        assert ret2 == 0

        from openpyxl import load_workbook
        wb = load_workbook(summary_xlsx)
        ws = wb.active

        col_keys = [c[0] for c in _HAP_SUMMARY_COLUMNS]
        cm_col = col_keys.index('custom_match') + 1

        for r in range(2, ws.max_row + 1):
            pkg = ws.cell(row=r, column=1).value
            if pkg and pkg != 'TOTAL':
                assert ws.cell(row=r, column=cm_col).value == original_cm

    def test_b11_app_container(self):
        """B11: APP container with inner HiTLS HAP -> per-HAP custom_match."""
        from tests.fixtures.elf_builder import ELFBuilder

        elf_hitls = (ELFBuilder()
                     .add_dynsym('HITLS_Accept', defined=False)
                     .add_dynsym('HITLS_Connect', defined=False)
                     .build())
        elf_clean = (ELFBuilder()
                     .add_dynsym('printf', defined=False)
                     .build())

        app_path = os.path.join(self.tmpdir, 'container.app')
        with zipfile.ZipFile(app_path, 'w') as app_zf:
            for name, bname, mname, elf_data in [
                ('hitls_mod.hap', 'com.test.container', 'hitls_mod',
                 elf_hitls),
                ('clean_mod.hap', 'com.test.container', 'clean_mod',
                 elf_clean),
            ]:
                buf = io.BytesIO()
                with zipfile.ZipFile(buf, 'w') as hap_zf:
                    hap_zf.writestr("module.json", json.dumps({
                        "module": {"name": mname, "type": "entry",
                                   "deviceTypes": ["default"]},
                        "app": {"bundleName": bname,
                                "versionCode": 1, "versionName": "1.0.0",
                                "minAPIVersion": 11}
                    }))
                    hap_zf.writestr(f"libs/arm64-v8a/lib{mname}.so", elf_data)
                app_zf.writestr(name, buf.getvalue())

        out_dir = os.path.join(self.tmpdir, 'out_b11')
        ret = main(['hap', app_path, '-o', out_dir, '--json-only'])
        assert ret == 0

        json_files = sorted(f for f in os.listdir(out_dir)
                            if f.endswith('.json'))
        assert len(json_files) == 2

        results = {}
        for jf in json_files:
            with open(os.path.join(out_dir, jf)) as f:
                d = json.load(f)
            mname = d['meta']['package']['module_name']
            results[mname] = d

        assert 'openHiTLS' in results['hitls_mod']['meta']['package']['custom_match']
        assert results['clean_mod']['meta']['package']['custom_match'] == ''


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
