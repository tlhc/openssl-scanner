"""Tests for HiTLS compat integration into source report pipeline."""

import json
import os
import sys
import tempfile
from types import SimpleNamespace

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openssl_scanner.hitls_compat import HiTLSCompat
from openssl_scanner.source_analyzer import CallSite, SourceScanResult
from openssl_scanner.source_exporter import (
    SourceExcelExporter, SourceJsonExporter, SourceMergeExporter,
)

FIXTURE_PATH = os.path.join(
    os.path.dirname(__file__), 'fixtures', 'test_hitls_compat.json'
)


def _make_compat():
    compat = HiTLSCompat()
    compat.load(FIXTURE_PATH)
    return compat


def _make_result(target="/test/src"):
    return SourceScanResult(
        target=target,
        scan_time="2026-03-12T10:00:00",
        tool_version="1.0.0",
        total_files_scanned=2,
        files_with_calls=1,
        total_call_sites=3,
        unique_symbols=["SSL_CTX_new", "EVP_DigestInit_ex", "ENGINE_init"],
        symbols_by_category={"ssl_core": ["SSL_CTX_new"],
                             "crypto_evp": ["EVP_DigestInit_ex"],
                             "crypto_engine": ["ENGINE_init"]},
        call_sites=[
            CallSite(file_path="/test/src/main.c", file_name="main.c",
                     caller_function="init_tls", line_number=10, column=4,
                     ossl_symbol="SSL_CTX_new", category="ssl_core",
                     call_args="(TLS_method())", language="c"),
            CallSite(file_path="/test/src/main.c", file_name="main.c",
                     caller_function="do_hash", line_number=20, column=4,
                     ossl_symbol="EVP_DigestInit_ex", category="crypto_evp",
                     call_args="(ctx, EVP_sha256(), NULL)", language="c"),
            CallSite(file_path="/test/src/eng.c", file_name="eng.c",
                     caller_function="setup_eng", line_number=5, column=4,
                     ossl_symbol="ENGINE_init", category="crypto_engine",
                     call_args="(e)", language="c"),
        ],
        errors=[],
    )


class TestExcelWithHiTLS:

    def test_xlsx_with_hitls_has_extra_columns(self):
        compat = _make_compat()
        result = _make_result()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            SourceExcelExporter().export(result, path, hitls_compat=compat)
            from openssl_scanner import _vendor  # noqa: F401
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb["OpenSSL Call Sites"]
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            assert "HiTLS Status" in headers
            assert "HiTLS Equivalent" in headers
            hi_idx = headers.index("HiTLS Status")
            he_idx = headers.index("HiTLS Equivalent")
            assert hi_idx == 6
            assert he_idx == 7

            row2 = [cell.value for cell in list(ws.iter_rows(min_row=2, max_row=2))[0]]
            assert row2[hi_idx] == "available"
            assert row2[he_idx] == "HITLS_CFG_NewTLSConfig"

            ws_sym = wb["Symbol Summary"]
            sym_headers = [cell.value for cell in next(ws_sym.iter_rows(max_row=1))]
            assert "HiTLS Status" in sym_headers
            assert "HiTLS Equivalent" in sym_headers
            wb.close()
        finally:
            os.unlink(path)

    def test_xlsx_without_hitls_no_extra_columns(self):
        result = _make_result()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            SourceExcelExporter().export(result, path)
            from openssl_scanner import _vendor  # noqa: F401
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb["OpenSSL Call Sites"]
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            assert "HiTLS Status" not in headers
            assert "HiTLS Equivalent" not in headers
            assert len(headers) == 8
            wb.close()
        finally:
            os.unlink(path)


class TestJsonWithHiTLS:

    def test_json_with_hitls_has_fields(self):
        compat = _make_compat()
        result = _make_result()
        json_str = SourceJsonExporter().export(result, hitls_compat=compat)
        data = json.loads(json_str)

        cs0 = data['call_sites'][0]
        assert cs0['hitls_status'] == 'available'
        assert cs0['hitls_equiv'] == 'HITLS_CFG_NewTLSConfig'

        cs1 = data['call_sites'][1]
        assert cs1['hitls_status'] == 'available'
        assert cs1['hitls_equiv'] == 'CRYPT_EAL_MdInit'

        cs2 = data['call_sites'][2]
        assert cs2['hitls_status'] == 'not_available'
        assert cs2['hitls_equiv'] is None

        cov = data['summary']['hitls_coverage']
        assert cov['available'] == 2
        assert cov['partial'] == 0
        assert cov['not_available'] == 1
        assert cov['unknown'] == 0

    def test_json_without_hitls_no_fields(self):
        result = _make_result()
        json_str = SourceJsonExporter().export(result)
        data = json.loads(json_str)
        assert 'hitls_status' not in data['call_sites'][0]
        assert 'hitls_coverage' not in data['summary']


class TestMergeWithHiTLS:

    def test_merge_from_json_with_hitls(self):
        compat = _make_compat()
        result = _make_result()

        with tempfile.TemporaryDirectory() as tmp:
            json_path = os.path.join(tmp, 'proj.json')
            SourceJsonExporter().export(result, json_path)

            xlsx_path = os.path.join(tmp, 'merged.xlsx')
            merger = SourceMergeExporter()
            merger.merge_from_json([json_path], xlsx_path,
                                   hitls_compat=compat)

            from openssl_scanner import _vendor  # noqa: F401
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True)
            ws_sym = wb["Symbol Summary"]
            sym_headers = [cell.value for cell in next(ws_sym.iter_rows(max_row=1))]
            assert "HiTLS Status" in sym_headers
            assert "HiTLS Equivalent" in sym_headers
            wb.close()

    def test_merge_from_json_without_hitls(self):
        result = _make_result()

        with tempfile.TemporaryDirectory() as tmp:
            json_path = os.path.join(tmp, 'proj.json')
            SourceJsonExporter().export(result, json_path)

            xlsx_path = os.path.join(tmp, 'merged.xlsx')
            merger = SourceMergeExporter()
            merger.merge_from_json([json_path], xlsx_path)

            from openssl_scanner import _vendor  # noqa: F401
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True)
            ws_sym = wb["Symbol Summary"]
            sym_headers = [cell.value for cell in next(ws_sym.iter_rows(max_row=1))]
            assert "HiTLS Status" not in sym_headers
            wb.close()


class TestDiffWithHiTLS:

    def test_diff_xlsx_with_hitls(self):
        from openssl_scanner.source_diff import (
            DiffResult, ProjectDelta, SymbolDelta, DiffStatus,
            MetricDelta, SourceDiffExcelExporter,
        )

        compat = _make_compat()
        pd = ProjectDelta(
            project="test",
            metrics=[MetricDelta("total_call_sites", 2, 3, 1)],
            symbol_delta=[
                SymbolDelta(DiffStatus.ADDED, "SSL_CTX_new", "ssl_core", 0, 5),
                SymbolDelta(DiffStatus.REMOVED, "ENGINE_init", "crypto_engine", 3, 0),
            ],
            has_call_site_changes=True,
        )
        result = DiffResult(old_label="old.json", new_label="new.json",
                            projects=[pd])

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            SourceDiffExcelExporter(hitls_compat=compat).export(result, path)
            from openssl_scanner import _vendor  # noqa: F401
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb["Symbol Delta"]
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            assert "HiTLS Status" in headers
            assert "HiTLS Equivalent" in headers

            row2 = [cell.value for cell in list(ws.iter_rows(min_row=2, max_row=2))[0]]
            hi_idx = headers.index("HiTLS Status")
            assert row2[hi_idx] in ('available', 'not_available', 'partial', 'unknown')
            wb.close()
        finally:
            os.unlink(path)

    def test_diff_xlsx_without_hitls(self):
        from openssl_scanner.source_diff import (
            DiffResult, ProjectDelta, SymbolDelta, DiffStatus,
            MetricDelta, SourceDiffExcelExporter,
        )

        pd = ProjectDelta(
            project="test",
            metrics=[MetricDelta("total_call_sites", 2, 3, 1)],
            symbol_delta=[
                SymbolDelta(DiffStatus.ADDED, "SSL_CTX_new", "ssl_core", 0, 5),
            ],
            has_call_site_changes=True,
        )
        result = DiffResult(old_label="old.json", new_label="new.json",
                            projects=[pd])

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            SourceDiffExcelExporter().export(result, path)
            from openssl_scanner import _vendor  # noqa: F401
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb["Symbol Delta"]
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            assert "HiTLS Status" not in headers
            wb.close()
        finally:
            os.unlink(path)

    def test_diff_json_with_hitls(self):
        from openssl_scanner.source_diff import (
            DiffResult, ProjectDelta, SymbolDelta, DiffStatus,
            MetricDelta, SourceDiffJsonExporter,
        )

        compat = _make_compat()
        pd = ProjectDelta(
            project="test",
            metrics=[MetricDelta("total_call_sites", 2, 3, 1)],
            symbol_delta=[
                SymbolDelta(DiffStatus.ADDED, "SSL_CTX_new", "ssl_core", 0, 5),
            ],
            has_call_site_changes=True,
        )
        result = DiffResult(old_label="old.json", new_label="new.json",
                            projects=[pd])

        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            path = f.name
        try:
            SourceDiffJsonExporter(hitls_compat=compat).export(result, path)
            with open(path) as f:
                data = json.load(f)
            sd = data['symbol_delta'][0]
            assert sd['hitls_status'] == 'available'
            assert sd['hitls_equiv'] == 'HITLS_CFG_NewTLSConfig'
        finally:
            os.unlink(path)

    def test_diff_json_without_hitls(self):
        from openssl_scanner.source_diff import (
            DiffResult, ProjectDelta, SymbolDelta, DiffStatus,
            MetricDelta, SourceDiffJsonExporter,
        )

        pd = ProjectDelta(
            project="test",
            metrics=[MetricDelta("total_call_sites", 2, 3, 1)],
            symbol_delta=[
                SymbolDelta(DiffStatus.ADDED, "SSL_CTX_new", "ssl_core", 0, 5),
            ],
            has_call_site_changes=True,
        )
        result = DiffResult(old_label="old.json", new_label="new.json",
                            projects=[pd])

        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            path = f.name
        try:
            SourceDiffJsonExporter().export(result, path)
            with open(path) as f:
                data = json.load(f)
            sd = data['symbol_delta'][0]
            assert 'hitls_status' not in sd
        finally:
            os.unlink(path)


class TestLoadHiTLSCompat:
    """GAP-1: Test _load_hitls_compat() CLI helper."""

    def test_disabled_returns_none(self):
        from openssl_scanner.__main__ import _load_hitls_compat
        args = SimpleNamespace(hitls_compat=False, hitls_map=None)
        assert _load_hitls_compat(args) is None

    def test_enabled_loads_builtin(self):
        from openssl_scanner.__main__ import _load_hitls_compat
        args = SimpleNamespace(hitls_compat=True, hitls_map=None)
        compat = _load_hitls_compat(args)
        assert compat is not None
        assert compat.is_loaded()

    def test_custom_path(self):
        from openssl_scanner.__main__ import _load_hitls_compat
        args = SimpleNamespace(hitls_compat=True, hitls_map=FIXTURE_PATH)
        compat = _load_hitls_compat(args)
        assert compat is not None
        status, _ = compat.lookup('SSL_CTX_new')
        assert status == 'available'

    def test_bad_path_raises(self):
        from openssl_scanner.__main__ import _load_hitls_compat
        args = SimpleNamespace(hitls_compat=True, hitls_map='/nonexistent.json')
        with pytest.raises(FileNotFoundError):
            _load_hitls_compat(args)


class TestMergeXlsxWithHiTLS:
    """GAP-3: Test merge() XLSX-to-XLSX path with HiTLS."""

    def test_merge_xlsx_with_hitls(self):
        compat = _make_compat()
        result = _make_result()

        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = os.path.join(tmp, 'proj.xlsx')
            SourceExcelExporter().export(result, xlsx_path)

            merged_path = os.path.join(tmp, 'merged.xlsx')
            merger = SourceMergeExporter()
            merger.merge([xlsx_path], merged_path, hitls_compat=compat)

            from openssl_scanner import _vendor  # noqa: F401
            from openpyxl import load_workbook
            wb = load_workbook(merged_path, read_only=True)
            ws_sym = wb["Symbol Summary"]
            headers = [cell.value for cell in next(ws_sym.iter_rows(max_row=1))]
            assert "HiTLS Status" in headers
            assert "HiTLS Equivalent" in headers
            wb.close()


class TestRemergeHiTLSXlsx:
    """BUG-3: Re-merging HiTLS-enabled XLSX should not corrupt columns."""

    def test_remerge_hitls_xlsx_preserves_call_args(self):
        compat = _make_compat()
        result = _make_result()

        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = os.path.join(tmp, 'proj.xlsx')
            SourceExcelExporter().export(result, xlsx_path, hitls_compat=compat)

            merged_path = os.path.join(tmp, 'merged.xlsx')
            merger = SourceMergeExporter()
            merger.merge([xlsx_path], merged_path)

            from openssl_scanner import _vendor  # noqa: F401
            from openpyxl import load_workbook
            wb = load_workbook(merged_path, read_only=True)
            ws = wb[wb.sheetnames[1]]
            row2 = [cell.value for cell in
                    list(ws.iter_rows(min_row=2, max_row=2))[0]]
            assert "(TLS_method())" in str(row2), (
                f"call_args lost after re-merge: {row2}")
            wb.close()
