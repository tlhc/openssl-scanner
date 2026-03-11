"""
R5: Source merge + symbol summary overflow tests.

Verifies that when per-project sheets overflow (split across multiple XLSX sheets),
the Summary sheet and Symbol Summary sheet still contain correct aggregated data.

Tests cover:
- Task 2: _merge_to_workbook with overflow
- Task 3: merge_from_json with overflow
- Task 4: Full round-trip (export -> _read_xlsx -> merge) with overflow
- Task 5: Edge cases at exact boundary (49 vs 50 rows at MAX_ROW=50)
"""

import json
import os
import sys
import tempfile
from unittest import mock

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openssl_scanner import _vendor  # noqa: F401
from openpyxl import load_workbook

from openssl_scanner.source_analyzer import CallSite, SourceScanResult
from openssl_scanner.source_exporter import (
    COLUMNS,
    SourceExcelExporter,
    SourceMergeExporter,
)


def _make_merge_row(idx, symbol=None, category=None, file_path=None):
    """Build a raw row list matching COLUMNS layout for _merge_to_workbook."""
    return [
        file_path or f"/tmp/src/file_{idx:04d}.c",
        f"file_{idx:04d}.c",
        f"func_{idx:04d}",
        idx + 1,
        symbol or f"SSL_sym_{idx:04d}",
        category or "ssl_core",
        f"(arg_{idx})",
        "dynamic-link",
    ]


def _make_call_site(idx, symbol=None, category=None, file_path=None):
    """Build a CallSite for SourceExcelExporter.export()."""
    fp = file_path or f"/tmp/src/file_{idx:04d}.c"
    return CallSite(
        file_path=fp,
        file_name=os.path.basename(fp),
        caller_function=f"func_{idx:04d}",
        line_number=idx + 1,
        column=4,
        ossl_symbol=symbol or f"SSL_sym_{idx:04d}",
        category=category or "ssl_core",
        call_args=f"(arg_{idx})",
        language="c",
    )


def _make_result(call_sites, target="/tmp/test"):
    """Build a SourceScanResult from a list of CallSite objects."""
    unique = sorted(set(cs.ossl_symbol for cs in call_sites))
    cats = {}
    for cs in call_sites:
        cats.setdefault(cs.category, set()).add(cs.ossl_symbol)
    symbols_by_category = {k: sorted(v) for k, v in cats.items()}
    return SourceScanResult(
        target=target,
        scan_time="2026-03-11T00:00:00",
        tool_version="1.0.0",
        total_files_scanned=max(1, len(call_sites)),
        files_with_calls=len(set(cs.file_path for cs in call_sites)) if call_sites else 0,
        total_call_sites=len(call_sites),
        unique_symbols=unique,
        symbols_by_category=symbols_by_category,
        call_sites=call_sites,
        errors=[],
    )


def _make_json_report(n_sites, symbol_prefix="SSL", category="ssl_core",
                      target="/tmp/proj", files_scanned=10):
    """Build a source_scan JSON report dict with n_sites call sites."""
    call_sites = []
    for i in range(n_sites):
        call_sites.append({
            "file_path": f"/tmp/proj/file_{i:04d}.c",
            "file_name": f"file_{i:04d}.c",
            "caller_function": f"func_{i:04d}",
            "line_number": i + 1,
            "column": 4,
            "ossl_symbol": f"{symbol_prefix}_sym_{i:04d}",
            "category": category,
            "call_args": f"(arg_{i})",
            "language": "c",
            "detection_method": "dynamic-link",
        })
    unique = sorted(set(cs["ossl_symbol"] for cs in call_sites))
    return {
        "meta": {
            "tool_version": "1.0.0",
            "report_type": "source_scan",
            "scan_time": "2026-03-11T00:00:00",
            "target": target,
        },
        "summary": {
            "total_files_scanned": files_scanned,
            "files_with_calls": len(set(cs["file_path"] for cs in call_sites)),
            "total_call_sites": len(call_sites),
            "unique_symbols_count": len(unique),
            "unique_symbols": unique,
            "symbols_by_category": {category: unique},
        },
        "call_sites": call_sites,
        "errors": [],
    }


def _count_project_sheets(wb, project_name):
    """Count how many sheets belong to a project (base + overflow)."""
    count = 0
    for sn in wb.sheetnames:
        if sn == project_name:
            count += 1
        elif sn.startswith(project_name[:25]) and "(" in sn:
            count += 1
    return count


def _count_data_rows_in_sheets(wb, sheet_prefix):
    """Count total data rows across all sheets with matching prefix."""
    total = 0
    for sn in wb.sheetnames:
        if sn == sheet_prefix or sn.startswith(sheet_prefix[:25]):
            ws = wb[sn]
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row=row_idx, column=1).value is not None:
                    total += 1
    return total


class TestMergeToWorkbookOverflow:
    """Task 2: _merge_to_workbook with overflow produces correct aggregation."""

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_summary_sheet_project_rows(self, tmp_path):
        """Summary sheet has 3 project rows + 1 TOTAL row."""
        path = str(tmp_path / "merge.xlsx")
        project_data = [
            {'name': 'ProjA', 'files_scanned': 10,
             'rows': [_make_merge_row(i) for i in range(80)]},
            {'name': 'ProjB', 'files_scanned': 20,
             'rows': [_make_merge_row(i, symbol=f"EVP_b_{i:04d}",
                                      category="crypto_evp") for i in range(120)]},
            {'name': 'ProjC', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"BIO_c_{i:04d}",
                                      category="crypto_bio") for i in range(10)]},
        ]
        SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path, read_only=True)
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))

        assert rows[0][0] == "Project"
        assert rows[1][0] == "ProjA"
        assert rows[2][0] == "ProjB"
        assert rows[3][0] == "ProjC"
        assert rows[4][0] == "TOTAL"
        assert len(rows) == 5
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_summary_call_site_counts(self, tmp_path):
        """Summary sheet per-project call site counts = [80, 120, 10]."""
        path = str(tmp_path / "merge.xlsx")
        project_data = [
            {'name': 'ProjA', 'files_scanned': 10,
             'rows': [_make_merge_row(i) for i in range(80)]},
            {'name': 'ProjB', 'files_scanned': 20,
             'rows': [_make_merge_row(i, symbol=f"EVP_b_{i:04d}",
                                      category="crypto_evp") for i in range(120)]},
            {'name': 'ProjC', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"BIO_c_{i:04d}",
                                      category="crypto_bio") for i in range(10)]},
        ]
        SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path, read_only=True)
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[1][3] == 80, f"ProjA call sites: {rows[1][3]}"
        assert rows[2][3] == 120, f"ProjB call sites: {rows[2][3]}"
        assert rows[3][3] == 10, f"ProjC call sites: {rows[3][3]}"
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_summary_total_call_sites(self, tmp_path):
        """Summary TOTAL call sites = 210."""
        path = str(tmp_path / "merge.xlsx")
        project_data = [
            {'name': 'ProjA', 'files_scanned': 10,
             'rows': [_make_merge_row(i) for i in range(80)]},
            {'name': 'ProjB', 'files_scanned': 20,
             'rows': [_make_merge_row(i, symbol=f"EVP_b_{i:04d}",
                                      category="crypto_evp") for i in range(120)]},
            {'name': 'ProjC', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"BIO_c_{i:04d}",
                                      category="crypto_bio") for i in range(10)]},
        ]
        SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path, read_only=True)
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        total_row = rows[4]
        assert total_row[0] == "TOTAL"
        assert total_row[3] == 210, f"TOTAL call sites: {total_row[3]}"
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_per_project_sheet_counts(self, tmp_path):
        """ProjA=2 sheets, ProjB=3 sheets, ProjC=1 sheet."""
        path = str(tmp_path / "merge.xlsx")
        project_data = [
            {'name': 'ProjA', 'files_scanned': 10,
             'rows': [_make_merge_row(i) for i in range(80)]},
            {'name': 'ProjB', 'files_scanned': 20,
             'rows': [_make_merge_row(i, symbol=f"EVP_b_{i:04d}",
                                      category="crypto_evp") for i in range(120)]},
            {'name': 'ProjC', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"BIO_c_{i:04d}",
                                      category="crypto_bio") for i in range(10)]},
        ]
        SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        proj_a_sheets = [s for s in wb.sheetnames
                         if s == "ProjA" or s.startswith("ProjA (")]
        proj_b_sheets = [s for s in wb.sheetnames
                         if s == "ProjB" or s.startswith("ProjB (")]
        proj_c_sheets = [s for s in wb.sheetnames
                         if s == "ProjC" or s.startswith("ProjC (")]

        assert len(proj_a_sheets) == 2, \
            f"ProjA sheets: expected 2, got {proj_a_sheets}"
        assert len(proj_b_sheets) == 3, \
            f"ProjB sheets: expected 3, got {proj_b_sheets}"
        assert len(proj_c_sheets) == 1, \
            f"ProjC sheets: expected 1, got {proj_c_sheets}"
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_per_project_total_data_rows(self, tmp_path):
        """Total data rows across sheets for each project: [80, 120, 10]."""
        path = str(tmp_path / "merge.xlsx")
        project_data = [
            {'name': 'ProjA', 'files_scanned': 10,
             'rows': [_make_merge_row(i) for i in range(80)]},
            {'name': 'ProjB', 'files_scanned': 20,
             'rows': [_make_merge_row(i, symbol=f"EVP_b_{i:04d}",
                                      category="crypto_evp") for i in range(120)]},
            {'name': 'ProjC', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"BIO_c_{i:04d}",
                                      category="crypto_bio") for i in range(10)]},
        ]
        SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)

        a_rows = _count_data_rows_in_sheets(wb, "ProjA")
        b_rows = _count_data_rows_in_sheets(wb, "ProjB")
        c_rows = _count_data_rows_in_sheets(wb, "ProjC")

        assert a_rows == 80, f"ProjA data rows: {a_rows}"
        assert b_rows == 120, f"ProjB data rows: {b_rows}"
        assert c_rows == 10, f"ProjC data rows: {c_rows}"
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_symbol_summary_exists(self, tmp_path):
        """Symbol Summary sheet exists with correct unique symbol count."""
        path = str(tmp_path / "merge.xlsx")
        project_data = [
            {'name': 'ProjA', 'files_scanned': 10,
             'rows': [_make_merge_row(i) for i in range(80)]},
            {'name': 'ProjB', 'files_scanned': 20,
             'rows': [_make_merge_row(i, symbol=f"EVP_b_{i:04d}",
                                      category="crypto_evp") for i in range(120)]},
            {'name': 'ProjC', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"BIO_c_{i:04d}",
                                      category="crypto_bio") for i in range(10)]},
        ]
        SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        assert "Symbol Summary" in wb.sheetnames

        ws = wb["Symbol Summary"]
        sym_rows = list(ws.iter_rows(min_row=2, values_only=True))
        unique_syms = {r[0] for r in sym_rows if r[0] is not None}

        expected_unique = 80 + 120 + 10
        assert len(unique_syms) == expected_unique, \
            f"Symbol Summary unique syms: expected {expected_unique}, got {len(unique_syms)}"
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_symbol_summary_projects_column(self, tmp_path):
        """Symbol Summary 'Projects' column correctly lists which projects use each symbol."""
        path = str(tmp_path / "merge.xlsx")

        shared_rows_a = [_make_merge_row(i, symbol="SSL_connect",
                                         category="ssl_core") for i in range(5)]
        shared_rows_b = [_make_merge_row(i, symbol="SSL_connect",
                                         category="ssl_core") for i in range(3)]
        unique_a = [_make_merge_row(i + 5, symbol=f"SSL_only_a_{i:04d}")
                    for i in range(75)]
        unique_b = [_make_merge_row(i + 5, symbol=f"EVP_only_b_{i:04d}",
                                    category="crypto_evp") for i in range(117)]
        unique_c = [_make_merge_row(i, symbol=f"BIO_only_c_{i:04d}",
                                    category="crypto_bio") for i in range(10)]

        project_data = [
            {'name': 'ProjA', 'files_scanned': 10,
             'rows': shared_rows_a + unique_a},
            {'name': 'ProjB', 'files_scanned': 20,
             'rows': shared_rows_b + unique_b},
            {'name': 'ProjC', 'files_scanned': 5, 'rows': unique_c},
        ]
        SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        ws = wb["Symbol Summary"]
        sym_rows = list(ws.iter_rows(min_row=2, values_only=True))

        ssl_connect_row = None
        for r in sym_rows:
            if r[0] == "SSL_connect":
                ssl_connect_row = r
                break
        assert ssl_connect_row is not None, "SSL_connect not in Symbol Summary"
        assert ssl_connect_row[2] == 8, \
            f"SSL_connect calls: expected 8, got {ssl_connect_row[2]}"
        assert ssl_connect_row[5] == 2, \
            f"SSL_connect projects: expected 2, got {ssl_connect_row[5]}"
        project_list_str = ssl_connect_row[6]
        assert "ProjA" in project_list_str
        assert "ProjB" in project_list_str
        assert "ProjC" not in project_list_str

        a_only = [r for r in sym_rows if r[0] and r[0].startswith("SSL_only_a_")]
        for r in a_only:
            assert r[5] == 1, f"{r[0]} projects: {r[5]}"
            assert r[6] == "ProjA"

        c_only = [r for r in sym_rows if r[0] and r[0].startswith("BIO_only_c_")]
        for r in c_only:
            assert r[5] == 1, f"{r[0]} projects: {r[5]}"
            assert r[6] == "ProjC"
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_return_stats(self, tmp_path):
        """_merge_to_workbook returns correct stats dict."""
        path = str(tmp_path / "merge.xlsx")
        project_data = [
            {'name': 'ProjA', 'files_scanned': 10,
             'rows': [_make_merge_row(i) for i in range(80)]},
            {'name': 'ProjB', 'files_scanned': 20,
             'rows': [_make_merge_row(i, symbol=f"EVP_b_{i:04d}",
                                      category="crypto_evp") for i in range(120)]},
            {'name': 'ProjC', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"BIO_c_{i:04d}",
                                      category="crypto_bio") for i in range(10)]},
        ]
        stats = SourceMergeExporter()._merge_to_workbook(project_data, path)

        assert len(stats['sheets']) == 3
        assert stats['sheets'][0]['call_sites'] == 80
        assert stats['sheets'][1]['call_sites'] == 120
        assert stats['sheets'][2]['call_sites'] == 10
        assert stats['total_symbols'] == 80 + 120 + 10


class TestMergeFromJsonOverflow:
    """Task 3: merge_from_json with overflow produces correct aggregation."""

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_merge_from_json_summary(self, tmp_path):
        """Two JSON reports with >50 call sites each merge correctly."""
        report_a = _make_json_report(80, symbol_prefix="SSL",
                                     category="ssl_core", target="/tmp/a",
                                     files_scanned=15)
        report_b = _make_json_report(70, symbol_prefix="EVP",
                                     category="crypto_evp", target="/tmp/b",
                                     files_scanned=25)

        path_a = str(tmp_path / "proj_a.json")
        path_b = str(tmp_path / "proj_b.json")
        with open(path_a, 'w') as f:
            json.dump(report_a, f)
        with open(path_b, 'w') as f:
            json.dump(report_b, f)

        out_path = str(tmp_path / "merged.xlsx")
        stats = SourceMergeExporter().merge_from_json([path_a, path_b], out_path)

        assert stats['sheets'][0]['call_sites'] == 80
        assert stats['sheets'][1]['call_sites'] == 70
        assert stats['total_symbols'] == 80 + 70

        wb = load_workbook(out_path, read_only=True)
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[1][3] == 80
        assert rows[2][3] == 70
        total_row = rows[3]
        assert total_row[0] == "TOTAL"
        assert total_row[3] == 150
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_merge_from_json_per_project_sheets(self, tmp_path):
        """Per-project sheets overflow correctly from JSON input."""
        report_a = _make_json_report(80, symbol_prefix="SSL",
                                     category="ssl_core", target="/tmp/a")
        report_b = _make_json_report(70, symbol_prefix="EVP",
                                     category="crypto_evp", target="/tmp/b")

        path_a = str(tmp_path / "proj_a.json")
        path_b = str(tmp_path / "proj_b.json")
        with open(path_a, 'w') as f:
            json.dump(report_a, f)
        with open(path_b, 'w') as f:
            json.dump(report_b, f)

        out_path = str(tmp_path / "merged.xlsx")
        SourceMergeExporter().merge_from_json([path_a, path_b], out_path)

        wb = load_workbook(out_path)
        a_sheets = [s for s in wb.sheetnames
                    if s == "proj_a" or s.startswith("proj_a (")]
        b_sheets = [s for s in wb.sheetnames
                    if s == "proj_b" or s.startswith("proj_b (")]

        assert len(a_sheets) == 2, f"proj_a sheets: {a_sheets}"
        assert len(b_sheets) == 2, f"proj_b sheets: {b_sheets}"
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_merge_from_json_symbol_summary(self, tmp_path):
        """Symbol Summary from JSON merge has all unique symbols."""
        report_a = _make_json_report(80, symbol_prefix="SSL",
                                     category="ssl_core", target="/tmp/a")
        report_b = _make_json_report(70, symbol_prefix="EVP",
                                     category="crypto_evp", target="/tmp/b")

        path_a = str(tmp_path / "proj_a.json")
        path_b = str(tmp_path / "proj_b.json")
        with open(path_a, 'w') as f:
            json.dump(report_a, f)
        with open(path_b, 'w') as f:
            json.dump(report_b, f)

        out_path = str(tmp_path / "merged.xlsx")
        SourceMergeExporter().merge_from_json([path_a, path_b], out_path)

        wb = load_workbook(out_path)
        ws = wb["Symbol Summary"]
        sym_rows = list(ws.iter_rows(min_row=2, values_only=True))
        unique_syms = {r[0] for r in sym_rows if r[0] is not None}

        assert len(unique_syms) == 150, \
            f"Symbol Summary unique: expected 150, got {len(unique_syms)}"
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_merge_from_json_files_scanned(self, tmp_path):
        """merge_from_json preserves total_files_scanned (unlike merge from XLSX)."""
        report_a = _make_json_report(60, files_scanned=15, target="/tmp/a")
        report_b = _make_json_report(60, files_scanned=25, target="/tmp/b")

        path_a = str(tmp_path / "proj_a.json")
        path_b = str(tmp_path / "proj_b.json")
        with open(path_a, 'w') as f:
            json.dump(report_a, f)
        with open(path_b, 'w') as f:
            json.dump(report_b, f)

        out_path = str(tmp_path / "merged.xlsx")
        stats = SourceMergeExporter().merge_from_json([path_a, path_b], out_path)

        assert stats['sheets'][0]['files_scanned'] == 15
        assert stats['sheets'][1]['files_scanned'] == 25

        wb = load_workbook(out_path, read_only=True)
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[1][1] == 15
        assert rows[2][1] == 25
        total_row = rows[3]
        assert total_row[1] == 40
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_merge_from_json_data_row_totals(self, tmp_path):
        """All data rows survive across overflow sheets."""
        report_a = _make_json_report(80, symbol_prefix="SSL",
                                     category="ssl_core", target="/tmp/a")
        report_b = _make_json_report(70, symbol_prefix="EVP",
                                     category="crypto_evp", target="/tmp/b")

        path_a = str(tmp_path / "proj_a.json")
        path_b = str(tmp_path / "proj_b.json")
        with open(path_a, 'w') as f:
            json.dump(report_a, f)
        with open(path_b, 'w') as f:
            json.dump(report_b, f)

        out_path = str(tmp_path / "merged.xlsx")
        SourceMergeExporter().merge_from_json([path_a, path_b], out_path)

        wb = load_workbook(out_path)
        a_total = _count_data_rows_in_sheets(wb, "proj_a")
        b_total = _count_data_rows_in_sheets(wb, "proj_b")

        assert a_total == 80, f"proj_a total rows: {a_total}"
        assert b_total == 70, f"proj_b total rows: {b_total}"
        wb.close()


class TestExportReadXlsxMergeRoundTrip:
    """Task 4: Full pipeline export -> _read_xlsx -> merge with overflow."""

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_round_trip_data_integrity(self, tmp_path):
        """Export 2 projects with >50 sites each, merge, verify totals."""
        sites_a = [_make_call_site(i, symbol=f"SSL_a_{i:04d}")
                   for i in range(80)]
        sites_b = [_make_call_site(i, symbol=f"EVP_b_{i:04d}",
                                   category="crypto_evp")
                   for i in range(70)]

        path_a = str(tmp_path / "projA.xlsx")
        path_b = str(tmp_path / "projB.xlsx")
        SourceExcelExporter().export(_make_result(sites_a), path_a)
        SourceExcelExporter().export(_make_result(sites_b), path_b)

        wb_a = load_workbook(path_a, read_only=True)
        a_sheets = [s for s in wb_a.sheetnames
                    if s == "OpenSSL Call Sites"
                    or s.startswith("OpenSSL Call Sites (")]
        assert len(a_sheets) == 2, \
            f"projA overflow sheets: {a_sheets}"
        wb_a.close()

        merger = SourceMergeExporter()
        _, _, rows_a = merger._read_xlsx(path_a)
        _, _, rows_b = merger._read_xlsx(path_b)
        assert len(rows_a) == 80, f"_read_xlsx projA: {len(rows_a)}"
        assert len(rows_b) == 70, f"_read_xlsx projB: {len(rows_b)}"

        merged_path = str(tmp_path / "merged.xlsx")
        stats = merger.merge([path_a, path_b], merged_path)

        assert stats['sheets'][0]['call_sites'] == 80
        assert stats['sheets'][1]['call_sites'] == 70
        assert stats['total_symbols'] == 80 + 70

        wb = load_workbook(merged_path, read_only=True)
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        total_row = rows[-1]
        assert total_row[0] == "TOTAL"
        assert total_row[3] == 150

        ws_sym = wb["Symbol Summary"]
        sym_rows = list(ws_sym.iter_rows(min_row=2, values_only=True))
        unique = {r[0] for r in sym_rows if r[0] is not None}
        assert len(unique) == 150
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_round_trip_symbol_values(self, tmp_path):
        """Verify exact symbol names survive the round-trip."""
        sites = [_make_call_site(i, symbol=f"SSL_rt_{i:04d}") for i in range(60)]
        xlsx_path = str(tmp_path / "rt.xlsx")
        SourceExcelExporter().export(_make_result(sites), xlsx_path)

        merger = SourceMergeExporter()
        _, _, rows = merger._read_xlsx(xlsx_path)

        expected_syms = {f"SSL_rt_{i:04d}" for i in range(60)}
        actual_syms = {r[4] for r in rows}
        assert actual_syms == expected_syms

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_round_trip_merged_symbol_summary_projects(self, tmp_path):
        """Merged Symbol Summary correctly tracks projects in round-trip."""
        shared_sym = "SSL_connect"
        sites_a = ([_make_call_site(0, symbol=shared_sym)] +
                   [_make_call_site(i + 1, symbol=f"SSL_a_{i:04d}")
                    for i in range(59)])
        sites_b = ([_make_call_site(0, symbol=shared_sym)] +
                   [_make_call_site(i + 1, symbol=f"EVP_b_{i:04d}",
                                   category="crypto_evp")
                    for i in range(59)])

        path_a = str(tmp_path / "projA.xlsx")
        path_b = str(tmp_path / "projB.xlsx")
        SourceExcelExporter().export(_make_result(sites_a), path_a)
        SourceExcelExporter().export(_make_result(sites_b), path_b)

        merged_path = str(tmp_path / "merged.xlsx")
        SourceMergeExporter().merge([path_a, path_b], merged_path)

        wb = load_workbook(merged_path)
        ws = wb["Symbol Summary"]
        sym_rows = list(ws.iter_rows(min_row=2, values_only=True))

        ssl_connect_row = None
        for r in sym_rows:
            if r[0] == shared_sym:
                ssl_connect_row = r
                break
        assert ssl_connect_row is not None
        assert ssl_connect_row[5] == 2, \
            f"SSL_connect projects count: {ssl_connect_row[5]}"
        assert "projA" in ssl_connect_row[6]
        assert "projB" in ssl_connect_row[6]
        wb.close()


class TestBoundaryConditions:
    """Task 5: Edge cases at exact boundary (49 vs 50 rows at MAX_ROW=50)."""

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_49_rows_no_overflow(self, tmp_path):
        """49 data rows + 1 header = 50 total rows, no overflow."""
        path = str(tmp_path / "b49.xlsx")
        project_data = [
            {'name': 'Exact49', 'files_scanned': 5,
             'rows': [_make_merge_row(i) for i in range(49)]},
        ]
        stats = SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        sheets = [s for s in wb.sheetnames
                  if s == "Exact49" or s.startswith("Exact49 (")]
        assert len(sheets) == 1, f"Expected 1 sheet, got {sheets}"

        ws = wb["Exact49"]
        assert ws.max_row == 50

        assert stats['sheets'][0]['call_sites'] == 49

        ws_sym = wb["Symbol Summary"]
        sym_rows = list(ws_sym.iter_rows(min_row=2, values_only=True))
        unique = {r[0] for r in sym_rows if r[0] is not None}
        assert len(unique) == 49
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_50_rows_triggers_overflow(self, tmp_path):
        """50 data rows: first sheet holds 49 (header+49=50), overflow gets 1."""
        path = str(tmp_path / "b50.xlsx")
        project_data = [
            {'name': 'Overflow50', 'files_scanned': 5,
             'rows': [_make_merge_row(i) for i in range(50)]},
        ]
        stats = SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        sheets = [s for s in wb.sheetnames
                  if s == "Overflow50" or s.startswith("Overflow50 (")]
        assert len(sheets) == 2, f"Expected 2 sheets, got {sheets}"

        ws1 = wb["Overflow50"]
        assert ws1.max_row == 50

        overflow_name = [s for s in wb.sheetnames if s.startswith("Overflow50 (")][0]
        ws2 = wb[overflow_name]
        data_rows_ws2 = sum(1 for i in range(2, ws2.max_row + 1)
                            if ws2.cell(row=i, column=1).value is not None)
        assert data_rows_ws2 == 1, f"Overflow sheet data rows: {data_rows_ws2}"

        assert stats['sheets'][0]['call_sites'] == 50

        ws_sym = wb["Symbol Summary"]
        sym_rows = list(ws_sym.iter_rows(min_row=2, values_only=True))
        unique = {r[0] for r in sym_rows if r[0] is not None}
        assert len(unique) == 50
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_all_three_projects_at_49(self, tmp_path):
        """All 3 projects with 49 rows each: no overflow, correct totals."""
        path = str(tmp_path / "all49.xlsx")
        project_data = [
            {'name': f'Proj{c}', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"{c}_sym_{i:04d}")
                      for i in range(49)]}
            for c in ['A', 'B', 'C']
        ]
        stats = SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        for c in ['A', 'B', 'C']:
            name = f"Proj{c}"
            sheets = [s for s in wb.sheetnames
                      if s == name or s.startswith(f"{name} (")]
            assert len(sheets) == 1, f"{name}: expected 1 sheet, got {sheets}"

        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        total = rows[-1]
        assert total[3] == 147

        assert stats['total_symbols'] == 147
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_all_three_projects_at_50(self, tmp_path):
        """All 3 projects with 50 rows each: all overflow, correct totals."""
        path = str(tmp_path / "all50.xlsx")
        project_data = [
            {'name': f'Proj{c}', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"{c}_sym_{i:04d}")
                      for i in range(50)]}
            for c in ['A', 'B', 'C']
        ]
        stats = SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        for c in ['A', 'B', 'C']:
            name = f"Proj{c}"
            sheets = [s for s in wb.sheetnames
                      if s == name or s.startswith(f"{name} (")]
            assert len(sheets) == 2, f"{name}: expected 2 sheets, got {sheets}"

        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        total = rows[-1]
        assert total[3] == 150

        assert stats['total_symbols'] == 150

        ws_sym = wb["Symbol Summary"]
        sym_rows = list(ws_sym.iter_rows(min_row=2, values_only=True))
        unique = {r[0] for r in sym_rows if r[0] is not None}
        assert len(unique) == 150
        wb.close()

    @mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_summary_correct_with_mixed_overflow(self, tmp_path):
        """Mix: ProjA=49 (no overflow), ProjB=50 (overflow), ProjC=51 (overflow).
        Summary must be correct regardless of per-project overflow state."""
        path = str(tmp_path / "mixed.xlsx")
        project_data = [
            {'name': 'ProjA', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"A_sym_{i:04d}")
                      for i in range(49)]},
            {'name': 'ProjB', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"B_sym_{i:04d}")
                      for i in range(50)]},
            {'name': 'ProjC', 'files_scanned': 5,
             'rows': [_make_merge_row(i, symbol=f"C_sym_{i:04d}")
                      for i in range(51)]},
        ]
        stats = SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)

        a_sheets = [s for s in wb.sheetnames
                    if s == "ProjA" or s.startswith("ProjA (")]
        b_sheets = [s for s in wb.sheetnames
                    if s == "ProjB" or s.startswith("ProjB (")]
        c_sheets = [s for s in wb.sheetnames
                    if s == "ProjC" or s.startswith("ProjC (")]
        assert len(a_sheets) == 1
        assert len(b_sheets) == 2
        assert len(c_sheets) == 2

        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[1][3] == 49
        assert rows[2][3] == 50
        assert rows[3][3] == 51
        total = rows[4]
        assert total[3] == 150

        assert stats['total_symbols'] == 49 + 50 + 51

        ws_sym = wb["Symbol Summary"]
        sym_rows = list(ws_sym.iter_rows(min_row=2, values_only=True))
        unique = {r[0] for r in sym_rows if r[0] is not None}
        assert len(unique) == 150
        wb.close()
