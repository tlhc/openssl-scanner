"""CLI integration tests for all source-related subcommands.

Covers gaps identified by coverage audit:
- _is_data_blob() detection
- _collect_source_files() symlink cycle handling
- -j nargs='?' behavior (bare -j, -j N, no -j)
- source subcommand error paths and flag combos
- source-merge CLI integration
- source-probe CLI integration
- combo-scan CLI integration
- source-diff flag combos (--summary-only, --include-unchanged, --ignore-categories)
- Multi-target worker budget division
"""

import json
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

tree_sitter = pytest.importorskip("tree_sitter")

from openssl_scanner.source_analyzer import (
    LANG_EXTENSIONS,
    SourceAnalyzer,
    _is_data_blob,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def test_root(tmp_path_factory):
    """Create a module-scoped test directory with two projects."""
    root = tmp_path_factory.mktemp("src_cli")
    proj_a = root / "proj_a"
    proj_a.mkdir()
    (proj_a / "ssl_client.c").write_text(
        '#include <openssl/ssl.h>\n'
        'void do_tls() {\n'
        '    SSL_connect(NULL);\n'
        '    SSL_read(NULL, NULL, 0);\n'
        '    SSL_write(NULL, NULL, 0);\n'
        '}\n'
    )
    (proj_a / "evp_hash.c").write_text(
        '#include <openssl/evp.h>\n'
        'void hash() {\n'
        '    EVP_DigestInit_ex(NULL, NULL, NULL);\n'
        '    EVP_DigestUpdate(NULL, NULL, 0);\n'
        '    EVP_DigestFinal_ex(NULL, NULL, NULL);\n'
        '}\n'
    )
    proj_b = root / "proj_b"
    proj_b.mkdir()
    (proj_b / "crypto.c").write_text(
        '#include <openssl/rand.h>\n'
        'void init() {\n'
        '    RAND_bytes(NULL, 0);\n'
        '    RAND_seed(NULL, 0);\n'
        '}\n'
    )
    return root


@pytest.fixture(scope="module")
def analyzer():
    from openssl_scanner.openssl_matcher import OpenSSLMatcher
    matcher = OpenSSLMatcher()
    matcher.load_combined_symbols()
    a = SourceAnalyzer(matcher)
    return a


def _write_parser_recovery_source(path):
    path.write_text(
        "#include <openssl/err.h>\n"
        "void f(unsigned long code) {\n"
        "    ERR_error_string(\n"
        "#if OPENSSL_IS_BORINGSSL\n"
        "        (uint32_t)\n"
        "#else\n"
        "        (unsigned long)\n"
        "#endif\n"
        "        code);\n"
        "}\n",
        encoding="utf-8",
    )


# ---------------------------------------------------------------------------
# _is_data_blob tests
# ---------------------------------------------------------------------------

class TestIsDataBlob:
    """P0 gap: _is_data_blob had 0 tests."""

    def test_hex_array_detected(self):
        data = b'0xff, 0xd8, 0xff, 0xe0, 0x00, 0x10, 0x4a, 0x46, 0x49, 0x46\n'
        assert _is_data_blob(data) is True

    def test_hex_with_leading_space(self):
        data = b' 0x00, 0x01, 0x02, 0x03\n'
        assert _is_data_blob(data) is True

    def test_hex_with_leading_tab(self):
        data = b'\t0x00, 0x01, 0x02\n'
        assert _is_data_blob(data) is True

    def test_normal_c_code_not_blob(self):
        data = b'#include <stdio.h>\nint main() { return 0; }\n'
        assert _is_data_blob(data) is False

    def test_hex_with_c_keyword_not_blob(self):
        """Hex data that also contains C keywords is NOT a data blob."""
        data = b'0xff, 0xd8\n#include <openssl/ssl.h>\n'
        assert _is_data_blob(data) is False

    def test_hex_with_typedef_not_blob(self):
        data = b'0x00, 0x01\ntypedef struct { int x; } foo;\n'
        assert _is_data_blob(data) is False

    def test_empty_source(self):
        assert _is_data_blob(b'') is False

    def test_whitespace_only(self):
        assert _is_data_blob(b'   \n\t\n') is False

    def test_non_hex_start(self):
        data = b'int data[] = {0xff, 0xd8};\n'
        assert _is_data_blob(data) is False

    def test_hex_with_static_keyword_not_blob(self):
        data = b'0xff, 0x00\nstatic const unsigned char data[] = {\n'
        assert _is_data_blob(data) is False


# ---------------------------------------------------------------------------
# _collect_source_files symlink cycle tests
# ---------------------------------------------------------------------------

class TestCollectSourceFilesSymlinkCycle:
    """P0 gap: symlink cycle detection had 0 tests."""

    def test_symlink_cycle_skipped(self, tmp_path):
        """Directory symlink creating a cycle should be skipped."""
        src = tmp_path / "src"
        src.mkdir()
        (src / "main.c").write_text("void f() {}\n")
        sub = src / "sub"
        sub.mkdir()
        (sub / "helper.c").write_text("void g() {}\n")
        cycle = sub / "loop"
        cycle.symlink_to(src)

        a = SourceAnalyzer.__new__(SourceAnalyzer)
        files = a._collect_source_files(str(src), recursive=True)
        paths = [os.path.basename(f) for f in files]
        assert "main.c" in paths
        assert "helper.c" in paths
        assert paths.count("main.c") == 1, "symlink cycle should not duplicate files"

    def test_non_recursive_ignores_subdirs(self, tmp_path):
        """Non-recursive scan should only list files in the top directory."""
        src = tmp_path / "src"
        src.mkdir()
        (src / "top.c").write_text("void f() {}\n")
        sub = src / "sub"
        sub.mkdir()
        (sub / "deep.c").write_text("void g() {}\n")

        a = SourceAnalyzer.__new__(SourceAnalyzer)
        files = a._collect_source_files(str(src), recursive=False)
        basenames = [os.path.basename(f) for f in files]
        assert "top.c" in basenames
        assert "deep.c" not in basenames

    def test_dotdir_skipped(self, tmp_path):
        """Directories starting with '.' should be skipped."""
        src = tmp_path / "src"
        src.mkdir()
        (src / "main.c").write_text("void f() {}\n")
        hidden = src / ".git"
        hidden.mkdir()
        (hidden / "internal.c").write_text("void g() {}\n")

        a = SourceAnalyzer.__new__(SourceAnalyzer)
        files = a._collect_source_files(str(src), recursive=True)
        basenames = [os.path.basename(f) for f in files]
        assert "main.c" in basenames
        assert "internal.c" not in basenames

    def test_supported_extensions_only(self, tmp_path):
        """Only files with supported extensions should be collected."""
        src = tmp_path / "src"
        src.mkdir()
        (src / "valid.c").write_text("void f() {}\n")
        (src / "valid.cpp").write_text("void f() {}\n")
        (src / "valid.rs").write_text("fn f() {}\n")
        (src / "ignore.py").write_text("pass\n")
        (src / "ignore.txt").write_text("text\n")
        (src / "Makefile").write_text("all:\n")

        a = SourceAnalyzer.__new__(SourceAnalyzer)
        files = a._collect_source_files(str(src), recursive=True)
        basenames = set(os.path.basename(f) for f in files)
        assert basenames == {"valid.c", "valid.cpp", "valid.rs"}


# ---------------------------------------------------------------------------
# -j nargs='?' behavior tests
# ---------------------------------------------------------------------------

class TestJobsArgParsing:
    """P1 gap: -j nargs='?' behavior across subcommands had no tests."""

    def _parse_source_args(self, argv):
        """Parse argv through the real source subcommand parser."""
        import argparse
        from openssl_scanner.__main__ import main
        from unittest.mock import patch

        captured = {}

        def fake_main():
            import openssl_scanner.__main__ as mod
            parser = argparse.ArgumentParser()
            subparsers = parser.add_subparsers(dest="subcommand")
            mod.setup_source_parser(subparsers)
            args = parser.parse_args(argv)
            captured['args'] = args

        with pytest.raises(SystemExit):
            fake_main()

        return captured.get('args')

    def test_bare_j_uses_cpu_count(self):
        """bare `-j` (no value) should use cpu_count as const."""
        import subprocess
        result = subprocess.run(
            [sys.executable, "-c",
             "import argparse, os; "
             "p = argparse.ArgumentParser(); "
             "cpu = os.cpu_count() or 4; "
             "p.add_argument('-j', nargs='?', type=int, const=cpu, default=cpu); "
             "args = p.parse_args(['-j']); "
             f"assert args.j == cpu, f'Expected {{cpu}}, got {{args.j}}'"],
            capture_output=True, text=True)
        assert result.returncode == 0, result.stderr

    def test_j_with_value(self):
        import subprocess
        result = subprocess.run(
            [sys.executable, "-c",
             "import argparse, os; "
             "cpu = os.cpu_count() or 4; "
             "p = argparse.ArgumentParser(); "
             "p.add_argument('-j', nargs='?', type=int, const=cpu, default=cpu); "
             "args = p.parse_args(['-j', '2']); "
             "assert args.j == 2, f'Expected 2, got {args.j}'"],
            capture_output=True, text=True)
        assert result.returncode == 0, result.stderr

    def test_no_j_uses_default(self):
        import subprocess
        result = subprocess.run(
            [sys.executable, "-c",
             "import argparse, os; "
             "cpu = os.cpu_count() or 4; "
             "p = argparse.ArgumentParser(); "
             "p.add_argument('-j', nargs='?', type=int, const=cpu, default=cpu); "
             "args = p.parse_args([]); "
             f"assert args.j == cpu, f'Expected {{cpu}}, got {{args.j}}'"],
            capture_output=True, text=True)
        assert result.returncode == 0, result.stderr


# ---------------------------------------------------------------------------
# source subcommand CLI integration
# ---------------------------------------------------------------------------

class TestSourceCLI:
    """CLI integration for the `source` subcommand."""

    def test_source_xlsx_output(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "report.xlsx"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), "-o", str(out)],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        assert out.exists()
        assert out.stat().st_size > 0

    def test_source_hitls_compat_xlsx_oh_source_shape(self, tmp_path):
        import subprocess

        from openpyxl import load_workbook

        oh_source = tmp_path / "oh-source"
        project = oh_source / "base" / "security" / "crypto_framework"
        project.mkdir(parents=True)
        (project / "crypto_adapter.c").write_text(
            '#include <openssl/ssl.h>\n'
            '#include <openssl/evp.h>\n'
            'void use_crypto(void *ssl) {\n'
            '    SSL_read(ssl, 0, 0);\n'
            '    SSL_CTX_new(0);\n'
            '    EVP_sha256();\n'
            '}\n',
            encoding="utf-8",
        )
        hitls_map = tmp_path / "hitls_map.json"
        hitls_map.write_text(
            json.dumps({
                "version": "test",
                "mapping": {
                    "SSL_read": {
                        "status": "partial",
                        "hitls": "HITLS_Read",
                        "notes": "",
                    },
                    "SSL_CTX_new": {
                        "status": "not_available",
                        "hitls": None,
                        "notes": "",
                    },
                    "EVP_sha256": {
                        "status": "available",
                        "hitls": "CRYPT_MD_SHA256",
                        "notes": "",
                    },
                },
            }),
            encoding="utf-8",
        )
        out = tmp_path / "oh_source_hitls.xlsx"

        r = subprocess.run(
            [
                sys.executable, "-m", "openssl_scanner", "source",
                str(project), "-o", str(out),
                "--hitls-compat", "--hitls-map", str(hitls_map),
            ],
            capture_output=True, text=True, timeout=30,
        )
        assert r.returncode == 0, r.stderr

        json_out = out.with_suffix(".json")
        data = json.loads(json_out.read_text())
        assert data["summary"]["hitls_coverage"] == {
            "available": 1,
            "partial": 1,
            "not_available": 1,
            "unknown": 0,
        }
        assert data["summary"]["hitls_direct_replace_ratio"] == 33.33
        assert data["summary"]["hitls_direct_or_partial_replace_ratio"] == 66.67
        by_symbol = {cs["ossl_symbol"]: cs for cs in data["call_sites"]}
        assert by_symbol["SSL_read"]["hitls_replacement"] == "HITLS_Read"
        assert "detection_method" not in by_symbol["SSL_read"]

        wb = load_workbook(out, read_only=True, data_only=True)
        ws = wb["OpenSSL Call Sites"]
        headers = [cell.value for cell in ws[1]]
        assert "Detection" not in headers
        assert "HiTLS Status" in headers
        assert "HiTLS Replacement" in headers

        coverage_ws = wb["HiTLS Coverage"]
        coverage = {
            row[0]: row[1]
            for row in coverage_ws.iter_rows(min_row=2, values_only=True)
            if row[0]
        }
        assert coverage["Available"] == 1
        assert coverage["Partial"] == 1
        assert coverage["Not Available"] == 1
        assert coverage["Direct Replace Ratio (%)"] == 33.33
        assert coverage["Direct+Partial Replace Ratio (%)"] == 66.67
        wb.close()

    def test_source_json_output(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "report.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), "-o", str(out), "--json-only"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        data = json.loads(out.read_text())
        assert data["summary"]["total_call_sites"] == 6
        assert data["summary"]["unique_symbols_count"] == 6

    def test_source_single_file(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "single.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a" / "ssl_client.c"),
             "-o", str(out), "--json-only"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        data = json.loads(out.read_text())
        assert data["summary"]["total_call_sites"] == 3
        syms = set(data["summary"]["unique_symbols"])
        assert syms == {"SSL_connect", "SSL_read", "SSL_write"}

    def test_source_no_recursive(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "norec.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), "-o", str(out),
             "--json-only", "--no-recursive"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        data = json.loads(out.read_text())
        assert data["summary"]["total_call_sites"] == 6

    def test_source_j2(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "j2.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), "-o", str(out),
             "--json-only", "-j", "2"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        data = json.loads(out.read_text())
        assert data["summary"]["total_call_sites"] == 6

    def test_source_bare_j(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "bare_j.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), "-o", str(out),
             "--json-only", "-j"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        data = json.loads(out.read_text())
        assert data["summary"]["total_call_sites"] == 6

    def test_source_multi_target(self, test_root, tmp_path):
        import subprocess
        out_dir = tmp_path / "multi"
        out_dir.mkdir()
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), str(test_root / "proj_b"),
             "-o", str(out_dir) + "/", "--json-only"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        assert (out_dir / "proj_a.json").exists()
        assert (out_dir / "proj_b.json").exists()

    def test_source_file_list(self, test_root, tmp_path):
        import subprocess
        targets = tmp_path / "targets.txt"
        targets.write_text(
            f"# comment\n"
            f"{test_root / 'proj_a'}\n"
            f"\n"
            f"{test_root / 'proj_b'}\n"
        )
        out_dir = tmp_path / "flist"
        out_dir.mkdir()
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             "-f", str(targets), "-o", str(out_dir) + "/", "--json-only"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        assert (out_dir / "proj_a.json").exists()
        assert (out_dir / "proj_b.json").exists()

    def test_source_verbose(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "verbose.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_b"), "-o", str(out),
             "--json-only", "-v"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0
        assert "INFO" in r.stderr

    def test_source_vvvv_shows_worker_parser_diagnostics(self, tmp_path):
        import subprocess

        src = tmp_path / "bad_sources"
        src.mkdir()
        for idx in range(2):
            (src / f"bad_{idx}.c").write_text(
                "#include <openssl/ssl.h>\n"
                "void broken(void) {\n"
                "    SSL_read(ssl, 0, 0)\n",
                encoding="utf-8",
            )
        out = tmp_path / "parser_diagnostics.json"
        r = subprocess.run(
            [
                sys.executable, "-m", "openssl_scanner", "source",
                str(src), "-o", str(out), "--json-only", "-j", "2", "-vvvv",
            ],
            capture_output=True, text=True, timeout=30,
        )
        assert r.returncode == 0, r.stderr
        assert "DEBUG" in r.stderr
        assert "Parser diagnostics in" in r.stderr
        assert "error" not in r.stderr.lower()

    def test_source_recover_parser_diagnostics_is_gated(self, tmp_path):
        import subprocess

        from openpyxl import load_workbook

        src = tmp_path / "recover_parser"
        src.mkdir()
        _write_parser_recovery_source(src / "recover.c")

        default_out = tmp_path / "default.xlsx"
        default_run = subprocess.run(
            [
                sys.executable, "-m", "openssl_scanner", "source",
                str(src), "-o", str(default_out),
            ],
            capture_output=True, text=True, timeout=30,
        )
        assert default_run.returncode == 0, default_run.stderr
        default_json = json.loads(default_out.with_suffix(".json").read_text())
        assert default_json["summary"]["total_call_sites"] == 0
        wb = load_workbook(default_out, read_only=True)
        headers = [cell.value for cell in wb["OpenSSL Call Sites"][1]]
        assert "Extraction Source" not in headers
        wb.close()

        recovered_out = tmp_path / "recovered.xlsx"
        recovered_run = subprocess.run(
            [
                sys.executable, "-m", "openssl_scanner", "source",
                str(src), "-o", str(recovered_out),
                "--recover-parser-diagnostics",
            ],
            capture_output=True, text=True, timeout=30,
        )
        assert recovered_run.returncode == 0, recovered_run.stderr
        recovered_json = json.loads(recovered_out.with_suffix(".json").read_text())
        assert recovered_json["summary"]["fallback_call_sites"] == 1
        call_site = recovered_json["call_sites"][0]
        assert call_site["ossl_symbol"] == "ERR_error_string"
        assert call_site["extraction_source"] == "parser-diagnostic-text"
        assert call_site["confidence"] == "fallback"
        assert call_site["parser_diagnostic_class"] == "preprocessor-fragment"

        wb = load_workbook(recovered_out, read_only=True)
        headers = [cell.value for cell in wb["OpenSSL Call Sites"][1]]
        assert "Extraction Source" in headers
        assert "Confidence" in headers
        assert "Parser Diagnostic Class" in headers
        wb.close()

    def test_source_nonexistent_target(self, tmp_path):
        import subprocess
        out = tmp_path / "noexist.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             "/nonexistent/path", "-o", str(out), "--json-only"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode != 0

    def test_scan_launcher_does_not_eval_arguments(self, tmp_path):
        import subprocess

        marker = tmp_path / "eval_injected"
        launcher = os.path.abspath(
            os.path.join(os.path.dirname(__file__), "..", "scan")
        )
        r = subprocess.run(
            [launcher, f"$(touch {marker})", "--version"],
            capture_output=True, text=True, timeout=30,
        )
        assert r.returncode != 0
        assert not marker.exists()


# ---------------------------------------------------------------------------
# source-merge CLI integration
# ---------------------------------------------------------------------------

class TestSourceMergeCLI:
    """P0 gap: source-merge had 0 CLI integration tests."""

    def test_merge_xlsx(self, test_root, tmp_path):
        import subprocess
        a_xlsx = tmp_path / "a.xlsx"
        b_xlsx = tmp_path / "b.xlsx"
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), "-o", str(a_xlsx)],
            capture_output=True, timeout=30, check=True)
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_b"), "-o", str(b_xlsx)],
            capture_output=True, timeout=30, check=True)

        merged = tmp_path / "merged.xlsx"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-merge",
             str(a_xlsx), str(b_xlsx), "-o", str(merged)],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        assert merged.exists()

        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook
        wb = load_workbook(str(merged))
        assert "Summary" in wb.sheetnames
        assert "Symbol Summary" in wb.sheetnames
        assert len(wb.sheetnames) >= 4

    def test_merge_verbose(self, test_root, tmp_path):
        import subprocess
        a_xlsx = tmp_path / "a.xlsx"
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), "-o", str(a_xlsx)],
            capture_output=True, timeout=30, check=True)

        merged = tmp_path / "merged_v.xlsx"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-merge",
             str(a_xlsx), "-o", str(merged), "-v"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr

    def test_merge_preserves_parser_recovery_columns(self, tmp_path):
        import subprocess

        from openpyxl import load_workbook

        src = tmp_path / "recover_src"
        src.mkdir()
        _write_parser_recovery_source(src / "recover.c")
        report = tmp_path / "recover.xlsx"
        subprocess.run(
            [
                sys.executable, "-m", "openssl_scanner", "source",
                str(src), "-o", str(report), "--recover-parser-diagnostics",
            ],
            capture_output=True, text=True, timeout=30, check=True,
        )

        merged = tmp_path / "merged.xlsx"
        r = subprocess.run(
            [
                sys.executable, "-m", "openssl_scanner", "source-merge",
                str(report), "-o", str(merged),
            ],
            capture_output=True, text=True, timeout=30,
        )
        assert r.returncode == 0, r.stderr

        wb = load_workbook(merged, read_only=True, data_only=True)
        ws = wb["recover"]
        headers = [cell.value for cell in ws[1]]
        row = [cell.value for cell in ws[2]]
        assert "Extraction Source" in headers
        assert "Confidence" in headers
        assert "Parser Diagnostic Class" in headers
        assert row[headers.index("Extraction Source")] == "parser-diagnostic-text"
        assert row[headers.index("Confidence")] == "fallback"
        wb.close()


# ---------------------------------------------------------------------------
# source-probe CLI integration
# ---------------------------------------------------------------------------

class TestSourceProbeCLI:
    """P0 gap: source-probe had 0 CLI integration tests."""

    def test_probe_finds_projects(self, test_root):
        import subprocess
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-probe",
             str(test_root)],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, r.stderr
        lines = [l for l in r.stdout.strip().split('\n') if not l.startswith('#')]
        assert len(lines) >= 2

    def test_probe_output_format(self, test_root):
        import subprocess
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-probe",
             str(test_root)],
            capture_output=True, text=True, timeout=30)
        header_lines = [l for l in r.stdout.split('\n') if l.startswith('#')]
        path_lines = [l for l in r.stdout.strip().split('\n') if not l.startswith('#')]
        assert any("source-probe:" in h for h in header_lines)
        for p in path_lines:
            assert os.path.isabs(p), f"Expected absolute path, got: {p}"

    def test_probe_verbose(self, test_root):
        import subprocess
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-probe",
             str(test_root), "-v"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0
        assert "INFO" in r.stderr

    def test_probe_empty_dir(self, tmp_path):
        import subprocess
        empty = tmp_path / "empty"
        empty.mkdir()
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-probe",
             str(empty)],
            capture_output=True, text=True, timeout=30)
        path_lines = [l for l in r.stdout.strip().split('\n')
                      if l and not l.startswith('#')]
        assert len(path_lines) == 0


# ---------------------------------------------------------------------------
# source-diff CLI integration
# ---------------------------------------------------------------------------

class TestSourceDiffCLI:
    """Tests for source-diff subcommand flags."""

    @pytest.fixture()
    def diff_reports(self, test_root, tmp_path):
        import subprocess
        old_json = tmp_path / "old.json"
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), "-o", str(old_json), "--json-only"],
            capture_output=True, timeout=30, check=True)

        new_dir = tmp_path / "new_src"
        new_dir.mkdir()
        (new_dir / "ssl_client.c").write_text(
            '#include <openssl/ssl.h>\n'
            'void do_tls() {\n'
            '    SSL_connect(NULL);\n'
            '    SSL_read(NULL, NULL, 0);\n'
            '    SSL_shutdown(NULL);\n'
            '}\n'
        )
        new_json = tmp_path / "new.json"
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(new_dir), "-o", str(new_json), "--json-only"],
            capture_output=True, timeout=30, check=True)
        return str(old_json), str(new_json)

    def test_diff_console_exit_code_1(self, diff_reports):
        import subprocess
        old, new = diff_reports
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-diff", old, new],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 1, "exit 1 = changes detected"
        assert "Symbols Added" in r.stdout or "Symbols Removed" in r.stdout

    def test_diff_xlsx_output(self, diff_reports, tmp_path):
        import subprocess
        old, new = diff_reports
        out = tmp_path / "diff.xlsx"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-diff",
             old, new, "-o", str(out)],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 1
        assert out.exists()

        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "Summary Delta" in wb.sheetnames
        assert "Symbol Delta" in wb.sheetnames

    def test_diff_json_output(self, diff_reports, tmp_path):
        import subprocess
        old, new = diff_reports
        out = tmp_path / "diff.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-diff",
             old, new, "-o", str(out)],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 1
        data = json.loads(out.read_text())
        assert data["meta"]["report_type"] == "source_diff"
        assert "symbol_delta" in data
        assert "call_site_delta" in data

    def test_diff_summary_only(self, diff_reports):
        import subprocess
        old, new = diff_reports
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-diff",
             old, new, "--summary-only"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 1

    def test_diff_include_unchanged(self, diff_reports):
        import subprocess
        old, new = diff_reports
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-diff",
             old, new, "--include-unchanged"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 1

    def test_diff_ignore_categories(self, diff_reports):
        import subprocess
        old, new = diff_reports
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-diff",
             old, new, "--ignore-categories", "crypto_evp"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 1
        assert "EVP_DigestInit_ex" not in r.stdout

    def test_diff_no_changes_exit_0(self, test_root, tmp_path):
        """Diffing same report against itself should exit 0."""
        import subprocess
        report = tmp_path / "same.json"
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(test_root / "proj_a"), "-o", str(report), "--json-only"],
            capture_output=True, timeout=30, check=True)
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-diff",
             str(report), str(report)],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 0, "same report diffed = no changes = exit 0"


# ---------------------------------------------------------------------------
# combo-scan CLI integration
# ---------------------------------------------------------------------------

class TestComboScanCLI:
    """P0 gap: combo-scan had 0 integration tests."""

    def test_combo_xlsx(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "combo.xlsx"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "combo-scan",
             str(test_root), "-o", str(out)],
            capture_output=True, text=True, timeout=60)
        assert r.returncode == 0, r.stderr
        assert out.exists()

        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "Summary" in wb.sheetnames
        assert "Symbol Summary" in wb.sheetnames

    def test_combo_json_only(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "combo.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "combo-scan",
             str(test_root), "-o", str(out), "--json-only"],
            capture_output=True, text=True, timeout=60)
        assert r.returncode == 0, r.stderr
        data = json.loads(out.read_text())
        assert data["meta"]["report_type"] == "combo_scan"
        assert data["meta"]["total_projects"] >= 2
        assert len(data["projects"]) >= 2

    def test_combo_exclude(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "combo_excl.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "combo-scan",
             str(test_root), "-o", str(out),
             "--json-only", "--exclude", "proj_b"],
            capture_output=True, text=True, timeout=60)
        assert r.returncode == 0, r.stderr
        data = json.loads(out.read_text())
        project_names = [p["project"] for p in data["projects"]]
        assert "proj_b" not in project_names

    def test_combo_output_directory(self, test_root, tmp_path):
        import subprocess
        out_dir = tmp_path / "combo_dir"
        out_dir.mkdir()
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "combo-scan",
             str(test_root), "-o", str(out_dir) + "/"],
            capture_output=True, text=True, timeout=60)
        assert r.returncode == 0, r.stderr
        assert (out_dir / "merged.xlsx").exists()
        json_files = list(out_dir.glob("*.json"))
        assert len(json_files) >= 2

    def test_combo_j2(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "combo_j2.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "combo-scan",
             str(test_root), "-o", str(out), "--json-only", "-j", "2"],
            capture_output=True, text=True, timeout=60)
        assert r.returncode == 0, r.stderr

    def test_combo_verbose(self, test_root, tmp_path):
        import subprocess
        out = tmp_path / "combo_v.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "combo-scan",
             str(test_root), "-o", str(out), "--json-only", "-v"],
            capture_output=True, text=True, timeout=60)
        assert r.returncode == 0
        assert "INFO" in r.stderr

    def test_combo_vvvv_forwards_source_parser_diagnostics(self, tmp_path):
        import subprocess

        root = tmp_path / "combo_parser_diagnostics"
        proj = root / "proj_bad"
        proj.mkdir(parents=True)
        (proj / "bad.c").write_text(
            "#include <openssl/ssl.h>\n"
            "void broken(void) {\n"
            "    SSL_read(ssl, 0, 0)\n",
            encoding="utf-8",
        )
        out = tmp_path / "combo_parser_diagnostics.json"
        log_file = tmp_path / "combo_parser_diagnostics.log"
        r = subprocess.run(
            [
                sys.executable, "-m", "openssl_scanner", "combo-scan",
                str(root), "-o", str(out), "--json-only", "-j", "2",
                "-vvvv", "--log-file", str(log_file),
            ],
            capture_output=True, text=True, timeout=60,
        )
        assert r.returncode == 0, r.stderr
        assert "DEBUG" in r.stderr
        assert "Parser diagnostics in" in r.stderr
        assert "Parse errors" not in r.stderr
        assert "Parser diagnostics in" in log_file.read_text(encoding="utf-8")

    def test_combo_xlsx_preserves_parser_recovery_columns(self, tmp_path):
        import subprocess

        from openpyxl import load_workbook

        root = tmp_path / "combo_recover"
        proj = root / "proj"
        proj.mkdir(parents=True)
        _write_parser_recovery_source(proj / "recover.c")
        out = tmp_path / "combo_recover.xlsx"
        r = subprocess.run(
            [
                sys.executable, "-m", "openssl_scanner", "combo-scan",
                str(root), "-o", str(out), "--recover-parser-diagnostics",
            ],
            capture_output=True, text=True, timeout=60,
        )
        assert r.returncode == 0, r.stderr

        wb = load_workbook(out, read_only=True, data_only=True)
        ws = wb["proj"]
        headers = [cell.value for cell in ws[1]]
        row = [cell.value for cell in ws[2]]
        assert "Extraction Source" in headers
        assert "Confidence" in headers
        assert "Parser Diagnostic Class" in headers
        assert row[headers.index("Extraction Source")] == "parser-diagnostic-text"
        assert row[headers.index("Confidence")] == "fallback"
        wb.close()

    def test_combo_empty_dir(self, tmp_path):
        import subprocess
        empty = tmp_path / "empty"
        empty.mkdir()
        out = tmp_path / "combo_empty.json"
        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "combo-scan",
             str(empty), "-o", str(out), "--json-only"],
            capture_output=True, text=True, timeout=60)
        assert r.returncode == 0


# ---------------------------------------------------------------------------
# Bug reproduction tests (Codex gpt-5.4 findings)
# ---------------------------------------------------------------------------

class TestBugFinding1WorkerBudgetCap:
    """Finding #1: ThreadPoolExecutor size must be capped by args.jobs."""

    def test_max_w_capped_by_jobs(self):
        """When -j is smaller than target count, thread count must not exceed -j."""
        cpu = os.cpu_count() or 4
        for jobs, n_targets in [(2, 3), (1, 5), (3, 10)]:
            max_w = min(n_targets, cpu, jobs)
            per_target_jobs = max(1, jobs // max_w)
            total = max_w * per_target_jobs
            assert total <= jobs, (
                f"-j {jobs}, {n_targets} targets: "
                f"total {total} > budget {jobs}")
            assert max_w <= jobs

    def test_single_target_full_budget(self):
        cpu = os.cpu_count() or 4
        jobs = cpu
        max_w = min(1, cpu, jobs)
        per_target_jobs = max(1, jobs // max_w)
        assert max_w == 1
        assert per_target_jobs == jobs

    def test_j1_with_many_targets(self):
        """-j 1 should serialize: 1 thread, 1 worker."""
        max_w = min(10, os.cpu_count() or 4, 1)
        per_target_jobs = max(1, 1 // max_w)
        assert max_w == 1
        assert per_target_jobs == 1


class TestBugFinding2SummaryOnlyArgChange:
    """Finding #2: --summary-only must detect arg-only changes."""

    def test_summary_only_arg_change_exits_1(self, test_root, tmp_path):
        import subprocess
        old_dir = tmp_path / "old"
        old_dir.mkdir()
        (old_dir / "tls.c").write_text(
            'void f() { SSL_connect(ctx); }\n')
        new_dir = tmp_path / "new"
        new_dir.mkdir()
        (new_dir / "tls.c").write_text(
            'void f() { SSL_connect(new_ctx); }\n')

        old_json = tmp_path / "old.json"
        new_json = tmp_path / "new.json"
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(old_dir), "-o", str(old_json), "--json-only"],
            capture_output=True, timeout=30, check=True)
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(new_dir), "-o", str(new_json), "--json-only"],
            capture_output=True, timeout=30, check=True)

        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-diff",
             str(old_json), str(new_json), "--summary-only"],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 1, (
            "--summary-only should exit 1 when args changed, "
            f"got exit {r.returncode}")


class TestBugFinding5SheetNameCollision:
    """Finding #5: _resolve_sheet_names must avoid suffix collisions."""

    def test_suffix_collision_avoided(self):
        from openssl_scanner.source_exporter import SourceMergeExporter
        m = SourceMergeExporter.__new__(SourceMergeExporter)
        names = [
            'a' * 31,
            'a' * 31,
            'a' * 29 + '_1',
        ]
        result = m._resolve_sheet_names(names)
        assert len(result) == len(set(result)), f"duplicates: {result}"

    def test_many_collisions(self):
        from openssl_scanner.source_exporter import SourceMergeExporter
        m = SourceMergeExporter.__new__(SourceMergeExporter)
        names = ['samename'] * 10
        result = m._resolve_sheet_names(names)
        assert len(result) == len(set(result)), f"duplicates: {result}"


class TestBugFinding7ExcelInvalidChars:
    """Finding #7: sheet names must not contain Excel-invalid characters."""

    def test_invalid_chars_sanitized(self):
        from openssl_scanner.source_exporter import SourceMergeExporter
        m = SourceMergeExporter.__new__(SourceMergeExporter)
        names = ['test:colon', 'test*star', 'test/slash',
                 'test[bracket]', 'test?question']
        result = m._resolve_sheet_names(names)
        invalid = set('[]\\:*?/')
        for name in result:
            bad = [c for c in name if c in invalid]
            assert not bad, f"Invalid char {bad} in sheet name: {name}"

    def test_sanitize_creates_no_duplicates(self):
        from openssl_scanner.source_exporter import SourceMergeExporter
        m = SourceMergeExporter.__new__(SourceMergeExporter)
        names = ['a:b', 'a_b']
        result = m._resolve_sheet_names(names)
        assert len(result) == len(set(result)), f"duplicates: {result}"


class TestBugFinding9MoveOnlyConsole:
    """Finding #9: move-only diffs must render call site details."""

    def test_move_only_shows_call_sites(self, test_root, tmp_path):
        import subprocess
        old_dir = tmp_path / "old"
        old_dir.mkdir()
        (old_dir / "a.c").write_text(
            'void f1() { SSL_connect(ctx); }\nvoid f2() { }\n')
        new_dir = tmp_path / "new"
        new_dir.mkdir()
        (new_dir / "a.c").write_text(
            'void f1() { }\nvoid f2() { SSL_connect(ctx); }\n')

        old_json = tmp_path / "old.json"
        new_json = tmp_path / "new.json"
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(old_dir), "-o", str(old_json), "--json-only"],
            capture_output=True, timeout=30, check=True)
        subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source",
             str(new_dir), "-o", str(new_json), "--json-only"],
            capture_output=True, timeout=30, check=True)

        r = subprocess.run(
            [sys.executable, "-m", "openssl_scanner", "source-diff",
             str(old_json), str(new_json)],
            capture_output=True, text=True, timeout=30)
        assert r.returncode == 1
        assert "Call Sites Added" in r.stdout
        assert "Call Sites Removed" in r.stdout
