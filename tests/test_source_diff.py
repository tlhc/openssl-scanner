"""Tests for source scan result diff engine."""

import json
import os
import tempfile

import pytest

from openssl_scanner.source_diff import (
    DiffStatus, MetricDelta, ArgDelta, CallSiteDelta, SymbolDelta,
    FileDelta, ProjectDelta, DiffResult, load_report, diff_single,
    diff_combo, SourceDiffJsonExporter, SourceDiffExcelExporter,
    format_console, DIFF_COLORS,
    _match_call_sites_within_group, _derive_group_status,
)


def _make_call_site(file_path="src/tls.c", caller="init_ssl", symbol="SSL_CTX_new",
                    category="ssl_core", line=10, column=4, args="(NULL)",
                    language="c"):
    return {
        "file_path": file_path,
        "file_name": os.path.basename(file_path),
        "caller_function": caller,
        "line_number": line,
        "column": column,
        "ossl_symbol": symbol,
        "category": category,
        "call_args": args,
        "language": language,
        "detection_method": "dynamic-link",
    }


def _make_source_scan_json(target="/tmp/proj", call_sites=None):
    if call_sites is None:
        call_sites = []
    unique = sorted(set(cs["ossl_symbol"] for cs in call_sites))
    cats: dict = {}
    for cs in call_sites:
        cats.setdefault(cs["category"], [])
        if cs["ossl_symbol"] not in cats[cs["category"]]:
            cats[cs["category"]].append(cs["ossl_symbol"])
    return {
        "meta": {
            "tool_version": "1.0.0",
            "report_type": "source_scan",
            "scan_time": "2026-03-01T00:00:00",
            "target": target,
        },
        "summary": {
            "total_files_scanned": 5,
            "files_with_calls": len(set(cs["file_path"] for cs in call_sites)),
            "total_call_sites": len(call_sites),
            "unique_symbols_count": len(unique),
            "unique_symbols": unique,
            "symbols_by_category": cats,
        },
        "call_sites": call_sites,
        "errors": [],
    }


def _make_combo_scan_json(projects=None):
    if projects is None:
        projects = []
    entries = []
    for p in projects:
        name = p.get("project", "proj")
        target = p.get("target", f"/tmp/{name}")
        call_sites = p.get("call_sites", [])
        entries.append({
            "project": name,
            "target": target,
            "total_files_scanned": p.get("total_files_scanned", 1),
            "files_with_calls": p.get("files_with_calls", 1),
            "total_call_sites": len(call_sites),
            "unique_symbols": sorted(set(cs["ossl_symbol"] for cs in call_sites)),
            "symbols_by_category": {},
            "call_sites": call_sites,
        })
    return {
        "meta": {
            "report_type": "combo_scan",
            "merge_time": "2026-03-01T00:00:00",
            "total_projects": len(entries),
            "total_call_sites": sum(e["total_call_sites"] for e in entries),
            "total_unique_symbols": 0,
        },
        "projects": entries,
    }


def _write_json(tmp_path, name, data):
    path = os.path.join(str(tmp_path), name)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f)
    return path


class TestDataclasses:

    def test_diff_status_values(self):
        assert DiffStatus.ADDED.value == "added"
        assert DiffStatus.REMOVED.value == "removed"
        assert DiffStatus.CHANGED.value == "changed"
        assert DiffStatus.MOVED.value == "moved"
        assert DiffStatus.UNCHANGED.value == "unchanged"

    def test_diff_status_is_str(self):
        assert isinstance(DiffStatus.ADDED, str)
        assert DiffStatus.ADDED == "added"

    def test_metric_delta_fields(self):
        m = MetricDelta(name="total_call_sites", old_value=10, new_value=15, delta=5)
        assert m.name == "total_call_sites"
        assert m.old_value == 10
        assert m.new_value == 15
        assert m.delta == 5

    def test_call_site_delta_defaults(self):
        d = CallSiteDelta(
            status=DiffStatus.ADDED,
            identity_key=("a.c", "func", "SSL_new"),
            old_count=0,
            new_count=1,
        )
        assert d.old_lines == []
        assert d.new_lines == []
        assert d.category == ""

    def test_diff_result_is_empty_true(self):
        dr = DiffResult(old_label="v1", new_label="v2", projects=[])
        assert dr.is_empty() is True

    def test_diff_result_is_empty_with_unchanged_only(self):
        proj = ProjectDelta(
            project="test",
            call_site_delta=[
                CallSiteDelta(
                    status=DiffStatus.UNCHANGED,
                    identity_key=("a.c", "f", "SSL_new"),
                    old_count=1, new_count=1,
                ),
            ],
            symbol_delta=[
                SymbolDelta(
                    status=DiffStatus.UNCHANGED,
                    symbol="SSL_new", category="ssl_core",
                    old_count=1, new_count=1,
                ),
            ],
            file_delta=[
                FileDelta(
                    status=DiffStatus.UNCHANGED,
                    file_path="a.c",
                    old_call_count=1, new_call_count=1,
                ),
            ],
        )
        dr = DiffResult(old_label="v1", new_label="v2", projects=[proj])
        assert dr.is_empty() is True

    def test_diff_result_is_empty_false_call_site(self):
        proj = ProjectDelta(
            project="test",
            call_site_delta=[
                CallSiteDelta(
                    status=DiffStatus.ADDED,
                    identity_key=("a.c", "f", "SSL_new"),
                    old_count=0, new_count=1,
                ),
            ],
        )
        dr = DiffResult(old_label="v1", new_label="v2", projects=[proj])
        assert dr.is_empty() is False

    def test_diff_result_is_empty_false_symbol(self):
        proj = ProjectDelta(
            project="test",
            symbol_delta=[
                SymbolDelta(
                    status=DiffStatus.REMOVED,
                    symbol="SSL_free", category="ssl_core",
                    old_count=2, new_count=0,
                ),
            ],
        )
        dr = DiffResult(old_label="v1", new_label="v2", projects=[proj])
        assert dr.is_empty() is False

    def test_diff_result_is_empty_false_file(self):
        proj = ProjectDelta(
            project="test",
            file_delta=[
                FileDelta(
                    status=DiffStatus.ADDED,
                    file_path="new.c",
                    old_call_count=0, new_call_count=3,
                ),
            ],
        )
        dr = DiffResult(old_label="v1", new_label="v2", projects=[proj])
        assert dr.is_empty() is False


class TestLoadReport:

    def test_load_source_scan(self, tmp_path):
        cs = _make_call_site(file_path="/tmp/proj/src/tls.c", line=10)
        data = _make_source_scan_json(target="/tmp/proj", call_sites=[cs])
        path = _write_json(tmp_path, "report.json", data)

        result = load_report(path)
        assert result["report_type"] == "source_scan"
        assert len(result["projects"]) == 1
        proj = result["projects"][0]
        assert proj["project"] == "proj"
        assert len(proj["call_sites"]) == 1
        assert proj["call_sites"][0]["file_path"] == "src/tls.c"

    def test_load_combo_scan(self, tmp_path):
        cs = _make_call_site(file_path="/tmp/mylib/conn.c", line=5)
        data = _make_combo_scan_json(projects=[{
            "project": "mylib",
            "target": "/tmp/mylib",
            "call_sites": [cs],
        }])
        path = _write_json(tmp_path, "combo.json", data)

        result = load_report(path)
        assert result["report_type"] == "combo_scan"
        assert len(result["projects"]) == 1
        proj = result["projects"][0]
        assert proj["project"] == "mylib"
        assert proj["call_sites"][0]["file_path"] == "conn.c"

    def test_load_with_prefix_stripping(self, tmp_path):
        cs = _make_call_site(file_path="/home/user/work/proj/src/tls.c", line=10)
        data = _make_source_scan_json(target="/home/user/work/proj", call_sites=[cs])
        path = _write_json(tmp_path, "report.json", data)

        result = load_report(path, prefix="/home/user/work")
        proj = result["projects"][0]
        assert proj["call_sites"][0]["file_path"] == "src/tls.c"

    def test_load_nonexistent_file(self):
        with pytest.raises(FileNotFoundError):
            load_report("/nonexistent/path/report.json")

    def test_load_invalid_json(self, tmp_path):
        path = os.path.join(str(tmp_path), "bad.json")
        with open(path, "w") as f:
            f.write("{invalid json")
        with pytest.raises(json.JSONDecodeError):
            load_report(path)

    def test_load_empty_call_sites(self, tmp_path):
        data = _make_source_scan_json(target="/tmp/proj", call_sites=[])
        path = _write_json(tmp_path, "empty.json", data)

        result = load_report(path)
        proj = result["projects"][0]
        assert proj["call_sites"] == []
        assert proj["summary"]["total_call_sites"] == 0


class TestDiffSingle:

    def test_identical_reports_empty_diff(self):
        cs = _make_call_site(file_path="src/tls.c", line=10)
        old_data = {"project": "proj", "call_sites": [cs], "summary": {
            "total_files_scanned": 5, "files_with_calls": 1, "total_call_sites": 1,
        }}
        new_data = {"project": "proj", "call_sites": [cs], "summary": {
            "total_files_scanned": 5, "files_with_calls": 1, "total_call_sites": 1,
        }}

        result = diff_single(old_data, new_data)
        assert result.project == "proj"
        assert result.call_site_delta == []

        sym_statuses = {s.status for s in result.symbol_delta}
        assert sym_statuses == {DiffStatus.UNCHANGED}

        file_statuses = {f.status for f in result.file_delta}
        assert file_statuses == {DiffStatus.UNCHANGED}

    def test_identical_with_include_unchanged(self):
        cs = _make_call_site(file_path="src/tls.c", line=10)
        old_data = {"project": "proj", "call_sites": [cs], "summary": {}}
        new_data = {"project": "proj", "call_sites": [cs], "summary": {}}

        result = diff_single(old_data, new_data, include_unchanged=True)
        assert len(result.call_site_delta) == 1
        assert result.call_site_delta[0].status == DiffStatus.UNCHANGED

    def test_added_symbol(self):
        old_cs = _make_call_site(file_path="src/tls.c", symbol="SSL_CTX_new", line=10)
        new_cs1 = _make_call_site(file_path="src/tls.c", symbol="SSL_CTX_new", line=10)
        new_cs2 = _make_call_site(file_path="src/tls.c", symbol="SSL_connect",
                                  caller="do_connect", line=20)

        old_data = {"project": "proj", "call_sites": [old_cs], "summary": {}}
        new_data = {"project": "proj", "call_sites": [new_cs1, new_cs2], "summary": {}}

        result = diff_single(old_data, new_data)
        added_cs = [d for d in result.call_site_delta if d.status == DiffStatus.ADDED]
        assert len(added_cs) == 1
        assert added_cs[0].identity_key[2] == "SSL_connect"

        added_sym = [s for s in result.symbol_delta if s.status == DiffStatus.ADDED]
        assert len(added_sym) == 1
        assert added_sym[0].symbol == "SSL_connect"

    def test_removed_symbol(self):
        old_cs1 = _make_call_site(file_path="src/tls.c", symbol="SSL_CTX_new", line=10)
        old_cs2 = _make_call_site(file_path="src/tls.c", symbol="SSL_free",
                                  caller="cleanup", line=50)
        new_cs = _make_call_site(file_path="src/tls.c", symbol="SSL_CTX_new", line=10)

        old_data = {"project": "proj", "call_sites": [old_cs1, old_cs2], "summary": {}}
        new_data = {"project": "proj", "call_sites": [new_cs], "summary": {}}

        result = diff_single(old_data, new_data)
        removed_cs = [d for d in result.call_site_delta if d.status == DiffStatus.REMOVED]
        assert len(removed_cs) == 1
        assert removed_cs[0].identity_key[2] == "SSL_free"

        removed_sym = [s for s in result.symbol_delta if s.status == DiffStatus.REMOVED]
        assert len(removed_sym) == 1
        assert removed_sym[0].symbol == "SSL_free"

    def test_changed_count(self):
        old_cs = _make_call_site(file_path="src/tls.c", symbol="SSL_read",
                                 caller="read_data", line=30)
        new_cs1 = _make_call_site(file_path="src/tls.c", symbol="SSL_read",
                                  caller="read_data", line=30)
        new_cs2 = _make_call_site(file_path="src/tls.c", symbol="SSL_read",
                                  caller="read_data", line=45)

        old_data = {"project": "proj", "call_sites": [old_cs], "summary": {}}
        new_data = {"project": "proj", "call_sites": [new_cs1, new_cs2], "summary": {}}

        result = diff_single(old_data, new_data)
        changed = [d for d in result.call_site_delta if d.status == DiffStatus.CHANGED]
        assert len(changed) == 1
        assert changed[0].old_count == 1
        assert changed[0].new_count == 2

    def test_moved_call_same_count_different_line(self):
        old_cs = _make_call_site(file_path="src/tls.c", symbol="SSL_write",
                                 caller="send_data", line=100)
        new_cs = _make_call_site(file_path="src/tls.c", symbol="SSL_write",
                                 caller="send_data", line=120)

        old_data = {"project": "proj", "call_sites": [old_cs], "summary": {}}
        new_data = {"project": "proj", "call_sites": [new_cs], "summary": {}}

        result = diff_single(old_data, new_data)
        moved = [d for d in result.call_site_delta if d.status == DiffStatus.MOVED]
        assert len(moved) == 1
        assert moved[0].old_lines == [100]
        assert moved[0].new_lines == [120]

    def test_file_delta_aggregation(self):
        old_cs = _make_call_site(file_path="src/old.c", symbol="EVP_EncryptInit",
                                 category="crypto_evp", line=10)
        new_cs = _make_call_site(file_path="src/new.c", symbol="EVP_DecryptInit",
                                 category="crypto_evp", line=20)

        old_data = {"project": "proj", "call_sites": [old_cs], "summary": {}}
        new_data = {"project": "proj", "call_sites": [new_cs], "summary": {}}

        result = diff_single(old_data, new_data)
        added_files = [f for f in result.file_delta if f.status == DiffStatus.ADDED]
        removed_files = [f for f in result.file_delta if f.status == DiffStatus.REMOVED]
        assert len(added_files) == 1
        assert added_files[0].file_path == "src/new.c"
        assert len(removed_files) == 1
        assert removed_files[0].file_path == "src/old.c"

    def test_category_delta(self):
        old_cs = _make_call_site(symbol="SSL_CTX_new", category="ssl_core", line=10)
        new_cs1 = _make_call_site(symbol="SSL_CTX_new", category="ssl_core", line=10)
        new_cs2 = _make_call_site(file_path="src/hash.c", symbol="SHA256_Init",
                                  category="crypto_hash", caller="hash_data", line=5)

        old_data = {"project": "proj", "call_sites": [old_cs], "summary": {}}
        new_data = {"project": "proj", "call_sites": [new_cs1, new_cs2], "summary": {}}

        result = diff_single(old_data, new_data)
        added_sym = [s for s in result.symbol_delta if s.status == DiffStatus.ADDED]
        added_cats = {s.category for s in added_sym}
        assert "crypto_hash" in added_cats

    def test_ignore_categories(self):
        cs1 = _make_call_site(symbol="SSL_CTX_new", category="ssl_core", line=10)
        cs2 = _make_call_site(file_path="src/err.c", symbol="ERR_print_errors",
                              category="crypto_err", caller="handle_error", line=5)
        cs3 = _make_call_site(file_path="src/bio.c", symbol="BIO_new",
                              category="crypto_bio", caller="setup_bio", line=15)

        old_data = {"project": "proj", "call_sites": [cs1, cs2], "summary": {}}
        new_data = {"project": "proj", "call_sites": [cs1, cs3], "summary": {}}

        result = diff_single(old_data, new_data, ignore_categories={"crypto_err", "crypto_bio"})
        assert result.call_site_delta == []
        assert all(s.status == DiffStatus.UNCHANGED for s in result.symbol_delta)

    def test_summary_only_empties_call_sites(self):
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=10)
        new_cs = _make_call_site(symbol="SSL_connect", caller="do_connect", line=20)

        old_data = {"project": "proj", "call_sites": [old_cs], "summary": {}}
        new_data = {"project": "proj", "call_sites": [new_cs], "summary": {}}

        result = diff_single(old_data, new_data, summary_only=True)
        assert result.call_site_delta == []
        assert len(result.symbol_delta) > 0
        assert len(result.file_delta) > 0

    def test_metrics_computed(self):
        old_data = {
            "project": "proj",
            "call_sites": [
                _make_call_site(symbol="SSL_CTX_new", line=10),
            ],
            "summary": {
                "total_files_scanned": 10,
                "files_with_calls": 1,
                "total_call_sites": 1,
            },
        }
        new_data = {
            "project": "proj",
            "call_sites": [
                _make_call_site(symbol="SSL_CTX_new", line=10),
                _make_call_site(file_path="src/conn.c", symbol="SSL_connect",
                                caller="do_connect", line=20),
            ],
            "summary": {
                "total_files_scanned": 12,
                "files_with_calls": 2,
                "total_call_sites": 2,
            },
        }

        result = diff_single(old_data, new_data)
        metric_map = {m.name: m for m in result.metrics}
        assert metric_map["total_files_scanned"].delta == 2
        assert metric_map["files_with_calls"].delta == 1
        assert metric_map["total_call_sites"].delta == 1
        assert metric_map["unique_symbols"].delta == 1

    def test_sort_order_added_before_removed(self):
        old_cs = _make_call_site(file_path="src/a.c", symbol="SSL_free",
                                 caller="cleanup", line=50)
        new_cs = _make_call_site(file_path="src/b.c", symbol="SSL_new",
                                 caller="setup", line=10)

        old_data = {"project": "proj", "call_sites": [old_cs], "summary": {}}
        new_data = {"project": "proj", "call_sites": [new_cs], "summary": {}}

        result = diff_single(old_data, new_data)
        statuses = [d.status for d in result.call_site_delta]
        assert statuses[0] == DiffStatus.ADDED
        assert statuses[1] == DiffStatus.REMOVED


class TestLoadAndDiffIntegration:

    def test_load_and_diff_source_scan(self, tmp_path):
        old_cs = _make_call_site(file_path="/tmp/proj/src/tls.c",
                                 symbol="SSL_CTX_new", line=10)
        new_cs1 = _make_call_site(file_path="/tmp/proj/src/tls.c",
                                  symbol="SSL_CTX_new", line=10)
        new_cs2 = _make_call_site(file_path="/tmp/proj/src/tls.c",
                                  symbol="SSL_connect", caller="do_connect", line=30)

        old_json = _make_source_scan_json(target="/tmp/proj", call_sites=[old_cs])
        new_json = _make_source_scan_json(target="/tmp/proj", call_sites=[new_cs1, new_cs2])

        old_path = _write_json(tmp_path, "old.json", old_json)
        new_path = _write_json(tmp_path, "new.json", new_json)

        old_report = load_report(old_path)
        new_report = load_report(new_path)

        result = diff_single(
            old_report["projects"][0],
            new_report["projects"][0],
        )

        assert result.project == "proj"
        added = [d for d in result.call_site_delta if d.status == DiffStatus.ADDED]
        assert len(added) == 1
        assert added[0].identity_key[2] == "SSL_connect"


class TestDiffCombo:

    def _load_combo(self, tmp_path, projects, name="combo.json"):
        data = _make_combo_scan_json(projects=projects)
        path = _write_json(tmp_path, name, data)
        return load_report(path)

    def test_identical_combo(self, tmp_path):
        cs = _make_call_site(file_path="/tmp/curl/lib/tls.c",
                             symbol="SSL_CTX_new", line=10)
        projs = [{"project": "curl", "target": "/tmp/curl",
                  "call_sites": [cs]}]

        old_report = self._load_combo(tmp_path, projs, "old.json")
        new_report = self._load_combo(tmp_path, projs, "new.json")

        result = diff_combo(old_report, new_report, include_unchanged=True)
        assert len(result.projects) == 1
        assert result.projects[0].project == "curl"
        for csd in result.projects[0].call_site_delta:
            assert csd.status == DiffStatus.UNCHANGED

    def test_added_project(self, tmp_path):
        cs_curl = _make_call_site(file_path="/tmp/curl/lib/tls.c",
                                  symbol="SSL_CTX_new", line=10)
        cs_nginx = _make_call_site(file_path="/tmp/nginx/src/ssl.c",
                                   symbol="SSL_connect", caller="do_ssl", line=5)

        old_report = self._load_combo(
            tmp_path, [{"project": "curl", "target": "/tmp/curl",
                        "call_sites": [cs_curl]}], "old.json")
        new_report = self._load_combo(
            tmp_path, [{"project": "curl", "target": "/tmp/curl",
                        "call_sites": [cs_curl]},
                       {"project": "nginx", "target": "/tmp/nginx",
                        "call_sites": [cs_nginx]}], "new.json")

        result = diff_combo(old_report, new_report)
        proj_names = {p.project for p in result.projects}
        assert "nginx" in proj_names

        nginx_proj = [p for p in result.projects if p.project == "nginx"][0]
        added_syms = [s for s in nginx_proj.symbol_delta
                      if s.status == DiffStatus.ADDED]
        assert len(added_syms) == 1
        assert added_syms[0].symbol == "SSL_connect"

    def test_removed_project(self, tmp_path):
        cs_curl = _make_call_site(file_path="/tmp/curl/lib/tls.c",
                                  symbol="SSL_CTX_new", line=10)
        cs_nginx = _make_call_site(file_path="/tmp/nginx/src/ssl.c",
                                   symbol="SSL_connect", caller="do_ssl", line=5)

        old_report = self._load_combo(
            tmp_path, [{"project": "curl", "target": "/tmp/curl",
                        "call_sites": [cs_curl]},
                       {"project": "nginx", "target": "/tmp/nginx",
                        "call_sites": [cs_nginx]}], "old.json")
        new_report = self._load_combo(
            tmp_path, [{"project": "curl", "target": "/tmp/curl",
                        "call_sites": [cs_curl]}], "new.json")

        result = diff_combo(old_report, new_report)
        nginx_proj = [p for p in result.projects if p.project == "nginx"][0]
        removed_syms = [s for s in nginx_proj.symbol_delta
                        if s.status == DiffStatus.REMOVED]
        assert len(removed_syms) == 1
        assert removed_syms[0].symbol == "SSL_connect"

    def test_changed_project(self, tmp_path):
        old_cs = _make_call_site(file_path="/tmp/curl/lib/tls.c",
                                 symbol="SSL_CTX_new", line=10)
        new_cs1 = _make_call_site(file_path="/tmp/curl/lib/tls.c",
                                  symbol="SSL_CTX_new", line=10)
        new_cs2 = _make_call_site(file_path="/tmp/curl/lib/tls.c",
                                  symbol="SSL_read", caller="read_cb", line=50)

        old_report = self._load_combo(
            tmp_path, [{"project": "curl", "target": "/tmp/curl",
                        "call_sites": [old_cs]}], "old.json")
        new_report = self._load_combo(
            tmp_path, [{"project": "curl", "target": "/tmp/curl",
                        "call_sites": [new_cs1, new_cs2]}], "new.json")

        result = diff_combo(old_report, new_report)
        assert len(result.projects) == 1
        curl_proj = result.projects[0]
        added = [s for s in curl_proj.symbol_delta
                 if s.status == DiffStatus.ADDED]
        assert any(s.symbol == "SSL_read" for s in added)

    def test_mixed_projects(self, tmp_path):
        cs_curl = _make_call_site(file_path="/tmp/curl/lib/tls.c",
                                  symbol="SSL_CTX_new", line=10)
        cs_nginx = _make_call_site(file_path="/tmp/nginx/src/ssl.c",
                                   symbol="SSL_connect", caller="do_ssl", line=5)
        cs_wget = _make_call_site(file_path="/tmp/wget/src/http.c",
                                  symbol="EVP_DigestInit", caller="hash",
                                  category="crypto_evp", line=8)

        old_report = self._load_combo(
            tmp_path, [
                {"project": "curl", "target": "/tmp/curl",
                 "call_sites": [cs_curl]},
                {"project": "nginx", "target": "/tmp/nginx",
                 "call_sites": [cs_nginx]},
            ], "old.json")
        new_report = self._load_combo(
            tmp_path, [
                {"project": "curl", "target": "/tmp/curl",
                 "call_sites": [cs_curl]},
                {"project": "wget", "target": "/tmp/wget",
                 "call_sites": [cs_wget]},
            ], "new.json")

        result = diff_combo(old_report, new_report)
        proj_names = {p.project for p in result.projects}
        assert "curl" in proj_names
        assert "nginx" in proj_names
        assert "wget" in proj_names

        nginx_proj = [p for p in result.projects if p.project == "nginx"][0]
        assert all(s.status == DiffStatus.REMOVED
                   for s in nginx_proj.symbol_delta)

        wget_proj = [p for p in result.projects if p.project == "wget"][0]
        assert all(s.status == DiffStatus.ADDED
                   for s in wget_proj.symbol_delta)


class TestJsonExporter:

    def test_export_single_roundtrip(self, tmp_path):
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=10)
        new_cs = _make_call_site(symbol="SSL_connect", caller="do_connect", line=20)

        old_data = {"project": "proj", "call_sites": [old_cs],
                    "summary": {"total_files_scanned": 5, "files_with_calls": 1,
                                "total_call_sites": 1}}
        new_data = {"project": "proj", "call_sites": [new_cs],
                    "summary": {"total_files_scanned": 6, "files_with_calls": 1,
                                "total_call_sites": 1}}

        pd = diff_single(old_data, new_data)
        dr = DiffResult(old_label="old.json", new_label="new.json", projects=[pd])

        out_path = os.path.join(str(tmp_path), "diff.json")
        SourceDiffJsonExporter().export(dr, out_path)

        with open(out_path, "r", encoding="utf-8") as f:
            exported = json.load(f)

        assert exported["meta"]["report_type"] == "source_diff"
        assert "summary_delta" in exported
        assert "symbol_delta" in exported
        assert "file_delta" in exported
        assert "call_site_delta" in exported

        added_syms = [s for s in exported["symbol_delta"]
                      if s["status"] == "added"]
        assert any(s["symbol"] == "SSL_connect" for s in added_syms)

    def test_export_combo_roundtrip(self, tmp_path):
        cs1 = _make_call_site(file_path="lib/tls.c", symbol="SSL_CTX_new", line=10)
        cs2 = _make_call_site(file_path="src/ssl.c", symbol="SSL_connect",
                              caller="do_ssl", line=5)

        pd1 = diff_single(
            {"project": "curl", "call_sites": [cs1], "summary": {}},
            {"project": "curl", "call_sites": [cs1], "summary": {}},
            include_unchanged=True,
        )
        pd2 = diff_single(
            {"project": "nginx", "call_sites": [], "summary": {}},
            {"project": "nginx", "call_sites": [cs2], "summary": {}},
        )
        dr = DiffResult(old_label="old_combo.json", new_label="new_combo.json",
                        projects=[pd1, pd2], is_combo=True)

        out_path = os.path.join(str(tmp_path), "combo_diff.json")
        SourceDiffJsonExporter().export(dr, out_path)

        with open(out_path, "r", encoding="utf-8") as f:
            exported = json.load(f)

        assert exported["meta"]["report_type"] == "combo_diff"
        assert "projects" in exported
        assert len(exported["projects"]) == 2
        proj_names = [p["project"] for p in exported["projects"]]
        assert "curl" in proj_names
        assert "nginx" in proj_names

    def test_export_meta_fields(self, tmp_path):
        pd = diff_single(
            {"project": "proj", "call_sites": [], "summary": {}},
            {"project": "proj", "call_sites": [], "summary": {}},
        )
        dr = DiffResult(old_label="v1.json", new_label="v2.json", projects=[pd])

        out_path = os.path.join(str(tmp_path), "meta_test.json")
        SourceDiffJsonExporter().export(dr, out_path)

        with open(out_path, "r", encoding="utf-8") as f:
            exported = json.load(f)

        meta = exported["meta"]
        assert "report_type" in meta
        assert meta["old_report"] == "v1.json"
        assert meta["new_report"] == "v2.json"
        assert "diff_time" in meta
        assert "tool_version" in meta


class TestConsoleFormatter:

    def test_format_has_summary(self):
        old_data = {
            "project": "proj",
            "call_sites": [_make_call_site(symbol="SSL_CTX_new", line=10)],
            "summary": {"total_files_scanned": 10, "files_with_calls": 1,
                        "total_call_sites": 1},
        }
        new_data = {
            "project": "proj",
            "call_sites": [
                _make_call_site(symbol="SSL_CTX_new", line=10),
                _make_call_site(symbol="SSL_connect", caller="do_connect", line=20),
            ],
            "summary": {"total_files_scanned": 12, "files_with_calls": 2,
                        "total_call_sites": 2},
        }

        pd = diff_single(old_data, new_data)
        dr = DiffResult(old_label="old.json", new_label="new.json", projects=[pd])

        output = format_console(dr)
        assert "old.json" in output
        assert "new.json" in output
        assert "total_files_scanned" in output or "Files Scanned" in output

    def test_format_no_changes(self):
        cs = _make_call_site(symbol="SSL_CTX_new", line=10)
        pd = diff_single(
            {"project": "proj", "call_sites": [cs], "summary": {}},
            {"project": "proj", "call_sites": [cs], "summary": {}},
            include_unchanged=True,
        )
        dr = DiffResult(old_label="old.json", new_label="new.json", projects=[pd])

        output = format_console(dr)
        assert "No changes detected" in output

    def test_format_shows_symbols(self):
        old_cs = _make_call_site(symbol="DES_ecb_encrypt",
                                 category="crypto_legacy", line=10)
        new_cs = _make_call_site(symbol="EVP_MAC_init",
                                 category="crypto_evp", caller="mac_setup", line=20)

        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json", projects=[pd])

        output = format_console(dr)
        assert "EVP_MAC_init" in output
        assert "DES_ecb_encrypt" in output

    def test_format_combo(self):
        cs1 = _make_call_site(symbol="SSL_CTX_new", line=10)
        cs2 = _make_call_site(file_path="src/ssl.c", symbol="SSL_connect",
                              caller="do_ssl", line=5)

        pd1 = diff_single(
            {"project": "curl", "call_sites": [cs1], "summary": {}},
            {"project": "curl", "call_sites": [cs1], "summary": {}},
        )
        pd2 = diff_single(
            {"project": "nginx", "call_sites": [], "summary": {}},
            {"project": "nginx", "call_sites": [cs2], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd1, pd2])

        output = format_console(dr)
        assert "curl" in output
        assert "nginx" in output


class TestExcelExporter:

    def _make_single_diff(self):
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=10)
        new_cs1 = _make_call_site(symbol="SSL_CTX_new", line=10)
        new_cs2 = _make_call_site(symbol="SSL_connect", caller="do_connect", line=20)
        new_cs3 = _make_call_site(file_path="src/hash.c", symbol="SHA256_Init",
                                  category="crypto_hash", caller="hash_data", line=5)

        old_data = {
            "project": "proj",
            "call_sites": [old_cs],
            "summary": {"total_files_scanned": 5, "files_with_calls": 1,
                         "total_call_sites": 1},
        }
        new_data = {
            "project": "proj",
            "call_sites": [new_cs1, new_cs2, new_cs3],
            "summary": {"total_files_scanned": 7, "files_with_calls": 2,
                         "total_call_sites": 3},
        }

        pd = diff_single(old_data, new_data)
        return DiffResult(old_label="old.json", new_label="new.json", projects=[pd])

    def _make_combo_diff(self):
        cs_curl = _make_call_site(file_path="lib/tls.c", symbol="SSL_CTX_new", line=10)
        cs_nginx = _make_call_site(file_path="src/ssl.c", symbol="SSL_connect",
                                   caller="do_ssl", line=5)

        pd1 = diff_single(
            {"project": "curl", "call_sites": [cs_curl], "summary": {
                "total_files_scanned": 10, "files_with_calls": 1,
                "total_call_sites": 1}},
            {"project": "curl", "call_sites": [cs_curl], "summary": {
                "total_files_scanned": 10, "files_with_calls": 1,
                "total_call_sites": 1}},
            include_unchanged=True,
        )
        pd2 = diff_single(
            {"project": "nginx", "call_sites": [], "summary": {
                "total_files_scanned": 0, "files_with_calls": 0,
                "total_call_sites": 0}},
            {"project": "nginx", "call_sites": [cs_nginx], "summary": {
                "total_files_scanned": 5, "files_with_calls": 1,
                "total_call_sites": 1}},
        )
        return DiffResult(old_label="old_combo.json", new_label="new_combo.json",
                          projects=[pd1, pd2], is_combo=True)

    def _load_wb(self, path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook
        return load_workbook(path)

    def test_export_creates_xlsx(self, tmp_path):
        """Single project diff creates XLSX file."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)
        assert os.path.isfile(out_path)

    def test_sheet_names_single(self, tmp_path):
        """Single project: 4 sheets, no Project Delta."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        names = wb.sheetnames
        assert "Summary Delta" in names
        assert "Symbol Delta" in names
        assert "File Delta" in names
        assert "Call Site Delta" in names
        assert "Project Delta" not in names
        assert len(names) == 4

    def test_combo_has_project_sheet(self, tmp_path):
        """Multi-project diff creates 5th Project Delta sheet."""
        dr = self._make_combo_diff()
        out_path = os.path.join(str(tmp_path), "combo_diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        assert "Project Delta" in wb.sheetnames
        assert len(wb.sheetnames) == 5

    def test_summary_sheet_has_metrics(self, tmp_path):
        """Summary Delta sheet contains metric rows with correct delta values."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Summary Delta"]
        header_row = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert header_row == ["Metric", "Old", "New", "Delta"]

        metrics_found = {}
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name and not name.startswith("["):
                metrics_found[name] = ws.cell(row=row, column=4).value

        assert "Files Scanned" in metrics_found
        assert metrics_found["Files Scanned"] == 2

    def test_summary_sheet_has_category_breakdown(self, tmp_path):
        """Summary Delta sheet has category breakdown rows below metrics."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Summary Delta"]

        all_values = []
        for row in range(1, ws.max_row + 1):
            all_values.append(ws.cell(row=row, column=1).value)

        assert any(v and v.startswith("[") for v in all_values), \
            "Expected category rows like [ssl_core]"

    def test_symbol_sheet_has_data(self, tmp_path):
        """Symbol Delta sheet contains symbol rows with status."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Symbol Delta"]
        header = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert header == ["OpenSSL Symbol", "Category", "Status", "Old Calls",
                          "New Calls", "Delta"]

        symbols = []
        for row in range(2, ws.max_row + 1):
            sym = ws.cell(row=row, column=1).value
            if sym:
                symbols.append(sym)
        assert "SSL_connect" in symbols
        assert "SHA256_Init" in symbols

    def test_file_sheet_has_data(self, tmp_path):
        """File Delta sheet contains file rows with 9-column layout."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["File Delta"]
        header = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert header == ["File Path", "Status", "Old Calls", "New Calls", "Delta",
                          "Old Symbols", "New Symbols", "Added Symbols",
                          "Removed Symbols"]

        file_paths = []
        for row in range(2, ws.max_row + 1):
            fp = ws.cell(row=row, column=1).value
            if fp:
                file_paths.append(fp)
        assert len(file_paths) > 0

    def test_callsite_sheet_has_data(self, tmp_path):
        """Call Site Delta sheet contains call site rows."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Call Site Delta"]
        header = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert header == ["Status", "File Path", "Caller Function", "OpenSSL Symbol",
                          "Category", "Old Line", "New Line", "Old Args", "New Args"]

        statuses = []
        for row in range(2, ws.max_row + 1):
            s = ws.cell(row=row, column=1).value
            if s:
                statuses.append(s)
        assert len(statuses) > 0

    def test_summary_only(self, tmp_path):
        """summary_only=True produces empty Call Site Delta sheet (header only)."""
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=10)
        new_cs = _make_call_site(symbol="SSL_connect", caller="do_connect", line=20)

        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
            summary_only=True,
        )
        dr = DiffResult(old_label="old.json", new_label="new.json", projects=[pd])

        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Call Site Delta"]
        assert ws.max_row == 1

    def test_status_fills(self, tmp_path):
        """Status cells have conditional fill colors."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Symbol Delta"]

        fills_found = set()
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                rgb = cell.fill.fgColor.rgb
                if isinstance(rgb, str) and len(rgb) >= 6:
                    fills_found.add(rgb[-6:])

        assert DIFF_COLORS["added"].upper() in {c.upper() for c in fills_found} or \
               DIFF_COLORS["added"] in {c.lower() for c in fills_found}, \
               f"Expected green fill for ADDED status, found fills: {fills_found}"

    def test_delta_font_colors(self, tmp_path):
        """Delta column uses colored font (green for positive, red for negative)."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Symbol Delta"]

        font_colors = set()
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=6)
            if cell.font and cell.font.color and cell.font.color.rgb:
                rgb = cell.font.color.rgb
                if isinstance(rgb, str) and len(rgb) >= 6:
                    font_colors.add(rgb[-6:])

        assert len(font_colors) > 0, "Expected colored fonts in delta column"

    def test_project_sheet_content(self, tmp_path):
        """Project Delta sheet has project rows with 9-column layout."""
        dr = self._make_combo_diff()
        out_path = os.path.join(str(tmp_path), "combo_diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Project Delta"]
        header = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert header == ["Project", "Status", "Old Calls", "New Calls", "Delta",
                          "Old Symbols", "New Symbols", "Added Symbols",
                          "Removed Symbols"]

        projects = []
        for row in range(2, ws.max_row + 1):
            proj = ws.cell(row=row, column=1).value
            if proj:
                projects.append(proj)
        assert "nginx" in projects

    def test_export_creates_parent_dirs(self, tmp_path):
        """Export creates parent directories if they don't exist."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "subdir", "deep", "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)
        assert os.path.isfile(out_path)

    def test_empty_diff_export(self, tmp_path):
        """Empty diff result still creates valid XLSX."""
        cs = _make_call_site(symbol="SSL_CTX_new", line=10)
        pd = diff_single(
            {"project": "proj", "call_sites": [cs], "summary": {}},
            {"project": "proj", "call_sites": [cs], "summary": {}},
        )
        dr = DiffResult(old_label="v1", new_label="v2", projects=[pd])

        out_path = os.path.join(str(tmp_path), "empty.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        assert "Summary Delta" in wb.sheetnames

    def test_unchanged_excluded_by_default(self, tmp_path):
        """UNCHANGED symbols excluded from Symbol Delta sheet by default."""
        cs = _make_call_site(symbol="SSL_CTX_new", line=10)
        pd = diff_single(
            {"project": "proj", "call_sites": [cs], "summary": {}},
            {"project": "proj", "call_sites": [cs], "summary": {}},
        )
        dr = DiffResult(old_label="v1", new_label="v2", projects=[pd])

        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Symbol Delta"]
        statuses = []
        for row in range(2, ws.max_row + 1):
            s = ws.cell(row=row, column=3).value
            if s:
                statuses.append(s)
        assert "unchanged" not in statuses

    def test_include_unchanged_flag(self, tmp_path):
        """include_unchanged=True shows UNCHANGED entries in symbol sheet."""
        cs = _make_call_site(symbol="SSL_CTX_new", line=10)
        pd = diff_single(
            {"project": "proj", "call_sites": [cs], "summary": {}},
            {"project": "proj", "call_sites": [cs], "summary": {}},
            include_unchanged=True,
        )
        dr = DiffResult(old_label="v1", new_label="v2", projects=[pd])

        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter(include_unchanged=True).export(dr, out_path)

        wb = self._load_wb(out_path)
        ws = wb["Symbol Delta"]
        statuses = []
        for row in range(2, ws.max_row + 1):
            s = ws.cell(row=row, column=3).value
            if s:
                statuses.append(s)
        assert "unchanged" in statuses


class TestCLI:

    def test_source_diff_json_output(self, tmp_path):
        from openssl_scanner.__main__ import main

        cs1 = [_make_call_site(symbol="SSL_CTX_new", line=10)]
        cs2 = [
            _make_call_site(symbol="SSL_CTX_new", line=10),
            _make_call_site(symbol="SSL_connect", caller="do_connect", line=20),
        ]
        old_path = _write_json(tmp_path, "old.json", _make_source_scan_json(call_sites=cs1))
        new_path = _write_json(tmp_path, "new.json", _make_source_scan_json(call_sites=cs2))
        out_path = os.path.join(str(tmp_path), "diff.json")

        rc = main(['source-diff', old_path, new_path, '-o', out_path])
        assert rc == 1
        assert os.path.exists(out_path)
        with open(out_path, "r") as f:
            data = json.load(f)
        assert "call_site_delta" in data
        assert "symbol_delta" in data
        symbols = [s["symbol"] for s in data["symbol_delta"]]
        assert "SSL_connect" in symbols

    def test_exit_code_0_no_changes(self, tmp_path):
        from openssl_scanner.__main__ import main

        cs = [_make_call_site(symbol="SSL_CTX_new", line=10)]
        old_path = _write_json(tmp_path, "old.json", _make_source_scan_json(call_sites=cs))
        new_path = _write_json(tmp_path, "new.json", _make_source_scan_json(call_sites=cs))
        out_path = os.path.join(str(tmp_path), "diff.json")

        rc = main(['source-diff', old_path, new_path, '-o', out_path])
        assert rc == 0

    def test_exit_code_1_has_changes(self, tmp_path):
        from openssl_scanner.__main__ import main

        cs1 = [_make_call_site(symbol="SSL_CTX_new", line=10)]
        cs2 = [_make_call_site(symbol="SSL_connect", caller="do_connect", line=20)]
        old_path = _write_json(tmp_path, "old.json", _make_source_scan_json(call_sites=cs1))
        new_path = _write_json(tmp_path, "new.json", _make_source_scan_json(call_sites=cs2))
        out_path = os.path.join(str(tmp_path), "diff.json")

        rc = main(['source-diff', old_path, new_path, '-o', out_path])
        assert rc == 1

    def test_exit_code_2_bad_input(self, tmp_path):
        from openssl_scanner.__main__ import main

        rc = main(['source-diff', '/nonexistent/old.json', '/nonexistent/new.json',
                    '-o', os.path.join(str(tmp_path), 'diff.json')])
        assert rc == 2

    def test_source_diff_xlsx_output(self, tmp_path):
        from openssl_scanner.__main__ import main

        cs1 = [_make_call_site(symbol="SSL_CTX_new", line=10)]
        cs2 = [
            _make_call_site(symbol="SSL_CTX_new", line=10),
            _make_call_site(symbol="EVP_DigestInit", caller="hash_it", line=30,
                            category="crypto_evp"),
        ]
        old_path = _write_json(tmp_path, "old.json", _make_source_scan_json(call_sites=cs1))
        new_path = _write_json(tmp_path, "new.json", _make_source_scan_json(call_sites=cs2))
        out_path = os.path.join(str(tmp_path), "diff.xlsx")

        rc = main(['source-diff', old_path, new_path, '-o', out_path])
        assert rc == 1
        assert os.path.exists(out_path)

    def test_summary_only_flag(self, tmp_path):
        from openssl_scanner.__main__ import main

        cs1 = [_make_call_site(symbol="SSL_CTX_new", line=10)]
        cs2 = [
            _make_call_site(symbol="SSL_CTX_new", line=10),
            _make_call_site(symbol="SSL_connect", caller="do_connect", line=20),
        ]
        old_path = _write_json(tmp_path, "old.json", _make_source_scan_json(call_sites=cs1))
        new_path = _write_json(tmp_path, "new.json", _make_source_scan_json(call_sites=cs2))
        out_path = os.path.join(str(tmp_path), "diff.json")

        rc = main(['source-diff', old_path, new_path, '-o', out_path, '--summary-only'])
        assert rc == 1
        with open(out_path, "r") as f:
            data = json.load(f)
        assert data["call_site_delta"] == []
        assert len(data["symbol_delta"]) > 0

    def test_ignore_categories_flag(self, tmp_path):
        from openssl_scanner.__main__ import main

        cs1 = [
            _make_call_site(symbol="SSL_CTX_new", line=10, category="ssl_core"),
            _make_call_site(symbol="EVP_DigestInit", caller="hash_it", line=30,
                            category="crypto_evp"),
        ]
        cs2 = [
            _make_call_site(symbol="SSL_CTX_new", line=10, category="ssl_core"),
        ]
        old_path = _write_json(tmp_path, "old.json", _make_source_scan_json(call_sites=cs1))
        new_path = _write_json(tmp_path, "new.json", _make_source_scan_json(call_sites=cs2))
        out_path = os.path.join(str(tmp_path), "diff.json")

        rc = main(['source-diff', old_path, new_path, '-o', out_path,
                    '--ignore-categories', 'crypto_evp'])
        assert rc == 0
        with open(out_path, "r") as f:
            data = json.load(f)
        categories = [s["category"] for s in data["symbol_delta"]]
        assert "crypto_evp" not in categories

    def test_cli_combo_json_output(self, tmp_path):
        from openssl_scanner.__main__ import main

        old_data = _make_combo_scan_json(projects=[{
            "project": "curl",
            "call_sites": [_make_call_site(symbol="SSL_CTX_new", line=10)],
        }])
        new_data = _make_combo_scan_json(projects=[
            {
                "project": "curl",
                "call_sites": [
                    _make_call_site(symbol="SSL_CTX_new", line=10),
                    _make_call_site(symbol="SSL_connect", caller="do_ssl", line=20),
                ],
            },
            {
                "project": "nginx",
                "call_sites": [_make_call_site(file_path="src/ssl.c",
                               symbol="EVP_DigestInit", caller="hash_it",
                               line=5, category="crypto_evp")],
            },
        ])
        old_path = _write_json(tmp_path, "old_combo.json", old_data)
        new_path = _write_json(tmp_path, "new_combo.json", new_data)
        out_path = os.path.join(str(tmp_path), "combo_diff.json")

        rc = main(['source-diff', old_path, new_path, '-o', out_path])
        assert rc == 1
        with open(out_path, "r") as f:
            data = json.load(f)
        assert data["meta"]["report_type"] == "combo_diff"
        assert "projects" in data
        proj_names = [p["project"] for p in data["projects"]]
        assert "curl" in proj_names
        assert "nginx" in proj_names


class TestFileDeltaNewFields:

    def test_file_delta_has_symbol_fields(self):
        old_cs = [
            _make_call_site(symbol="SSL_CTX_new", line=10),
            _make_call_site(symbol="SSL_connect", caller="do_conn", line=20),
        ]
        new_cs = [
            _make_call_site(symbol="SSL_CTX_new", line=10),
            _make_call_site(symbol="EVP_DigestInit", caller="hash_it",
                            line=30, category="crypto_evp"),
        ]
        pd = diff_single(
            {"project": "proj", "call_sites": old_cs, "summary": {}},
            {"project": "proj", "call_sites": new_cs, "summary": {}},
        )
        file_map = {fd.file_path: fd for fd in pd.file_delta}
        fd = file_map["src/tls.c"]
        assert fd.old_symbols == 2
        assert fd.new_symbols == 2
        assert "EVP_DigestInit" in fd.added_symbols
        assert "SSL_connect" in fd.removed_symbols

    def test_file_delta_added_file_symbols(self):
        new_cs = [
            _make_call_site(file_path="src/new.c", symbol="BIO_new",
                            category="crypto_bio", line=5),
        ]
        pd = diff_single(
            {"project": "proj", "call_sites": [], "summary": {}},
            {"project": "proj", "call_sites": new_cs, "summary": {}},
        )
        assert len(pd.file_delta) == 1
        fd = pd.file_delta[0]
        assert fd.status == DiffStatus.ADDED
        assert fd.old_symbols == 0
        assert fd.new_symbols == 1
        assert fd.added_symbols == ["BIO_new"]
        assert fd.removed_symbols == []

    def test_file_delta_json_includes_symbol_fields(self, tmp_path):
        old_cs = [_make_call_site(symbol="SSL_CTX_new", line=10)]
        new_cs = [
            _make_call_site(symbol="SSL_CTX_new", line=10),
            _make_call_site(symbol="SSL_connect", caller="do_conn", line=20),
        ]
        pd = diff_single(
            {"project": "proj", "call_sites": old_cs, "summary": {}},
            {"project": "proj", "call_sites": new_cs, "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json", projects=[pd])
        out_path = os.path.join(str(tmp_path), "diff.json")
        SourceDiffJsonExporter().export(dr, out_path)

        with open(out_path, "r") as f:
            data = json.load(f)
        fd_list = data["file_delta"]
        assert len(fd_list) > 0
        assert "old_symbols" in fd_list[0]
        assert "added_symbols" in fd_list[0]

    def test_xlsx_file_sheet_symbol_columns(self, tmp_path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        old_cs = [_make_call_site(symbol="DES_ecb_encrypt",
                                  category="crypto_legacy", line=10)]
        new_cs = [
            _make_call_site(symbol="EVP_EncryptInit",
                            caller="do_encrypt", line=20,
                            category="crypto_evp"),
            _make_call_site(symbol="EVP_EncryptUpdate",
                            caller="do_encrypt", line=25,
                            category="crypto_evp"),
        ]

        pd = diff_single(
            {"project": "proj", "call_sites": old_cs, "summary": {}},
            {"project": "proj", "call_sites": new_cs, "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json", projects=[pd])
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = load_workbook(out_path)
        ws = wb["File Delta"]
        for row in range(2, ws.max_row + 1):
            fp = ws.cell(row=row, column=1).value
            if fp == "src/tls.c":
                old_syms = ws.cell(row=row, column=6).value
                new_syms = ws.cell(row=row, column=7).value
                added = ws.cell(row=row, column=8).value
                removed = ws.cell(row=row, column=9).value
                assert old_syms == 1
                assert new_syms == 2
                assert "EVP_EncryptInit" in (added or "")
                assert "DES_ecb_encrypt" in (removed or "")
                break
        else:
            pytest.fail("src/tls.c not found in File Delta sheet")


class TestProjectDeltaSymbolColumns:

    def test_xlsx_project_sheet_added_removed(self, tmp_path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        cs_curl_old = _make_call_site(file_path="lib/tls.c",
                                      symbol="SSL_CTX_new", line=10)
        cs_curl_new = _make_call_site(file_path="lib/tls.c",
                                      symbol="SSL_connect", caller="do_ssl",
                                      line=20)
        cs_nginx = _make_call_site(file_path="src/ssl.c",
                                   symbol="EVP_DigestInit",
                                   caller="hash_it", line=5,
                                   category="crypto_evp")

        pd1 = diff_single(
            {"project": "curl", "call_sites": [cs_curl_old], "summary": {}},
            {"project": "curl", "call_sites": [cs_curl_new], "summary": {}},
        )
        pd2 = diff_single(
            {"project": "nginx", "call_sites": [], "summary": {}},
            {"project": "nginx", "call_sites": [cs_nginx], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd1, pd2], is_combo=True)
        out_path = os.path.join(str(tmp_path), "combo.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = load_workbook(out_path)
        ws = wb["Project Delta"]
        for row in range(2, ws.max_row + 1):
            proj = ws.cell(row=row, column=1).value
            if proj == "nginx":
                added = ws.cell(row=row, column=8).value
                assert "EVP_DigestInit" in (added or "")
                break
        else:
            pytest.fail("nginx not found in Project Delta sheet")


class TestConsoleFilesChanged:

    def test_format_shows_changed_files(self):
        old_cs = [
            _make_call_site(symbol="SSL_CTX_new", line=10),
        ]
        new_cs = [
            _make_call_site(symbol="SSL_CTX_new", line=10),
            _make_call_site(symbol="SSL_connect", caller="do_conn", line=20),
        ]
        pd = diff_single(
            {"project": "proj", "call_sites": old_cs, "summary": {}},
            {"project": "proj", "call_sites": new_cs, "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json", projects=[pd])
        output = format_console(dr)
        assert "Files Changed" in output
        assert "src/tls.c" in output
        assert "1 -> 2 calls" in output


class TestJsonMetaScanTime:

    def test_scan_time_propagated(self, tmp_path):
        dr = DiffResult(
            old_label="old.json", new_label="new.json",
            old_scan_time="2026-02-15T10:00:00",
            new_scan_time="2026-03-01T10:00:00",
        )
        out_path = os.path.join(str(tmp_path), "meta.json")
        SourceDiffJsonExporter().export(dr, out_path)

        with open(out_path, "r") as f:
            data = json.load(f)
        assert data["meta"]["old_scan_time"] == "2026-02-15T10:00:00"
        assert data["meta"]["new_scan_time"] == "2026-03-01T10:00:00"

    def test_scan_time_absent_when_empty(self, tmp_path):
        dr = DiffResult(old_label="old.json", new_label="new.json")
        out_path = os.path.join(str(tmp_path), "meta.json")
        SourceDiffJsonExporter().export(dr, out_path)

        with open(out_path, "r") as f:
            data = json.load(f)
        assert "old_scan_time" not in data["meta"]
        assert "new_scan_time" not in data["meta"]

    def test_load_report_extracts_scan_time(self, tmp_path):
        report = _make_source_scan_json(call_sites=[])
        path = _write_json(tmp_path, "report.json", report)
        loaded = load_report(path)
        assert loaded["scan_time"] == "2026-03-01T00:00:00"


class TestExcelExporterAutoFilter:

    def _load_wb(self, path):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook
        return load_workbook(path)

    def _make_single_diff(self):
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=10)
        new_cs = _make_call_site(symbol="SSL_connect", caller="do_connect", line=20)
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs],
             "summary": {"total_files_scanned": 5, "files_with_calls": 1,
                         "total_call_sites": 1}},
            {"project": "proj", "call_sites": [old_cs, new_cs],
             "summary": {"total_files_scanned": 7, "files_with_calls": 1,
                         "total_call_sites": 2}},
        )
        return DiffResult(old_label="old.json", new_label="new.json", projects=[pd])

    def _make_combo_diff(self):
        cs_curl = _make_call_site(file_path="lib/tls.c", symbol="SSL_CTX_new", line=10)
        cs_nginx = _make_call_site(file_path="src/ssl.c", symbol="SSL_connect",
                                   caller="do_ssl", line=5)
        pd1 = diff_single(
            {"project": "curl", "call_sites": [cs_curl],
             "summary": {"total_files_scanned": 10, "files_with_calls": 1,
                         "total_call_sites": 1}},
            {"project": "curl", "call_sites": [cs_curl],
             "summary": {"total_files_scanned": 10, "files_with_calls": 1,
                         "total_call_sites": 1}},
            include_unchanged=True,
        )
        pd2 = diff_single(
            {"project": "nginx", "call_sites": [],
             "summary": {"total_files_scanned": 0, "files_with_calls": 0,
                         "total_call_sites": 0}},
            {"project": "nginx", "call_sites": [cs_nginx],
             "summary": {"total_files_scanned": 5, "files_with_calls": 1,
                         "total_call_sites": 1}},
        )
        return DiffResult(old_label="old.json", new_label="new.json",
                          projects=[pd1, pd2], is_combo=True)

    def _assert_auto_filter(self, ws, expected_col, sheet_label=""):
        """Assert auto_filter ref starts at A1, ends at expected column, row >= 2."""
        import re
        ref = ws.auto_filter.ref
        assert ref is not None, f"{sheet_label} auto_filter.ref should not be None"
        m = re.match(r'^A1:([A-Z]+)(\d+)$', ref)
        assert m, f"{sheet_label} auto_filter.ref has unexpected format: {ref}"
        assert m.group(1) == expected_col, (
            f"{sheet_label} expected column {expected_col}, got {m.group(1)}")
        assert int(m.group(2)) >= 2, (
            f"{sheet_label} expected row >= 2, got {m.group(2)}")

    def test_auto_filter_single_summary(self, tmp_path):
        """Summary Delta sheet has auto_filter set."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        self._assert_auto_filter(wb["Summary Delta"], "D", "Summary Delta")

    def test_auto_filter_single_file(self, tmp_path):
        """File Delta sheet has auto_filter with 9 columns (A-I)."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        self._assert_auto_filter(wb["File Delta"], "I", "File Delta")

    def test_auto_filter_single_callsite(self, tmp_path):
        """Call Site Delta (single) has auto_filter with 9 columns (A-I)."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        self._assert_auto_filter(wb["Call Site Delta"], "I", "Call Site Delta")

    def test_auto_filter_single_symbol(self, tmp_path):
        """Symbol Delta sheet has auto_filter with 6 columns (A-F)."""
        dr = self._make_single_diff()
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        self._assert_auto_filter(wb["Symbol Delta"], "F", "Symbol Delta")

    def test_auto_filter_combo_callsite(self, tmp_path):
        """Call Site Delta (combo) has auto_filter with 10 columns (A-J)."""
        dr = self._make_combo_diff()
        out_path = os.path.join(str(tmp_path), "combo.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        self._assert_auto_filter(wb["Call Site Delta"], "J", "Call Site Delta")

    def test_auto_filter_combo_project(self, tmp_path):
        """Project Delta sheet has auto_filter with 9 columns (A-I)."""
        dr = self._make_combo_diff()
        out_path = os.path.join(str(tmp_path), "combo.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        self._assert_auto_filter(wb["Project Delta"], "I", "Project Delta")

    def test_auto_filter_empty_diff(self, tmp_path):
        """Empty diff (no data rows) should not set auto_filter."""
        pd = diff_single(
            {"project": "proj", "call_sites": [],
             "summary": {"total_files_scanned": 0, "files_with_calls": 0,
                         "total_call_sites": 0}},
            {"project": "proj", "call_sites": [],
             "summary": {"total_files_scanned": 0, "files_with_calls": 0,
                         "total_call_sites": 0}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json", projects=[pd])
        out_path = os.path.join(str(tmp_path), "empty.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)

        wb = self._load_wb(out_path)
        for sheet_name in ["Symbol Delta", "File Delta", "Call Site Delta"]:
            ws = wb[sheet_name]
            assert ws.auto_filter.ref is None, (
                f"{sheet_name} should not have auto_filter when empty"
            )


class TestArgDelta:

    def _make_entry(self, line=10, args="(NULL)"):
        return {"line_number": line, "call_args": args}

    def test_exact_match_unchanged(self):
        old = [self._make_entry(10, "(NULL)")]
        new = [self._make_entry(10, "(NULL)")]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.UNCHANGED
        assert result[0].old_line == 10
        assert result[0].new_line == 10
        assert result[0].old_args == "(NULL)"
        assert result[0].new_args == "(NULL)"

    def test_same_line_different_args(self):
        old = [self._make_entry(42, "(TLSv1_2_method())")]
        new = [self._make_entry(42, "(TLS_method())")]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.CHANGED
        assert result[0].old_args == "(TLSv1_2_method())"
        assert result[0].new_args == "(TLS_method())"

    def test_same_args_different_line(self):
        old = [self._make_entry(10, "(NULL)")]
        new = [self._make_entry(20, "(NULL)")]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.MOVED
        assert result[0].old_line == 10
        assert result[0].new_line == 20

    def test_nearest_line_fallback(self):
        old = [self._make_entry(10, "(EVP_sha1())")]
        new = [self._make_entry(15, "(EVP_sha256())")]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.CHANGED
        assert result[0].old_line == 10
        assert result[0].new_line == 15
        assert result[0].old_args == "(EVP_sha1())"
        assert result[0].new_args == "(EVP_sha256())"

    def test_added_call_within_group(self):
        old = []
        new = [self._make_entry(10, "(NULL)")]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.ADDED
        assert result[0].new_line == 10
        assert result[0].old_line is None

    def test_removed_call_within_group(self):
        old = [self._make_entry(10, "(NULL)")]
        new = []
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.REMOVED
        assert result[0].old_line == 10
        assert result[0].new_line is None

    def test_mixed_group_matching(self):
        old = [
            self._make_entry(10, "(NULL)"),
            self._make_entry(20, "(EVP_sha1())"),
            self._make_entry(30, "(ctx)"),
        ]
        new = [
            self._make_entry(10, "(NULL)"),
            self._make_entry(20, "(EVP_sha256())"),
            self._make_entry(40, "(ctx)"),
            self._make_entry(50, "(new_arg)"),
        ]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 4
        statuses = {ad.status for ad in result}
        assert DiffStatus.UNCHANGED in statuses
        assert DiffStatus.CHANGED in statuses
        assert DiffStatus.MOVED in statuses
        assert DiffStatus.ADDED in statuses

    def test_group_status_all_unchanged(self):
        deltas = [
            ArgDelta(status=DiffStatus.UNCHANGED, old_line=10, new_line=10),
            ArgDelta(status=DiffStatus.UNCHANGED, old_line=20, new_line=20),
        ]
        assert _derive_group_status(deltas) == DiffStatus.UNCHANGED

    def test_group_status_moved_only(self):
        deltas = [
            ArgDelta(status=DiffStatus.UNCHANGED, old_line=10, new_line=10),
            ArgDelta(status=DiffStatus.MOVED, old_line=20, new_line=30),
        ]
        assert _derive_group_status(deltas) == DiffStatus.MOVED

    def test_group_status_changed_args(self):
        deltas = [
            ArgDelta(status=DiffStatus.UNCHANGED, old_line=10, new_line=10),
            ArgDelta(status=DiffStatus.CHANGED, old_line=20, new_line=20,
                     old_args="(EVP_sha1())", new_args="(EVP_sha256())"),
        ]
        assert _derive_group_status(deltas) == DiffStatus.CHANGED

    def test_group_status_changed_count(self):
        deltas = [
            ArgDelta(status=DiffStatus.UNCHANGED, old_line=10, new_line=10),
            ArgDelta(status=DiffStatus.ADDED, new_line=20),
        ]
        assert _derive_group_status(deltas) == DiffStatus.CHANGED

    def test_json_has_arg_deltas(self, tmp_path):
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLSv1_2_method())")
        new_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLS_method())")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd])
        out_path = os.path.join(str(tmp_path), "diff.json")
        SourceDiffJsonExporter().export(dr, out_path)
        with open(out_path) as f:
            data = json.load(f)
        csd_list = data["call_site_delta"]
        assert len(csd_list) == 1
        assert "arg_deltas" in csd_list[0]
        ads = csd_list[0]["arg_deltas"]
        assert len(ads) == 1
        assert ads[0]["status"] == "changed"
        assert ads[0]["old_args"] == "(TLSv1_2_method())"
        assert ads[0]["new_args"] == "(TLS_method())"

    def test_json_backward_compat(self, tmp_path):
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=10,
                                 args="(TLSv1_2_method())")
        new_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLS_method())")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd])
        out_path = os.path.join(str(tmp_path), "diff.json")
        SourceDiffJsonExporter().export(dr, out_path)
        with open(out_path) as f:
            data = json.load(f)
        csd = data["call_site_delta"][0]
        assert "old_count" in csd
        assert "new_count" in csd
        assert "old_lines" in csd
        assert "new_lines" in csd

    def test_xlsx_new_column_headers(self, tmp_path):
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLSv1_2_method())")
        new_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLS_method())")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd])
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb["Call Site Delta"]
        header = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert header == ["Status", "File Path", "Caller Function",
                          "OpenSSL Symbol", "Category", "Old Line",
                          "New Line", "Old Args", "New Args"]

    def test_xlsx_per_call_site_rows(self, tmp_path):
        old_cs1 = _make_call_site(symbol="SSL_CTX_new", line=10, args="(NULL)")
        old_cs2 = _make_call_site(symbol="SSL_CTX_new", line=20, args="(NULL)")
        new_cs1 = _make_call_site(symbol="SSL_CTX_new", line=10, args="(NULL)")
        new_cs2 = _make_call_site(symbol="SSL_CTX_new", line=20,
                                  args="(TLS_method())")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs1, old_cs2], "summary": {}},
            {"project": "proj", "call_sites": [new_cs1, new_cs2], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd])
        out_path = os.path.join(str(tmp_path), "diff.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb["Call Site Delta"]
        data_rows = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value:
                data_rows += 1
        assert data_rows == 2

    def test_console_shows_args_changed(self):
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLSv1_2_method())")
        new_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLS_method())")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd])
        text = format_console(dr)
        assert "Args Changed" in text
        assert "(TLSv1_2_method())" in text
        assert "(TLS_method())" in text

    def test_empty_args_handled(self):
        old = [{"line_number": 10}]
        new = [{"line_number": 10}]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.UNCHANGED
        assert result[0].old_args == ""
        assert result[0].new_args == ""

    def test_both_empty_lists(self):
        result = _match_call_sites_within_group([], [])
        assert len(result) == 0
        assert _derive_group_status(result) == DiffStatus.UNCHANGED

    def test_duplicate_same_line_same_args(self):
        old = [self._make_entry(10, "(NULL)"), self._make_entry(10, "(NULL)")]
        new = [self._make_entry(10, "(NULL)"), self._make_entry(10, "(NULL)")]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 2
        assert all(ad.status == DiffStatus.UNCHANGED for ad in result)
        assert _derive_group_status(result) == DiffStatus.UNCHANGED

    def test_five_old_zero_new(self):
        old = [self._make_entry(i * 10, f"(arg{i})") for i in range(5)]
        new = []
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 5
        assert all(ad.status == DiffStatus.REMOVED for ad in result)
        assert _derive_group_status(result) == DiffStatus.REMOVED

    def test_zero_old_five_new(self):
        old = []
        new = [self._make_entry(i * 10, f"(arg{i})") for i in range(5)]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 5
        assert all(ad.status == DiffStatus.ADDED for ad in result)
        assert _derive_group_status(result) == DiffStatus.ADDED

    def test_tls_upgrade_real_scenario(self):
        """Real-world: TLS version upgrade changes method argument."""
        old = [
            self._make_entry(42, "(TLSv1_2_client_method())"),
            self._make_entry(85, "(NULL)"),
        ]
        new = [
            self._make_entry(42, "(TLS_client_method())"),
            self._make_entry(85, "(NULL)"),
        ]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 2
        unchanged = [ad for ad in result if ad.status == DiffStatus.UNCHANGED]
        changed = [ad for ad in result if ad.status == DiffStatus.CHANGED]
        assert len(unchanged) == 1
        assert unchanged[0].old_line == 85
        assert len(changed) == 1
        assert changed[0].old_line == 42
        assert changed[0].old_args == "(TLSv1_2_client_method())"
        assert changed[0].new_args == "(TLS_client_method())"
        assert _derive_group_status(result) == DiffStatus.CHANGED

    def test_hash_algorithm_migration_scenario(self):
        """Real-world: SHA-1 to SHA-256 migration across multiple call sites."""
        old = [
            self._make_entry(10, "(EVP_sha1())"),
            self._make_entry(20, "(EVP_sha1())"),
            self._make_entry(30, "(EVP_sha256())"),
        ]
        new = [
            self._make_entry(10, "(EVP_sha256())"),
            self._make_entry(20, "(EVP_sha256())"),
            self._make_entry(30, "(EVP_sha256())"),
        ]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 3
        changed = [ad for ad in result if ad.status == DiffStatus.CHANGED]
        unchanged = [ad for ad in result if ad.status == DiffStatus.UNCHANGED]
        assert len(changed) == 2
        assert len(unchanged) == 1
        assert unchanged[0].old_line == 30
        for ad in changed:
            assert ad.old_args == "(EVP_sha1())"
            assert ad.new_args == "(EVP_sha256())"

    def test_cipher_suite_string_change(self):
        """Real-world: cipher suite string update."""
        old = [self._make_entry(100,
               '("ECDHE-RSA-AES128-GCM-SHA256:AES256-SHA")')]
        new = [self._make_entry(100,
               '("TLS_AES_128_GCM_SHA256:TLS_CHACHA20_POLY1305_SHA256")')]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.CHANGED
        assert "ECDHE" in result[0].old_args
        assert "TLS_AES" in result[0].new_args

    def test_code_refactor_moves_call_with_same_args(self):
        """Real-world: function body reorganized, calls moved to new lines."""
        old = [
            self._make_entry(10, "(ctx, EVP_sha256(), NULL)"),
            self._make_entry(25, "(ctx, EVP_sha256(), NULL)"),
        ]
        new = [
            self._make_entry(50, "(ctx, EVP_sha256(), NULL)"),
            self._make_entry(75, "(ctx, EVP_sha256(), NULL)"),
        ]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 2
        assert all(ad.status == DiffStatus.MOVED for ad in result)
        assert _derive_group_status(result) == DiffStatus.MOVED

    def test_e2e_diff_single_with_args_change(self, tmp_path):
        """End-to-end: diff_single produces correct arg_deltas in CallSiteDelta."""
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLSv1_2_method())")
        new_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLS_method())")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        assert len(pd.call_site_delta) == 1
        csd = pd.call_site_delta[0]
        assert csd.status == DiffStatus.CHANGED
        assert len(csd.arg_deltas) == 1
        assert csd.arg_deltas[0].status == DiffStatus.CHANGED
        assert csd.arg_deltas[0].old_args == "(TLSv1_2_method())"
        assert csd.arg_deltas[0].new_args == "(TLS_method())"

    def test_e2e_unchanged_excluded_from_diff_single(self):
        """End-to-end: identical call sites excluded from diff_single by default."""
        cs = _make_call_site(symbol="SSL_CTX_new", line=10, args="(NULL)")
        pd = diff_single(
            {"project": "proj", "call_sites": [cs], "summary": {}},
            {"project": "proj", "call_sites": [cs], "summary": {}},
        )
        assert len(pd.call_site_delta) == 0

    def test_e2e_unchanged_included_when_requested(self):
        """End-to-end: identical call sites included with include_unchanged=True."""
        cs = _make_call_site(symbol="SSL_CTX_new", line=10, args="(NULL)")
        pd = diff_single(
            {"project": "proj", "call_sites": [cs], "summary": {}},
            {"project": "proj", "call_sites": [cs], "summary": {}},
            include_unchanged=True,
        )
        assert len(pd.call_site_delta) == 1
        csd = pd.call_site_delta[0]
        assert csd.status == DiffStatus.UNCHANGED
        assert len(csd.arg_deltas) == 1
        assert csd.arg_deltas[0].status == DiffStatus.UNCHANGED

    def test_combo_xlsx_with_arg_deltas(self, tmp_path):
        """Combo mode XLSX produces correct per-call-site rows with Project column."""
        old_cs = _make_call_site(file_path="lib/tls.c", symbol="SSL_CTX_new",
                                 line=10, args="(TLSv1_2_method())")
        new_cs = _make_call_site(file_path="lib/tls.c", symbol="SSL_CTX_new",
                                 line=10, args="(TLS_method())")
        pd = diff_single(
            {"project": "curl", "call_sites": [old_cs],
             "summary": {"total_files_scanned": 5, "files_with_calls": 1,
                         "total_call_sites": 1}},
            {"project": "curl", "call_sites": [new_cs],
             "summary": {"total_files_scanned": 5, "files_with_calls": 1,
                         "total_call_sites": 1}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd], is_combo=True)
        out_path = os.path.join(str(tmp_path), "combo.xlsx")
        SourceDiffExcelExporter().export(dr, out_path)
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb["Call Site Delta"]
        header = [ws.cell(row=1, column=c).value for c in range(1, 11)]
        assert header == ["Project", "Status", "File Path", "Caller Function",
                          "OpenSSL Symbol", "Category", "Old Line", "New Line",
                          "Old Args", "New Args"]
        assert ws.cell(row=2, column=1).value == "curl"
        assert ws.cell(row=2, column=2).value == "changed"
        assert ws.cell(row=2, column=9).value == "(TLSv1_2_method())"
        assert ws.cell(row=2, column=10).value == "(TLS_method())"

    def test_console_no_args_changed_for_unchanged(self):
        """Console output does not show Args Changed when no args actually differ."""
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=10, args="(NULL)")
        new_cs = _make_call_site(symbol="SSL_CTX_new", line=20, args="(NULL)")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd])
        text = format_console(dr)
        assert "Args Changed" not in text

    def test_json_combo_has_arg_deltas(self, tmp_path):
        """Combo diff JSON includes arg_deltas in each project's call_site_delta."""
        old_cs = _make_call_site(file_path="lib/tls.c", symbol="SSL_CTX_new",
                                 line=42, args="(TLSv1_2_method())")
        new_cs = _make_call_site(file_path="lib/tls.c", symbol="SSL_CTX_new",
                                 line=42, args="(TLS_method())")
        pd = diff_single(
            {"project": "curl", "call_sites": [old_cs], "summary": {}},
            {"project": "curl", "call_sites": [new_cs], "summary": {}},
        )
        dr = DiffResult(old_label="old.json", new_label="new.json",
                        projects=[pd], is_combo=True)
        out_path = os.path.join(str(tmp_path), "combo.json")
        SourceDiffJsonExporter().export(dr, out_path)
        with open(out_path) as f:
            data = json.load(f)
        proj = data["projects"][0]
        assert "arg_deltas" in proj["call_site_delta"][0]
        ad = proj["call_site_delta"][0]["arg_deltas"][0]
        assert ad["status"] == "changed"
        assert ad["old_args"] == "(TLSv1_2_method())"

    def test_same_line_two_different_calls(self):
        """Multiple calls on same line with different args (macro expansion).

        old: (10,"sha1"), (10,"sha256")
        new: (10,"sha256"), (10,"sha512")
        Phase 1: old[1] (10,"sha256") matches new[0] (10,"sha256") -> UNCHANGED
        Phase 2: old[0] (10,"sha1") matches new[1] (10,"sha512") same line -> CHANGED
        """
        old = [
            self._make_entry(10, "(ctx, EVP_sha1())"),
            self._make_entry(10, "(ctx, EVP_sha256())"),
        ]
        new = [
            self._make_entry(10, "(ctx, EVP_sha256())"),
            self._make_entry(10, "(ctx, EVP_sha512())"),
        ]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 2
        unchanged = [ad for ad in result if ad.status == DiffStatus.UNCHANGED]
        changed = [ad for ad in result if ad.status == DiffStatus.CHANGED]
        assert len(unchanged) == 1
        assert unchanged[0].old_args == "(ctx, EVP_sha256())"
        assert len(changed) == 1
        assert changed[0].old_args == "(ctx, EVP_sha1())"
        assert changed[0].new_args == "(ctx, EVP_sha512())"

    def test_result_sorted_by_line(self):
        """ArgDelta results are sorted by line number for deterministic output."""
        old = [
            self._make_entry(50, "(a)"),
            self._make_entry(10, "(b)"),
            self._make_entry(30, "(c)"),
        ]
        new = [
            self._make_entry(10, "(b)"),
            self._make_entry(30, "(c)"),
            self._make_entry(50, "(d)"),
        ]
        result = _match_call_sites_within_group(old, new)
        lines = []
        for ad in result:
            line = ad.new_line if ad.new_line is not None else ad.old_line
            lines.append(line)
        assert lines == sorted(lines), f"Results not sorted by line: {lines}"

    def test_phase4_different_line_different_args(self):
        """Phase 4: both line and args differ, nearest line pairing."""
        old = [
            self._make_entry(10, "(EVP_sha1())"),
            self._make_entry(100, "(EVP_md5())"),
        ]
        new = [
            self._make_entry(12, "(EVP_sha256())"),
            self._make_entry(98, "(EVP_sha3_256())"),
        ]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 2
        assert all(ad.status == DiffStatus.CHANGED for ad in result)
        matched_pairs = {(ad.old_line, ad.new_line) for ad in result}
        assert (10, 12) in matched_pairs
        assert (100, 98) in matched_pairs

    def test_none_line_number_handled(self):
        """line_number=None preserved in output, normalized to 0 only for matching."""
        old = [{"line_number": None, "call_args": "(NULL)"}]
        new = [{"line_number": 10, "call_args": "(NULL)"}]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.MOVED
        assert result[0].old_line is None
        assert result[0].new_line == 10

    def test_none_call_args_handled(self):
        """call_args=None should not crash the matcher."""
        old = [{"line_number": 10, "call_args": None}]
        new = [{"line_number": 10, "call_args": None}]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.UNCHANGED
        assert result[0].old_args == ""

    def test_duplicate_count_mismatch_same_line(self):
        """2 identical calls on old side, 1 on new side: 1 UNCHANGED + 1 REMOVED."""
        old = [
            self._make_entry(10, "(NULL)"),
            self._make_entry(10, "(NULL)"),
        ]
        new = [
            self._make_entry(10, "(NULL)"),
        ]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 2
        statuses = [ad.status for ad in result]
        assert statuses.count(DiffStatus.UNCHANGED) == 1
        assert statuses.count(DiffStatus.REMOVED) == 1

    def test_duplicate_count_mismatch_reverse(self):
        """1 call on old side, 2 identical calls on new side: 1 UNCHANGED + 1 ADDED."""
        old = [
            self._make_entry(10, "(NULL)"),
        ]
        new = [
            self._make_entry(10, "(NULL)"),
            self._make_entry(10, "(NULL)"),
        ]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 2
        statuses = [ad.status for ad in result]
        assert statuses.count(DiffStatus.UNCHANGED) == 1
        assert statuses.count(DiffStatus.ADDED) == 1

    def test_derive_group_status_moved_and_added(self):
        """MOVED + ADDED -> CHANGED group status."""
        deltas = [
            ArgDelta(status=DiffStatus.MOVED, old_line=10, new_line=20),
            ArgDelta(status=DiffStatus.ADDED, new_line=30),
        ]
        assert _derive_group_status(deltas) == DiffStatus.CHANGED

    def test_derive_group_status_moved_and_removed(self):
        """MOVED + REMOVED -> CHANGED group status."""
        deltas = [
            ArgDelta(status=DiffStatus.MOVED, old_line=10, new_line=20),
            ArgDelta(status=DiffStatus.REMOVED, old_line=30),
        ]
        assert _derive_group_status(deltas) == DiffStatus.CHANGED

    def test_semantic_improvement_same_line_diff_args_is_changed(self):
        """Key semantic improvement: same line + different args = CHANGED (not UNCHANGED).

        Previously, status was derived only from count/lines:
            same count + same lines => UNCHANGED (even if args differed!)
        Now status is derived from per-call-site matching:
            same line + different args => Phase 2 CHANGED
        """
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLSv1_2_method())")
        new_cs = _make_call_site(symbol="SSL_CTX_new", line=42,
                                 args="(TLS_method())")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        assert len(pd.call_site_delta) == 1
        assert pd.call_site_delta[0].status == DiffStatus.CHANGED

    def test_semantic_diff_line_moved_args_changed_is_changed(self):
        """Different line + different args = CHANGED (not MOVED).

        Previously: same count + different lines => MOVED (ignoring args).
        Now: different line + different args => Phase 4 CHANGED (correct).
        """
        old_cs = _make_call_site(symbol="EVP_DigestInit", line=10,
                                 args="(ctx, EVP_sha1())")
        new_cs = _make_call_site(symbol="EVP_DigestInit", line=20,
                                 args="(ctx, EVP_sha256())")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        assert len(pd.call_site_delta) == 1
        assert pd.call_site_delta[0].status == DiffStatus.CHANGED

    def test_semantic_diff_line_moved_args_same_is_moved(self):
        """Different line + same args = MOVED (preserved from old behavior)."""
        old_cs = _make_call_site(symbol="SSL_CTX_new", line=10,
                                 args="(TLS_method())")
        new_cs = _make_call_site(symbol="SSL_CTX_new", line=50,
                                 args="(TLS_method())")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        assert len(pd.call_site_delta) == 1
        assert pd.call_site_delta[0].status == DiffStatus.MOVED

    def test_code_shift_all_lines_move_by_n(self):
        """When all lines shift by N but args stay the same, all should be MOVED.

        Scenario: a header include is added, shifting all lines down by 5.
        Before the Phase 2/3 swap, this incorrectly produced CHANGED because
        Phase 2 (same-line) matched cross-pairs before Phase 3 (same-args).
        """
        old_entries = [
            {"line_number": 10, "call_args": "(TLS_method())"},
            {"line_number": 20, "call_args": "(ctx)"},
            {"line_number": 30, "call_args": "(ctx, cert_file, SSL_FILETYPE_PEM)"},
        ]
        new_entries = [
            {"line_number": 15, "call_args": "(TLS_method())"},
            {"line_number": 25, "call_args": "(ctx)"},
            {"line_number": 35, "call_args": "(ctx, cert_file, SSL_FILETYPE_PEM)"},
        ]
        result = _match_call_sites_within_group(old_entries, new_entries)
        assert len(result) == 3
        for ad in result:
            assert ad.status == DiffStatus.MOVED, (
                f"Expected MOVED for line {ad.old_line}->{ad.new_line}, "
                f"got {ad.status.value}"
            )
        assert result[0].old_line == 10 and result[0].new_line == 15
        assert result[1].old_line == 20 and result[1].new_line == 25
        assert result[2].old_line == 30 and result[2].new_line == 35

    def test_phase_priority_args_over_line(self):
        """Phase 2 (same-args) takes priority over Phase 3 (same-line).

        old: line 10 arg_A, line 20 arg_B
        new: line 20 arg_A, line 10 arg_B

        If line-matching ran first: (10,arg_A)-(10,arg_B)=CHANGED, (20,arg_B)-(20,arg_A)=CHANGED
        With args-matching first:   (10,arg_A)-(20,arg_A)=MOVED,   (20,arg_B)-(10,arg_B)=MOVED
        """
        old_entries = [
            {"line_number": 10, "call_args": "(arg_A)"},
            {"line_number": 20, "call_args": "(arg_B)"},
        ]
        new_entries = [
            {"line_number": 20, "call_args": "(arg_A)"},
            {"line_number": 10, "call_args": "(arg_B)"},
        ]
        result = _match_call_sites_within_group(old_entries, new_entries)
        assert len(result) == 2
        statuses = {ad.status for ad in result}
        assert statuses == {DiffStatus.MOVED}, (
            f"Expected all MOVED, got {[ad.status.value for ad in result]}"
        )

    def test_mixed_changed_and_moved_groups_xlsx(self):
        """XLSX correctly renders mixed CHANGED + MOVED groups as per-call-site rows."""
        old_cs = [
            _make_call_site(symbol="SSL_CTX_new", line=10, args="(TLSv1_2_method())"),
            _make_call_site(symbol="EVP_DigestInit", line=50,
                            args="(ctx, EVP_sha256())", category="crypto_evp"),
            _make_call_site(symbol="EVP_DigestInit", line=60,
                            args="(ctx2, EVP_sha256())", category="crypto_evp"),
        ]
        new_cs = [
            _make_call_site(symbol="SSL_CTX_new", line=10, args="(TLS_method())"),
            _make_call_site(symbol="EVP_DigestInit", line=55,
                            args="(ctx, EVP_sha256())", category="crypto_evp"),
            _make_call_site(symbol="EVP_DigestInit", line=65,
                            args="(ctx2, EVP_sha256())", category="crypto_evp"),
        ]
        pd = diff_single(
            {"project": "p", "call_sites": old_cs, "summary": {}},
            {"project": "p", "call_sites": new_cs, "summary": {}},
        )
        # SSL_CTX_new: CHANGED (args differ), EVP_DigestInit: MOVED (lines shifted)
        by_sym = {csd.identity_key[2]: csd for csd in pd.call_site_delta}
        assert by_sym["SSL_CTX_new"].status == DiffStatus.CHANGED
        assert by_sym["EVP_DigestInit"].status == DiffStatus.MOVED

        dr = DiffResult(projects=[pd], old_label="v1", new_label="v2")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            xlsx_path = f.name
        try:
            SourceDiffExcelExporter().export(dr, xlsx_path)
            from openssl_scanner._vendor.openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb["Call Site Delta"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            assert len(rows) == 3
            statuses = [r[0] for r in rows]
            assert "changed" in statuses
            assert "moved" in statuses
        finally:
            os.unlink(xlsx_path)

    def test_added_removed_group_derive_changed(self):
        """Group with only ADDED + REMOVED arg_deltas derives to CHANGED."""
        ads = [
            ArgDelta(status=DiffStatus.ADDED, new_line=10, new_args="(X)"),
            ArgDelta(status=DiffStatus.REMOVED, old_line=20, old_args="(Y)"),
        ]
        assert _derive_group_status(ads) == DiffStatus.CHANGED

    def test_derive_unchanged_plus_added_is_changed(self):
        """Group with UNCHANGED + ADDED derives to CHANGED, not ADDED."""
        ads = [
            ArgDelta(status=DiffStatus.UNCHANGED, old_line=10, new_line=10,
                     old_args="(A)", new_args="(A)"),
            ArgDelta(status=DiffStatus.ADDED, new_line=20, new_args="(B)"),
        ]
        assert _derive_group_status(ads) == DiffStatus.CHANGED

    def test_derive_unchanged_plus_removed_is_changed(self):
        """Group with UNCHANGED + REMOVED derives to CHANGED, not REMOVED."""
        ads = [
            ArgDelta(status=DiffStatus.UNCHANGED, old_line=10, new_line=10,
                     old_args="(A)", new_args="(A)"),
            ArgDelta(status=DiffStatus.REMOVED, old_line=20, old_args="(B)"),
        ]
        assert _derive_group_status(ads) == DiffStatus.CHANGED

    def test_e2e_unchanged_plus_added_call(self):
        """Adding a second call to same group: 1 UNCHANGED + 1 ADDED = CHANGED."""
        old_cs = [_make_call_site(symbol="SSL_read", line=10, args="(ctx, buf, n)")]
        new_cs = [
            _make_call_site(symbol="SSL_read", line=10, args="(ctx, buf, n)"),
            _make_call_site(symbol="SSL_read", line=20, args="(ctx2, buf2, n2)"),
        ]
        pd = diff_single(
            {"project": "p", "call_sites": old_cs, "summary": {}},
            {"project": "p", "call_sites": new_cs, "summary": {}},
        )
        assert len(pd.call_site_delta) == 1
        csd = pd.call_site_delta[0]
        assert csd.status == DiffStatus.CHANGED
        statuses = [ad.status for ad in csd.arg_deltas]
        assert DiffStatus.UNCHANGED in statuses
        assert DiffStatus.ADDED in statuses

    def test_removed_entries_sort_by_old_line(self):
        """REMOVED arg_deltas (new_line=None) sort by old_line ascending."""
        old_entries = [
            {"line_number": 30, "call_args": "(C)"},
            {"line_number": 10, "call_args": "(A)"},
            {"line_number": 20, "call_args": "(B)"},
        ]
        result = _match_call_sites_within_group(old_entries, [])
        lines = [ad.old_line for ad in result]
        assert lines == [10, 20, 30]
        assert all(ad.status == DiffStatus.REMOVED for ad in result)

    def test_xlsx_formula_injection_in_args(self):
        """XLSX escapes formula-injection chars (= + - @) in arg values."""
        old_cs = _make_call_site(args='=HYPERLINK("http://evil.com")')
        new_cs = _make_call_site(args="+cmd|exit")
        pd = diff_single(
            {"project": "p", "call_sites": [old_cs], "summary": {}},
            {"project": "p", "call_sites": [new_cs], "summary": {}},
        )
        dr = DiffResult(projects=[pd], old_label="v1", new_label="v2")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            xlsx_path = f.name
        try:
            SourceDiffExcelExporter().export(dr, xlsx_path)
            from openssl_scanner._vendor.openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb["Call Site Delta"]
            row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            old_args_val = row[7]
            new_args_val = row[8]
            assert old_args_val.startswith("'"), (
                f"Expected leading quote for = injection: {old_args_val}")
            assert new_args_val.startswith("'"), (
                f"Expected leading quote for + injection: {new_args_val}")
        finally:
            os.unlink(xlsx_path)

    def test_none_line_number_e2e_diff_single(self):
        """None line_number in JSON should not crash sorted() in _compute_call_site_delta."""
        old_cs1 = _make_call_site(symbol="SSL_read", line=None, args="(ssl, buf, len)")
        old_cs2 = _make_call_site(symbol="SSL_read", line=10, args="(ssl, buf, len)")
        new_cs1 = _make_call_site(symbol="SSL_read", line=10, args="(ssl, buf, len)")
        new_cs2 = _make_call_site(symbol="SSL_read", line=20, args="(ssl, buf, len)")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs1, old_cs2], "summary": {}},
            {"project": "proj", "call_sites": [new_cs1, new_cs2], "summary": {}},
        )
        assert len(pd.call_site_delta) == 1
        csd = pd.call_site_delta[0]
        assert len(csd.arg_deltas) == 2

    def test_none_vs_zero_line_collision(self):
        """None and 0 are equivalent in matching but None preserved in output."""
        old = [{"line_number": None, "call_args": "(buf)"}]
        new = [{"line_number": 0, "call_args": "(buf)"}]
        result = _match_call_sites_within_group(old, new)
        assert len(result) == 1
        assert result[0].status == DiffStatus.UNCHANGED
        assert result[0].old_line is None
        assert result[0].new_line == 0

    def test_json_preserves_null_line(self):
        """JSON arg_deltas should have null for None lines, not 0."""
        old_cs = _make_call_site(symbol="SSL_read", line=None, args="(ssl, buf, n)")
        new_cs = _make_call_site(symbol="SSL_read", line=10, args="(ssl, buf, n)")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        from openssl_scanner.source_diff import SourceDiffJsonExporter
        exporter = SourceDiffJsonExporter()
        data = exporter._serialize_project(pd)
        ad = data["call_site_delta"][0]["arg_deltas"][0]
        assert ad["old_line"] is None, f"Expected null, got {ad['old_line']}"
        assert ad["new_line"] == 10

    def test_console_none_line_shows_question_mark(self):
        """Console output should show L? for unknown lines, not LNone."""
        old_cs = _make_call_site(symbol="SSL_read", line=None, args="(old)")
        new_cs = _make_call_site(symbol="SSL_read", line=None, args="(new)")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        from openssl_scanner.source_diff import DiffResult
        dr = DiffResult(
            old_label="old.json", new_label="new.json",
            projects=[pd],
        )
        text = format_console(dr)
        assert "LNone" not in text, f"Should not contain LNone: {text}"
        assert "L?" in text, f"Should contain L? for unknown line: {text}"

    def test_xlsx_none_line_blank_cell(self):
        """XLSX should have blank cell for None line, not 0."""
        import tempfile, os
        old_cs = _make_call_site(symbol="SSL_read", line=None, args="(old)")
        new_cs = _make_call_site(symbol="SSL_read", line=10, args="(new)")
        pd = diff_single(
            {"project": "proj", "call_sites": [old_cs], "summary": {}},
            {"project": "proj", "call_sites": [new_cs], "summary": {}},
        )
        from openssl_scanner.source_diff import DiffResult, SourceDiffExcelExporter
        dr = DiffResult(
            old_label="old.json", new_label="new.json",
            projects=[pd],
        )
        xlsx_path = tempfile.mktemp(suffix=".xlsx")
        try:
            SourceDiffExcelExporter().export(dr, xlsx_path)
            from openssl_scanner._vendor.openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb["Call Site Delta"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                status = row[0]
                old_line = row[5]
                if status == "changed":
                    assert old_line is None or old_line == "", (
                        f"Expected blank for None line, got {old_line}")
                    break
        finally:
            os.unlink(xlsx_path)
