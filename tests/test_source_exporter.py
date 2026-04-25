"""Tests for source scan exporters (XLSX and JSON)."""

# ruff: noqa: I001

import json
import os
import tempfile

from openssl_scanner.source_analyzer import CallSite, SourceScanResult
from openssl_scanner.source_exporter import (
    COLUMNS,
    SourceExcelExporter,
    SourceJsonExporter,
)


def _make_result(call_sites=None):
    if call_sites is None:
        call_sites = []
    unique = sorted({cs.ossl_symbol for cs in call_sites})
    return SourceScanResult(
        target="/tmp/test",
        scan_time="2026-01-01T00:00:00",
        tool_version="1.0.0",
        total_files_scanned=1,
        files_with_calls=1 if call_sites else 0,
        total_call_sites=len(call_sites),
        unique_symbols=unique,
        symbols_by_category={},
        call_sites=call_sites,
        errors=[],
    )


def _make_sites(n):
    sites = []
    for i in range(n):
        sites.append(CallSite(
            file_path=f"/tmp/test/file{i}.c",
            file_name=f"file{i}.c",
            caller_function=f"func{i}",
            line_number=i + 1,
            column=4,
            ossl_symbol="SSL_connect",
            category="ssl_core",
            call_args=f"(arg{i})",
            language="c",
        ))
    return sites


class TestExcelExporter:
    def test_export_with_data(self):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        result = _make_result(_make_sites(5))
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name

        try:
            exporter = SourceExcelExporter()
            exporter.export(result, path)

            wb = load_workbook(path)
            ws = wb.active
            assert ws.title == "OpenSSL Call Sites"
            assert ws.max_row == 6
            assert ws.max_column == len(COLUMNS)

            headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
            assert headers == [
                'File Path', 'File Name', 'Caller Function',
                'Line', 'OpenSSL Symbol', 'Category', 'Call Arguments'
            ]

            assert ws.cell(row=2, column=5).value == "SSL_connect"
        finally:
            os.unlink(path)

    def test_export_empty(self):
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        result = _make_result([])
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name

        try:
            exporter = SourceExcelExporter()
            exporter.export(result, path)

            wb = load_workbook(path)
            ws = wb.active
            assert ws.max_row == 1
        finally:
            os.unlink(path)


class TestJsonExporter:
    def test_export_with_data(self):
        result = _make_result(_make_sites(3))
        exporter = SourceJsonExporter()

        with tempfile.NamedTemporaryFile(suffix='.json', delete=False, mode='w') as f:
            path = f.name

        try:
            json_str = exporter.export(result, path)
            data = json.loads(json_str)

            assert data['meta']['report_type'] == 'source_scan'
            assert data['summary']['total_call_sites'] == 3
            assert len(data['call_sites']) == 3
            assert data['call_sites'][0]['ossl_symbol'] == 'SSL_connect'
        finally:
            os.unlink(path)

    def test_export_returns_string(self):
        result = _make_result([])
        exporter = SourceJsonExporter()
        json_str = exporter.export(result)
        data = json.loads(json_str)
        assert data['summary']['total_call_sites'] == 0
