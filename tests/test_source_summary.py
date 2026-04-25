"""Tests for source-summary report aggregation and export."""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from openssl_scanner import _vendor  # noqa: F401


def _write_manifest(manifest_path: Path) -> None:
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(
        """<?xml version="1.0" encoding="UTF-8"?>
<manifest>
  <remote name="origin" fetch="https://gitcode.com/openharmony" />
  <default remote="origin" />
  <project name="security_crypto_framework" path="base/security/crypto_framework" />
  <project name="communication_request" path="foundation/communication/request" />
</manifest>
""",
        encoding="utf-8",
    )


def _write_source_scan_report(
    report_path: Path,
    target: Path,
    *,
    unique_symbols: list[str],
    call_sites: list[dict],
    files_scanned: int = 10,
    files_with_calls: int = 2,
    total_call_sites: int | None = None,
    symbol_coverage: dict | None = None,
    direct_ratio: float = 0.0,
    partial_ratio: float = 0.0,
    report_type: str = "source_scan",
) -> None:
    if total_call_sites is None:
        total_call_sites = len(call_sites)
    if symbol_coverage is None:
        symbol_coverage = {
            "available": 0,
            "partial": 0,
            "not_available": len(unique_symbols),
            "unknown": 0,
        }

    payload = {
        "meta": {
            "tool_version": "1.0.0",
            "report_type": report_type,
            "scan_time": "2026-04-16T17:00:00",
            "target": str(target),
        },
        "summary": {
            "total_files_scanned": files_scanned,
            "files_with_calls": files_with_calls,
            "total_call_sites": total_call_sites,
            "unique_symbols_count": len(unique_symbols),
            "unique_symbols": unique_symbols,
            "symbols_by_category": {},
            "hitls_coverage": symbol_coverage,
            "hitls_direct_replace_ratio": direct_ratio,
            "hitls_direct_or_partial_replace_ratio": partial_ratio,
        },
        "call_sites": call_sites,
        "errors": [],
    }
    report_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def _sample_call_site(
    symbol: str,
    status: str,
    *,
    category: str = "crypto_evp",
    file_path: str = "/tmp/src/a.c",
    file_name: str = "a.c",
    replacement: str | None = None,
) -> dict:
    return {
        "file_path": file_path,
        "file_name": file_name,
        "caller_function": "main",
        "line_number": 10,
        "column": 2,
        "ossl_symbol": symbol,
        "category": category,
        "call_args": "()",
        "language": "c",
        "detection_method": "call",
        "hitls_status": status,
        "hitls_equiv": replacement,
        "hitls_replacement": replacement,
    }


def test_source_summary_loads_source_scan_and_computes_dual_ratios(tmp_path):
    source_root = tmp_path / "oh-source"
    repo_path = source_root / "base" / "security" / "crypto_framework"
    repo_path.mkdir(parents=True)
    (repo_path / "README.md").write_text(
        "# crypto framework\nCore crypto subsystem.\n",
        encoding="utf-8",
    )
    manifest_path = source_root / ".repo" / "manifest.xml"
    _write_manifest(manifest_path)

    report_dir = tmp_path / "reports"
    report_dir.mkdir()
    _write_source_scan_report(
        report_dir / "crypto_framework.json",
        repo_path,
        unique_symbols=["EVP_MD_CTX_new", "EVP_sha256", "OPENSSL_free"],
        symbol_coverage={
            "available": 1,
            "partial": 1,
            "not_available": 1,
            "unknown": 0,
        },
        direct_ratio=33.33,
        partial_ratio=66.67,
        call_sites=[
            _sample_call_site("EVP_MD_CTX_new", "available"),
            _sample_call_site("EVP_sha256", "partial", replacement="CRYPT_EAL_Md(...)"),
            _sample_call_site("EVP_sha256", "partial", replacement="CRYPT_EAL_Md(...)"),
            _sample_call_site("OPENSSL_free", "not_available"),
            _sample_call_site("OPENSSL_free", "unknown"),
        ],
    )
    _write_source_scan_report(
        report_dir / "ignored.json",
        repo_path,
        unique_symbols=["SSL_new"],
        call_sites=[_sample_call_site("SSL_new", "not_available")],
        report_type="combo_scan",
    )

    from openssl_scanner.source_summary import SourceSummaryBuilder

    builder = SourceSummaryBuilder(
        source_root=str(source_root),
        manifest_path=str(manifest_path),
    )
    rows, blockers = builder.build_from_inputs([str(report_dir)])

    assert len(rows) == 1
    row = rows[0]
    assert row["仓库名称"] == "security_crypto_framework"
    assert row["仓库路径"] == "base/security/crypto_framework"
    assert row["地址"] == "https://gitcode.com/openharmony/security_crypto_framework"
    assert row["HM标签"] == "HM自研组件"
    assert row["使用OSSL接口数量"] == 3
    assert row["OpenSSL调用点数量"] == 5
    assert row["可直接替换接口数"] == 1
    assert row["可部分替换接口数"] == 1
    assert row["不可替换接口数"] == 1
    assert row["直接替换率(符号)"] == 33.33
    assert row["直接+部分替换率(符号)"] == 66.67
    assert row["可直接替换调用点数"] == 1
    assert row["可部分替换调用点数"] == 2
    assert row["不可替换调用点数"] == 1
    assert row["未知调用点数"] == 1
    assert row["直接替换率(调用点)"] == 20.0
    assert row["直接+部分替换率(调用点)"] == 60.0
    assert "OPENSSL_free" in row["Top阻塞接口"]
    assert len(blockers) == 2


def test_source_summary_workbook_and_tsv_include_expected_sheets(tmp_path):
    source_root = tmp_path / "oh-source"
    repo_a = source_root / "foundation" / "communication" / "request"
    repo_b = source_root / "base" / "security" / "crypto_framework"
    repo_a.mkdir(parents=True)
    repo_b.mkdir(parents=True)
    (repo_a / "README.md").write_text("request repo\n", encoding="utf-8")
    (repo_b / "README.md").write_text("crypto repo\n", encoding="utf-8")
    manifest_path = source_root / ".repo" / "manifest.xml"
    _write_manifest(manifest_path)

    report_dir = tmp_path / "reports"
    report_dir.mkdir()
    _write_source_scan_report(
        report_dir / "request.json",
        repo_a,
        unique_symbols=["SHA256_Update", "SHA256_Final"],
        symbol_coverage={
            "available": 0,
            "partial": 1,
            "not_available": 1,
            "unknown": 0,
        },
        direct_ratio=0.0,
        partial_ratio=50.0,
        call_sites=[
            _sample_call_site("SHA256_Update", "partial", replacement="CRYPT_EAL_Md(...)"),
            _sample_call_site("SHA256_Final", "not_available"),
        ],
    )
    _write_source_scan_report(
        report_dir / "crypto_framework.json",
        repo_b,
        unique_symbols=[],
        call_sites=[],
        files_with_calls=0,
    )

    from openssl_scanner.source_summary import SourceSummaryBuilder

    builder = SourceSummaryBuilder(
        source_root=str(source_root),
        manifest_path=str(manifest_path),
    )
    rows, blockers = builder.build_from_inputs([str(report_dir)])
    output_xlsx = tmp_path / "summary.xlsx"
    output_tsv = tmp_path / "nonzero.tsv"
    builder.write_workbook(rows, blockers, str(output_xlsx))
    builder.write_nonzero_index(rows, str(output_tsv))

    wb = load_workbook(str(output_xlsx), data_only=True)
    assert wb.sheetnames == ["AllRepos", "OpenSSLUsed", "TopBlockers"]

    all_rows = list(wb["AllRepos"].iter_rows(values_only=True))
    used_rows = list(wb["OpenSSLUsed"].iter_rows(values_only=True))
    blocker_rows = list(wb["TopBlockers"].iter_rows(values_only=True))
    assert all_rows[0][0] == "仓库名称"
    assert any(row[0] == "communication_request" for row in all_rows[1:])
    assert any(row[0] == "communication_request" for row in used_rows[1:])
    assert all(row[0] != "security_crypto_framework" for row in used_rows[1:])
    assert blocker_rows[0][0] == "仓库名称"
    assert blocker_rows[1][2] == "SHA256_Final"

    lines = output_tsv.read_text(encoding="utf-8").strip().splitlines()
    assert len(lines) == 1
    fields = lines[0].split("\t")
    assert fields[0] == "2"
    assert fields[1] == "2"
    assert fields[2] == "communication_request"
    assert fields[6] == "0.0"
    assert fields[7] == "50.0"
    assert fields[8] == "0.0"
    assert fields[9] == "50.0"


def test_source_summary_cli_generates_xlsx_and_tsv(tmp_path):
    source_root = tmp_path / "oh-source"
    repo_path = source_root / "foundation" / "communication" / "request"
    repo_path.mkdir(parents=True)
    (repo_path / "README.md").write_text("request repo\n", encoding="utf-8")
    manifest_path = source_root / ".repo" / "manifest.xml"
    _write_manifest(manifest_path)

    report_dir = tmp_path / "reports"
    report_dir.mkdir()
    _write_source_scan_report(
        report_dir / "request.json",
        repo_path,
        unique_symbols=["SHA256_Update", "SHA256_Final"],
        symbol_coverage={
            "available": 0,
            "partial": 1,
            "not_available": 1,
            "unknown": 0,
        },
        direct_ratio=0.0,
        partial_ratio=50.0,
        call_sites=[
            _sample_call_site("SHA256_Update", "partial", replacement="CRYPT_EAL_Md(...)"),
            _sample_call_site("SHA256_Final", "not_available"),
        ],
    )

    output_xlsx = tmp_path / "summary.xlsx"
    output_tsv = tmp_path / "summary.tsv"

    import subprocess

    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "openssl_scanner",
            "source-summary",
            str(report_dir),
            "-o",
            str(output_xlsx),
            "--nonzero-index",
            str(output_tsv),
            "--source-root",
            str(source_root),
            "--manifest",
            str(manifest_path),
        ],
        capture_output=True,
        text=True,
        timeout=30,
    )

    assert result.returncode == 0, result.stderr
    assert output_xlsx.exists()
    assert output_tsv.exists()
    assert "Summary generated" in result.stdout
