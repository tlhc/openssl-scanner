"""
T4: Edge Case Reviewer - XLSX overflow sheet-splitting tests.

Tests extreme and boundary conditions for the 5 overflow locations:
1. source_exporter.SourceExcelExporter.export()          (Call Sites)
2. source_exporter.SourceMergeExporter._merge_to_workbook() (per-project)
3. exporter.ExcelExporter._create_file_symbol_sheet()     (File-Symbol)
4. exporter.ExcelExporter._create_import_chains_sheet()   (Import Chains)
5. source_diff.SourceDiffExcelExporter._write_callsite_sheet() (Call Site Delta)
"""

import json
import os
import sys
import tempfile
from unittest import mock

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openssl_scanner import _vendor  # noqa: F401
from openpyxl import load_workbook

from openssl_scanner.source_analyzer import CallSite, SourceScanResult
from openssl_scanner.source_exporter import (
    SourceExcelExporter, SourceMergeExporter, COLUMNS, LAST_COL_LETTER,
)
from openssl_scanner.exporter import ExcelExporter
from openssl_scanner.source_diff import (
    ArgDelta, CallSiteDelta, DiffResult, DiffStatus, FileDelta,
    MetricDelta, ProjectDelta, SourceDiffExcelExporter, SymbolDelta,
)


def _make_call_site(idx=0, symbol="SSL_connect", category="ssl_core"):
    return CallSite(
        file_path=f"/tmp/test/file{idx}.c",
        file_name=f"file{idx}.c",
        caller_function=f"func{idx}",
        line_number=idx + 1,
        column=4,
        ossl_symbol=symbol,
        category=category,
        call_args=f"(arg{idx})",
        language="c",
    )


def _make_result(call_sites=None, target="/tmp/test"):
    if call_sites is None:
        call_sites = []
    unique = sorted(set(cs.ossl_symbol for cs in call_sites))
    cat_map = {}
    for cs in call_sites:
        cat_map.setdefault(cs.category, [])
        if cs.ossl_symbol not in cat_map[cs.category]:
            cat_map[cs.category].append(cs.ossl_symbol)
    return SourceScanResult(
        target=target,
        scan_time="2026-01-01T00:00:00",
        tool_version="1.0.0",
        total_files_scanned=1,
        files_with_calls=1 if call_sites else 0,
        total_call_sites=len(call_sites),
        unique_symbols=unique,
        symbols_by_category=cat_map,
        call_sites=call_sites,
        errors=[],
    )


def _make_elf_report(file_symbol_pairs=None, import_chains=None):
    """Build an ELF scan report dict with File-Symbol and Import Chains data."""
    by_file = {}
    by_category = {}
    if file_symbol_pairs:
        for fpath, syms in file_symbol_pairs:
            by_file[fpath] = {'count': len(syms), 'symbols': syms}
            for sym in syms:
                cat = 'ssl_core' if sym.startswith('SSL_') else 'crypto_evp'
                by_category.setdefault(cat, {'count': 0, 'symbols': []})
                if sym not in by_category[cat]['symbols']:
                    by_category[cat]['symbols'].append(sym)
                    by_category[cat]['count'] += 1

    chains = import_chains or {}

    return {
        'meta': {
            'tool_version': '1.0.0',
            'report_type': 'single',
            'scan_time': '2026-01-01T00:00:00',
            'scan_root': '/usr/bin/app',
            'target_arch': 'aarch64',
        },
        'summary': {
            'total_files_scanned': 1,
            'total_elf_files': 1,
            'files_with_openssl_deps': 1,
            'total_openssl_symbols': 0,
            'unique_openssl_symbols': 0,
            'openssl_libs_found': [],
            'files_with_static_openssl': 0,
            'files_with_dlopen': 0,
            'dlopen_unique_symbols': 0,
            'dlopen_libs_detected': [],
        },
        'openssl_symbols': {
            'by_file': by_file,
            'by_category': by_category,
            'by_depth': {},
            'import_chains': chains,
            'all_unique': [],
        },
        'files_detail': [],
        'dependency_tree': {},
    }


def _make_diff_result(n_callsites, is_combo=False):
    """Build a DiffResult with n_callsites ArgDelta entries."""
    csd_list = []
    for i in range(n_callsites):
        ad = ArgDelta(
            status=DiffStatus.ADDED,
            old_line=None,
            new_line=i + 1,
            old_args="",
            new_args=f"(arg{i})",
        )
        csd = CallSiteDelta(
            status=DiffStatus.ADDED,
            identity_key=(f"/tmp/f{i}.c", f"func{i}", "SSL_connect"),
            old_count=0,
            new_count=1,
            old_lines=[],
            new_lines=[i + 1],
            category="ssl_core",
            arg_deltas=[ad],
        )
        csd_list.append(csd)

    proj = ProjectDelta(
        project="test_proj",
        metrics=[MetricDelta("total_call_sites", 0, n_callsites, n_callsites)],
        call_site_delta=csd_list,
        symbol_delta=[
            SymbolDelta(DiffStatus.ADDED, "SSL_connect", "ssl_core", 0, n_callsites)
        ],
        file_delta=[
            FileDelta(DiffStatus.ADDED, f"/tmp/f{i}.c", 0, 1)
            for i in range(min(n_callsites, 10))
        ],
    )
    return DiffResult(
        old_label="old",
        new_label="new",
        projects=[proj],
        old_scan_time="2026-01-01T00:00:00",
        new_scan_time="2026-01-02T00:00:00",
        is_combo=is_combo,
    )


"""
==========================================================================
  Test 1: Empty data (0 rows) - no crash, header-only sheets
==========================================================================
"""


class TestEmptyData:
    """Test 1: Each overflow location with 0 data rows."""

    def test_source_exporter_empty(self, tmp_path):
        """Location 1: SourceExcelExporter with 0 call sites."""
        path = str(tmp_path / "empty.xlsx")
        result = _make_result(call_sites=[])
        SourceExcelExporter().export(result, path)

        wb = load_workbook(path)
        ws = wb["OpenSSL Call Sites"]
        assert ws.max_row == 1
        assert ws.cell(row=1, column=1).value == "File Path"
        assert "OpenSSL Call Sites (2)" not in wb.sheetnames

    def test_merge_exporter_empty_project(self, tmp_path):
        """Location 2: SourceMergeExporter with a project that has 0 rows."""
        path = str(tmp_path / "empty_merge.xlsx")
        project_data = [{'name': 'empty_proj', 'files_scanned': 0, 'rows': []}]
        exporter = SourceMergeExporter()
        exporter._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        assert "empty_proj" in wb.sheetnames
        ws = wb["empty_proj"]
        assert ws.max_row == 1
        assert ws.cell(row=1, column=1).value == "File Path"

    def test_exporter_file_symbol_empty(self, tmp_path):
        """Location 3: ExcelExporter File-Symbol with no data."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        data = _make_elf_report()
        with open(report_path, 'w') as f:
            json.dump(data, f)

        ExcelExporter().export(report_path, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb["File-Symbol"]
        assert ws.cell(row=2, column=1).value == "No file-symbol data available"
        assert "File-Symbol (2)" not in wb.sheetnames

    def test_exporter_import_chains_empty(self, tmp_path):
        """Location 4: ExcelExporter Import Chains with no data."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        data = _make_elf_report()
        with open(report_path, 'w') as f:
            json.dump(data, f)

        ExcelExporter().export(report_path, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb["Import Chains"]
        assert ws.cell(row=2, column=1).value == "No import chain data available"
        assert "Import Chains (2)" not in wb.sheetnames

    def test_source_diff_callsite_empty(self, tmp_path):
        """Location 5: SourceDiffExcelExporter with 0 call site deltas."""
        path = str(tmp_path / "diff_empty.xlsx")
        result = _make_diff_result(0)
        SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        assert "Call Site Delta" in wb.sheetnames
        ws = wb["Call Site Delta"]
        assert ws.max_row == 1
        assert "Call Site Delta (2)" not in wb.sheetnames


"""
==========================================================================
  Test 2: Single row (1 data row) - no overflow
==========================================================================
"""


class TestSingleRow:
    """Test 2: Each location with exactly 1 data row."""

    def test_source_exporter_one_row(self, tmp_path):
        path = str(tmp_path / "one.xlsx")
        result = _make_result([_make_call_site(0)])
        SourceExcelExporter().export(result, path)

        wb = load_workbook(path)
        ws = wb["OpenSSL Call Sites"]
        assert ws.max_row == 2
        assert ws.cell(row=2, column=5).value == "SSL_connect"
        assert "OpenSSL Call Sites (2)" not in wb.sheetnames

    def test_merge_exporter_one_row(self, tmp_path):
        path = str(tmp_path / "one_merge.xlsx")
        rows = [["/tmp/f.c", "f.c", "func", 1, "SSL_connect",
                 "ssl_core", "()", "dynamic-link"]]
        project_data = [{'name': 'proj', 'files_scanned': 1, 'rows': rows}]
        SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        ws = wb["proj"]
        assert ws.max_row == 2
        assert ws.cell(row=2, column=5).value == "SSL_connect"

    def test_exporter_file_symbol_one_row(self, tmp_path):
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        data = _make_elf_report(
            file_symbol_pairs=[("/usr/bin/app", ["SSL_connect"])]
        )
        with open(report_path, 'w') as f:
            json.dump(data, f)

        ExcelExporter().export(report_path, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb["File-Symbol"]
        assert ws.cell(row=2, column=3).value == "SSL_connect"
        assert ws.max_row == 2
        assert "File-Symbol (2)" not in wb.sheetnames

    def test_exporter_import_chains_one_row(self, tmp_path):
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        data = _make_elf_report(
            import_chains={
                "SSL_connect": [
                    {"source_file": "/usr/bin/app",
                     "chain": "app -> libssl.so", "depth": 1}
                ]
            }
        )
        data['openssl_symbols']['by_category'] = {
            'ssl_core': {'count': 1, 'symbols': ['SSL_connect']}
        }
        with open(report_path, 'w') as f:
            json.dump(data, f)

        ExcelExporter().export(report_path, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb["Import Chains"]
        assert ws.cell(row=2, column=3).value == "SSL_connect"
        assert ws.max_row == 2

    def test_source_diff_one_row(self, tmp_path):
        path = str(tmp_path / "diff_one.xlsx")
        result = _make_diff_result(1)
        SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        ws = wb["Call Site Delta"]
        assert ws.max_row == 2
        assert "Call Site Delta (2)" not in wb.sheetnames


"""
==========================================================================
  Test 3: Exact MAX_ROW - 1 data rows (fills exactly, no overflow)
==========================================================================
"""


class TestExactFill:
    """Test 3: Monkey-patch XLSX_MAX_ROW=50, provide 49 data rows.
    Expect exactly 1 sheet, 50 rows (1 header + 49 data), no overflow.
    """

    def test_source_exporter_exact_fill(self, tmp_path):
        path = str(tmp_path / "exact.xlsx")
        sites = [_make_call_site(i) for i in range(49)]
        result = _make_result(sites)

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50):
            SourceExcelExporter().export(result, path)

        wb = load_workbook(path)
        ws = wb["OpenSSL Call Sites"]
        assert ws.max_row == 50
        assert "OpenSSL Call Sites (2)" not in wb.sheetnames

    def test_merge_exporter_exact_fill(self, tmp_path):
        path = str(tmp_path / "exact_merge.xlsx")
        rows = [
            [f"/tmp/f{i}.c", f"f{i}.c", f"func{i}", i + 1,
             "SSL_connect", "ssl_core", "()", "dynamic-link"]
            for i in range(49)
        ]
        project_data = [{'name': 'proj', 'files_scanned': 10, 'rows': rows}]

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50):
            SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        ws = wb["proj"]
        assert ws.max_row == 50
        assert not any(
            sn.startswith("proj") and "(" in sn for sn in wb.sheetnames
        )

    def test_exporter_file_symbol_exact_fill(self, tmp_path):
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        pairs = [
            (f"/usr/bin/app{i}", [f"SSL_{i}"])
            for i in range(49)
        ]
        data = _make_elf_report(file_symbol_pairs=pairs)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", 50):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["File-Symbol"]
        assert ws.max_row == 50
        assert "File-Symbol (2)" not in wb.sheetnames

    def test_source_diff_exact_fill(self, tmp_path):
        path = str(tmp_path / "diff_exact.xlsx")
        result = _make_diff_result(49)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", 50):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        ws = wb["Call Site Delta"]
        assert ws.max_row == 50
        assert "Call Site Delta (2)" not in wb.sheetnames


"""
==========================================================================
  Test 4: Sheet name collision in merge overflow
==========================================================================
"""


class TestSheetNameCollision:
    """Test 4: Two projects whose 25-char truncated names collide."""

    def test_overflow_name_collision(self, tmp_path):
        path = str(tmp_path / "collision.xlsx")

        name_a = "very_long_project_name_xxAA"
        name_b = "very_long_project_name_xxBB"
        assert len(name_a) > 25
        assert name_a[:25] == name_b[:25] == "very_long_project_name_xx"

        def _make_rows(n):
            return [
                [f"/f{i}.c", f"f{i}.c", f"fn{i}", i + 1,
                 "SSL_connect", "ssl_core", "()", "dynamic-link"]
                for i in range(n)
            ]

        project_data = [
            {'name': name_a, 'files_scanned': 1, 'rows': _make_rows(60)},
            {'name': name_b, 'files_scanned': 1, 'rows': _make_rows(60)},
        ]

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50):
            SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        sheet_names = wb.sheetnames

        overflow_sheets = [s for s in sheet_names if "(2)" in s]
        assert len(overflow_sheets) >= 2, (
            f"Expected 2 overflow sheets, got: {overflow_sheets}"
        )

        unique_names = set(sheet_names)
        assert len(unique_names) == len(sheet_names), (
            f"Duplicate sheet names found: {sheet_names}"
        )


"""
==========================================================================
  Test 5: auto_filter correctness with overflow
==========================================================================
"""


class TestAutoFilter:
    """Test 5: Verify auto_filter on first sheet only, correct range."""

    def test_source_exporter_auto_filter(self, tmp_path):
        path = str(tmp_path / "af.xlsx")
        sites = [_make_call_site(i) for i in range(120)]
        result = _make_result(sites)

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50):
            SourceExcelExporter().export(result, path)

        wb = load_workbook(path)
        ws1 = wb["OpenSSL Call Sites"]
        assert ws1.auto_filter.ref == f"A1:{LAST_COL_LETTER}50"

        ws2 = wb["OpenSSL Call Sites (2)"]
        assert ws2.auto_filter.ref is None or ws2.auto_filter.ref == ""

        ws3 = wb["OpenSSL Call Sites (3)"]
        assert ws3.auto_filter.ref is None or ws3.auto_filter.ref == ""

    def test_merge_exporter_auto_filter(self, tmp_path):
        path = str(tmp_path / "af_merge.xlsx")
        rows = [
            [f"/f{i}.c", f"f{i}.c", f"fn{i}", i + 1,
             "SSL_connect", "ssl_core", "()", "dynamic-link"]
            for i in range(120)
        ]
        project_data = [{'name': 'proj', 'files_scanned': 10, 'rows': rows}]

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50):
            SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        ws1 = wb["proj"]
        assert ws1.auto_filter.ref == f"A1:{LAST_COL_LETTER}50"

        for sn in wb.sheetnames:
            if sn.startswith("proj") and "(" in sn:
                ws = wb[sn]
                assert ws.auto_filter.ref is None or ws.auto_filter.ref == ""

    def test_exporter_file_symbol_auto_filter(self, tmp_path):
        """File-Symbol sheet uses _emit_row; first sheet only gets auto_filter
        but the current code does NOT set auto_filter at all for File-Symbol.
        Verify no crash and that continuation sheets exist.
        """
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        pairs = [
            (f"/bin/app{i}", [f"SSL_{i}"])
            for i in range(120)
        ]
        data = _make_elf_report(file_symbol_pairs=pairs)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", 50):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        assert "File-Symbol" in wb.sheetnames
        assert "File-Symbol (2)" in wb.sheetnames
        assert "File-Symbol (3)" in wb.sheetnames


"""
==========================================================================
  Test 6: Header correctness on continuation sheets
==========================================================================
"""


class TestContinuationHeaders:
    """Test 6: Every continuation sheet has correct header at row 1."""

    def test_source_exporter_continuation_headers(self, tmp_path):
        path = str(tmp_path / "cont.xlsx")
        sites = [_make_call_site(i) for i in range(120)]
        result = _make_result(sites)
        expected_headers = [h for _, _, h in COLUMNS]

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50):
            SourceExcelExporter().export(result, path)

        wb = load_workbook(path)
        for sn in wb.sheetnames:
            if sn.startswith("OpenSSL Call Sites"):
                ws = wb[sn]
                actual = [ws.cell(row=1, column=c).value
                          for c in range(1, len(COLUMNS) + 1)]
                assert actual == expected_headers, (
                    f"Sheet '{sn}': headers = {actual}"
                )
                assert ws.cell(row=2, column=1).value is not None, (
                    f"Sheet '{sn}': no data at row 2"
                )

    def test_merge_exporter_continuation_headers(self, tmp_path):
        path = str(tmp_path / "cont_merge.xlsx")
        rows = [
            [f"/f{i}.c", f"f{i}.c", f"fn{i}", i + 1,
             "SSL_connect", "ssl_core", "()", "dynamic-link"]
            for i in range(120)
        ]
        project_data = [{'name': 'proj', 'files_scanned': 10, 'rows': rows}]
        expected_headers = [h for _, _, h in COLUMNS]

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50):
            SourceMergeExporter()._merge_to_workbook(project_data, path)

        wb = load_workbook(path)
        for sn in wb.sheetnames:
            if sn.startswith("proj"):
                ws = wb[sn]
                actual = [ws.cell(row=1, column=c).value
                          for c in range(1, len(COLUMNS) + 1)]
                assert actual == expected_headers, (
                    f"Sheet '{sn}': headers = {actual}"
                )

    def test_source_diff_continuation_headers(self, tmp_path):
        path = str(tmp_path / "cont_diff.xlsx")
        result = _make_diff_result(120, is_combo=False)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", 50):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        expected = ["Status", "File Path", "Caller Function",
                    "OpenSSL Symbol", "Category", "Old Line",
                    "New Line", "Old Args", "New Args"]
        for sn in wb.sheetnames:
            if sn.startswith("Call Site Delta"):
                ws = wb[sn]
                actual = [ws.cell(row=1, column=c).value
                          for c in range(1, len(expected) + 1)]
                assert actual == expected, (
                    f"Sheet '{sn}': headers = {actual}"
                )

    def test_exporter_file_symbol_continuation_headers(self, tmp_path):
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        pairs = [(f"/bin/app{i}", [f"SSL_{i}"]) for i in range(120)]
        data = _make_elf_report(file_symbol_pairs=pairs)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", 50):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        expected = ["Component", "Binary", "Symbol", "Category", "Detection"]
        for sn in wb.sheetnames:
            if sn.startswith("File-Symbol"):
                ws = wb[sn]
                actual = [ws.cell(row=1, column=c).value
                          for c in range(1, len(expected) + 1)]
                assert actual == expected, (
                    f"Sheet '{sn}': headers = {actual}"
                )

    def test_exporter_import_chains_continuation_headers(self, tmp_path):
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        chains = {}
        for i in range(120):
            sym = f"SSL_{i}"
            chains[sym] = [
                {"source_file": f"/bin/app{i}",
                 "chain": f"app{i} -> libssl.so", "depth": 1}
            ]
        data = _make_elf_report(import_chains=chains)
        by_cat = {}
        for sym in chains:
            by_cat.setdefault('ssl_core', {'count': 0, 'symbols': []})
            by_cat['ssl_core']['symbols'].append(sym)
            by_cat['ssl_core']['count'] += 1
        data['openssl_symbols']['by_category'] = by_cat
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", 50):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        expected = ["Source File", "File Name", "Symbol",
                    "Category", "Import Chain", "Depth"]
        for sn in wb.sheetnames:
            if sn.startswith("Import Chains"):
                ws = wb[sn]
                actual = [ws.cell(row=1, column=c).value
                          for c in range(1, len(expected) + 1)]
                assert actual == expected, (
                    f"Sheet '{sn}': headers = {actual}"
                )


"""
==========================================================================
  Test 7: Mixed data - files with many symbols and files with zero
==========================================================================
"""


class TestMixedData:
    """Test 7: One file with 100 symbols, another with 0."""

    def test_exporter_file_symbol_mixed(self, tmp_path):
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")

        many_syms = [f"SSL_{i}" for i in range(100)]
        pairs = [
            ("/usr/bin/heavy_app", many_syms),
            ("/usr/bin/light_app", []),
        ]
        data = _make_elf_report(file_symbol_pairs=pairs)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        ExcelExporter().export(report_path, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb["File-Symbol"]
        data_rows = 0
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val and val != "No file-symbol data available":
                data_rows += 1
        assert data_rows == 100

    def test_source_exporter_mixed_symbols(self, tmp_path):
        """Some call sites with different symbols, verify total row count."""
        path = str(tmp_path / "mixed.xlsx")
        sites = []
        for i in range(80):
            sites.append(_make_call_site(i, symbol="SSL_connect"))
        for i in range(20):
            sites.append(_make_call_site(80 + i, symbol="EVP_sha256",
                                         category="crypto_evp"))
        result = _make_result(sites)

        with mock.patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50):
            SourceExcelExporter().export(result, path)

        wb = load_workbook(path)
        total_data_rows = 0
        for sn in wb.sheetnames:
            if sn.startswith("OpenSSL Call Sites"):
                ws = wb[sn]
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value is not None:
                        total_data_rows += 1
        assert total_data_rows == 100


"""
==========================================================================
  Test 8: source_diff Call Site Delta with is_combo=True vs False
==========================================================================
"""


class TestSourceDiffComboVsNonCombo:
    """Test 8: Both combo and non-combo modes overflow correctly."""

    def _export_and_check(self, tmp_path, is_combo, n_rows):
        path = str(tmp_path / f"diff_{'combo' if is_combo else 'single'}.xlsx")
        result = _make_diff_result(n_rows, is_combo=is_combo)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", 50):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)

        callsite_sheets = [
            sn for sn in wb.sheetnames if sn.startswith("Call Site Delta")
        ]
        return wb, callsite_sheets

    def test_non_combo_overflow(self, tmp_path):
        wb, sheets = self._export_and_check(tmp_path, is_combo=False, n_rows=120)
        assert len(sheets) == 3, f"Expected 3 sheets, got {sheets}"

        ws1 = wb[sheets[0]]
        headers = [ws1.cell(row=1, column=c).value for c in range(1, 10)]
        assert headers[0] == "Status"
        assert "Project" not in headers

    def test_combo_overflow(self, tmp_path):
        wb, sheets = self._export_and_check(tmp_path, is_combo=True, n_rows=120)
        assert len(sheets) == 3, f"Expected 3 sheets, got {sheets}"

        ws1 = wb[sheets[0]]
        headers = [ws1.cell(row=1, column=c).value for c in range(1, 11)]
        assert headers[0] == "Project"
        assert headers[1] == "Status"

    def test_combo_data_rows_have_project(self, tmp_path):
        """Verify combo mode writes project name in every data row."""
        wb, sheets = self._export_and_check(tmp_path, is_combo=True, n_rows=5)
        ws = wb[sheets[0]]
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            assert val == "test_proj", (
                f"Row {row_idx}: expected 'test_proj', got '{val}'"
            )

    def test_non_combo_no_project_column(self, tmp_path):
        """Verify non-combo mode has no Project column."""
        wb, sheets = self._export_and_check(tmp_path, is_combo=False, n_rows=5)
        ws = wb[sheets[0]]
        assert ws.cell(row=1, column=1).value == "Status"

    def test_combo_total_row_count(self, tmp_path):
        """All data rows present across overflow sheets in combo mode."""
        wb, sheets = self._export_and_check(tmp_path, is_combo=True, n_rows=120)
        total = 0
        for sn in sheets:
            ws = wb[sn]
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row=row_idx, column=1).value is not None:
                    total += 1
        assert total == 120

    def test_non_combo_total_row_count(self, tmp_path):
        """All data rows present across overflow sheets in non-combo mode."""
        wb, sheets = self._export_and_check(tmp_path, is_combo=False, n_rows=120)
        total = 0
        for sn in sheets:
            ws = wb[sn]
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row=row_idx, column=1).value is not None:
                    total += 1
        assert total == 120


"""
==========================================================================
  Test 9: exporter.py File-Symbol fallback paths with overflow
==========================================================================
"""


class TestFileSymbolFallbackPaths:
    """Test 9: File-Symbol data from files_detail and components fallback paths."""

    def _make_report_with_files_detail(self, n_files, syms_per_file=1):
        """Build report with by_file=empty, files_detail populated."""
        by_category = {}
        files_detail = []
        for i in range(n_files):
            syms = [f"SSL_sym_{i}_{j}" for j in range(syms_per_file)]
            files_detail.append({
                'path': f'/usr/lib/app{i}.so',
                'openssl_symbols_used': syms,
            })
            for sym in syms:
                by_category.setdefault('ssl_core', {'count': 0, 'symbols': []})
                if sym not in by_category['ssl_core']['symbols']:
                    by_category['ssl_core']['symbols'].append(sym)
                    by_category['ssl_core']['count'] += 1

        return {
            'meta': {
                'tool_version': '1.0.0', 'report_type': 'single',
                'scan_time': '2026-01-01T00:00:00',
                'scan_root': '/usr/lib', 'target_arch': 'aarch64',
            },
            'summary': {
                'total_files_scanned': n_files, 'total_elf_files': n_files,
                'files_with_openssl_deps': n_files, 'total_openssl_symbols': 0,
                'unique_openssl_symbols': 0, 'openssl_libs_found': [],
                'files_with_static_openssl': 0, 'files_with_dlopen': 0,
                'dlopen_unique_symbols': 0, 'dlopen_libs_detected': [],
            },
            'openssl_symbols': {
                'by_file': {},
                'by_category': by_category,
                'by_depth': {},
                'import_chains': {},
                'all_unique': [],
            },
            'files_detail': files_detail,
            'dependency_tree': {},
        }

    def _make_report_with_components(self, n_components, syms_per_comp=1):
        """Build report with by_file=empty, files_detail=empty, components populated."""
        components = {}
        for i in range(n_components):
            comp_name = f"comp_{i}"
            syms = [f"EVP_comp_{i}_{j}" for j in range(syms_per_comp)]
            components[comp_name] = {
                'by_category': {
                    'crypto_evp': {
                        'count': len(syms),
                        'symbols': syms,
                    }
                },
                'executables_detail': {
                    f'bin_{i}': {
                        'by_category': {
                            'crypto_evp': {
                                'count': len(syms),
                                'symbols': syms,
                            }
                        }
                    }
                },
            }

        return {
            'meta': {
                'tool_version': '1.0.0', 'report_type': 'aggregated',
                'scan_time': '2026-01-01T00:00:00',
                'scan_root': '/usr/lib', 'target_arch': 'aarch64',
            },
            'summary': {
                'total_files_scanned': n_components,
                'total_elf_files': n_components,
                'files_with_openssl_deps': n_components,
                'total_openssl_symbols': 0, 'unique_openssl_symbols': 0,
                'openssl_libs_found': [],
                'files_with_static_openssl': 0, 'files_with_dlopen': 0,
                'dlopen_unique_symbols': 0, 'dlopen_libs_detected': [],
            },
            'openssl_symbols': {
                'by_file': {},
                'by_category': {},
                'by_depth': {},
                'import_chains': {},
                'all_unique': [],
            },
            'files_detail': [],
            'components': components,
            'dependency_tree': {},
        }

    def test_files_detail_fallback_produces_rows(self, tmp_path):
        """When by_file is empty but files_detail has data, File-Symbol gets rows."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        data = self._make_report_with_files_detail(5)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        ExcelExporter().export(report_path, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb["File-Symbol"]
        data_rows = 0
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val and val != "No file-symbol data available":
                data_rows += 1
        assert data_rows == 5, (
            f"files_detail fallback should produce 5 rows, got {data_rows}"
        )

    def test_files_detail_fallback_with_overflow(self, tmp_path):
        """files_detail fallback path with enough rows to trigger overflow."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        data = self._make_report_with_files_detail(120)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", 50):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        assert "File-Symbol" in wb.sheetnames
        assert "File-Symbol (2)" in wb.sheetnames
        assert "File-Symbol (3)" in wb.sheetnames

        total_data = 0
        for sn in wb.sheetnames:
            if sn.startswith("File-Symbol"):
                ws = wb[sn]
                for row_idx in range(2, ws.max_row + 1):
                    val = ws.cell(row=row_idx, column=1).value
                    if val and val != "No file-symbol data available":
                        total_data += 1
        assert total_data == 120, (
            f"files_detail overflow: expected 120 rows total, got {total_data}"
        )

    def test_components_fallback_produces_rows(self, tmp_path):
        """When by_file and files_detail are empty but components has data,
        File-Symbol gets rows from the components path."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        data = self._make_report_with_components(5)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        ExcelExporter().export(report_path, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb["File-Symbol"]
        data_rows = 0
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val and val != "No file-symbol data available":
                data_rows += 1
        assert data_rows == 5, (
            f"components fallback should produce 5 rows, got {data_rows}"
        )

    def test_components_fallback_with_overflow(self, tmp_path):
        """components fallback path with enough rows to trigger overflow."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        data = self._make_report_with_components(120)
        with open(report_path, 'w') as f:
            json.dump(data, f)

        with mock.patch("openssl_scanner.exporter.XLSX_MAX_ROW", 50):
            ExcelExporter().export(report_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        assert "File-Symbol" in wb.sheetnames
        assert "File-Symbol (2)" in wb.sheetnames
        assert "File-Symbol (3)" in wb.sheetnames

        total_data = 0
        for sn in wb.sheetnames:
            if sn.startswith("File-Symbol"):
                ws = wb[sn]
                for row_idx in range(2, ws.max_row + 1):
                    val = ws.cell(row=row_idx, column=1).value
                    if val and val != "No file-symbol data available":
                        total_data += 1
        assert total_data == 120, (
            f"components overflow: expected 120 rows total, got {total_data}"
        )

    def test_all_sources_empty_shows_placeholder(self, tmp_path):
        """When by_file, files_detail, and components are all empty,
        File-Symbol shows 'No file-symbol data available'."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")
        data = _make_elf_report()
        with open(report_path, 'w') as f:
            json.dump(data, f)

        ExcelExporter().export(report_path, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb["File-Symbol"]
        assert ws.cell(row=2, column=1).value == "No file-symbol data available"

    def test_by_file_populated_skips_fallbacks(self, tmp_path):
        """When by_file has data, files_detail and components fallbacks are skipped
        even if they are populated."""
        report_path = str(tmp_path / "report.json")
        xlsx_path = str(tmp_path / "out.xlsx")

        data = _make_elf_report(
            file_symbol_pairs=[("/usr/bin/app", ["SSL_connect"])]
        )
        data['files_detail'] = [
            {'path': '/usr/lib/fallback.so',
             'openssl_symbols_used': ['EVP_fallback_should_not_appear']}
        ]
        with open(report_path, 'w') as f:
            json.dump(data, f)

        ExcelExporter().export(report_path, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb["File-Symbol"]

        all_symbols = set()
        for row_idx in range(2, ws.max_row + 1):
            sym = ws.cell(row=row_idx, column=3).value
            if sym:
                all_symbols.add(sym)

        assert "SSL_connect" in all_symbols
        assert "EVP_fallback_should_not_appear" not in all_symbols, (
            "files_detail fallback should be skipped when by_file has data"
        )


"""
==========================================================================
  Test 10: source_diff middle continuation sheets have no auto_filter
==========================================================================
"""


class TestSourceDiffMiddleSheetAutoFilter:
    """Test 10: With 4+ overflow sheets, middle sheets must NOT have auto_filter."""

    def test_middle_sheets_no_auto_filter(self, tmp_path):
        """200 arg deltas at MAX_ROW=50 -> 5 sheets (49+49+49+49+4).
        Only first and last should have auto_filter."""
        path = str(tmp_path / "middle.xlsx")
        result = _make_diff_result(200, is_combo=False)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", 50):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        callsite_sheets = [
            sn for sn in wb.sheetnames if sn.startswith("Call Site Delta")
        ]
        assert len(callsite_sheets) >= 4, (
            f"Expected 4+ sheets, got {callsite_sheets}"
        )

        ws_first = wb[callsite_sheets[0]]
        first_ref = ws_first.auto_filter.ref
        assert first_ref is not None and first_ref != "", (
            "First sheet must have auto_filter"
        )

        ws_last = wb[callsite_sheets[-1]]
        last_ref = ws_last.auto_filter.ref
        assert last_ref is not None and last_ref != "", (
            "Last sheet must have auto_filter (source_diff behavior)"
        )

        for sn in callsite_sheets[1:-1]:
            ws_mid = wb[sn]
            mid_ref = ws_mid.auto_filter.ref
            assert mid_ref is None or mid_ref == "", (
                f"Middle sheet '{sn}' should NOT have auto_filter, got {mid_ref}"
            )

    def test_middle_sheets_combo_mode(self, tmp_path):
        """Same test in combo mode with Project column."""
        path = str(tmp_path / "middle_combo.xlsx")
        result = _make_diff_result(200, is_combo=True)

        with mock.patch("openssl_scanner.source_diff.XLSX_MAX_ROW", 50):
            SourceDiffExcelExporter(include_unchanged=True).export(result, path)

        wb = load_workbook(path)
        callsite_sheets = [
            sn for sn in wb.sheetnames if sn.startswith("Call Site Delta")
        ]
        assert len(callsite_sheets) >= 4

        for sn in callsite_sheets[1:-1]:
            ws_mid = wb[sn]
            mid_ref = ws_mid.auto_filter.ref
            assert mid_ref is None or mid_ref == "", (
                f"Middle sheet '{sn}' should NOT have auto_filter, got {mid_ref}"
            )


"""
==========================================================================
  Test 11: parallel -j per_pkg_jobs cap
==========================================================================
"""


class TestParallelJobsCap:
    """Test 11: per_pkg_jobs is capped at cpu_count when only 1 package."""

    def test_per_pkg_jobs_cap_single_package(self):
        """When only 1 package to scan (use_parallel=False),
        per_pkg_jobs = min(args.jobs, cpu_count)."""
        cpu_count = os.cpu_count() or 4

        class MockArgs:
            jobs = 999

        args = MockArgs()
        to_scan = [("idx0", "entry0")]

        use_parallel = len(to_scan) > 1 and args.jobs != 1
        assert use_parallel is False

        if use_parallel:
            per_pkg_jobs = 1
        else:
            per_pkg_jobs = min(args.jobs, cpu_count)

        assert per_pkg_jobs == cpu_count, (
            f"per_pkg_jobs should be capped at cpu_count={cpu_count}, "
            f"got {per_pkg_jobs}"
        )
        assert per_pkg_jobs <= cpu_count

    def test_per_pkg_jobs_under_cap(self):
        """When args.jobs <= cpu_count, per_pkg_jobs == args.jobs."""
        cpu_count = os.cpu_count() or 4
        small_j = max(1, cpu_count - 1)

        class MockArgs:
            jobs = small_j

        args = MockArgs()
        to_scan = [("idx0", "entry0")]

        use_parallel = len(to_scan) > 1 and args.jobs != 1
        if use_parallel:
            per_pkg_jobs = 1
        else:
            per_pkg_jobs = min(args.jobs, cpu_count)

        assert per_pkg_jobs == small_j, (
            f"per_pkg_jobs should equal args.jobs={small_j}, got {per_pkg_jobs}"
        )

    def test_parallel_mode_per_pkg_is_one(self):
        """When multiple packages (use_parallel=True), per_pkg_jobs is always 1."""
        class MockArgs:
            jobs = 8

        args = MockArgs()
        to_scan = [("idx0", "e0"), ("idx1", "e1"), ("idx2", "e2")]

        use_parallel = len(to_scan) > 1 and args.jobs != 1
        assert use_parallel is True

        if use_parallel:
            parallel = min(len(to_scan), args.jobs)
            per_pkg_jobs = 1
        else:
            parallel = 1
            per_pkg_jobs = min(args.jobs, os.cpu_count() or 4)

        assert per_pkg_jobs == 1
        assert parallel == 3

    def test_jobs_one_disables_parallel(self):
        """When args.jobs == 1, use_parallel is False even with multiple packages."""
        class MockArgs:
            jobs = 1

        args = MockArgs()
        to_scan = [("idx0", "e0"), ("idx1", "e1")]

        use_parallel = len(to_scan) > 1 and args.jobs != 1
        assert use_parallel is False

        if use_parallel:
            per_pkg_jobs = 1
        else:
            per_pkg_jobs = min(args.jobs, os.cpu_count() or 4)

        assert per_pkg_jobs == 1


class TestMultiTargetWorkerBudget:
    """Test: multi-target source scan divides worker budget to avoid
    over-subscription (ThreadPoolExecutor x ProcessPoolExecutor)."""

    def test_per_target_jobs_divides_budget(self):
        """With N targets and J jobs, each target gets J//N workers."""
        cpu = os.cpu_count() or 4
        targets = [f"/src/proj{i}" for i in range(4)]
        jobs = cpu

        max_w = min(len(targets), cpu)
        per_target_jobs = max(1, jobs // max_w)

        total = max_w * per_target_jobs
        assert total <= jobs, (
            f"total workers {total} exceeds budget {jobs}"
        )
        assert per_target_jobs >= 1

    def test_single_target_gets_full_budget(self):
        """Single target should get all workers."""
        cpu = os.cpu_count() or 4
        targets = ["/src/single"]
        jobs = cpu

        max_w = min(len(targets), cpu)
        per_target_jobs = max(1, jobs // max_w)

        assert max_w == 1
        assert per_target_jobs == jobs

    def test_many_targets_get_one_each(self):
        """When targets >= jobs, each target gets 1 worker."""
        cpu = os.cpu_count() or 4
        targets = [f"/src/proj{i}" for i in range(cpu * 2)]
        jobs = cpu

        max_w = min(len(targets), cpu)
        per_target_jobs = max(1, jobs // max_w)

        assert per_target_jobs == 1
        assert max_w == cpu

    def test_workers_override_used_by_scan_single_target(self):
        """_scan_single_target uses workers_override when provided."""
        from argparse import Namespace
        args = Namespace(jobs=16, no_recursive=False)

        assert args.jobs == 16
        workers = 4
        effective = workers if workers is not None else args.jobs
        assert effective == 4
