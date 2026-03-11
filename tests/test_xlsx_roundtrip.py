"""Tests for _read_xlsx overflow sheet round-trip correctness.

Verifies that SourceMergeExporter._read_xlsx reads ALL overflow continuation
sheets and that no data is lost in the write -> read -> merge pipeline.
"""

import os
import re
import tempfile
from unittest.mock import patch

import pytest

tree_sitter = pytest.importorskip("tree_sitter")

from openssl_scanner.source_analyzer import CallSite, SourceScanResult
from openssl_scanner.source_exporter import (
    COLUMNS,
    SourceExcelExporter,
    SourceMergeExporter,
    XLSX_MAX_ROW,
)


def _make_call_site(index, symbol=None, category=None, file_path=None,
                    caller=None, line=None, args=None):
    """Create a CallSite with deterministic, distinct values per index."""
    return CallSite(
        file_path=file_path or f"/tmp/src/file_{index:04d}.c",
        file_name=os.path.basename(file_path) if file_path else f"file_{index:04d}.c",
        caller_function=caller or f"func_{index:04d}",
        line_number=line if line is not None else (index + 1),
        column=4,
        ossl_symbol=symbol or f"SSL_sym_{index:04d}",
        category=category or "ssl_core",
        call_args=args or f"(arg_{index})",
        language="c",
    )


def _make_result(call_sites, target="/tmp/test_proj"):
    unique = sorted(set(cs.ossl_symbol for cs in call_sites))
    cats = {}
    for cs in call_sites:
        cats.setdefault(cs.category, set()).add(cs.ossl_symbol)
    symbols_by_category = {k: sorted(v) for k, v in cats.items()}
    return SourceScanResult(
        target=target,
        scan_time="2026-03-11T00:00:00",
        tool_version="1.0.0",
        total_files_scanned=max(1, len(call_sites)),
        files_with_calls=len(set(cs.file_path for cs in call_sites)) if call_sites else 0,
        total_call_sites=len(call_sites),
        unique_symbols=unique,
        symbols_by_category=symbols_by_category,
        call_sites=call_sites,
        errors=[],
    )


class TestReadXlsxOverflow:
    """Test 1: _read_xlsx reads all overflow sheets."""

    @patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_reads_all_overflow_sheets(self, tmp_path):
        """120 call sites at MAX_ROW=50 -> 3 sheets (49+49+22 data rows).
        _read_xlsx must return all 120 rows."""
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        sites = [_make_call_site(i) for i in range(120)]
        result = _make_result(sites)

        xlsx_path = str(tmp_path / "overflow.xlsx")
        SourceExcelExporter().export(result, xlsx_path)

        wb = load_workbook(xlsx_path, read_only=True)
        call_sheets = [s for s in wb.sheetnames
                       if s == "OpenSSL Call Sites"
                       or s.startswith("OpenSSL Call Sites (")]
        assert len(call_sheets) == 3, f"Expected 3 sheets, got {call_sheets}"

        sheet_row_counts = []
        for sn in call_sheets:
            ws = wb[sn]
            data_rows = [r for i, r in enumerate(ws.iter_rows(values_only=True))
                         if i > 0 and any(v is not None for v in r)]
            sheet_row_counts.append(len(data_rows))
        wb.close()

        assert sheet_row_counts == [49, 49, 22], \
            f"Sheet row distribution wrong: {sheet_row_counts}"

        merger = SourceMergeExporter()
        rows = merger._read_xlsx(xlsx_path)
        _name, _fs, read_rows = rows

        assert len(read_rows) == 120, \
            f"Expected 120 rows from _read_xlsx, got {len(read_rows)}"

        assert read_rows[0][4] == sites[0].ossl_symbol, \
            f"First row symbol mismatch: {read_rows[0][4]} != {sites[0].ossl_symbol}"
        assert read_rows[-1][4] == sites[-1].ossl_symbol, \
            f"Last row symbol mismatch: {read_rows[-1][4]} != {sites[-1].ossl_symbol}"


class TestReadXlsxSingleSheet:
    """Test 2: _read_xlsx with NO overflow (backward compat)."""

    @patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_single_sheet_no_overflow(self, tmp_path):
        """30 call sites at MAX_ROW=50 -> 1 sheet. _read_xlsx returns 30."""
        sites = [_make_call_site(i) for i in range(30)]
        result = _make_result(sites)

        xlsx_path = str(tmp_path / "single.xlsx")
        SourceExcelExporter().export(result, xlsx_path)

        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True)
        call_sheets = [s for s in wb.sheetnames
                       if s == "OpenSSL Call Sites"
                       or s.startswith("OpenSSL Call Sites (")]
        assert len(call_sheets) == 1
        wb.close()

        merger = SourceMergeExporter()
        _name, _fs, read_rows = merger._read_xlsx(xlsx_path)
        assert len(read_rows) == 30


class TestFullMergePipelineWithOverflow:
    """Test 3: Full merge pipeline with overflow -> verify totals."""

    @patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_merge_two_projects_with_overflow(self, tmp_path):
        """Project A: 80 sites (2 sheets), Project B: 40 sites (1 sheet).
        Merged workbook must have correct Summary counts."""
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import load_workbook

        sites_a = [_make_call_site(i, symbol=f"SSL_a_{i:04d}") for i in range(80)]
        sites_b = [_make_call_site(i, symbol=f"EVP_b_{i:04d}", category="crypto_evp")
                   for i in range(40)]

        path_a = str(tmp_path / "projA.xlsx")
        path_b = str(tmp_path / "projB.xlsx")
        SourceExcelExporter().export(_make_result(sites_a), path_a)
        SourceExcelExporter().export(_make_result(sites_b), path_b)

        wb_a = load_workbook(path_a, read_only=True)
        a_call_sheets = [s for s in wb_a.sheetnames
                         if s == "OpenSSL Call Sites"
                         or s.startswith("OpenSSL Call Sites (")]
        assert len(a_call_sheets) == 2, f"Project A should have 2 call sheets, got {a_call_sheets}"
        wb_a.close()

        merged_path = str(tmp_path / "merged.xlsx")
        merger = SourceMergeExporter()
        stats = merger.merge([path_a, path_b], merged_path)

        assert stats['sheets'][0]['call_sites'] == 80, \
            f"Project A call_sites: expected 80, got {stats['sheets'][0]['call_sites']}"
        assert stats['sheets'][1]['call_sites'] == 40, \
            f"Project B call_sites: expected 40, got {stats['sheets'][1]['call_sites']}"

        wb = load_workbook(merged_path, read_only=True)
        ws_sum = wb["Summary"]
        summary_rows = list(ws_sum.iter_rows(values_only=True))
        total_row = summary_rows[-1]
        assert total_row[0] == "TOTAL"
        assert total_row[3] == 120, \
            f"TOTAL call sites: expected 120, got {total_row[3]}"

        ws_sym = wb["Symbol Summary"]
        sym_rows = list(ws_sym.iter_rows(values_only=True))
        unique_syms = {r[0] for r in sym_rows[1:] if r[0] is not None}
        assert len(unique_syms) == 120, \
            f"Symbol Summary unique symbols: expected 120, got {len(unique_syms)}"
        assert stats['total_symbols'] == 120
        wb.close()


class TestDataIntegrityRoundTrip:
    """Test 4: Cell values survive export -> _read_xlsx round-trip."""

    @patch("openssl_scanner.source_exporter.XLSX_MAX_ROW", 50)
    def test_all_columns_preserved(self, tmp_path):
        """60 call sites with distinct values: export -> _read_xlsx -> compare."""
        sites = [
            _make_call_site(
                i,
                symbol=f"API_{i:04d}",
                category="ssl_core" if i % 2 == 0 else "crypto_evp",
                file_path=f"/proj/src/mod_{i:04d}.c",
                caller=f"caller_{i:04d}",
                line=100 + i,
                args=f"(x_{i}, y_{i})",
            )
            for i in range(60)
        ]
        result = _make_result(sites)

        xlsx_path = str(tmp_path / "integrity.xlsx")
        SourceExcelExporter().export(result, xlsx_path)

        merger = SourceMergeExporter()
        _name, _fs, read_rows = merger._read_xlsx(xlsx_path)

        assert len(read_rows) == 60, f"Expected 60 rows, got {len(read_rows)}"

        for i, (site, row) in enumerate(zip(sites, read_rows)):
            assert row[0] == site.file_path, \
                f"Row {i} file_path: {row[0]} != {site.file_path}"
            assert row[1] == site.file_name, \
                f"Row {i} file_name: {row[1]} != {site.file_name}"
            assert row[2] == site.caller_function, \
                f"Row {i} caller: {row[2]} != {site.caller_function}"
            assert row[3] == site.line_number, \
                f"Row {i} line: {row[3]} != {site.line_number}"
            assert row[4] == site.ossl_symbol, \
                f"Row {i} symbol: {row[4]} != {site.ossl_symbol}"
            assert row[5] == site.category, \
                f"Row {i} category: {row[5]} != {site.category}"
            assert row[6] == site.call_args, \
                f"Row {i} args: {row[6]} != {site.call_args}"


class TestReadXlsxIgnoresUnrelatedSheets:
    """Test 5: _read_xlsx only reads Call Sites sheets, not others."""

    def test_ignores_symbol_summary_and_random(self, tmp_path):
        """XLSX with 'OpenSSL Call Sites', 'OpenSSL Call Sites (2)',
        'Symbol Summary', 'Random Sheet'. Only 2 call site sheets read."""
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "OpenSSL Call Sites"
        for col_idx, (_, _, title) in enumerate(COLUMNS, 1):
            ws1.cell(row=1, column=col_idx, value=title)
        for i in range(10):
            ws1.cell(row=i + 2, column=1, value=f"/path/file_{i}.c")
            ws1.cell(row=i + 2, column=2, value=f"file_{i}.c")
            ws1.cell(row=i + 2, column=3, value=f"func_{i}")
            ws1.cell(row=i + 2, column=4, value=i + 1)
            ws1.cell(row=i + 2, column=5, value=f"SSL_sym_{i}")
            ws1.cell(row=i + 2, column=6, value="ssl_core")
            ws1.cell(row=i + 2, column=7, value=f"(arg_{i})")
            ws1.cell(row=i + 2, column=8, value="dynamic-link")

        ws2 = wb.create_sheet("OpenSSL Call Sites (2)")
        for col_idx, (_, _, title) in enumerate(COLUMNS, 1):
            ws2.cell(row=1, column=col_idx, value=title)
        for i in range(5):
            idx = i + 10
            ws2.cell(row=i + 2, column=1, value=f"/path/file_{idx}.c")
            ws2.cell(row=i + 2, column=2, value=f"file_{idx}.c")
            ws2.cell(row=i + 2, column=3, value=f"func_{idx}")
            ws2.cell(row=i + 2, column=4, value=idx + 1)
            ws2.cell(row=i + 2, column=5, value=f"SSL_sym_{idx}")
            ws2.cell(row=i + 2, column=6, value="ssl_core")
            ws2.cell(row=i + 2, column=7, value=f"(arg_{idx})")
            ws2.cell(row=i + 2, column=8, value="dynamic-link")

        ws_sym = wb.create_sheet("Symbol Summary")
        ws_sym.cell(row=1, column=1, value="OpenSSL Symbol")
        ws_sym.cell(row=2, column=1, value="SHOULD_NOT_APPEAR")

        ws_rnd = wb.create_sheet("Random Sheet")
        ws_rnd.cell(row=1, column=1, value="Noise")
        ws_rnd.cell(row=2, column=1, value="GARBAGE_DATA")

        xlsx_path = str(tmp_path / "mixed_sheets.xlsx")
        wb.save(xlsx_path)
        wb.close()

        merger = SourceMergeExporter()
        _name, _fs, read_rows = merger._read_xlsx(xlsx_path)

        assert len(read_rows) == 15, \
            f"Expected 15 rows (10+5), got {len(read_rows)}"

        all_symbols = [r[4] for r in read_rows]
        assert "SHOULD_NOT_APPEAR" not in all_symbols
        assert "GARBAGE_DATA" not in [r[0] for r in read_rows]


class TestNamingPatternCorrectness:
    """Test 6: _read_xlsx pattern only matches exact names."""

    def test_does_not_match_similar_names(self, tmp_path):
        """Sheets 'OpenSSL Call Sites Summary' or 'OpenSSL Call Sites_backup'
        must NOT be read. Only exact 'OpenSSL Call Sites' and
        'OpenSSL Call Sites (N)' should match."""
        from openssl_scanner import _vendor  # noqa: F401
        from openpyxl import Workbook

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "OpenSSL Call Sites"
        for col_idx, (_, _, title) in enumerate(COLUMNS, 1):
            ws1.cell(row=1, column=col_idx, value=title)
        ws1.cell(row=2, column=1, value="/path/real.c")
        ws1.cell(row=2, column=2, value="real.c")
        ws1.cell(row=2, column=3, value="real_func")
        ws1.cell(row=2, column=4, value=1)
        ws1.cell(row=2, column=5, value="SSL_connect")
        ws1.cell(row=2, column=6, value="ssl_core")
        ws1.cell(row=2, column=7, value="(ctx)")
        ws1.cell(row=2, column=8, value="dynamic-link")

        ws_bad1 = wb.create_sheet("OpenSSL Call Sites Summary")
        ws_bad1.cell(row=1, column=1, value="Header")
        ws_bad1.cell(row=2, column=1, value="BAD1")
        ws_bad1.cell(row=2, column=5, value="FAKE_SYMBOL_1")

        ws_bad2 = wb.create_sheet("OpenSSL Call Sites_backup")
        ws_bad2.cell(row=1, column=1, value="Header")
        ws_bad2.cell(row=2, column=1, value="BAD2")
        ws_bad2.cell(row=2, column=5, value="FAKE_SYMBOL_2")

        ws_bad3 = wb.create_sheet("OpenSSL Call SitesX")
        ws_bad3.cell(row=1, column=1, value="Header")
        ws_bad3.cell(row=2, column=1, value="BAD3")
        ws_bad3.cell(row=2, column=5, value="FAKE_SYMBOL_3")

        ws_good = wb.create_sheet("OpenSSL Call Sites (2)")
        for col_idx, (_, _, title) in enumerate(COLUMNS, 1):
            ws_good.cell(row=1, column=col_idx, value=title)
        ws_good.cell(row=2, column=1, value="/path/real2.c")
        ws_good.cell(row=2, column=2, value="real2.c")
        ws_good.cell(row=2, column=3, value="real_func2")
        ws_good.cell(row=2, column=4, value=2)
        ws_good.cell(row=2, column=5, value="SSL_read")
        ws_good.cell(row=2, column=6, value="ssl_core")
        ws_good.cell(row=2, column=7, value="(buf, len)")
        ws_good.cell(row=2, column=8, value="dynamic-link")

        xlsx_path = str(tmp_path / "tricky_names.xlsx")
        wb.save(xlsx_path)
        wb.close()

        merger = SourceMergeExporter()
        _name, _fs, read_rows = merger._read_xlsx(xlsx_path)

        assert len(read_rows) == 2, \
            f"Expected 2 rows (from 2 valid sheets), got {len(read_rows)}"

        symbols = {r[4] for r in read_rows}
        assert symbols == {"SSL_connect", "SSL_read"}, \
            f"Unexpected symbols: {symbols}"

        for r in read_rows:
            assert "FAKE_SYMBOL" not in str(r), \
                f"Leaked data from invalid sheet: {r}"

    def test_pattern_regex_validation(self):
        """Direct validation of the naming pattern logic used in _read_xlsx."""
        base_title = "OpenSSL Call Sites"

        def matches(sheet_name):
            return sheet_name == base_title or sheet_name.startswith(base_title + " (")

        assert matches("OpenSSL Call Sites") is True
        assert matches("OpenSSL Call Sites (2)") is True
        assert matches("OpenSSL Call Sites (10)") is True

        assert matches("OpenSSL Call Sites Summary") is False
        assert matches("OpenSSL Call Sites_backup") is False
        assert matches("OpenSSL Call SitesX") is False
        assert matches("OpenSSL Call Sites2") is False
        assert matches("xOpenSSL Call Sites") is False
        assert matches("Symbol Summary") is False
        assert matches("Random Sheet") is False
